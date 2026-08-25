# -*- coding: utf-8 -*-
# ============================================================================
# PROCESADOR CUSTOM ANALYTICS — el eje de la DEMANDA (Fase 0, MULTIPAÍS)
# ----------------------------------------------------------------------------
# Qué hace:
#   Lee el export "Custom Analytics" del Seller (Analytics → panel dimensión ASIN
#   → UN SOLO marketplace → Descargar .xlsx) del buzón informes/custom_analytics/ y
#   carga en `demanda_asin` la demanda por ASIN: visitas, sesiones, conversión,
#   unidades pedidas/enviadas, ventas, reembolsos y —lo que ningún otro informe
#   trae— LA BUY BOX (ratio de oferta destacada). Hasta hoy el sistema sabía qué
#   stock hay y qué se vendió; NO sabía cuánta gente llegó ni quién se llevó la caja.
#
# 🔴 EL CAJÓN: PELÍCULA DE LECTURAS (cambió el 10-ago-2026) — §1.6 de CLAUDE.md
#   🔴 ESTO ERA "FOTO POR VENTANA" Y DEJÓ DE SERLO. La migración
#   `2026-08-07_demanda_asin_contador.sql` demostró que el informe NO cubre una ventana:
#   es un CONTADOR ACUMULADO desde un punto de partida fijo y desconocido. Medido el
#   7-ago-2026: 1.605 comparaciones ASIN×métrica entre las lecturas del 30-jul y del
#   7-ago, CERO bajadas. Un acumulado que nunca baja no es "lo que pasó del día X al Y":
#   la ventana era una etiqueta inventada por quien subía el fichero.
#
#   Así que cada carga apila UNA LECTURA del contador, fechada con `leido_at`:
#     1) el país lo da el SELECTOR; la FECHA la da el FICHERO (no se declara),
#     2) en UNA transacción:
#          DELETE FROM demanda_asin
#           WHERE pais=<sel> AND leido_at=<del fichero>;      -- IGUALDAD
#          INSERT de todas las filas;
#     3) commit si aplicar, rollback si ensayo.
#   El DELETE por IGUALDAD no es "borrar el histórico": recierra LA MISMA lectura si se
#   recarga (idempotente) y deja intactas las demás lecturas y países. Es lo que dice el
#   COMMENT de la tabla. Con BETWEEN se cargaría el mundo por delante.
#
#   🔴 Y LAS CIFRAS DE UN PERIODO NO SE LEEN DE UNA FILA: salen de RESTAR dos lecturas.
#   Eso vive en `v_demanda_asin_ultima`, no aquí. El cargador no interpreta.
#
# 🔴 EL PAÍS LO MANDA EL SELECTOR; LA FECHA, EL FICHERO (§3.5, §6.6)
#   El fichero no dice de qué marketplace es (cabeceras en español los tres, URLs a
#   amazon.com los tres). El país entra por el input PAIS y se VERIFICA cruzando las
#   unidades por ASIN con transacciones_movimientos (guarda 6.6): si el declarado no es
#   el de menor error, ABORTA.
#   La fecha del dato NO se declara: es `wb.properties.created` (cuándo lo generó
#   Amazon) → `leido_at`. Sin ella se ABORTA (guarda 6.5): es el eje del dato y no se
#   puede inventar.
#
# 🔴 Y EN ENSAYO SE INVENTARÍA EL BUZÓN ANTES DE NADA (§3): se abre CADA .xlsx y se dice
#   qué `leido_at` le tocaría Y qué huella tienen sus DATOS. Nació el 10-ago-2026 de un
#   error concreto: dar por idénticos dos ficheros porque pesaban lo mismo AL BYTE. Los
#   eTag de Storage decían que no. Un .xlsx es un ZIP y la fecha es una cadena de
#   longitud fija: mismo peso y distinto contenido es exactamente lo que pasa aquí.
#   DEL TAMAÑO NO SE DEDUCE NADA.
#   Y se compara por los DOS lados: la fecha repetida es el duplicado inofensivo (se
#   recierra solo); el peligroso es la fecha DISTINTA con datos IDÉNTICOS, que entra
#   como dos lecturas y mete un tramo de cero movimiento en la serie.
#
# 🔒 EL CARGADOR NO INTERPRETA (§3.2)
#   Los tres RATIOS (conversión, ratio de oferta destacada = buy box, ratio de
#   reembolsos) vienen 0-1 AUNQUE la cabecera de Amazon diga "(%)". Se guardan TAL
#   CUAL. Nada de multiplicar por 100 al cargar: quien lo pinte decide.
#
# 🔴 CÓMO SE ABRE EL .xlsx (§0.1 — el fallo grave de la v1 de este encargo)
#   Amazon escribe el ASIN como fórmula `=HYPERLINK("…/dp/B0…","B0…")` SIN valor
#   cacheado. `data_only=True` pide el valor cacheado y devuelve None → TODOS los ASIN
#   saldrían vacíos. Se abre con data_only=False (por defecto) y el ASIN se saca de la
#   fórmula con regex. Medido byte a byte contra los 3 ficheros reales el 31-jul.
#
# 🔒 Escritura por lotes (execute_values), calcado del patrón de la casa. Ni comisión
#   ni ratios se tocan. Encoding: no aplica (openpyxl lee el .xlsx binario).
#
# 🔒 SELLO DE FRESCURA: al aplicar escribe una fila en informes_subidos
#   (tipo='custom_analytics') con los 10 totales del cuadre en resumen_json — donde
#   vive lo que no es una fila. Ojo: la RPC frescura_informes() lee la fecha del dato
#   de `max(demanda_asin.leido_at)::date`, NO de aquí (el sello es registro/auditoría).
#   `fecha_dato_desde` y `fecha_dato_hasta` de informes_subidos van las DOS a la fecha
#   de la lectura: una lectura es un INSTANTE, no un rango, y poner un rango inventado
#   ahí sería volver a contar el cuento de la ventana por la puerta de atrás.
# ============================================================================

import os, sys, io, re, hashlib, unicodedata
from datetime import date, datetime, timezone
from collections import Counter, defaultdict
from statistics import median

import psycopg2
from psycopg2.extras import Json, execute_values
import openpyxl

# Del patrón común solo se reutiliza lo de FOTO que aplica: Aborta y las lecturas de
# Storage con reintento. NO barrer_sobrantes/archivar_foto (esto es FOTO POR VENTANA,
# no FOTO: el borrado es por ventana exacta, no "lo que no viene en el fichero").
from foto_comun import Aborta, conectar_bd, listar_buzon, descargar_buzon, refrescar_vistas

# ---------------------------------------------------------------------------
# 0) Configuración (secrets de GitHub; jamás credenciales en el código)
# ---------------------------------------------------------------------------
SUPABASE_URL  = os.environ.get('SUPABASE_URL', 'https://ogfbjjdxcltzpygzuyla.supabase.co')
SUPABASE_KEY  = os.environ.get('SUPABASE_KEY', '')   # llave de servicio: LEER el Storage cerrado
DB_URL        = os.environ.get('DB_URL', '')         # postgres del ENTORNO (staging o prod)
MODO          = os.environ.get('MODO', 'ensayo').strip().lower()       # ensayo | aplicar
ENTORNO       = os.environ.get('ENTORNO', 'staging').strip().lower()   # staging | produccion
PAIS          = os.environ.get('PAIS', '').strip().upper()             # ES | IT | FR (selector)
FICHERO       = os.environ.get('FICHERO', '').strip()                  # nombre EXACTO; vacío = más reciente
# 🔴 FORZAR: la salida de la ZONA GRIS de la guarda 6.14, y NADA MÁS. Zona gris = la
#   comparación contra la lectura anterior no puede probar nada (quedan cuatro ASIN
#   comunes, o la referencia es de hace meses). Entonces la guarda no dice "esto está
#   mal": dice "no lo sé", aborta, y pide que un humano lo mire y relance.
# 🔒 LO QUE **NO** LEVANTA, y por eso es seguro que exista: los cuatro criterios de
#   aborto duro (negativos, ≥5% de bajadas, un desplome de más de la mitad, las nueve
#   métricas a la vez). Esos son EVIDENCIA de que el fichero es de otra cosa, y un
#   interruptor que apague la evidencia convierte la guarda en una sugerencia.
#   Un `forzar` que valga para todo es un `ON_ERROR_STOP=0` con otro nombre (§3).
# Vacío = 'no'. Lo manda el .yml; la app de Elena no lo manda y no le hace falta:
#   `workflow_dispatch` aplica el default del propio fichero.
FORZAR        = os.environ.get('FORZAR', 'no').strip().lower() in ('si', 'sí', 'yes', 'true', '1')
# (Aquí se leían PERIODO_DESDE y PERIODO_HASTA, obsoletos desde el modelo contador y
#  borrados el 10-ago-2026 al cerrarse la secuencia de tres pasos: primero el procesador
#  dejó de usarlos y lo dijo en el log, luego la ficha de moloka-app-v2 dejó de mandarlos,
#  y solo entonces desaparecen de aquí y del .yml. El orden no era una preferencia: un
#  `workflow_dispatch` rechaza el disparo ENTERO si le llega un input que el .yml no
#  declara, así que hacerlo al revés le habría dado un 422 al botón de Elena.)

BUCKET, CARPETA = 'informes', 'custom_analytics'
# 🔒 Escalabilidad (§8): la lista de países vive en UN solo sitio por lado. Añadir DE
# o PL es tocar esto + el choice del .yml + las opciones de la ficha v2. Nada más.
PAISES_VALIDOS = ('ES', 'IT', 'FR')

# Guarda 6.6: días mínimos de transacciones con los que fiarse del cruce de país.
# ⚠️ NÚMERO NUEVO, Y HAY QUE DECIRLO. El modelo de ventana pedía que la intersección
#   cubriera el 40% de la ventana DECLARADA; sin ventana declarada ese porcentaje no
#   tiene contra qué medirse. Se sustituye por un suelo ABSOLUTO en días. 30 se elige
#   por ser un mes natural completo — suficiente para una mediana de cuotas sobre los 12
#   ASIN más vendidos — y NO está medido contra el caso real, porque la primera carga del
#   modelo contador todavía no ha ocurrido. Es la duda de diseño de este PR: se anota
#   aquí y se decide en frío cuando haya dos lecturas reales que mirar.
DIAS_MIN_CRUCE_PAIS = 30

# ---------------------------------------------------------------------------
# Guarda 6.14 · los tres números de la ZONA GRIS. No son umbrales de aborto: son la
# línea a partir de la cual la comparación DEJA DE PROBAR NADA y hay que llamar a un
# humano. Los tres se explican enteros aquí para que nadie tenga que adivinar de dónde
# salieron — que es la mitad de §3 de CLAUDE.md.
# ---------------------------------------------------------------------------
# 🔴 EL CORTE DE DÍAS, Y DE DÓNDE SALE (11-ago-2026). Es el único de los tres que no es
#   una cuenta redonda del encargo, así que va con su derivación completa. La banda está
#   MEDIDA por los dos extremos; el número de dentro es una línea convencional, y decirlo
#   es parte del trabajo:
#     · SUELO 1 — la cadencia real de la serie: **8 días**. Es el único hueco que existe
#       hoy en `demanda_asin` de producción (ES: 30-jul 18:06 → 7-ago 18:03 = 7 d 23 h
#       57 m). FR e IT tienen una sola lectura, así que no aportan hueco.
#     · SUELO 2 — el retraso del panel: **9 días** (medido el 10-ago-2026: "datos
#       disponibles hasta el 1/8/2026"). Por debajo de eso dos lecturas pueden traer
#       cifras IDÉNTICAS sin que pase nada raro — de eso ya se ocupa la guarda 6.15.
#       El corte tiene que quedar MUY por encima de 9 o gritaría en la operativa normal.
#     · TECHO — los **92 días** del «Custom date range» del panel (§2 de CLAUDE.md). Es
#       el único límite duro de lo que puede contener un fichero exportado con otro rango.
#       Con un hueco de ese orden, una ventana pirata acumula tanto como la referencia y
#       las cuatro pruebas de abajo dejan de distinguir una cosa de la otra.
#       🔬 Y esto NO es teoría, está medido: `2-ago_DISCONTINUO → 7-ago` da **0 bajadas
#       sobre 2.214 y las 9 métricas subiendo** — o sea, la misma firma exacta que una
#       carga buena. Cuando la referencia se queda muy por detrás en acumulado, cualquier
#       fichero pasa por bueno. Ése es el agujero que este corte tapa a medias.
#     · 31 = un mes natural, dentro de la banda 9…92. Casi 4× la cadencia observada (cabe
#       una avería o unas vacaciones sin molestar a nadie) y un tercio del techo.
# 🔒 Lo que hace barato equivocarse por lo bajo: NO tira la carga. Pide un `forzar`, o
#   sea un clic y un humano mirando. Por eso se elige el lado prudente de la banda.
CORTE_REF_DIAS = 31
# Los otros dos son del encargo del 11-ago-2026, tal cual: con menos ASIN comunes que
# esto, los totales se calculan sobre cuatro filas y "suben" no significa nada.
MIN_ASIN_COMUNES     = 20      # absoluto: por debajo, no hay muestra
MIN_FRACCION_COMUNES = 0.30    # relativo a la lectura anterior

# ---------------------------------------------------------------------------
# Guarda 6.14 · criterio 3: QUÉ MÉTRICAS PUEDEN DESPLOMARSE Y CUÁLES NO.
# 🔴 EL CRITERIO SE PARTE POR LA NATURALEZA DEL DATO, NO POR UN SUELO NUMÉRICO.
#   Decisión de Fernando, 11-ago-2026, y el porqué en una frase: **Amazon recalcula
#   PEDIDOS, no TRÁFICO.** Una cancelación o una devolución mueve unidades y euros —es la
#   vida normal de un marketplace—, pero nadie devuelve una visita. Así que una caída de
#   más de la mitad en unidades o en euros NO dice nada por sí sola, y en tráfico sí.
# 🔒 LAS SEIS DE PEDIDO QUEDAN FUERA del criterio 3 (unidades y euros de enviadas,
#   pedidas y reembolsadas). No quedan sin vigilar: el criterio 2 (≥5% de las
#   comparaciones) ya coge el caso masivo, que es el que delata otro rango.
# 🔬 MEDIDO SOBRE CUATRO PARES, y es lo que hace que esto no sea una opinión:
#     · ES 30-jul → 7-ago (producción, por SQL): 321 ASIN comunes, 0 bajadas en las nueve
#     · ES 1-ago → 7-ago (el fichero contra la base): 0 de 2.889
#     · IT, el del falso rojo: 2 bajadas, LAS DOS de pedido, 0 de tráfico
#     · ES 1-ago → 2-ago DISCONTINUO (el malo): de sus 1.465 desplomes, 610 son de
#       tráfico y 855 de pedido
#   Con el criterio de abajo: caso malo 527 desplomes, casos buenos 0.
# 🔴 Y EL SUELO NUMÉRICO SE DESCARTÓ CON DATOS, no por gusto: exigir 50 uds / 500 € para
#   mirar una bajada dejaría exentas el **89,3%** de las celdas de FR y el **78,4%** de las
#   de IT (medido el 11-ago-2026 sobre las nueve columnas de la última lectura de cada
#   país: FR 1.764 celdas, IT 1.008). Eso no es afinar un criterio: es apagarlo con un
#   número inventado.
# ⚠️ EL LÍMITE, dicho aquí y en §2 de CLAUDE.md: **no está probado que el tráfico no pueda
#   bajar legítimamente.** Una fusión de fichas o una depuración de tráfico inválido por
#   parte de Amazon lo haría. La evidencia es 3 pares buenos con CERO desplomes de tráfico
#   y 1 malo con 610 — suficiente para elegir, no para dar el asunto por cerrado. Si un día
#   salta un desplome de tráfico con todo lo demás en orden, ÉSE es el caso a estudiar, y
#   el aborto lo dice con esas palabras.
COLS_TRAFICO = ('visitas', 'sesiones', 'buybox_visiones')
# El 100 no es un suelo de importancia, es un suelo de RUIDO: por debajo, un porcentaje
# sobre un puñado de visitas no es un porcentaje. Va solo sobre tráfico.
MIN_TRAFICO_DESPLOME = 100

# ---------------------------------------------------------------------------
# Las 18 columnas medidas (§3.1). canon → cabecera NORMALIZada (sin acentos, minúsculas,
# espacios colapsados: la col 18 trae DOS espacios antes del paréntesis). Resolución por
# NOMBRE, jamás por posición (§4.4): el panel del 28-jul traía 8 columnas y estos 18.
# ---------------------------------------------------------------------------
CANON_A_CABECERA = {
    'asin':                 'asin',
    'nombre_producto':      'nombre del producto',
    'resenas':              'recuento de resenas de producto',
    'estrellas':            'valoraciones en estrellas generales',
    'visitas':              'visitas',
    'conversion':           'tasa de conversion (%)',
    'unidades_enviadas':    'unidades enviadas',
    'precio_venta_medio':   'precio de venta medio (€)',
    'ventas_enviadas_eur':  'ventas de unidades enviadas (€)',
    'inventario_disponible':'unidades de inventario disponibles',
    'buybox_ratio':         'ratio de oferta destacada',
    'buybox_visiones':      'visiones de ofertas destacadas',
    'reembolsado_eur':      'importe reembolsado (€)',
    'unidades_reembolsadas':'unidades reembolsadas',
    'reembolsos_ratio':     'ratio de reembolsos (%)',
    'sesiones':             'sesiones',
    'unidades_pedidas':     'unidades pedidas',
    'facturacion_pedida_eur':'facturacion neta de productos pedidos (€)',
}
# Tipo de cada columna (para tipar y para el cuadre). El ASIN y el nombre son texto.
COLS_ENTERAS = {'resenas', 'visitas', 'unidades_enviadas', 'inventario_disponible',
                'buybox_visiones', 'unidades_reembolsadas', 'sesiones', 'unidades_pedidas'}
COLS_NUMERICAS = {'estrellas', 'conversion', 'precio_venta_medio', 'ventas_enviadas_eur',
                  'buybox_ratio', 'reembolsado_eur', 'reembolsos_ratio', 'facturacion_pedida_eur'}
# Las 10 ADITIVAS: su suma tiene que cuadrar con la fila Total al céntimo (§3.3).
COLS_ADITIVAS = ('visitas', 'unidades_enviadas', 'ventas_enviadas_eur', 'inventario_disponible',
                 'buybox_visiones', 'reembolsado_eur', 'unidades_reembolsadas', 'sesiones',
                 'unidades_pedidas', 'facturacion_pedida_eur')
# Las NUEVE que de verdad ACUMULAN. `inventario_disponible` se cae de la lista: es un NIVEL
# (lo que hay ahora), no un acumulado, y baja legítimamente cada vez que se vende algo.
# Ratios y medias (conversión, buy box, estrellas, precio medio) tampoco acumulan nada.
# 🔒 UNA sola definición: la usan la guarda 6.14 (que ABORTA la carga) y el inventario del
# ensayo (que informa). Si fueran dos listas, el ensayo podría dar por bueno lo que la
# carga rechaza.
COLS_ACUMULADAS = tuple(c for c in COLS_ADITIVAS if c != 'inventario_disponible')

# Columnas de la tabla en el orden del INSERT (id/procesado_at aparte).
# 🔒 Son 22, y con `id` y `procesado_at` dan las 24 que tiene la tabla tras la migración
#   del 10-ago-2026. Fuera `periodo_desde`/`periodo_hasta` (la ventana inventada) y fuera
#   `dias` (era GENERATED sobre las dos). `exportado_at` no se ha ido: se llama `leido_at`,
#   que es lo que siempre fue — el instante en que Amazon generó la lectura.
COLS_DB = ['pais', 'asin', 'nombre_producto',
           'resenas', 'estrellas', 'visitas', 'sesiones', 'conversion',
           'unidades_pedidas', 'unidades_enviadas', 'precio_venta_medio',
           'ventas_enviadas_eur', 'facturacion_pedida_eur', 'buybox_ratio',
           'buybox_visiones', 'reembolsado_eur', 'unidades_reembolsadas',
           'reembolsos_ratio', 'inventario_disponible', 'fichero', 'leido_at', 'crudo']

RE_ASIN = re.compile(r'/dp/([A-Z0-9]{10})')

# 🔴 Nombres que PARECEN de usar y tirar. No decide nada: sirve para GRITAR cuando uno de
#   estos se carga como lectura buena, que es exactamente lo que pasó el 10-ago-2026.
#   `PRUEBA_ES/IT/FR.xlsx` resultaron ser las lecturas VERDADERAS del 30-jul 18:06.
#   ⚠️ Y OJO CON EL NOMBRE EN GENERAL: los del buzón se los puso a mano quien subió los
#   ficheros el 8-ago-2026; los originales de Amazon son `metric-data (7)`…`(14)`, sin
#   fecha y sin etiqueta. Un sufijo como `_01ago` o `_DISCONTINUO` es una anotación
#   humana, no un dato: no hay registro de por qué se puso.
#   Lo que decide qué es una lectura son sus DATOS y su `leido_at`, nunca cómo se llame el
#   fichero. El aviso existe para que nadie los borre del buzón creyendo que son basura.
RE_NOMBRE_DESECHABLE = re.compile(
    r'(?:^|[^A-Za-z0-9])(prueba|test|tmp|temp|borrar|copia|copy)(?:[^A-Za-z0-9]|$)', re.I)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _clean(v):
    return ('' if v is None else str(v)).replace('﻿', '').replace('\xa0', ' ').strip()

def _sin_acentos(s):
    return ''.join(c for c in unicodedata.normalize('NFD', str(s)) if unicodedata.category(c) != 'Mn')

def _norm(s):
    """Cabecera normalizada: sin acentos, minúsculas, espacios colapsados a uno (la col 18
    trae DOS espacios), sin ':' final. Comparar literal fallaría por esos dos espacios."""
    t = _sin_acentos(_clean(s)).lower().rstrip(':').strip()
    return re.sub(r'\s+', ' ', t)

def _txt(v):
    s = _clean(v)
    return s or None

def _num(v):
    """Valor del .xlsx → (float|None, fallo_bool). openpyxl ya devuelve número; se acepta
    también cadena europea por si Amazon cambia el formato (lo caza la guarda 6.4)."""
    if v is None:
        return None, False
    if isinstance(v, bool):   # True/False no es un número aquí
        return None, True
    if isinstance(v, (int, float)):
        return float(v), False
    s = _clean(v)
    if s in ('', '-'):
        return None, False
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return float(s), False
    except ValueError:
        return None, True

def _ent(v):
    f, fallo = _num(v)
    return (int(round(f)) if f is not None else None), fallo

def _bt(s):
    return _clean(s)

def _fecha_utc(v):
    """Fecha de las propiedades del .xlsx → UTC. openpyxl las devuelve sin zona y Amazon
    exporta en UTC; sin esto, dos fechas iguales compararían distinto."""
    if isinstance(v, datetime) and v.tzinfo is None:
        v = v.replace(tzinfo=timezone.utc)
    return v

def huella_serie(filas):
    """Huella de UNA LECTURA, pensada para comparar FICHERO contra BASE.
    md5 de `asin:visitas:unidades_pedidas`, unido por '|' y ORDENADO por asin.
    `filas` = iterable de (asin, visitas, unidades_pedidas).

    🔴 REGLA, NO DESCRIPCIÓN: ESTA HUELLA SOLO USA COLUMNAS ENTERAS.
    No se le añaden euros (`ventas_enviadas_eur`, `facturacion_pedida_eur`,
    `reembolsado_eur`) ni ratios, por muy tentador que sea "para que sea más específica".
    El motivo: esta huella cruza los dos mundos, y un `numeric` vuelve de la base como
    **Decimal** mientras el .xlsx da **float**. Su texto no coincide, y entonces la huella
    de la base y la del fichero NO cuadran NUNCA — la guarda dejaría de saltar, o saltaría
    siempre, y en los dos casos sin dar error. Es la misma trampa que costó el bug de la
    guarda 6.14 (§2 de CLAUDE.md: `43.98 < Decimal('43.98')` es True).
    ⚠️ Y conviene decir la verdad de cómo salieron estas tres: se eligieron el 10-ago-2026
    por ser las que identificaban una lectura con lo que había a mano, NO pensando en el
    ida y vuelta a la base. Que sean enteras fue SUERTE. Por eso queda escrito como
    restricción: la próxima vez no puede depender de que alguien acierte sin saberlo.
    Los tres enteros van y vuelven exactos. La huella ancha de 16 métricas
    (`huella_datos`) se queda para comparar fichero contra fichero, donde no hay base de
    por medio y los floats se comparan entre floats.

    🔒 ES REPRODUCIBLE EN SQL, que es como se ancló:
        md5(string_agg(asin||':'||visitas||':'||unidades_pedidas, '|' order by asin))
    Contrastada el 10-ago-2026 contra las CUATRO lecturas de producción:
        ES 30-jul 1873d6d6f35624654f4bcb6b06a52d64 · ES 7-ago 77255419bf48109c5b4dc876e020d9c0
        FR 30-jul 36fc9021dbfd09b27033ebd8cf981f4e · IT 30-jul 545beb2d25fa595e157982c565acb07a
    Las cuatro salen iguales por los dos caminos. Si alguien toca esta función, se
    contrasta otra vez contra esos cuatro números.

    🔒 Una fila con `visitas` o `unidades_pedidas` a NULL NO entra, igual que en SQL: allí
    la concatenación con NULL da NULL y `string_agg` lo descarta. Si aquí entrara, las dos
    huellas dejarían de ser la misma huella."""
    trozos = []
    for asin, visitas, uds in sorted(filas, key=lambda f: f[0]):
        if visitas is None or uds is None:
            continue
        trozos.append(f"{asin}:{visitas}:{uds}")
    return hashlib.md5('|'.join(trozos).encode('utf-8')).hexdigest()


def _created_de_wb(wb):
    """LA FECHA DEL DATO, leída tal como la lee el cargador. UNA sola función para los dos
    sitios que la necesitan —el cargador y el inventario del ensayo— porque si el
    inventario la leyera por su cuenta podría ANUNCIAR una fecha y cargarse otra, y
    entonces el inventario no probaría nada. Es la lección de `sql/huella_acl.sql`: una
    misma cosa medida con dos códigos no es una medida."""
    return _fecha_utc(getattr(wb.properties, 'created', None))


# ---------------------------------------------------------------------------
# 1) PARSEO — recibe BYTES, devuelve filas + totales. NO toca Storage ni la base.
#    Así se ejecuta contra los 3 ficheros locales las veces que haga falta (§2.1).
# ---------------------------------------------------------------------------
def analizar(bytes_xlsx, pais, fichero):
    if pais not in PAISES_VALIDOS:
        raise Aborta(f"[PAIS] {pais!r} no es ES/IT/FR. El país lo manda el selector y no se "
                     f"asume: sin país determinado, se ABORTA (§3.5).")

    # 🔴 data_only=False (por defecto): con True el ASIN sale None (§0.1). read_only para
    # no cargar la hoja entera en memoria; con read_only, max_row/max_col pueden ser None,
    # así que se ITERA y se cuenta a mano (§4.1).
    wb = openpyxl.load_workbook(io.BytesIO(bytes_xlsx), read_only=True)
    ws = wb['metric-data'] if 'metric-data' in wb.sheetnames else wb.active
    creator = _clean(getattr(wb.properties, 'creator', '') or '')
    leido_at = _created_de_wb(wb)
    # Las OTRAS dos propiedades del documento. No se guardan en la tabla: son para el
    # inventario del ensayo, donde desempatan dos ficheros con los mismos datos y distinta
    # fecha (§3). Un fichero reescrito trae `modified` posterior a `created` y, casi
    # siempre, un `lastModifiedBy` que Amazon no pone.
    modificado = _fecha_utc(getattr(wb.properties, 'modified', None))
    ultimo_autor = _clean(getattr(wb.properties, 'lastModifiedBy', '') or '')

    # 🔴 ¿DICE EL FICHERO EN ALGÚN SITIO DE QUÉ PERIODO ES? Es la pregunta que decide si la
    #   serie mide el mercado o solo la cadencia de Amazon. Medido el 10-ago-2026: el panel
    #   va NUEVE DÍAS por detrás ("datos disponibles hasta el 1/8/2026" con fecha 10-ago) y
    #   permite elegir rango, así que `leido_at` —que es cuándo se EXPORTÓ— no es la fecha
    #   de los datos. Si el rango viniera dentro, sería la llave de verdad.
    #   Se recoge TODO lo que podría llevarlo y se imprime en el inventario, en vez de
    #   suponer que no está: las otras hojas del libro y las propiedades del documento.
    hojas = list(wb.sheetnames)
    props_doc = {}
    for _k in ('title', 'subject', 'description', 'keywords', 'category', 'identifier'):
        _v = _clean(getattr(wb.properties, _k, '') or '')
        if _v:
            props_doc[_k] = _v
    pistas_periodo = []
    for _hn in hojas:
        if _hn == ws.title:
            continue                       # la hoja de datos ya se lee entera más abajo
        try:
            for _fila in wb[_hn].iter_rows(min_row=1, max_row=10, values_only=True):
                for _c in (_fila or ()):
                    _t = _clean(_c)
                    if _t:
                        pistas_periodo.append(f"[{_hn}] {_t}")
        except Exception:                  # una hoja ilegible no tumba el parseo
            pistas_periodo.append(f"[{_hn}] (no se ha podido leer)")

    filas = list(ws.iter_rows(values_only=True))
    wb.close()

    # 🔴 Guarda 6.5 · LA LECTURA — ocupa el hueco que dejó la guarda del PERIODO.
    #   Antes, sin periodo se abortaba porque el periodo era el eje del dato. Ahora el eje
    #   es CUÁNDO se leyó el contador, y esa fecha no la declara nadie: la trae el fichero
    #   en `wb.properties.created`. Si no viene, no hay lectura que apilar.
    # 🔒 Y no es una guarda decorativa: `leido_at` es NOT NULL en la tabla desde la
    #   migración del 10-ago-2026. Sin este aborto, un .xlsx sin `created` reventaría
    #   dentro del INSERT por lotes, a mitad de transacción y con un error de Postgres que
    #   no dice de qué fichero se trata. Mejor parar aquí y decir qué pasa.
    if not isinstance(leido_at, datetime):
        raise Aborta(
            f"[Guarda 6.5 · LECTURA] El .xlsx no trae fecha de creación "
            f"(wb.properties.created = {leido_at!r}), y esa fecha ES el dato: identifica la "
            f"lectura del contador y es la llave junto con el país y el ASIN. NO se inventa "
            f"ni se sustituye por 'ahora' (dos lecturas distintas quedarían fechadas igual). "
            f"Vuelve a descargar el export del Seller y súbelo otra vez.")

    avisos = []

    # Guarda 6.1: origen. No aborta (el exportador podría cambiar de nombre), grita.
    if 'custom analytics' not in _norm(creator):
        avisos.append(f"[Guarda 6.1] El creator del .xlsx es {creator!r}, sin 'Custom "
                      f"Analytics'. ¿Es de verdad un export de Custom Analytics? (Entra igual.)")

    if not filas or not any(_clean(c) for c in (filas[0] or ())):
        raise Aborta("[Guarda 6.3] El fichero no tiene ni cabecera legible. "
                     "¿Export de Custom Analytics? Abortando.")

    cabecera = list(filas[0])
    cab_norm = [_norm(c) for c in cabecera]

    # Resolver cada columna canónica por NOMBRE normalizado (§4.4).
    canon_idx = {}
    for canon, cab_esperada in CANON_A_CABECERA.items():
        for i, n in enumerate(cab_norm):
            if n == cab_esperada:
                canon_idx[canon] = i
                break

    # Guarda 6.2: sin ASIN no hay nada que cargar → ABORTA con la cabecera real entera.
    if 'asin' not in canon_idx:
        raise Aborta(
            f"[Guarda 6.2] No aparece la columna 'ASIN' en el fichero. NO se aproxima: se "
            f"ABORTA.\n   Cabecera real ({len(cabecera)} cols): {cabecera}")
    idx_asin = canon_idx['asin']

    # Columnas del canon que NO vienen (quedan NULL, no aborta) y columnas del fichero que
    # NO conocemos (se GRITAN: una métrica nueva que nadie ve es una métrica perdida, §4.4).
    ausentes = [c for c in CANON_A_CABECERA if c not in canon_idx]
    conocidas = {CANON_A_CABECERA[c] for c in CANON_A_CABECERA}
    desconocidas = [cabecera[i] for i, n in enumerate(cab_norm) if n not in conocidas]
    if ausentes:
        avisos.append(f"[columnas ausentes] No vienen en el panel (quedan NULL): "
                      f"{sorted(ausentes)}. (El panel se configuró con otras métricas.)")
    if desconocidas:
        avisos.append(f"[columna NUEVA sin canon] El fichero trae columnas que NO conozco → "
                      f"NO se tipan (viven solo en `crudo`): {desconocidas}. Añádelas al "
                      f"procesador si hacen falta como columna.")

    # --- Recorrer las filas de datos (la 2 en adelante). La fila 'Total' se aparta. ---
    datos = []
    total_declarado = None
    fallos = Counter()      # canon → nº de valores que no parsean
    intentos = Counter()    # canon → nº de valores no vacíos (para el % de la guarda 6.4)

    for pos, fila in enumerate(filas[1:], start=2):   # 'pos' = nº de fila real del .xlsx
        a_raw = fila[idx_asin] if idx_asin < len(fila) else None
        a_clean = _clean(a_raw)

        # La fila Total (primera fila de datos, ASIN='Total', nombre vacío): se aparta.
        if a_clean.lower() == 'total':
            total_declarado = fila
            continue

        m = RE_ASIN.search(str(a_raw or ''))
        if not m:
            # Colas de filas totalmente vacías que openpyxl a veces devuelve: se ignoran.
            if not any(_clean(c) for c in fila):
                continue
            raise Aborta(
                f"[Guarda 6.2] Fila {pos}: ASIN no reconocible en la celda {a_raw!r}. "
                f"No se descarta en silencio: se ABORTA. (Se esperaba una fórmula "
                f"=HYPERLINK(\"…/dp/B0…\") o la fila 'Total'.)")
        asin = m.group(1)

        registro = {'pais': pais, 'asin': asin, 'fichero': fichero, 'leido_at': leido_at}
        crudo = {}
        for i, h in enumerate(cabecera):     # crudo = fila ENTERA (despensa, §5): todo, tal cual
            crudo[str(h)] = fila[i] if i < len(fila) else None

        for canon, ci in canon_idx.items():
            if canon in ('asin',):
                continue
            val = fila[ci] if ci < len(fila) else None
            if canon == 'nombre_producto':
                registro[canon] = _txt(val)
            elif canon in COLS_ENTERAS:
                v, fallo = _ent(val)
                registro[canon] = v
                if _clean(val) not in ('', '-'):
                    intentos[canon] += 1
                    if fallo:
                        fallos[canon] += 1
            elif canon in COLS_NUMERICAS:
                v, fallo = _num(val)
                registro[canon] = v
                if _clean(val) not in ('', '-'):
                    intentos[canon] += 1
                    if fallo:
                        fallos[canon] += 1
        registro['crudo'] = crudo
        datos.append(registro)

    # Guarda 6.3: anti-vacío (el del 28-jul vino así: cabecera y cero filas).
    if not datos:
        raise Aborta("[Guarda 6.3] 0 filas de datos bajo la cabecera: el export vino VACÍO. "
                     "Vuelve a generarlo en el Seller con el panel de la dimensión ASIN.")

    # Guarda 6.4: valores que no parsean → se CUENTAN; si alguna columna pasa del 5% de
    # fallos, se GRITA (señal de que Amazon cambió el formato). Hoy sale a cero.
    for canon in sorted(set(fallos) | set(intentos)):
        n_int = intentos[canon]
        n_fal = fallos[canon]
        if n_int and n_fal / n_int > 0.05:
            avisos.append(f"[Guarda 6.4] La columna {canon!r} tiene {n_fal}/{n_int} valores que "
                          f"NO parsean ({100*n_fal/n_int:.1f}%). ¿Cambió Amazon el formato? "
                          f"(Esos valores quedan NULL.)")
        elif n_fal:
            avisos.append(f"[valores no numéricos] {canon!r}: {n_fal}/{n_int} no parsean (NULL).")

    # --- Totales del fichero (suma de las 10 aditivas sobre las filas de datos) ---
    totales_fichero = {}
    for canon in COLS_ADITIVAS:
        if canon in canon_idx:
            totales_fichero[canon] = round(sum((r.get(canon) or 0) for r in datos), 2)

    # Guarda 6.7: cuadre contra la fila Total (§3.3). La firma de Amazon: si alguna aditiva
    # no cuadra al céntimo, se ha leído mal el fichero → ABORTA.
    if total_declarado is None:
        avisos.append("[Guarda 6.7] El fichero NO trae fila 'Total': no se ha podido hacer el "
                      "cuadre de control contra la suma. (Se carga igual; la firma de Amazon "
                      "falta.)")
    else:
        descuadres = []
        for canon in COLS_ADITIVAS:
            if canon not in canon_idx:
                continue
            ci = canon_idx[canon]
            declarado, _ = _num(total_declarado[ci] if ci < len(total_declarado) else None)
            declarado = round(declarado or 0, 2)
            calculado = totales_fichero.get(canon, 0)
            if abs(declarado - calculado) > 0.01:
                descuadres.append(f"{canon}: fila Total={declarado} vs suma={calculado} "
                                  f"(dif {round(declarado - calculado, 2)})")
        if descuadres:
            raise Aborta(
                "[Guarda 6.7] El cuadre contra la fila 'Total' de Amazon NO da. Se ha leído mal "
                "el fichero (columnas desplazadas, filas perdidas…). NO se carga:\n        · "
                + "\n        · ".join(descuadres))

    # --- HUELLA DE LOS DATOS: md5 de ASIN + métricas, SIN nada de metadata ---
    # 🔴 Para qué. Dos exports de la MISMA lectura pueden traer `dcterms:created` distinto
    #   (es una cadena de longitud fija dentro del ZIP: cambiarla no mueve ni el tamaño del
    #   fichero). Entonces sus `leido_at` difieren, no se recierran, y entran en la serie
    #   como DOS lecturas con las mismas cifras. La resta entre ellas da CERO movimiento
    #   donde no lo hubo. Comparar fechas no lo detecta: hay que comparar los DATOS.
    # 🔒 Qué entra: `asin` + las 16 métricas, en orden fijo. NO entran ni `fichero`, ni
    #   `leido_at`, ni `pais`, ni `crudo` (metadata, y `crudo` arrastraría las cabeceras).
    #   NO entra `nombre_producto`: es descriptivo y Amazon lo retoca sin que el contador se
    #   mueva, así que ensuciaría la comparación con falsos "son distintos".
    # 🔒 Se ordena por ASIN: el md5 no puede depender del orden en que vengan las filas.
    _metricas = sorted(COLS_ENTERAS | COLS_NUMERICAS)
    _h = hashlib.md5()
    for r in sorted(datos, key=lambda x: x['asin']):
        _h.update(r['asin'].encode('utf-8'))
        for _c in _metricas:
            _v = r.get(_c)
            _h.update(b'|' + (b'' if _v is None else repr(_v).encode('utf-8')))
        _h.update(b'\n')

    return {
        'datos': datos,
        'n_asin': len(datos),
        'totales_fichero': totales_fichero,
        'creator': creator,
        'leido_at': leido_at,
        'modificado': modificado,
        'ultimo_autor': ultimo_autor,
        'huella_datos': _h.hexdigest(),
        'cabecera': [str(c) for c in cabecera],   # tal cual viene: delata otro panel (§3)
        'hojas': hojas,
        'props_doc': props_doc,
        'pistas_periodo': pistas_periodo[:40],
        'avisos': avisos,
        'columnas_ausentes': sorted(ausentes),
        'columnas_desconocidas': desconocidas,
    }


# ---------------------------------------------------------------------------
# 2) GUARDA 6.6 — EL PAÍS: cruzar la DEMANDA por ASIN con transacciones_movimientos.
#    El fichero no dice de qué marketplace es; el riesgo real es subir el de IT y
#    marcar ES. Se identifica por CUÁL de ES/IT/FR tiene menor error mediano.
#
#    🔴 SE COMPARAN CUOTAS, NO UNIDADES ABSOLUTAS (corrección de Fernando, 31-jul).
#    transacciones no cubre el mismo tramo que el fichero (empieza tarde en IT/FR y
#    acaba unos días antes), así que exigir cobertura total saltaba la guarda SIEMPRE
#    en el caso real. En su lugar se compara la CUOTA de cada ASIN (uds_asin / total
#    del conjunto), que es robusta a que los dos tramos no coincidan.
#
#    🔴 QUÉ CAMBIÓ CON EL MODELO CONTADOR (10-ago-2026). Antes el tramo de comparación
#    era la INTERSECCIÓN de la ventana DECLARADA con lo que cubre transacciones. Ya no
#    hay ventana declarada: el fichero es un acumulado desde un punto de partida fijo y
#    desconocido hasta `leido_at`. El análogo honesto es cruzar contra TODO lo que
#    transacciones tiene del país declarado hasta la fecha de la lectura:
#        [ min(fecha) del país declarado  →  min( max(fecha), leido_at ) ]
#    El extremo derecho se corta en `leido_at` porque el fichero no puede saber nada de
#    lo que pasó después de generarse.
#    Se SALTA la guarda (y se dice) si el fichero no trae 'Unidades pedidas' o si ese
#    tramo tiene menos de DIAS_MIN_CRUCE_PAIS días. NO caza una MEZCLA (un marketplace
#    por fichero; eso lo para el procedimiento).
#
#    MEDIDO CON EL TRAMO NUEVO (staging, ensayo del 10-ago-2026, run 31367604666,
#    fichero CA_ES_07ago.xlsx declarado ES):
#        ES = 2,0%   ·   IT = 97,6%   ·   FR = 90,4%
#    sobre un tramo de 220 días (2025-12-31 → 2026-08-07, cortado en la lectura).
#    O sea: el correcto por debajo del 3% y los incorrectos por encima del 90%. Separa
#    más que el modelo de ventana, donde el correcto daba 1,2% pero los incorrectos se
#    quedaban en ">75%". El suelo de DIAS_MIN_CRUCE_PAIS (30) no llegó a atar: sobraban
#    190 días.
#    ⚠️ La medición del modelo VIEJO era: correcto ES 1,2% · FR 11% · IT 11%, con ventana
#    declarada año→30-jul. Se deja escrita para poder comparar, no porque siga vigente.
# ---------------------------------------------------------------------------
def _puente_sku_asin(cur):
    """SKU→ASIN: listings_amazon ∪ productos(es_chase=false). El chase nace SIN ASIN,
    así que ya queda fuera. Listings manda si un SKU apareciera en los dos."""
    sku2asin = {}
    cur.execute(
        "SELECT sku, asin FROM productos "
        " WHERE coalesce(es_chase,false)=false AND asin IS NOT NULL AND btrim(asin)<>'' "
        "   AND sku IS NOT NULL AND btrim(sku)<>''")
    for sku, asin in cur.fetchall():
        sku2asin[_bt(sku)] = _bt(asin)
    cur.execute(
        "SELECT seller_sku, asin FROM listings_amazon "
        " WHERE asin IS NOT NULL AND btrim(asin)<>'' "
        "   AND seller_sku IS NOT NULL AND btrim(seller_sku)<>''")
    for sku, asin in cur.fetchall():
        sku2asin[_bt(sku)] = _bt(asin)   # listings pisa a productos (fuente dura del ASIN)
    return sku2asin


def _errores_de_cuota(cur, ini, fin, uds_fichero):
    """Error mediano de CUOTA del fichero contra CADA país, sobre el tramo [ini, fin].
    Devuelve ({pais: error|None}, tabla_legible).

    🔒 UNA SOLA IMPLEMENTACIÓN, y no es cosmética: la usan la guarda 6.6 (que decide si
    ABORTA la carga) y el etiquetado de países del inventario (que decide qué lecturas se
    comparan entre sí). Si el inventario etiquetara con otra fórmula, podría emparejar dos
    ficheros que la guarda considera de países distintos, y entonces sus comparaciones
    dirían cualquier cosa. Es la misma regla de `sql/huella_acl.sql`."""
    total_fichero = sum(uds_fichero.values())
    sku2asin = _puente_sku_asin(cur)

    # Unidades pedidas por (país, asin) sobre el tramo, para TODOS los candidatos.
    cur.execute(
        "SELECT pais, sku, sum(cantidad)::numeric FROM transacciones_movimientos "
        " WHERE tipo_norm='pedido' AND fecha BETWEEN %s AND %s "
        "   AND cantidad IS NOT NULL AND sku IS NOT NULL "
        " GROUP BY pais, sku;", (ini, fin))
    trans = defaultdict(lambda: defaultdict(float))   # pais → asin → uds
    for p, sku, uds in cur.fetchall():
        asin = sku2asin.get(_bt(sku))
        if asin:
            trans[p][asin] += float(uds)

    # CUOTAS: cuota de cada ASIN en el fichero (sobre el acumulado) y en cada candidato
    # (sobre el tramo). Comparar cuotas neutraliza que los dos tramos no coincidan.
    cuota_fichero = {a: u / total_fichero for a, u in uds_fichero.items()} if total_fichero else {}
    errores = {}
    for cand in PAISES_VALIDOS:
        total_cand = sum(trans[cand].values())
        if total_cand <= 0:
            errores[cand] = None
            continue
        # Los 12 ASIN de más unidades del candidato; error relativo de su CUOTA vs el fichero.
        top = sorted(trans[cand].items(), key=lambda kv: kv[1], reverse=True)[:12]
        errs = []
        for asin, tu in top:
            if tu <= 0:                       # no dividir por cero (medido: qty mín = 1)
                continue
            ts = tu / total_cand
            errs.append(abs(cuota_fichero.get(asin, 0.0) - ts) / ts)
        errores[cand] = median(errs) if errs else None

    tabla = " · ".join(
        f"{c}={'s/d' if errores[c] is None else format(100*errores[c], '.1f')+'%'}"
        for c in PAISES_VALIDOS)
    return errores, tabla


def etiquetar_pais(cur, leido_at, uds_fichero):
    """Para el INVENTARIO: ¿de qué marketplace es este fichero? Devuelve (pais|None, tabla).

    🔴 POR QUÉ HACE FALTA, medido el 10-ago-2026. El inventario agrupaba las lecturas por
    SOLAPE DE ASIN para decidir cuáles comparar entre sí, dando por hecho que cada
    marketplace tiene sus ASIN. Es FALSO: en Amazon EU el mismo ASIN vale para varios
    países. El solape medido entre el fichero de ES y el de FR fue del 86%, y entre FR e IT
    del 97%, así que los tres se agruparon como una sola serie y la comparación "de
    contador" acabó restando España contra Francia — 1.129 falsas bajadas.
    El país NO se adivina por solape: se identifica cruzando con transacciones, que es lo
    que ya hace la guarda 6.6. Aquí se reutiliza su misma fórmula para ETIQUETAR (sin
    abortar: el inventario informa, no decide).
    🔒 El tramo se toma sobre TODOS los países, no sobre uno declarado: aquí no hay
    declaración que respetar y el tramo tiene que ser el mismo para los tres candidatos."""
    if sum(uds_fichero.values()) <= 0:
        return (None, "sin 'Unidades pedidas': no hay con qué cruzar")
    cur.execute("SELECT min(fecha), max(fecha) FROM transacciones_movimientos;")
    fmin, fmax = cur.fetchone()
    if fmin is None:
        return (None, "transacciones vacía")
    ini, fin = fmin, min(fmax, leido_at.date())
    if (fin - ini).days + 1 < DIAS_MIN_CRUCE_PAIS:
        return (None, f"tramo de {(fin - ini).days + 1} días, menos de {DIAS_MIN_CRUCE_PAIS}")
    errores, tabla = _errores_de_cuota(cur, ini, fin, uds_fichero)
    validos = [c for c in PAISES_VALIDOS if errores[c] is not None]
    if not validos:
        return (None, f"sin cruce ({tabla})")
    return (min(validos, key=lambda c: errores[c]), tabla)


def guarda_pais(cur, pais_declarado, leido_at, uds_fichero):
    """Devuelve (veredicto, detalle). veredicto ∈ {'ok','grita','salta'}; si el país
    declarado NO gana, lanza Aborta. `uds_fichero` = {asin: unidades_pedidas} del fichero.
    Compara CUOTAS (no unidades) sobre todo lo que transacciones cubre del país declarado
    hasta la fecha de la lectura — robusto a que los dos tramos no coincidan."""
    # 0) Sin 'Unidades pedidas' en el fichero (el panel puede no traerla: el export del
    #    28-jul tenía 8 columnas) → uds_fichero todo a cero: no hay con qué cruzar.
    total_fichero = sum(uds_fichero.values())
    if total_fichero <= 0:
        return ('salta', "el fichero no trae 'Unidades pedidas' (columna ausente o suman 0): "
                         "no hay con qué cruzar el país. Guarda SALTADA.")

    # 1) Tramo de comparación: lo que transacciones cubre del declarado, hasta la lectura.
    cur.execute("SELECT min(fecha), max(fecha) FROM transacciones_movimientos WHERE pais=%s;",
                (pais_declarado,))
    fmin, fmax = cur.fetchone()
    if fmin is None:
        return ('salta', f"transacciones no tiene datos de {pais_declarado}: no hay con qué "
                         f"cruzar. Guarda SALTADA.")
    leido_dia = leido_at.date()
    ini, fin = fmin, min(fmax, leido_dia)
    dias_inter = (fin - ini).days + 1 if fin >= ini else 0
    if dias_inter < DIAS_MIN_CRUCE_PAIS:
        return ('salta', f"transacciones de {pais_declarado} solo cubre {dias_inter} días hasta "
                         f"la lectura ({fmin}→{fmax}, cortado en {leido_dia}): menos de "
                         f"{DIAS_MIN_CRUCE_PAIS}, demasiado poco para fiarse. Guarda SALTADA.")

    errores, tabla = _errores_de_cuota(cur, ini, fin, uds_fichero)

    if errores.get(pais_declarado) is None:
        return ('salta', f"no hay unidades cruzables para {pais_declarado} en la intersección "
                         f"(cuota por país: {tabla}). Guarda SALTADA.")

    ganador = min((c for c in PAISES_VALIDOS if errores[c] is not None), key=lambda c: errores[c])
    if ganador != pais_declarado:
        raise Aborta(
            f"[Guarda 6.6 · PAÍS] Se declaró {pais_declarado} pero el fichero cuadra con "
            f"{ganador}. Error mediano de CUOTA vs transacciones [{ini}→{fin}]: {tabla}. "
            f"O el selector de país va equivocado o subiste el fichero de otro marketplace. "
            f"NO se carga.")

    if errores[pais_declarado] > 0.25:
        return ('grita',
                f"{pais_declarado} GANA (cuota por país: {tabla}) pero su error pasa del 25%. "
                f"Ya no puede ser 'la ventana mal declarada' (no hay ventana): mira si el .xlsx "
                f"MEZCLA marketplaces, que es lo único que esta guarda no caza. (Entra igual.)")
    return ('ok', f"{pais_declarado} identificado por CUOTA (error mediano: {tabla}; "
                  f"tramo {ini}→{fin} = {dias_inter} días, cortado en la lectura {leido_dia}).")


# ---------------------------------------------------------------------------
# 3) INVENTARIO DEL BUZÓN — qué fecha trae CADA .xlsx. Solo en ensayo.
#
# 🔴 POR QUÉ EXISTE, con nombre y fecha. El 10-ago-2026 se dio por hecho que tres
#   parejas del buzón (PRUEBA_ES/IT/FR contra CA_ES/IT/FR_01ago) eran el mismo fichero
#   subido dos veces, y se dedujo de que pesaban lo mismo AL BYTE. Era FALSO: los eTag
#   de Storage —md5 del contenido, subida de una sola parte— diferían en las tres
#   parejas. Un .xlsx es un ZIP y `dcterms:created` es una cadena de longitud FIJA, así
#   que dos exports pueden pesar exactamente igual y traer fechas distintas.
#   Y la fecha no es un detalle: ES la llave (pais, leido_at, asin).
#       DEL TAMAÑO NO SE DEDUCE NADA. HAY QUE ABRIR EL FICHERO.
#
# 🔴 Y MIRA LOS DOS SENTIDOS, PORQUE SE ESCAPABA EL QUE IMPORTA. La primera versión de
#   este inventario solo comparaba FECHAS. Con eso detectaba el duplicado obvio —dos
#   ficheros con el mismo `leido_at`—, que además es el INOFENSIVO: el DELETE por
#   (pais, leido_at) lo recierra solo. Y se le escapaba justo el caso que sugieren los
#   eTag, que es el traicionero:
#       FECHAS DISTINTAS CON DATOS IDÉNTICOS.
#   Ahí el inventario habría impreso dos fechas distintas, habría dicho "ninguna se
#   repite" y habría dado luz verde a cargar dos veces la misma lectura. No se
#   recierran: entran como dos, con las mismas cifras, y la resta entre ellas da CERO
#   movimiento en un tramo donde sí lo hubo. Un dato falso con pinta de dato bueno.
#   Por eso junto a la fecha va una HUELLA DE LOS DATOS (§1, `huella_datos`), y el
#   veredicto tiene tres salidas:
#       · mismo leido_at                      → misma lectura, recarga idempotente, OK
#       · misma huella y leido_at distinto    → 🔴 EL PANEL NO REFRESCÓ (ver abajo)
#       · huella distinta y fecha distinta    → dos lecturas de verdad, adelante
#   Todo esto se imprime ANTES de cargar nada y en modo ensayo, que es donde se mira y
#   se decide, no después.
#
# 🔴 Y LA SALIDA DEL MEDIO RESULTÓ SER OTRA COSA, MÁS GRANDE. Al principio se leyó como
#   "el mismo fichero reescrito". No lo es: medido contra Drive el 10-ago-2026, las
#   exportaciones del 30-jul 18:06 y las del 1-ago 08:06 salieron de DOS descargas
#   distintas, en carpetas distintas, y devolvieron cifras IDÉNTICAS. O sea que no es un
#   problema de ficheros:
#       AMAZON NO REFRESCA EL PANEL CUANDO TÚ LO PIDES, SINO CUANDO ÉL QUIERE.
#   Así que la huella no sirve solo para cazar copias: es lo único que distingue
#   "no pasó nada en el mercado" de "Amazon todavía no lo ha contado". Sin ella, dos
#   lecturas sin refresco meten en la serie un tramo de cero movimiento que es cierto del
#   panel y FALSO del mercado.
#
# 📌 PENDIENTE DE MODELO, DECIDIDO EL 10-AGO-2026 PERO NO EN ESTE PR (§5 de CLAUDE.md).
#   Una lectura que no sea comparable con la serie —sin refresco, o con otro inicio de
#   acumulación— ENTRA Y SE MARCA: no se tira una descarga real. Pero marcarla no basta,
#   y esto está MEDIDO sobre `v_demanda_asin_ultima` en producción:
#     · Los DELTAS ya están a salvo: el `CASE WHEN visitas >= visitas_ant … ELSE NULL`
#       devuelve NULL en vez de un negativo.
#     · Pero la vista elige fila con `row_number() OVER (PARTITION BY pais, asin ORDER BY
#       leido_at DESC)` y `WHERE rn = 1`, y de esa fila sirve los ABSOLUTOS tal cual
#       (visitas, sesiones, unidades_pedidas, ventas_enviadas_eur). Así que una lectura
#       rara, si es la más reciente, se convierte en «la última buena» para todo lo que
#       cuelga: `v_trackeador_cola` expone visitas, sesiones, conversión y buy box.
#     Con el `metric-data (14)` dentro, el trackeador habría visto las visitas de ES al
#     ~1% de lo real y una conversión disparatada, con los deltas en NULL y nadie
#     mirándolos. Silencioso, que es la forma en que duele.
#   Así que la lectura marcada tiene que quedar FUERA del `rn = 1`: que la partición
#   ordene solo entre lecturas comparables. Eso toca la vista y, con ella, los cuatro
#   pisos de la cadena (`v_trackeador_cola`, `v_amazon_se_despierta`). Es otra migración,
#   va por la escalera y con su propio ensayo: NO se cuela en el PR del procesador.
#   Hoy el procesador DETECTA y GRITA en el ensayo, ABORTA en la carga si el contador
#   retrocede (guarda 6.14), y guarda la huella en el sello
#   (`informes_subidos.resumen_json`) para que ese PR tenga con qué empezar.
# ---------------------------------------------------------------------------
def inventario_lecturas(sb, xlsxs, pais, cur):
    """Abre TODOS los .xlsx del buzón y, por cada uno, dice qué `leido_at` le tocaría y
    qué huella tienen sus DATOS. No escribe nada: es para MIRAR antes de decidir.
    🔒 Parsea con `analizar()`, el MISMO código que carga. Si el inventario parseara por
    su cuenta podría anunciar una cosa y cargarse otra, y entonces no probaría nada."""
    print(f"\n--- INVENTARIO DEL BUZÓN: qué trae cada .xlsx ({len(xlsxs)}) ---", flush=True)
    print("    (se abre CADA fichero: ni el tamaño ni el nombre dicen nada)", flush=True)
    leidos = []                                  # un dict por fichero legible
    for o in xlsxs:
        nom = o.get('name') or '?'
        try:
            info_i = analizar(descargar_buzon(sb, BUCKET, f"{CARPETA}/{nom}"), pais, nom)
        except Aborta as e:
            print(f"        · {nom}: ABORTARÍA → {str(e).splitlines()[0]}", flush=True)
            continue
        except Exception as e:                   # un fichero ilegible no tumba el inventario
            print(f"        · {nom}: NO se ha podido leer ({type(e).__name__}: {e})", flush=True)
            continue
        porasin = {d['asin']: d for d in info_i['datos']}
        uds = {a: (d.get('unidades_pedidas') or 0) for a, d in porasin.items()}
        etiqueta, tabla_p = etiquetar_pais(cur, info_i['leido_at'], uds)
        leidos.append({'nombre': nom, 'leido_at': info_i['leido_at'],
                       'huella': info_i['huella_datos'], 'n': info_i['n_asin'],
                       'modificado': info_i['modificado'], 'autor': info_i['ultimo_autor'],
                       'cabecera': info_i['cabecera'], 'pais': etiqueta, 'porasin': porasin,
                       'hojas': info_i['hojas'], 'props_doc': info_i['props_doc'],
                       'pistas_periodo': info_i['pistas_periodo']})
        print(f"        · {nom}: leido_at={info_i['leido_at']}  ·  "
              f"datos={info_i['huella_datos'][:12]}…  ·  {info_i['n_asin']} ASIN", flush=True)
        print(f"              modificado={info_i['modificado']}  ·  "
              f"ultimo_autor={info_i['ultimo_autor'] or '(vacío)'}", flush=True)
        print(f"              país por cuota: {etiqueta or '(no identificable)'}  ({tabla_p})",
              flush=True)

    # 1) MISMA FECHA → misma lectura. El caso inofensivo: se recierra solo.
    por_fecha = defaultdict(list)
    for r in leidos:
        por_fecha[r['leido_at']].append(r['nombre'])
    for la, noms in sorted(por_fecha.items()):
        if len(noms) > 1:
            print(f"\n    ⚠️  MISMO leido_at ({la}): {', '.join(sorted(noms))}\n"
                  f"        Son LA MISMA lectura. Cargarlas en el mismo país se recierra "
                  f"(idempotente): ni duplica ni ensucia la serie.", flush=True)

    # 2) 🔴 MISMOS DATOS Y FECHA DISTINTA → el traicionero. Es lo que sugieren los eTag.
    por_huella = defaultdict(list)
    for r in leidos:
        por_huella[r['huella']].append(r)
    gemelas = 0
    for hu, grupo in sorted(por_huella.items()):
        if len(grupo) > 1 and len({r['leido_at'] for r in grupo}) > 1:
            gemelas += 1
            print(f"\n    🔴 MISMOS DATOS, FECHAS DISTINTAS (huella de datos {hu[:12]}…):",
                  flush=True)
            for r in sorted(grupo, key=lambda x: x['leido_at']):
                print(f"        · {r['nombre']}  →  leido_at={r['leido_at']}  ·  "
                      f"modificado={r['modificado']}  ·  autor={r['autor'] or '(vacío)'}",
                      flush=True)
            print("        EL PANEL DE AMAZON NO REFRESCÓ ENTRE LAS DOS. No es un fichero "
                  "duplicado ni una copia: son exportaciones de verdad, hechas en momentos "
                  "distintos, que devolvieron EL MISMO contador porque Amazon no lo había "
                  "actualizado todavía. Medido el 10-ago-2026 contra Drive: las del 30-jul "
                  "18:06 y las del 1-ago 08:06 salieron de dos descargas distintas, en "
                  "carpetas distintas, con cifras idénticas.", flush=True)
            print("        🔴 SI SE CARGAN LAS DOS, la resta entre ellas dice 'cero visitas en "
                  "dos días'. Eso es CIERTO del panel de Amazon y FALSO del mercado. Entra "
                  "UNA: la de fecha MÁS ANTIGUA, que es la lectura en la que el contador "
                  "valía eso de verdad; la posterior no midió nada nuevo.", flush=True)
            print("        📌 PENDIENTE DE MODELO (decidido el 10-ago-2026, no lo trae este "
                  "PR): la lectura no comparable ENTRA Y SE MARCA —no se tira una descarga "
                  "real— pero tiene que quedar FUERA del rn=1 de v_demanda_asin_ultima, no "
                  "solo fuera de la resta: esa vista sirve los ABSOLUTOS de la fila elegida y "
                  "v_trackeador_cola los expone. Toca la cadena de cuatro pisos y va en su "
                  "propio PR, con su ensayo.", flush=True)

    if leidos and not gemelas:
        print(f"\n    ✅ Ninguna pareja con los mismos datos y distinta fecha: los "
              f"{len(por_huella)} contenidos distintos son lecturas de verdad.", flush=True)

    # 3) 🔴 ¿SON LECTURAS DEL MISMO CONTADOR? La premisa que sostiene el modelo entero.
    #   El modelo v3 —`leido_at` como única fecha, y la resta entre lecturas como el
    #   movimiento del periodo— se apoya en que este informe es un ACUMULADO que NUNCA baja.
    #   Si baja una sola métrica de un solo ASIN entre dos lecturas, esa premisa no se
    #   sostiene para esos ficheros, y hay que verlo ANTES de cargar, no después.
    # 🔒 `inventario_disponible` NO entra: es un NIVEL (lo que hay), no un acumulado, y baja
    #   legítimamente cada vez que se vende algo. Meterlo daría bajadas falsas todo el rato.
    #   Tampoco entran ratios ni medias (conversión, buy box, estrellas, precio medio): no
    #   acumulan nada. Quedan las NUEVE que sí cuentan hacia arriba.
    # 🔴 LAS LECTURAS SE AGRUPAN POR PAÍS IDENTIFICADO, NO POR SOLAPE DE ASIN.
    #   La primera versión agrupaba por solape, dando por hecho que cada marketplace tiene
    #   sus ASIN. Es FALSO: en Amazon EU el mismo ASIN vale para varios países. Medido el
    #   10-ago-2026: ES contra FR solapaban al 86% y FR contra IT al 97%, así que los tres
    #   ficheros cayeron en la misma "serie" y la comparación acabó restando España contra
    #   Francia — 1.129 bajadas que no eran un contador retrocediendo, sino dos países
    #   distintos. El país se identifica cruzando con transacciones (`etiquetar_pais`),
    #   que es el método de la guarda 6.6. Un fichero que no se pueda etiquetar NO se
    #   compara con nadie, y se dice.
    ACUMULADAS = COLS_ACUMULADAS
    hay_bajadas = 0
    por_pais = defaultdict(list)
    sin_etiqueta = [r['nombre'] for r in leidos if not r['pais']]
    for r in leidos:
        if r['pais']:
            por_pais[r['pais']].append(r)
    if sin_etiqueta:
        print(f"\n    ⚠️  Sin país identificable, NO se comparan con nadie: "
              f"{', '.join(sorted(sin_etiqueta))}", flush=True)

    # 🔒 SE COMPARAN TODOS LOS PARES, no solo los consecutivos. Medido el 10-ago-2026: la
    #   cadena ES iba 30-jul → 1-ago → 2-ago → 7-ago, y el 2-ago resultó tener otro inicio
    #   de acumulación. Con solo los consecutivos, la pareja que de verdad se iba a cargar
    #   —30-jul contra 7-ago— NO se llegaba a comparar nunca: la cadena pasaba por el
    #   fichero roto y no encadenaba. Un par sin medir es un par del que no se sabe nada.
    for pais_s in sorted(por_pais):
        s = sorted(por_pais[pais_s], key=lambda x: x['leido_at'])
        if len(s) < 2:
            print(f"\n    · {pais_s}: una sola lectura ({s[0]['nombre']}), nada que comparar.",
                  flush=True)
            continue
        pares = [(s[a], s[b]) for a in range(len(s)) for b in range(a + 1, len(s))]
        for ant, pos in pares:
            a, p = ant['porasin'], pos['porasin']
            comunes, nuevos, faltan = set(a) & set(p), set(p) - set(a), set(a) - set(p)
            solape = 100 * len(comunes) / max(1, min(len(a), len(p)))
            print(f"\n    ── ¿MISMO CONTADOR? [{pais_s}]  {ant['nombre']} ({ant['leido_at']})"
                  f"  →  {pos['nombre']} ({pos['leido_at']}) ──", flush=True)
            print(f"        ASIN {len(a)} → {len(p)}  ·  comunes {len(comunes)} "
                  f"(solape {solape:.0f}%)  ·  nuevos {len(nuevos)}  ·  "
                  f"desaparecidos {len(faltan)}", flush=True)
            if faltan:
                print(f"        🔴 {len(faltan)} ASIN del anterior NO están en el posterior. "
                      f"Un acumulado no PIERDE ASIN: o el panel filtró, o no es la misma "
                      f"población. Ejemplos: {sorted(faltan)[:5]}", flush=True)
            if nuevos:
                print(f"        · {len(nuevos)} ASIN nuevos (esto sí es normal: un ASIN se "
                      f"estrena y aparece).", flush=True)
            bajadas = [(asin, m, a[asin][m], p[asin][m])
                       for asin in sorted(comunes) for m in ACUMULADAS
                       if a[asin].get(m) is not None and p[asin].get(m) is not None
                       and p[asin][m] < a[asin][m]]
            total_comp = len(comunes) * len(ACUMULADAS)
            if bajadas:
                hay_bajadas += len(bajadas)
                print(f"        🔴 EL CONTADOR RETROCEDE: {len(bajadas)} bajadas sobre "
                      f"{total_comp} comparaciones ASIN×métrica ({len(comunes)} ASIN × "
                      f"{len(ACUMULADAS)} acumuladas).", flush=True)
                for asin, m, va, vp in bajadas[:8]:
                    print(f"            · {asin} · {m}: {va} → {vp}", flush=True)
                if len(bajadas) > 8:
                    print(f"            · … y {len(bajadas) - 8} más", flush=True)
            else:
                print(f"        ✅ CERO bajadas en {total_comp} comparaciones ASIN×métrica: "
                      f"consistente con un acumulado.", flush=True)

    if hay_bajadas:
        print(f"\n    🔴🔴 PARA ANTES DE CARGAR. {hay_bajadas} bajadas en total. El modelo v3 "
              f"se validó justamente sobre que este contador NUNCA retrocede (medido el "
              f"7-ago-2026: 1.605 comparaciones, cero bajadas). Si retrocede, o no es "
              f"acumulado desde fecha fija, o no es la misma métrica — y entonces `leido_at` "
              f"como única fecha y la resta entre lecturas dejan de sostenerse para estos "
              f"ficheros. NO es un aviso de trámite.", flush=True)

    # 4) LAS CABECERAS: una columna distinta delata otro panel.
    por_cab = defaultdict(list)
    for r in leidos:
        por_cab[tuple(r['cabecera'])].append(r['nombre'])
    if len(por_cab) > 1:
        mayoritaria = max(por_cab, key=lambda k: len(por_cab[k]))
        print(f"\n    🔴 NO todos los .xlsx traen la misma cabecera ({len(por_cab)} distintas). "
              f"Una columna distinta delata otro panel o un rango declarado:", flush=True)
        for cab, noms in sorted(por_cab.items(), key=lambda kv: -len(kv[1])):
            marca = ' (mayoritaria)' if cab == mayoritaria else ''
            print(f"        · {len(cab)} col{marca}: {', '.join(sorted(noms))}", flush=True)
            if cab != mayoritaria:
                sobran = [c for c in cab if c not in mayoritaria]
                faltan_c = [c for c in mayoritaria if c not in cab]
                if sobran:
                    print(f"            trae de más: {sobran}", flush=True)
                if faltan_c:
                    print(f"            le faltan:   {faltan_c}", flush=True)
    elif leidos:
        print(f"\n    ✅ Los {len(leidos)} ficheros traen la MISMA cabecera "
              f"({len(next(iter(por_cab)))} columnas).", flush=True)

    # 5) 🔴 ¿TRAE EL FICHERO SU PERIODO? La pregunta que decide qué mide la serie.
    #   `leido_at` es cuándo se EXPORTÓ, no la fecha de los datos: medido el 10-ago-2026,
    #   el panel de Amazon iba NUEVE DÍAS por detrás ("datos disponibles hasta el 1/8/2026").
    #   Si el rango viniera dentro del .xlsx, sería la llave de verdad; si no viene, hay que
    #   decirlo en voz alta, porque entonces la resta entre dos lecturas mide la CADENCIA DE
    #   AMAZON, no lo que pasó en el mercado entre esas dos fechas.
    if leidos:
        r0 = leidos[0]
        print("\n    ── ¿DICE EL FICHERO DE QUÉ PERIODO ES? ──", flush=True)
        print(f"        hojas del libro: {r0['hojas']}", flush=True)
        print(f"        cabecera completa ({len(r0['cabecera'])} col): {r0['cabecera']}",
              flush=True)
        # 🔒 El veredicto NO puede darse por satisfecho con cualquier cadena. Medido el
        #   10-ago-2026: los ocho ficheros traen `title='Workbook'` y nada más, y con un
        #   "¿hay algo?" a secas eso bastaba para SUPRIMIR la conclusión — un diagnóstico
        #   que no llega nunca a su conclusión no sirve de nada. Se imprime todo lo
        #   encontrado (transparencia) pero solo cuenta como PISTA lo que podría llevar un
        #   periodo: algo con un dígito o con una palabra de rango.
        re_pista = re.compile(r'\d|periodo|period|rango|range|fecha|date|desde|hasta', re.I)
        pistas_reales = []
        for r in leidos:
            for k, v in (r.get('props_doc') or {}).items():
                print(f"        · {r['nombre']} · propiedad {k}={v!r}", flush=True)
                if re_pista.search(v):
                    pistas_reales.append(f"{r['nombre']} · {k}={v}")
            for p in (r.get('pistas_periodo') or []):
                print(f"        · {r['nombre']} · {p}", flush=True)
                if re_pista.search(p):
                    pistas_reales.append(f"{r['nombre']} · {p}")
        if pistas_reales:
            print(f"        ⚠️  Hay {len(pistas_reales)} cadena(s) que PODRÍAN llevar un "
                  f"periodo. Míralas: si alguna lo trae, esa es la llave de verdad y este "
                  f"modelo mejora. {pistas_reales[:5]}", flush=True)
        else:
            print("        🔴 NADA. Ni otras hojas, ni propiedades de documento, ni una "
                  "columna de periodo: el .xlsx NO dice de qué rango es. Consecuencias, "
                  "dichas en alto:\n"
                  "          · La única fecha del fichero es `created` = cuándo se exportó,\n"
                  "            y Amazon publica con días de retraso (9 el 10-ago-2026).\n"
                  "          · Por tanto restar dos lecturas NO da 'lo que pasó entre esas\n"
                  "            dos fechas': da lo que Amazon contó entre dos cortes suyos\n"
                  "            que no conocemos. La serie mide CADENCIA DE AMAZON, no de\n"
                  "            mercado. Sirve para tendencia y comparación entre ASIN; no\n"
                  "            para decir 'en agosto se vendieron X'.\n"
                  "          · Y el rango solo se garantiza por PROCEDIMIENTO: exportar\n"
                  "            siempre con «Desde el inicio de año». La guarda 6.14 es la\n"
                  "            red que caza al que no lo haga.", flush=True)
    return leidos


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------
def main():
    print("=== PROCESADOR CUSTOM ANALYTICS (DEMANDA · PELÍCULA DE LECTURAS) ===", flush=True)
    print(f"MODO: {MODO}  ·  ENTORNO: {ENTORNO}  ·  PAIS: {PAIS or '(sin selector)'}  ·  "
          f"FORZAR: {'SÍ' if FORZAR else 'no'}  ·  "
          f"fecha del dato: la trae el fichero (leido_at)", flush=True)
    print("=" * 60, flush=True)

    if MODO not in ('ensayo', 'aplicar'):
        sys.exit(f"MODO desconocido: {MODO!r} (usa 'ensayo' o 'aplicar')")
    if ENTORNO not in ('staging', 'produccion'):
        sys.exit(f"ENTORNO desconocido: {ENTORNO!r} (usa 'staging' o 'produccion')")
    if PAIS not in PAISES_VALIDOS:
        sys.exit(f"PAIS desconocido: {PAIS!r}. El país lo manda el selector (ES/IT/FR) y NO "
                 f"se asume: sin país no se carga (§3.5).")
    if not SUPABASE_KEY or not DB_URL:
        sys.exit("Faltan credenciales (SUPABASE_KEY / DB_URL). Revisa los secrets del workflow.")

    # (Aquí vivía el aviso de "recibido periodo_desde=… — IGNORADO". Cumplió su función
    #  durante la mudanza —una obsolescencia declarada no es una mentira, una silenciosa
    #  sí— y se va con los inputs que anunciaba: ya no los manda nadie.)

    # --- Bajar el fichero del buzón (Storage de PRODUCCIÓN) ---
    from supabase import create_client
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)
    objs = listar_buzon(sb, BUCKET, CARPETA)
    xlsxs = [o for o in objs if (o.get('name') or '').lower().endswith('.xlsx')]
    if not xlsxs:
        sys.exit(f"No hay ningún .xlsx en {BUCKET}/{CARPETA}/. Sube el export de Custom "
                 f"Analytics de {PAIS} y relanza.")
    xlsxs.sort(key=lambda o: (o.get('updated_at') or o.get('created_at') or ''), reverse=True)

    # Con ES/IT/FR en la misma carpeta "el más reciente" es una lotería: aquí pedir el
    # nombre EXACTO es lo NORMAL. Si se pide y no está → ABORTA (no cae al más reciente).
    if FICHERO:
        nombres = [o['name'] for o in xlsxs]
        if FICHERO not in nombres:
            print(f"\n❌ ABORTA (no se ha escrito nada):\n"
                  f"[Guarda fichero] Se pidió {FICHERO!r} y no está en {BUCKET}/{CARPETA}/.\n"
                  f"   Hay {len(nombres)} .xlsx en el buzón: {nombres}\n"
                  f"   No se cae al más reciente: cargaría un país/periodo distinto.", flush=True)
            sys.exit(1)
        fichero = FICHERO
        print(f"Fichero elegido (pedido a dedo): {fichero}", flush=True)
    else:
        fichero = xlsxs[0]['name']
        print(f"Fichero elegido (el más reciente de {len(xlsxs)}): {fichero}", flush=True)

    # 🔴 El nombre no decide qué es una lectura; los datos y el leido_at sí. Si el fichero
    #    que se va a cargar TIENE PINTA de desechable, se dice bien alto — porque se está
    #    cargando como bueno y alguien podría borrarlo del buzón creyendo que es basura.
    nombre_enganoso = bool(RE_NOMBRE_DESECHABLE.search(fichero))
    if nombre_enganoso:
        print(f"\n🔴 EL NOMBRE DE ESTE FICHERO MIENTE, Y SE CARGA IGUAL.\n"
              f"   '{fichero}' parece un fichero de usar y tirar, pero se está cargando como "
              f"una LECTURA BUENA: lo que decide qué es una lectura son sus DATOS y su "
              f"leido_at, no cómo se llame.\n"
              f"   Caso real (10-ago-2026): PRUEBA_ES/IT/FR.xlsx eran las lecturas VERDADERAS "
              f"del 30-jul 18:06. Los nombres del buzón se pusieron a mano al subirlos; los "
              f"que da Amazon son 'metric-data (N)', sin fecha y sin etiqueta.\n"
              f"   🔴 NO BORRES ESTE FICHERO DEL BUZÓN pensando que es basura de test.",
              flush=True)

    crudo_bytes = descargar_buzon(sb, BUCKET, f"{CARPETA}/{fichero}")

    # --- Parseo + guardas estructurales (antes de tocar la base) ---
    try:
        info = analizar(crudo_bytes, PAIS, fichero)
    except Aborta as e:
        print(f"\n❌ ABORTA (no se ha escrito nada):\n{e}", flush=True)
        sys.exit(1)

    datos = info['datos']
    leido_at = info['leido_at']
    print(f"\nASIN leídos: {info['n_asin']}  ·  país {PAIS}  ·  lectura {leido_at}", flush=True)
    print("   leido_at = wb.properties.created (cuándo generó Amazon el fichero)", flush=True)
    print(f"   totales del fichero: " + " · ".join(
        f"{k}={v}" for k, v in info['totales_fichero'].items()), flush=True)
    for a in info['avisos']:
        print(f"⚠️  {a}", flush=True)

    # Una lectura fechada en el FUTURO no la puede haber generado Amazon: o hay desajuste
    # de reloj o el fichero no es lo que parece. GRITA (no aborta): la fecha no la teclea
    # nadie, viene del propio .xlsx, y una lectura adelantada se colaría como "la última"
    # de la serie sin que nadie lo note.
    if leido_at.date() > date.today():
        print(f"\n⚠️  [lectura futura] leido_at ({leido_at.date()}) es posterior a hoy "
              f"({date.today()}). Se cargaría como la lectura MÁS RECIENTE de {PAIS} y taparía "
              f"a la de verdad en v_demanda_asin_ultima. Míralo. (Entra igual.)", flush=True)

    # --- Conectar al ENTORNO ---
    con = conectar_bd(DB_URL)
    con.autocommit = False
    cur = con.cursor()

    # Guarda 6.13: la tabla tiene que EXISTIR y estar CERRADA (RLS). NO se crea ni se
    # activa aquí (huevo y gallina §10): la crea la migración 2026-07-31_demanda_asin.sql.
    cur.execute("SELECT to_regclass('public.demanda_asin');")
    if cur.fetchone()[0] is None:
        print("\n❌ ABORTA: la tabla demanda_asin NO existe. La crea la migración "
              "2026-07-31_demanda_asin.sql (huevo y gallina §10): aplícala por la escalera y "
              "relanza. El procesador NO crea tablas nuevas.", flush=True)
        con.rollback(); cur.close(); con.close(); sys.exit(1)
    cur.execute("SELECT relrowsecurity FROM pg_class WHERE oid='public.demanda_asin'::regclass;")
    if not cur.fetchone()[0]:
        print("\n❌ ABORTA: RLS no está activa en demanda_asin. La activa la migración "
              "(regla del 29-jul: el procesador no toca la seguridad en cada carga). "
              "Aplica la migración y relanza.", flush=True)
        con.rollback(); cur.close(); con.close(); sys.exit(1)

    # Guarda 6.13 (cont.): y la tabla tiene que estar en el MODELO CONTADOR.
    # 🔴 Sin esto, apuntar a una base donde la migración del 10-ago no esté aplicada
    #   revienta DENTRO del INSERT por lotes, con un 'column "leido_at" does not exist' a
    #   mitad de transacción y después de haber hecho ya todo el trabajo. Es un aborto
    #   barato que convierte un error críptico en una instrucción.
    cur.execute("SELECT count(*) FROM information_schema.columns "
                " WHERE table_schema='public' AND table_name='demanda_asin' "
                "   AND column_name='leido_at';")
    if not cur.fetchone()[0]:
        print("\n❌ ABORTA: demanda_asin NO tiene la columna `leido_at`, o sea que esta base "
              "sigue en el modelo VIEJO (ventana declarada). Este procesador ya solo sabe "
              "escribir el modelo CONTADOR. Aplica por la escalera la migración "
              "2026-08-07_demanda_asin_contador.sql y relanza.", flush=True)
        con.rollback(); cur.close(); con.close(); sys.exit(1)

    # --- INVENTARIO DEL BUZÓN (solo ensayo) ---
    # 🔒 Va AQUÍ, después de conectar, y no antes de elegir el fichero como estaba: para
    #   etiquetar el país de cada lectura hace falta cruzar con transacciones, y eso pide
    #   la base. Sin país, las comparaciones de "mismo contador" acaban restando España
    #   contra Francia (medido el 10-ago-2026).
    # 🔒 No se hace en `aplicar` a propósito: ese es el modo del botón de Elena y no tiene
    #   que bajarse el buzón entero en cada carga. El ensayo es donde se MIRA y se decide.
    if MODO == 'ensayo':
        inventario_lecturas(sb, xlsxs, PAIS, cur)

    # --- Guarda 6.6: EL PAÍS ---
    uds_fichero = {r['asin']: (r.get('unidades_pedidas') or 0) for r in datos}
    try:
        veredicto, detalle = guarda_pais(cur, PAIS, leido_at, uds_fichero)
    except Aborta as e:
        print(f"\n❌ ABORTA (no se ha escrito nada):\n{e}", flush=True)
        con.rollback(); cur.close(); con.close(); sys.exit(1)
    marca = {'ok': '✅', 'grita': '⚠️', 'salta': '⏭️'}[veredicto]
    print(f"\n{marca} [Guarda 6.6 · país] {detalle}", flush=True)

    # --- Guarda 6.9: ASIN huérfanos (cuenta, no aborta) ---
    cur.execute("SELECT btrim(asin) FROM productos "
                " WHERE coalesce(es_chase,false)=false AND asin IS NOT NULL AND btrim(asin)<>'';")
    asin_prod = {r[0] for r in cur.fetchall()}
    huerfanos = [r['asin'] for r in datos if r['asin'] not in asin_prod]
    print(f"   ASIN huérfanos (no en productos, no chase): {len(huerfanos)} de {len(datos)} "
          f"(normal: ASIN que ya no listas o de otro vendedor en la misma ficha).", flush=True)

    # --- Guarda 6.11: país nuevo (radar fiscal; no aborta) ---
    cur.execute("SELECT DISTINCT pais FROM demanda_asin;")
    paises_bd = {r[0] for r in cur.fetchall()}
    if not paises_bd:
        print(f"\n1ª carga (tabla vacía): país inicial → {PAIS}", flush=True)
    elif PAIS not in paises_bd:
        print(f"\n🆕 PAÍS NUEVO en demanda_asin: {PAIS}. Posible NUEVA OBLIGACIÓN DE IVA — "
              f"revisar. (Entra igual.)", flush=True)

    # --- Guarda 6.8: LA SERIE de lecturas de ese país (grita, no borra) ---
    # Sustituye al viejo "solapamiento de ventanas": sin ventanas no hay solape que mirar.
    # Lo que sí puede pasar en una serie es apilar una lectura ANTERIOR a la última que ya
    # hay. Es legal —la tabla las admite en cualquier orden— pero casi siempre es un
    # despiste, y lo peligroso es que no se nota: v_demanda_asin_ultima seguiría enseñando
    # la de antes, así que la carga parecería no haber servido de nada.
    cur.execute(
        "SELECT leido_at, count(*), coalesce(sum(unidades_pedidas),0) "
        "  FROM demanda_asin WHERE pais=%s "
        " GROUP BY leido_at ORDER BY leido_at DESC;", (PAIS,))
    serie = cur.fetchall()
    if serie:
        print(f"\n   Lecturas de {PAIS} ya en la base: {len(serie)}", flush=True)
        for la, n, su in serie[:5]:
            cual = "   ← ESTA MISMA (se recierra)" if la == leido_at else ""
            print(f"        · {la}: {n} filas, {int(su)} uds pedidas acumuladas{cual}", flush=True)
        if len(serie) > 5:
            print(f"        · … y {len(serie) - 5} lectura(s) más", flush=True)
        ultima = serie[0][0]
        if leido_at < ultima:
            print(f"\n⚠️  [Guarda 6.8] Esta lectura ({leido_at}) es ANTERIOR a la última que ya "
                  f"tienes de {PAIS} ({ultima}). Se apila igual y no borra nada, pero "
                  f"v_demanda_asin_ultima seguirá mostrando la de {ultima}: para la pantalla, "
                  f"esta carga no cambiaría nada. ¿Es el fichero que querías subir?", flush=True)

    # --- Guarda 6.10: anti-encogimiento CONTRA LA LECTURA ANTERIOR ---
    # 🔴 El listón NO puede ser "lo que ya había de ESTA lectura": en una serie, una lectura
    #   nueva empieza SIEMPRE en 0 filas, así que la guarda no saltaría jamás y sería
    #   decorativa. El listón con sentido es la lectura ANTERIOR del mismo país: si el
    #   contador traía 195 ASIN y ahora trae 80, el export vino a medias.
    cur.execute("SELECT count(*) FROM demanda_asin WHERE pais=%s AND leido_at=%s;",
                (PAIS, leido_at))
    previas = cur.fetchone()[0]
    ref_n, ref_cual = 0, None
    for la, n, _su in serie:
        if la != leido_at:
            ref_n, ref_cual = n, la      # la más reciente que no es esta misma
            break
    if ref_n and len(datos) < ref_n * 0.5:
        print(f"\n❌ ABORTA (no se ha escrito nada):\n"
              f"[Guarda 6.10] La lectura anterior de {PAIS} ({ref_cual}) traía {ref_n} ASIN y "
              f"esta trae {len(datos)}: menos del 50%. Un informe a medias no da información "
              f"incompleta, da información FALSA. No se borra ni se escribe nada.", flush=True)
        con.rollback(); cur.close(); con.close(); sys.exit(1)

    # --- Guarda 6.15: EL PANEL NO HA REFRESCADO --------------------------------
    # 🔴 EL CASO QUE LA 6.14 NO VE, PORQUE NO BAJA: ES IGUAL.
    #   El panel de Custom Analytics publica con días de retraso — el 10-ago-2026 avisaba
    #   "datos disponibles hasta el 1/8/2026", NUEVE días — y hasta que su corte no avanza,
    #   dos exportaciones de días distintos devuelven EXACTAMENTE el mismo contador. No es
    #   un fichero duplicado: son dos descargas legítimas de un dato que no se ha movido.
    #   Medido el 10-ago-2026 contra Drive: las del 30-jul 18:06 y las del 1-ago 08:06
    #   salieron de dos sesiones distintas, en carpetas distintas, con cifras idénticas.
    # 🔴 SI ENTRARA, la serie tendría un tramo PLANO que no ocurrió: la resta entre las dos
    #   diría "cero visitas en dos días". Cierto del panel de Amazon y FALSO del mercado.
    #   Y la guarda 6.14 no lo caza por construcción: no baja nada, está todo igual.
    # 🔒 ABORTA, y aquí abortar no pierde nada: el fichero se queda en el buzón y se carga
    #   cuando el aviso del panel avance. No hay dato que rescatar — el dato ya está
    #   cargado, con su fecha buena, en la lectura anterior.
    #
    # 🔴 POR QUÉ CONTRA LA ÚLTIMA LECTURA Y NO CONTRA TODAS, que parece un hueco y no lo es.
    #   Son DOS casos distintos con DOS mecanismos distintos, y ninguno sobra:
    #     · MISMO FICHERO cargado dos veces → mismo `created` → mismo `leido_at` → la clave
    #       única (pais, leido_at, asin) lo convierte en una recarga IDEMPOTENTE. Ni
    #       siquiera llega aquí: no hay nada que comparar, se recierra y ya.
    #     · FICHERO NUEVO con datos idénticos (el panel no refrescó) → `leido_at` distinto,
    #       la clave única no lo ve, y ES ESTA GUARDA la que lo para. Y siempre es contra la
    #       ÚLTIMA: un panel que no avanza repite lo último, no algo de hace un mes.
    #   🔒 Si alguien "simplifica" quitando uno de los dos, se abre justo el agujero que el
    #   otro no cubre. Comparar contra TODAS las lecturas no añade nada y costaría una
    #   pasada por toda la serie en cada carga.
    # 📌 Esta es la versión de HOY. La fina —que la lectura entre MARCADA y quede fuera del
    #   `rn = 1` de v_demanda_asin_ultima, para no tirar una descarga real— pide columna
    #   nueva y tocar los cuatro pisos de la cadena: va en su PR y con su escalera.
    if ref_cual is not None:
        cur.execute("SELECT asin, visitas, unidades_pedidas FROM demanda_asin "
                    " WHERE pais=%s AND leido_at=%s;", (PAIS, ref_cual))
        h_previa = huella_serie(cur.fetchall())
        h_nueva = huella_serie([(r['asin'], r.get('visitas'), r.get('unidades_pedidas'))
                                for r in datos])
        if h_previa == h_nueva:
            print(f"\n❌ ABORTA (no se ha escrito nada):\n"
                  f"[Guarda 6.15 · EL PANEL NO HA REFRESCADO] Esta exportación trae los "
                  f"MISMOS datos que la lectura del {ref_cual}: Amazon no ha refrescado el "
                  f"panel.\n"
                  f"   · huella de serie de las dos: {h_nueva}\n"
                  f"   · asin+visitas+unidades_pedidas, idénticos ASIN por ASIN.\n"
                  f"   NO se carga. Si entrara, la serie tendría un tramo plano entre "
                  f"{ref_cual} y {leido_at} que no ocurrió: la resta diría 'cero movimiento' "
                  f"donde solo pasa que Amazon aún no lo ha contado.\n"
                  f"   🔑 QUÉ HACER: nada, y no se pierde nada. El fichero se queda en el "
                  f"buzón. Mira en el panel del Seller hasta qué fecha dice que hay datos y "
                  f"vuelve a intentarlo cuando ese aviso avance.", flush=True)
            con.rollback(); cur.close(); con.close(); sys.exit(1)
        print(f"\n✅ [Guarda 6.15] El panel SÍ ha refrescado desde la lectura anterior "
              f"({ref_cual}): huella de serie {h_previa[:12]}… → {h_nueva[:12]}…", flush=True)

    # --- Guarda 6.14: EL CONTADOR NO RETROCEDE ---------------------------------
    # 🔴 ESTA ES LA GUARDA QUE SOSTIENE EL MODELO ENTERO, y nació de un fichero real.
    #   El panel de Custom Analytics permite elegir el periodo, y por defecto viene con
    #   «Desde el inicio de año» (1-ene, inicio FIJO, fin móvil). Eso es lo que lo convierte
    #   en un contador acumulado y lo que hace que restar dos lecturas signifique algo.
    #   Pero si alguien exporta con OTRO rango —un «Custom date range» más corto—, el
    #   fichero resultante NO es una lectura de este contador: es otra ventana, y sus
    #   cifras son más pequeñas. Cargarlo mete restas NEGATIVAS en la serie.
    # 🔴 Caso real medido el 10-ago-2026 (`metric-data (14)`, subido como
    #   CA_ES_02ago_DISCONTINUO.xlsx): 246 ASIN contra los 321 de la lectura anterior, y
    #   1.583 bajadas sobre 2.214 comparaciones. Un ASIN cualquiera: visitas 35.400 → 428,
    #   unidades_enviadas 2.615 → 4. No es que el contador retrocediera: es que ese fichero
    #   empezaba a contar en otra fecha.
    # 🔒 ABORTA, no grita. Un acumulado que baja no es un dato incompleto: es un dato de
    #   OTRA cosa, y mezclado con la serie la envenena en silencio — que es justo lo que §1.4
    #   de CLAUDE.md dice de un informe caducado.
    #
    # 🔴 EL CRITERIO ES DE FERNANDO (11-ago-2026) Y ES EL TERCERO QUE SE ESCRIBE. Los dos
    #   anteriores fallaron por el MISMO sitio —confundir "ha bajado algo" con "este
    #   fichero es de otra cosa"—, y dejarlo escrito es lo que evita que vuelva el cuarto:
    #     · v1 «CUALQUIER bajada aborta» → rechazó una carga BUENA de IT en producción
    #       (run 31416495261) por DOS bajadas sobre 1.008: un solo ASIN, 9→8 unidades y
    #       19,99 € menos. O sea UNA CANCELACIÓN, la vida normal de un marketplace.
    #     · v2 «cualquier TOTAL acumulado que baje aborta» → el mismo falso rojo un piso
    #       más arriba: en un país pequeño, tres cancelaciones de 24,99 € hunden el total
    #       de facturación mientras las otras ocho métricas suben. Fue el fallo 3 de la
    #       revisión adversarial del 10-ago, y por eso este PR no se fusionó esa noche.
    #   v3 = CUATRO criterios de aborto DURO + una ZONA GRIS. ABORTA si CUALQUIERA:
    #     1. algún acumulado viene NEGATIVO            (mide el FICHERO: no usa referencia)
    #     2. ≥5% de las comparaciones ASIN×métrica bajan
    #     3. algún desplome de TRÁFICO: visitas, sesiones o buybox_visiones, con el valor
    #        anterior ≥100, se lleva MÁS DE LA MITAD (ver COLS_TRAFICO arriba: las seis de
    #        pedido quedan fuera, y el porqué está medido allí)
    #     4. bajan las NUEVE métricas acumuladas a la vez
    #   🔒 Ninguno de los cuatro es "algún total baja". Ése se ha ido A PROPÓSITO: era el
    #   fallo 3, y volver a meterlo "por si acaso" reabre el falso rojo de los países
    #   chicos. Si alguien lo echa de menos, el criterio 4 es su versión que sí distingue.
    #
    # 🔬 CALIBRADO CONTRA LOS FICHEROS REALES, no a ojo (§3 de CLAUDE.md). Las cuatro
    #   parejas que se pueden medir hoy, medidas el 11-ago-2026:
    #     pareja                                   bajadas       %    totales    neg → veredicto
    #     ES 30-jul → 7-ago             (buena)      0/2.889   0,0%  9/9 suben   0    CARGA
    #     IT 30-jul → metric-data (1)   (buena)      2/1.008   0,2%  9/9 suben   0    CARGA
    #     ES 30-jul → 2-ago DISCONTINUO (MALA)   1.583/2.214  71,5%  9/9 BAJAN   2    ABORTA
    #     ES 2-ago DISCONTINUO → 7-ago               0/2.214   0,0%  9/9 suben   0    ZONA GRIS
    #   Entre la peor carga BUENA (0,2%) y el fichero MALO (71,5%) hay dos órdenes de
    #   magnitud: el 5% cae en medio y no hay nada cerca por ninguno de los dos lados.
    #   La primera fila está contrastada por la OTRA VÍA, en SQL contra la base y no por
    #   el log: 2.889 comparaciones, 321 ASIN comunes, 0 bajadas, 0 desplomes, 0 totales
    #   que bajen. Los dos negativos del fichero malo son `facturacion_pedida_eur`:
    #   B0D19B54F2 = −12,90 y B07N5XFFS3 = −22,99. Y el mínimo de las NUEVE columnas en
    #   las CUATRO lecturas buenas de producción es 0 — nunca negativo. Por eso el
    #   criterio 1 vale y por eso NO necesita referencia: un acumulado desde el 1-ene no
    #   puede ser negativo, pero una ventana que empieza tarde sí (recoge la devolución de
    #   un pedido anterior a su propio inicio). El negativo no es un error de Amazon: es
    #   LA FIRMA de que el fichero empieza a contar en otra fecha.
    #
    # ⚠️ EL FALSO ROJO QUE TUVO EL CRITERIO 3 DURANTE MEDIA MAÑANA, y cómo se cerró. La
    #   primera versión miraba las nueve métricas, y "más de la mitad" sobre un valor de 1
    #   es una cancelación cualquiera: 1 → 0 es un −100%. Medido el 11-ago-2026:
    #     · FR: 23 ASIN con exactamente 1 unidad pedida = 9,7% de sus 238 unidades
    #     · IT: 23 ASIN con 1 unidad = 6,1% de sus 380 unidades
    #     · ES: 21 ASIN con 1 unidad = 0,2% de sus 13.768 unidades
    #   O sea: en FR, una de cada diez cancelaciones habría tirado una carga buena.
    # 🔒 NO SE ARREGLÓ CON UN SUELO, que era lo cómodo y lo equivocado: 50 uds / 500 €
    #   dejaba exentas el 89,3% de las celdas de FR y el 78,4% de las de IT, o sea apagar
    #   el criterio con un número inventado. Se arregló partiendo por la NATURALEZA del
    #   dato — Amazon recalcula pedidos, no tráfico — y eso es COLS_TRAFICO.
    #
    # 🔒 LA ZONA GRIS: cuando la comparación NO PUEDE PROBAR NADA, la guarda no dice "esto
    #   está mal", dice "no lo sé". Aborta y pide relanzar con `forzar`. Tres motivos, los
    #   tres con su número derivado arriba (CORTE_REF_DIAS, MIN_ASIN_COMUNES,
    #   MIN_FRACCION_COMUNES). Es la respuesta a los fallos 1 y 2 de la revisión: con
    #   cuatro ASIN comunes los totales se calculan sobre cuatro filas y "suben" no
    #   significa nada, y con una referencia de hace meses un export de otro rango puede
    #   SUBIR en todo (medido: 2-ago DISCONTINUO → 7-ago da 0 bajadas y 9/9 subiendo).
    # 🔒 `forzar` NO levanta los cuatro criterios duros. Sería un `ON_ERROR_STOP=0` con
    #   otro nombre: apagar la evidencia en vez de mirarla.
    # 🔒 Los criterios 2, 3 y 4 solo se aplican hacia ADELANTE (leido_at posterior a la
    #   referencia). Cargar una lectura anterior a la última ya la avisa la guarda 6.8, y
    #   compararlas al revés daría bajadas falsas por construcción. El criterio 1 NO
    #   depende de eso —mide el fichero—, así que también caza un DISCONTINUO metido por
    #   detrás: es la única parte del agujero «vieja detrás de nueva» que este PR tapa.
    #
    # 🔒 Y SE MIDE TODO ANTES DE DECIDIR NADA, aunque el criterio 1 por sí solo ya baste
    #   para abortar. Si el primer criterio que salta cortase la ejecución, el log del día
    #   malo enseñaría UNA línea y las otras tres sin medir — y entonces el veredicto no
    #   se puede comprobar, que es exactamente el defecto que este PR viene a quitar.
    #   Medir es barato: dos diccionarios en memoria sobre un fichero ya leído entero.

    # ── LA MEDICIÓN ────────────────────────────────────────────────────────────
    # CRITERIO 1 · NINGÚN ACUMULADO NEGATIVO. Mide el FICHERO: corre siempre.
    negativos = [(r['asin'], col, r.get(col))
                 for r in datos for col in COLS_ACUMULADAS
                 if r.get(col) is not None and float(r.get(col)) < 0]

    # CRITERIOS 2, 3 y 4 · CONTRA LA LECTURA ANTERIOR. Solo hacia adelante.
    hay_ref = ref_cual is not None and leido_at > ref_cual
    bajadas, desplomes, faltan_asin, gris = [], [], [], []
    if hay_ref:
        cur.execute(
            "SELECT asin, " + ", ".join(COLS_ACUMULADAS) + " FROM demanda_asin "
            " WHERE pais=%s AND leido_at=%s;", (PAIS, ref_cual))
        previo = {r[0]: r[1:] for r in cur.fetchall()}
        nuevo = {r['asin']: r for r in datos}
        comunes = set(previo) & set(nuevo)
        faltan_asin = sorted(set(previo) - set(nuevo))
        comparadas = len(comunes) * len(COLS_ACUMULADAS)
        hueco_dias = (leido_at - ref_cual).total_seconds() / 86400.0

        for asin in sorted(comunes):
            vals, fila_nueva = previo[asin], nuevo[asin]
            for i, col in enumerate(COLS_ACUMULADAS):
                va, vn = vals[i], fila_nueva.get(col)
                if va is None or vn is None:
                    continue
                # 🔴 float(...) EN LOS DOS LADOS, Y NO ES UNA COMPARACIÓN INOCENTE.
                #   La base devuelve `numeric` como Decimal EXACTO y el .xlsx da float
                #   BINARIO. En Python `43.98 < Decimal('43.98')` es True, porque el float
                #   43.98 vale en realidad 43.9799999… Medido el 10-ago-2026: 7 de 9 valores
                #   de euros reales del fichero daban BAJADA FALSA comparados así.
                #   Sin este float(), la guarda abortaría una carga buena en cuanto un ASIN
                #   pasara una semana sin vender: su importe seguiría igual y la guarda lo
                #   leería como retroceso. Una guarda que dice que no cuando debería decir
                #   que sí es tan mala como la que no salta.
                #   🔒 Se descubrió porque el inventario contó 1.583 bajadas y la guarda
                #   1.605 sobre la MISMA pareja. Veintidós de diferencia, y explicarlas al
                #   dígito (§1.3) es lo que destapó el bug.
                fa, fn = float(va), float(vn)
                if fn >= fa:
                    continue
                # CRITERIO 3: qué PARTE del valor anterior se lleva la bajada. Con el
                # anterior en 0 no hay proporción que calcular —bajar de 0 es irse a
                # negativo— y de eso ya se ocupa el criterio 1: cuenta como desplome.
                frac = (fa - fn) / fa if fa > 0 else 1.0
                bajadas.append((asin, col, va, vn, frac))

        # CRITERIO 3 · SOLO TRÁFICO, y solo por encima del suelo de ruido. El porqué está
        # entero arriba, en COLS_TRAFICO: Amazon recalcula pedidos, no tráfico.
        desplomes = [b for b in bajadas
                     if b[1] in COLS_TRAFICO
                     and float(b[2]) >= MIN_TRAFICO_DESPLOME
                     and b[4] > 0.50]
        pct_bajadas = (100.0 * len(bajadas) / comparadas) if comparadas else 0.0
        # Las dos "peores" se separan a propósito: la de tráfico es la que DECIDE, y la de
        # pedido se imprime para que se vea que se ha mirado y que NO cuenta. Un número que
        # no se enseña es un número del que luego nadie se acuerda.
        peor_trafico = max((b[4] for b in bajadas
                            if b[1] in COLS_TRAFICO
                            and float(b[2]) >= MIN_TRAFICO_DESPLOME), default=0.0)
        peor_pedido = max((b[4] for b in bajadas if b[1] not in COLS_TRAFICO), default=0.0)
        pct_comunes = (100.0 * len(comunes) / len(previo)) if previo else 0.0

        tot_antes, tot_ahora = {}, {}
        for i, col in enumerate(COLS_ACUMULADAS):
            # 🔴 float() en los DOS lados, por lo mismo que en la comparación de arriba:
            #   la base da Decimal y el .xlsx float binario (§2 de CLAUDE.md).
            tot_antes[col] = sum(float(previo[a][i])
                                 for a in comunes if previo[a][i] is not None)
            tot_ahora[col] = sum(float(nuevo[a].get(col))
                                 for a in comunes if nuevo[a].get(col) is not None)
        # Medio céntimo de holgura: sumar floats no da exacto, y un 1e-10 de ruido no
        # es un retroceso. Sin esto, el criterio 4 heredaría el bug de la guarda vieja.
        totales_abajo = [c for c in COLS_ACUMULADAS
                         if tot_ahora[c] < tot_antes[c] - 0.005]
        todas_abajo = len(totales_abajo) == len(COLS_ACUMULADAS)

        # ZONA GRIS: ¿puede esta comparación probar algo?
        if len(comunes) < MIN_ASIN_COMUNES:
            gris.append(f"solo hay {len(comunes)} ASIN comunes con la lectura anterior "
                        f"(el mínimo para que la comparación signifique algo es "
                        f"{MIN_ASIN_COMUNES})")
        if previo and len(comunes) < MIN_FRACCION_COMUNES * len(previo):
            gris.append(f"los {len(comunes)} ASIN comunes son solo el {pct_comunes:.1f}% "
                        f"de los {len(previo)} de la lectura anterior (mínimo "
                        f"{MIN_FRACCION_COMUNES * 100:.0f}%)")
        if hueco_dias > CORTE_REF_DIAS:
            gris.append(f"la lectura de referencia es de hace {hueco_dias:.0f} días "
                        f"(el corte son {CORTE_REF_DIAS}, un mes)")

    # ── LA MEDICIÓN, IMPRESA SIEMPRE ───────────────────────────────────────────
    # 🔒 Se imprime CARGUE O NO CARGUE, y aunque no haya una sola bajada. El motivo no es
    #   adorno: hasta el 10-ago-2026 esta guarda afirmaba una causa ("se exportó con OTRO
    #   PERIODO") que NO había medido, y era falsa — el propio procesador había impreso
    #   cinco segundos antes unos totales que subían todos. Un mensaje que da por segura
    #   una causa sin medirla hace perder el mismo tiempo que un verde falso. Con la tabla
    #   delante, el veredicto se COMPRUEBA en vez de creerse.
    if hay_ref:
        print(f"\n[Guarda 6.14] Comparación de {PAIS} contra la lectura anterior "
              f"({ref_cual}): hueco de {hueco_dias:.1f} días.", flush=True)
        print(f"   · TOTALES acumulados sobre los {len(comunes)} ASIN comunes:", flush=True)
        for col in COLS_ACUMULADAS:
            flecha = "BAJA" if col in totales_abajo else "sube/igual"
            print(f"        · {col}: {tot_antes[col]:g} → {tot_ahora[col]:g}  [{flecha}]",
                  flush=True)
    else:
        motivo_sin_ref = ("es la primera lectura de este país"
                          if ref_cual is None else
                          f"esta lectura NO es posterior a la última de {PAIS} "
                          f"({ref_cual}); de eso avisa la guarda 6.8, que grita y sigue")
        print(f"\n[Guarda 6.14] Sin comparación: {motivo_sin_ref}. Los criterios 2, 3 y 4 "
              f"no se pueden medir; el 1 sí, porque mide el fichero.", flush=True)

    print(f"   · LOS CUATRO CRITERIOS, con la cifra en que se basa cada uno:", flush=True)
    print(f"        1 · valores negativos en el fichero ... {len(negativos)}/"
          f"{len(datos) * len(COLS_ACUMULADAS)}   (aborta con ≥1)", flush=True)
    if hay_ref:
        print(f"        2 · bajadas ASIN×métrica ............. {len(bajadas)}/{comparadas}"
              f" = {pct_bajadas:.1f}%   (aborta con ≥5,0%)\n"
              f"        3 · desplomes de TRÁFICO ............ {len(desplomes)}   "
              f"(aborta con ≥1)\n"
              f"              · qué cuenta: {', '.join(COLS_TRAFICO)}, con el valor "
              f"anterior ≥{MIN_TRAFICO_DESPLOME} y una bajada de más de la mitad.\n"
              f"              · peor bajada de tráfico: {peor_trafico * 100:.1f}%   ·   "
              f"peor bajada de pedido: {peor_pedido * 100:.1f}% (NO cuenta aquí: Amazon "
              f"recalcula pedidos, no tráfico)\n"
              f"        4 · métricas cuyo TOTAL baja ........ {len(totales_abajo)}/"
              f"{len(COLS_ACUMULADAS)}   (aborta con las nueve)", flush=True)
        print(f"   · ¿PUEDE ESTA COMPARACIÓN PROBAR ALGO? (la zona gris):\n"
              f"        · ASIN comunes ...................... {len(comunes)} = "
              f"{pct_comunes:.1f}% de los {len(previo)} de la lectura anterior   "
              f"(mínimos: {MIN_ASIN_COMUNES} y {MIN_FRACCION_COMUNES * 100:.0f}%)\n"
              f"        · antigüedad de la referencia ....... {hueco_dias:.1f} días   "
              f"(corte: {CORTE_REF_DIAS})", flush=True)
        if faltan_asin:
            print(f"   · {len(faltan_asin)} ASIN de la lectura anterior NO vienen en ésta "
                  f"({100.0 - pct_comunes:.1f}%). Ejemplos: {faltan_asin[:5]}", flush=True)
    else:
        print(f"        2, 3 y 4 · no medibles sin lectura anterior.", flush=True)

    # ── EL VEREDICTO ───────────────────────────────────────────────────────────
    # Los motivos se arman FUERA del f-string a propósito: estas ramas corren el día que
    # haya un fichero malo de verdad, y un condicional anidado en una f-string es justo
    # donde se esconde el bug que nadie ve hasta entonces (§3: legible > compacto cuando
    # el código tarda meses en ejecutarse).
    alarmas = []
    if negativos:
        alarmas.append(f"criterio 1: {len(negativos)} valor(es) acumulado(s) por debajo "
                       f"de cero")
    if hay_ref:
        if pct_bajadas >= 5.0:
            alarmas.append(f"criterio 2: bajan {len(bajadas)} de {comparadas} "
                           f"comparaciones ({pct_bajadas:.1f}%, el listón es 5%)")
        if desplomes:
            alarmas.append(f"criterio 3: {len(desplomes)} desplome(s) de TRÁFICO se llevan "
                           f"más de la mitad del valor anterior (el peor, "
                           f"{peor_trafico * 100:.1f}%)")
        if todas_abajo:
            alarmas.append(f"criterio 4: bajan las {len(COLS_ACUMULADAS)} métricas "
                           f"acumuladas a la vez")

    if alarmas:
        if negativos:
            print(f"\n   Los {len(negativos)} negativos (criterio 1):", flush=True)
            for asin, col, v in negativos[:20]:
                print(f"        · {asin} · {col}: {v}", flush=True)
            if len(negativos) > 20:
                print(f"        · … y {len(negativos) - 20} más.", flush=True)
        if bajadas:
            print(f"\n   Las bajadas (criterios 2 y 3), las 20 primeras de "
                  f"{len(bajadas)}:", flush=True)
            for asin, col, va, vn, frac in bajadas[:20]:
                print(f"        · {asin} · {col}: {va} → {vn}  (−{frac * 100:.1f}%)",
                      flush=True)
            if len(bajadas) > 20:
                # 🔒 AQUÍ la lista se corta, y en la carga aceptada NO. No es incoherencia:
                #   esta carga no entra en la base, así que no hay nada que auditar después
                #   y con veinte se ve la forma. Cuando la carga SÍ entra, el log es el
                #   único sitio donde queda constancia y va entera.
                print(f"        · … y {len(bajadas) - 20} bajadas más. Se corta a "
                      f"propósito: esta carga NO entra en la base, así que no hay nada "
                      f"que auditar después.", flush=True)
        print(f"\n❌ ABORTA (no se ha escrito nada):\n"
              f"[Guarda 6.14 · RETROCESO] Este fichero no es la siguiente lectura del "
              f"mismo contador. {' · '.join(alarmas)}.", flush=True)
        print("   🔑 CAUSA MÁS PROBABLE, y está medida arriba, no supuesta: este .xlsx se "
              "exportó con OTRO RANGO. El panel deja elegir (hay un 'Custom date range' "
              "con tope de 92 días) y esta cañería SOLO admite «Desde el inicio de año» "
              "— inicio 1-ene fijo, fin móvil.\n"
              "   Vuelve al Seller, pon «Desde el inicio de año», re-exporta y sube ese.\n"
              "   ⚠️ `forzar` NO levanta esto. Solo sirve para la zona gris, que es otra "
              "cosa: allí la guarda no sabe, y aquí ha medido.", flush=True)
        # ⚠️ EL CASO A ESTUDIAR, y por eso se dice en el propio aborto.
        #   (Aquí vivía un aviso de "puede ser un falso rojo" para cuando el criterio 3
        #   saltaba por una cancelación de 1 → 0 unidades. Se fue el 11-ago-2026 con el
        #   criterio nuevo: las seis de pedido ya no entran en el criterio 3, así que ese
        #   falso rojo no puede ocurrir. Lo que queda es el hueco de conocimiento del otro
        #   lado, y no es lo mismo.)
        # 🔴 Lo que NO está probado: que el tráfico no pueda bajar legítimamente. Una
        #   fusión de fichas o una depuración de tráfico inválido por parte de Amazon lo
        #   haría. La evidencia son 3 pares buenos con CERO desplomes de tráfico y 1 malo
        #   con 610 — suficiente para elegir el criterio, no para dar el asunto por
        #   cerrado. Por eso: si el 3 salta SOLO, esto no se resuelve re-exportando.
        if desplomes and len(alarmas) == 1:
            peor_b = max((b for b in desplomes), key=lambda b: b[4])
            print(f"   ⚠️ OJO: el criterio 3 es el ÚNICO que ha saltado, y ÉSE es el caso "
                  f"que hay que ESTUDIAR, no despachar. {len(desplomes)} desplome(s) de "
                  f"tráfico con todo lo demás en orden — el peor, {peor_b[0]} · "
                  f"{peor_b[1]}: {peor_b[2]} → {peor_b[3]}.\n"
                  f"   Que el tráfico no pueda bajar legítimamente NO está probado: una "
                  f"fusión de fichas o una depuración de tráfico inválido de Amazon lo "
                  f"harían. La evidencia del criterio son 3 pares buenos con cero y 1 malo "
                  f"con 610 desplomes de tráfico (11-ago-2026).\n"
                  f"   🔑 Antes de re-exportar nada, enséñale esto a Fernando: puede ser el "
                  f"primer caso legítimo, y entonces lo que hay que cambiar es el criterio, "
                  f"no el fichero.", flush=True)
        # 🔴 …SALVO EN ENERO. A partir del 1-ene-2027 la causa de arriba puede ser la
        #   equivocada, y mandar a re-exportar con el MISMO periodo no arreglaría nada.
        #   «Desde el inicio de año» tiene el inicio FIJO en el 1 de enero: el contador se
        #   reinicia a cero, y la primera lectura del año nuevo es legítimamente MENOR que
        #   la última del anterior. Abortar está bien —esa resta sería basura—, pero el
        #   diagnóstico es otro. Ver §2 de CLAUDE.md.
        if hay_ref and leido_at.year != ref_cual.year:
            print(f"   ⚠️ OJO, HAY OTRA CAUSA POSIBLE Y AQUÍ ENCAJA: la lectura anterior "
                  f"es de {ref_cual.year} y esta de {leido_at.year}. El contador se "
                  f"REINICIA cada 1 de enero, así que esto puede ser el reinicio y no un "
                  f"fichero mal exportado. NO la fuerces: la resta entre dos años "
                  f"distintos no significa nada. Es el caso escrito en §2 de CLAUDE.md, "
                  f"pendiente de que el modelo guarde el año de acumulación.", flush=True)
        con.rollback(); cur.close(); con.close(); sys.exit(1)

    if gris and not FORZAR:
        print(f"\n❌ ABORTA (no se ha escrito nada):\n"
              f"[Guarda 6.14 · ZONA GRIS] Esta comparación NO PUEDE PROBAR NADA, así que "
              f"la guarda no dice que el fichero esté mal: dice que no lo sabe.",
              flush=True)
        for g in gris:
            print(f"        · {g}", flush=True)
        print(f"   🔑 POR QUÉ IMPORTA, y está medido (11-ago-2026): con la referencia muy "
              f"por detrás, un fichero exportado con OTRO RANGO sube en TODO y pasa por "
              f"bueno — la pareja `2-ago DISCONTINUO → 7-ago` da 0 bajadas sobre 2.214 y "
              f"las nueve métricas subiendo, la misma firma exacta que una carga limpia. Y "
              f"con cuatro ASIN comunes, los totales se calculan sobre cuatro filas: que "
              f"'suban' no dice nada de los otros 300.\n"
              f"   🔑 QUÉ HACER, por este orden:\n"
              f"      1) Mira la tabla de arriba y el nombre del fichero. ¿Es de verdad la "
              f"siguiente lectura de {PAIS}, exportada con «Desde el inicio de año»?\n"
              f"      2) Si falta alguna lectura intermedia por cargar, cárgala ANTES: el "
              f"hueco se cierra solo y la comparación vuelve a valer.\n"
              f"      3) Si estás seguro, relanza este mismo workflow con `forzar = si`. "
              f"Queda en el log que entró forzada y contra qué medición.\n"
              f"   ⚠️ `forzar` levanta SOLO esta zona gris. Los cuatro criterios de aborto "
              f"duro siguen puestos y se vuelven a medir igual.", flush=True)
        con.rollback(); cur.close(); con.close(); sys.exit(1)

    if gris:
        print(f"\n⚠️  [Guarda 6.14 · ZONA GRIS · FORZADA] Se ha lanzado con `forzar = si`, "
              f"así que ENTRA pese a que la comparación no prueba nada:", flush=True)
        for g in gris:
            print(f"        · {g}", flush=True)
        print(f"    Queda escrito aquí a propósito: si mañana la serie de {PAIS} no cuadra, "
              f"esta lectura ({leido_at}, fichero {fichero}) entró FORZADA y es la primera "
              f"a la que hay que mirar.", flush=True)

    if bajadas:
        # 🔒 AQUÍ LA LISTA VA ENTERA, no truncada. Y la diferencia con el aborto de arriba
        #   no es capricho: esta carga SÍ entra en la base, y el log es el ÚNICO sitio
        #   donde queda constancia de CUÁLES eran las bajadas. Un aviso incompleto sobre
        #   un dato que entra no sirve para auditarlo después.
        print(f"\n⚠️  [Guarda 6.14 · RETROCESO PUNTUAL] {len(bajadas)} bajadas sobre "
              f"{comparadas} comparaciones ({pct_bajadas:.1f}%), ningún desplome de "
              f"TRÁFICO, y no bajan las nueve a la vez. Eso NO es otro rango: es Amazon "
              f"recalculando (cancelaciones, devoluciones), que es la vida normal de un "
              f"marketplace — y lo que recalcula son PEDIDOS, no visitas. SE CARGA. Lista "
              f"COMPLETA:", flush=True)
        for asin, col, va, vn, frac in bajadas:
            print(f"        · {asin} · {col}: {va} → {vn}  (−{frac * 100:.1f}%)",
                  flush=True)
        print(f"    🔒 Y lo que cierra el argumento para cargar: `v_demanda_asin_ultima` YA "
              f"protege aguas abajo. Su `CASE WHEN visitas >= visitas_ant … ELSE NULL` "
              f"devuelve NULL en vez de un delta negativo (verificado el 10-ago-2026 "
              f"leyendo la definición en producción). Si la lista crece mucho de una "
              f"lectura a otra, míralo: eso ya no sería recálculo.", flush=True)
    elif faltan_asin:
        print(f"\n⚠️  [Guarda 6.14] Ni una sola bajada en {comparadas} comparaciones, pero "
              f"faltan {len(faltan_asin)} ASIN de la lectura anterior. Se carga: "
              f"`v_demanda_asin_ultima` va por (pais, asin) con rn=1, así que la última "
              f"lectura conocida de esos ASIN se queda donde está y no se inventa ningún "
              f"delta.", flush=True)
    elif hay_ref:
        print(f"\n✅ [Guarda 6.14] El contador no retrocede contra la lectura anterior "
              f"({ref_cual}): 0 bajadas en {comparadas} comparaciones ASIN×métrica "
              f"acumulada, ningún negativo, y no falta ningún ASIN.", flush=True)
    else:
        print(f"\n✅ [Guarda 6.14] Sin lectura anterior con la que comparar, pero el "
              f"criterio 1 se ha medido y ha pasado: ni un acumulado negativo en las "
              f"{len(datos) * len(COLS_ACUMULADAS)} celdas del fichero.", flush=True)

    # --- Carga PELÍCULA: DELETE de ESTA lectura por IGUALDAD + INSERT (misma transacción) ---
    # 🔒 El DELETE no contradice el cajón PELÍCULA (§1.6): no borra el histórico, recierra
    #   la MISMA lectura si se recarga. En una carga normal borra 0 filas.
    cur.execute("DELETE FROM demanda_asin WHERE pais=%s AND leido_at=%s;", (PAIS, leido_at))
    borradas = cur.rowcount

    plantilla = "(" + ", ".join(['%s'] * len(COLS_DB)) + ")"
    valores = [[(Json(r['crudo']) if c == 'crudo' else r.get(c)) for c in COLS_DB] for r in datos]
    execute_values(cur, f"INSERT INTO demanda_asin ({', '.join(COLS_DB)}) VALUES %s",
                   valores, template=plantilla, page_size=1000)
    insertadas = len(valores)

    # --- SELLO DE FRESCURA en informes_subidos (los 10 totales del cuadre) ---
    leido_dia = leido_at.date()
    resumen = {
        'pais': PAIS, 'tipo': 'custom_analytics', 'archivo': fichero,
        'leido_at': leido_at.isoformat(),
        'asin': len(datos), 'huerfanos': len(huerfanos),
        'totales': info['totales_fichero'],
        'lectura_anterior': ref_cual.isoformat() if ref_cual else None,
        # 🔴 La huella de los DATOS entra en el sello. Hoy no la usa nadie al cargar —no hay
        #    columna con la que compararla—, pero es lo único que distingue "no pasó nada" de
        #    "Amazon no lo ha refrescado". El PR que meta la columna en demanda_asin y haga
        #    que v_demanda_asin_ultima salte las lecturas sin refresco arranca desde aquí.
        'huella_datos': info['huella_datos'],
        # 🔴 Que el sello lo diga TAMBIÉN, y no solo el log: el log de un run caduca, la
        #    fila de informes_subidos se queda. Si el fichero se llama PRUEBA_* y es una
        #    lectura buena, esto es lo que lo dirá dentro de tres meses.
        'nombre_enganoso': nombre_enganoso,
        'aviso_fichero': (f"El fichero '{fichero}' TIENE NOMBRE DE DESECHABLE pero es una "
                          f"LECTURA BUENA: NO borrar del buzón. Lo que decide es leido_at + "
                          f"datos, no el nombre.") if nombre_enganoso else None,
        'guarda_pais': detalle, 'columnas_ausentes': info['columnas_ausentes'],
        'columnas_desconocidas': [str(c) for c in info['columnas_desconocidas']],
        'avisos': info['avisos'],
        'fuente': 'procesador_custom_analytics (Fase 0)',
    }
    # 🔒 Las dos fechas del sello van a la MISMA: una lectura es un instante, no un rango.
    cur.execute(
        "INSERT INTO informes_subidos "
        "(tipo, archivo_nombre, filas_procesadas, filas_validas, filas_descartadas, "
        " fecha_dato_desde, fecha_dato_hasta, resumen_json, procesado_at, notas) "
        "VALUES ('custom_analytics', %s, %s, %s, 0, %s, %s, %s, now(), %s);",
        (fichero, len(datos), insertadas, leido_dia, leido_dia, Json(resumen),
         f'procesador_custom_analytics Fase 0 · {PAIS} · lectura {leido_at.isoformat()}'))

    # --- Resumen ---
    verbo = 'se han' if MODO == 'aplicar' else 'se habrían'
    print(f"\n--- DEMANDA {PAIS} · lectura {leido_at} (PELÍCULA DE LECTURAS) ---")
    print(f"   · ASIN del fichero:              {len(datos)}")
    print(f"   · lectura anterior de {PAIS}:        "
          f"{ref_cual if ref_cual else '(ninguna: es la primera)'}"
          f"{f' con {ref_n} ASIN' if ref_n else ''}")
    print(f"   · ya había de ESTA lectura:      {previas}")
    print(f"   · BORRADOS de esta lectura ({verbo}): {borradas}")
    print(f"   · INSERTADOS ({verbo}):          {insertadas}")
    print(f"   · otras lecturas y países:       intactos (borrado por IGUALDAD de leido_at)")
    print(f"   · sello en informes_subidos ({verbo}): 1 fila (tipo='custom_analytics')")

    if MODO == 'aplicar':
        con.commit()
        # 🔒 Despues del commit y SOLO en `aplicar`: en `ensayo` no se ha escrito nada
        #    y refrescar seria copiar un estado que se acaba de deshacer. Un ensayo con
        #    efectos secundarios deja de ser un ensayo.
        # 🔴 ESTE PROCESADOR NO TENIA GANCHO HASTA HOY, y era el unico hueco: la copia
        #    `mv_demanda_asin_ultima` no se pondria al dia sola, asi que la pantalla
        #    ensenaria la demanda ANTERIOR con toda naturalidad. Y de paso ahora esta
        #    corrida tambien pone al dia la copia del Trackeador, que tambien mira la
        #    demanda.
        refrescar_vistas(con, 'custom_analytics')
        print(f"\n✅ APLICADO en {ENTORNO}: {insertadas} filas de {PAIS} · lectura {leido_at} en "
              f"demanda_asin (lectura recerrada por igualdad; el resto de la serie intacto; RLS "
              f"activo sin políticas; sello escrito).")
    else:
        con.rollback()
        print(f"\n🔎 ENSAYO: TODAS las guardas pasaron, NO se ha escrito nada. (El borrado de la "
              f"lectura, el volcado y el sello se han probado dentro de una transacción revertida.)")

    cur.close(); con.close()
    print(f"\n=== FIN · entorno={ENTORNO} · modo={MODO} · pais={PAIS} · "
          f"lectura={leido_at.isoformat()} · asin={len(datos)} · borrados={borradas} · "
          f"insertados={insertadas} · huerfanos={len(huerfanos)} ===", flush=True)


if __name__ == '__main__':
    main()
