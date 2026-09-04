# -*- coding: utf-8 -*-
"""Banco: la hoja de Excel del escaner hace LA MISMA CUENTA que la columna Decision.

EL FALLO QUE CIERRA (3-sep-2026). En la MISMA fila habia dos cuentas distintas:
   · la columna `Decision` y `escaner_detalle` salian de
     calc_rentabilidad(..., isd=ISD_PAIS[dom]);
   · las formulas VIVAS de la hoja calculaban la comision como ref_pct/100 x 1,03
     plano y el beneficio a partir de ahi.
En ES e IT las dos coincidian POR CASUALIDAD (COM_DIGITALES = 1,03 y el ISD es el
3 %). En FR no: alli el recargo cae tambien sobre la tarifa FBA y la hoja se dejaba
ese trozo, o sea que ensenaba ~1 punto de margen de MAS con un semaforo al lado
diciendo otra cosa. Una fila con dos cuentas es una fila que no se puede creer.

LAS TRES PARTES, Y POR QUE HACEN FALTA LAS TRES:
  (A) La IDENTIDAD, sobre las funciones REALES del escaner (sacadas del fichero con
      `ast`, por estructura, y ejecutadas):
          precio x pct_comision_celda + isd_sobre_fee  ==  calc_rentabilidad[...]
      Es la cuenta. Si alguien toca una de las dos, esto se pone rojo.
  (B) La ESTRUCTURA de la hoja: que la columna nueva este AL FINAL y que ninguna de
      las 29 de antes se haya movido de letra (hay consumidores que leen por letra),
      y que la formula de Com. Amazon use de verdad la columna nueva.
  (C) DE PUNTA A PUNTA: se corre el escaner ENTERO -- el fichero real, sin tocar una
      linea -- con un catalogo pequeno de perfil MIS_COMPRAS y con `keepa` y
      `supabase` sustituidos por dobles en memoria. Se lee el .xlsx que sale, se
      EVALUAN sus formulas y se comparan con `_paises_calc`. Sin red, sin secretos,
      sin tocar produccion, y sin escribir en escaner_memoria (perfil efimero).

SI (C) SE ROMPE POR UN CAMBIO AJENO (el escaner pide una tabla nueva, otra variable
de entorno...), lo que hay que tocar son los dobles de aqui abajo, NO la cuenta.
Los dobles son deliberadamente tontos: el de Supabase acepta cualquier metodo
encadenado y devuelve vacio.
"""
import ast
import io
import json
import os
import re
import sys
import types

RUTA = 'moloka_escaner_nube.py'
FUENTE = io.open(RUTA, encoding='utf-8').read()
ARBOL = ast.parse(FUENTE, RUTA)

fallos = []


def eq(nombre, obtenido, esperado):
    ok = obtenido == esperado
    if not ok:
        fallos.append(nombre)
    print(('OK ' if ok else 'XX ') + nombre
          + ('' if ok else '   got=%r exp=%r' % (obtenido, esperado)))


def casi(nombre, obtenido, esperado, tol):
    ok = abs(obtenido - esperado) <= tol
    if not ok:
        fallos.append(nombre)
    print(('OK ' if ok else 'XX ') + nombre
          + ('' if ok else '   got=%.10f exp=%.10f (tol %g)' % (obtenido, esperado, tol)))


# ===========================================================================
# (A) LA IDENTIDAD, sobre las funciones REALES del escaner
# ===========================================================================
# El escaner no se puede importar: en la linea 46 crea el cliente de Keepa con la
# clave del entorno y sale a la red. Se sacan los nodos que hacen falta POR
# ESTRUCTURA (nombre en el arbol, no grep) y se ejecutan.
def _nodo(nombre):
    for n in ARBOL.body:
        if isinstance(n, ast.FunctionDef) and n.name == nombre:
            return n
        if isinstance(n, ast.Assign) and any(
                isinstance(t, ast.Name) and t.id == nombre for t in n.targets):
            return n
    print('XX %s ya no esta en %s (o dejo de ser de nivel superior)' % (nombre, RUTA))
    sys.exit(1)


NOMBRES = ('SIN_ISD_HISTORICO', 'ISD_PAIS', 'calc_rentabilidad',
           'pct_comision_celda', 'isd_sobre_fee')
_ns = {'object': object}
exec(compile(ast.fix_missing_locations(ast.Module(body=[_nodo(n) for n in NOMBRES],
                                                  type_ignores=[])), RUTA, 'exec'), _ns)
ISD_PAIS = _ns['ISD_PAIS']
calc_rentabilidad = _ns['calc_rentabilidad']
pct_comision_celda = _ns['pct_comision_celda']
isd_sobre_fee = _ns['isd_sobre_fee']
print('extraidas de %s: %s' % (RUTA, ', '.join(NOMBRES)))
print()

ALMACEN, COM_DIGITALES = 0.15, 1.03


def com_celda(precio, ref_pct, fee, pais):
    """Lo que hacen las DOS celdas juntas: precio x %Comision + ISD s/ Fee Log."""
    return precio * pct_comision_celda(ref_pct, ISD_PAIS[pais]) + isd_sobre_fee(fee, ISD_PAIS[pais])


def com_funcion(precio, pa, ref_pct, fee, iva, pais):
    return calc_rentabilidad(precio, pa, ref_pct, fee, iva, almacen=ALMACEN,
                             com_digitales=COM_DIGITALES, isd=ISD_PAIS[pais])['com_amazon']


# --- (A1) La identidad, en los cuatro paises y sobre una rejilla de valores ---
_peor = 0.0
_n = 0
for pais in ('ES', 'IT', 'FR', 'DE'):
    for precio in (4.95, 12.13, 16.99, 21.90, 149.0):
        for ref in (8.0, 15.0, 45.0):
            for fee in (0.0, 2.70, 5.29, 11.4):
                _n += 1
                _peor = max(_peor, abs(com_celda(precio, ref, fee, pais)
                                       - com_funcion(precio, 5.0, ref, fee, 0.21, pais)))
casi('(A1) 🔴 las dos cuentas dan lo mismo en %d combinaciones (peor caso)' % _n, _peor, 0.0, 1e-12)

# --- (A2) Y que la rejilla DISTINGA de verdad las dos bases -------------------
# Si la vieja cuenta tambien cuadrara, esta rejilla no probaria nada.
_dif_fr = abs(com_celda(21.90, 15.0, 5.29, 'FR') - 21.90 * 15.0 / 100 * COM_DIGITALES)
casi('(A2) 🔴 en FR la cuenta VIEJA (x1,03 plano) NO cuadra: se deja el 3% de la FBA',
     _dif_fr, 5.29 * 0.03, 1e-12)
_dif_es = abs(com_celda(21.90, 15.0, 5.29, 'ES') - 21.90 * 15.0 / 100 * COM_DIGITALES)
casi('(A2) en ES la vieja y la nueva coinciden (por eso el fallo solo se veia en FR)',
     _dif_es, 0.0, 1e-12)

# --- (A3) El caso canonico de la factura, EN LAS CELDAS -----------------------
# Pedido 404-7912092-2024339 (FR): comision 1,82 + FBA 5,29 -> Amazon cobro 0,21.
_PRECIO, _REF, _FBA = 12.13, 15.0, 5.29
_comision = _PRECIO * _REF / 100
casi('(A3) la comision de la fila es la de la factura (1,82)', round(_comision, 2), 1.82, 0)
_isd_en_celdas = com_celda(_PRECIO, _REF, _FBA, 'FR') - _comision
casi('(A3) 🔴 el ISD que sale de las celdas es el de la factura (0,21)',
     round(_isd_en_celdas, 2), 0.21, 0)
casi('(A3) y sin redondear: 3% de (1,8195 + 5,29)', _isd_en_celdas,
     (_comision + _FBA) * 0.03, 1e-12)

# --- (A4) Las piezas, por separado -------------------------------------------
casi('(A4) el % de la celda sale de ISD_PAIS, no de COM_DIGITALES',
     pct_comision_celda(15.0, ISD_PAIS['ES']), 0.15 * 1.03, 1e-12)
eq('(A4) fuera de Francia el ISD s/ fee es CERO, no un hueco',
   [isd_sobre_fee(5.29, ISD_PAIS[p]) for p in ('ES', 'IT', 'DE')], [0.0, 0.0, 0.0])
casi('(A4) en Francia el ISD s/ fee es el 3% de la fee', isd_sobre_fee(5.29, ISD_PAIS['FR']),
     5.29 * 0.03, 1e-12)
casi('(A4) sin fee (None) no revienta y vale 0', isd_sobre_fee(None, ISD_PAIS['FR']), 0.0, 0)

# --- (A5) 🔴 Y QUE REACCIONEN A LA TABLA ------------------------------------
# Hoy 1 + pct = 1,03 = COM_DIGITALES, asi que un 1,03 escrito a mano daria los
# MISMOS numeros y todo lo de arriba seguiria verde: la coincidencia que escondio
# el fallo durante meses tambien puede esconder su vuelta. Se les pasa una tabla
# INVENTADA: si el numero no se mueve, es que no la estan leyendo.
_OTRA = {'pct': 0.07, 'incluye_fba': False}
casi('(A5) 🔴 el % de la celda se mueve con la tabla (no es un 1,03 a mano)',
     pct_comision_celda(15.0, _OTRA), 0.15 * 1.07, 1e-12)
casi('(A5) 🔴 el ISD s/ fee tambien se mueve con la tabla',
     isd_sobre_fee(5.29, {'pct': 0.07, 'incluye_fba': True}), 5.29 * 0.07, 1e-12)


# ===========================================================================
# (B) LA ESTRUCTURA DE LA HOJA: nadie se ha movido de letra
# ===========================================================================
print()
COLS_ANTES = [
    'Nombre', 'EAN', 'ASIN', 'Marca', 'PA (€)', 'País', 'Rank actual', 'Rank 90d',
    'Vendidos/mes', 'Precio venta (€)', 'Canal BB', 'Nº ofertas', '% Comisión',
    'Com. Amazon (€)', 'Fee Logística (€)', 'Almacén (€)', 'Promo activa',
    'Beneficio (€)', 'ROI', 'Margen', 'Decisión', 'En mi BD', 'EAN ambiguo',
    'Amazon (título)', 'Coincide', 'Cotejo', 'Cotejo (detalle)', 'Coherencia caja',
    'OcioStock',
    # Entro con ESTE encargo (#266/#267) y ya es una de las de antes: lo que este
    # banco protege es que NADIE se mueva de letra, no que la lista no crezca.
    'ISD s/ Fee Log. (€)']
ISD = 'ISD s/ Fee Log. (€)'
# La ultima que ha llegado (3-sep-2026, guarda del catalogo propio): dice de donde
# sale el IVA de cada fila. Su banco propio es test_escaner_catalogo_propio.py.
NUEVA = 'Origen IVA'
COLS = ast.literal_eval(_nodo('COLS').value)
eq('(B1) 🔴 las 30 columnas de antes siguen EN SU SITIO, en el mismo orden',
   COLS[:len(COLS_ANTES)], COLS_ANTES)
eq('(B1) 🔴 la columna nueva va la ULTIMA', COLS[len(COLS_ANTES):], [NUEVA])

# (B2) La formula de Com. Amazon usa de verdad la columna nueva. Se busca la
# f-string que monta la formula y se mira que cite las DOS columnas.
_formulas_con_isd = []
for n in ast.walk(ARBOL):
    if isinstance(n, ast.JoinedStr):
        _refs = {ast.literal_eval(s.slice)
                 for s in ast.walk(n)
                 if isinstance(s, ast.Subscript) and isinstance(s.value, ast.Name)
                 and s.value.id == 'L' and isinstance(s.slice, ast.Constant)}
        if ISD in _refs:
            _formulas_con_isd.append(_refs)
eq('(B2) 🔴 la formula de Com. Amazon suma la columna del ISD',
   any({'Precio venta (€)', '% Comisión', ISD} <= refs for refs in _formulas_con_isd), True)

# (B3) El % Comision pasa por pct_comision_celda (si no, la hoja volveria al x1,03).
_asig_pct = [n for n in ast.walk(ARBOL)
             if isinstance(n, ast.Assign)
             and any(isinstance(t, ast.Name) and t.id == 'pct' for t in n.targets)
             and any(isinstance(c, ast.Call) and isinstance(c.func, ast.Name)
                     and c.func.id == 'pct_comision_celda' for c in ast.walk(n.value))]
eq('(B3) 🔴 el % Comisión de la hoja sale de pct_comision_celda()', len(_asig_pct), 1)

# (B4) Y la cuenta vieja ya no multiplica por COM_DIGITALES en ningun sitio.
# Por ESTRUCTURA (un BinOp de multiplicacion), que un grep casaria tambien dentro
# de los comentarios -- y este fichero tiene comentarios que la citan.
_vieja = [n for n in ast.walk(ARBOL)
          if isinstance(n, ast.BinOp) and isinstance(n.op, ast.Mult)
          and any(isinstance(x, ast.Name) and x.id == 'COM_DIGITALES' for x in (n.left, n.right))]
eq('(B4) 🔴 ya no queda ningun `... * COM_DIGITALES` en el codigo', len(_vieja), 0)
eq('(B4) el motivo sigue escrito en el fichero (la otra mitad del ancla)',
   'COM_DIGITALES' in FUENTE, True)

# (B5) El hipervinculo de OcioStock ya no da por hecho que es la ultima columna.
_len_cols_como_columna = [k for n in ast.walk(ARBOL) if isinstance(n, ast.Call)
                          for k in n.keywords
                          if k.arg == 'column' and isinstance(k.value, ast.Call)
                          and isinstance(k.value.func, ast.Name) and k.value.func.id == 'len']
eq('(B5) 🔴 ningun `column=len(COLS)`: el enlace iria a la columna equivocada',
   len(_len_cols_como_columna), 0)


# ===========================================================================
# (C) DE PUNTA A PUNTA: el escaner real -> un .xlsx real -> sus formulas
# ===========================================================================
print()
CATALOGO = (
    "Nombre,EAN,Precio,Marca,Stock\n"
    "Funko Pop Gizmo,0889698498883,4.50,Funko,12\n"
    "Funko Pop Luke Red 5,0889698851909,5.20,Funko,4\n"
    "CASO CANONICO FR,0889698000004,2.00,Funko,9\n"
    "Ultra Pro Fundas 100,0074427811266,1.10,Ultra Pro,30\n"
    "SIN FEE FBA EN ES,0889698000005,3.00,Funko,7\n")
RECADO = {"proveedor": "MIS_COMPRAS", "marca": "TODAS", "modo": "todo",
          "rank_maximo": 200000, "incluir_sin_rank": False}
ASIN = {'0889698498883': 'B0GIZMO', '0889698851909': 'B0LUKE5',
        '0889698000004': 'B0CANON', '0074427811266': 'B0FUNDA',
        '0889698000005': 'B0SINFEE'}
# precio de venta, ref_pct, fee FBA, rank. La fila 'B0CANON'/FR es la factura:
# 12,13 x 15% = 1,8195 de comision, fee 5,29 -> ISD (1,8195+5,29) x 3% = 0,2133.
PAIS = {
    'B0GIZMO': {'ES': (16.99, 15.0, 3.51, 4200), 'IT': (17.49, 15.0, 3.62, 9100),
                'FR': (18.25, 15.0, 3.72, 6400)},
    'B0LUKE5': {'ES': (15.50, 15.0, 3.51, 30000), 'IT': (14.90, 15.0, 3.62, 51000),
                'FR': (21.90, 15.0, 5.29, 12500)},
    'B0CANON': {'ES': (12.13, 15.0, 5.29, 15000), 'IT': (12.13, 15.0, 5.29, 15000),
                'FR': (12.13, 15.0, 5.29, 15000)},
    'B0FUNDA': {'ES': (4.20, 15.0, 2.70, 88000), 'IT': (4.60, 8.0, 2.75, 120000),
                'FR': (4.95, 8.0, 2.80, 140000)},
    # 🔴 fee = None en ES: Keepa a veces no trae `fbaFees.pickAndPackFee`. Es la fila
    #    que enseñaba un Beneficio calculado con la celda de fee VACIA (leida como 0)
    #    mientras la columna Decision decia 'Sin datos'.
    'B0SINFEE': {'ES': (19.99, 15.0, None, 5000), 'IT': (19.99, 15.0, 3.60, 5000),
                 'FR': (19.99, 15.0, 3.70, 5000)},
}


def _stats(rank, precio, con_bb):
    cur = [-1] * 20
    cur[1] = int(round(precio * 100))          # NEW
    cur[3] = rank                               # SALES RANK
    if con_bb:
        cur[18] = int(round(precio * 100))      # BUY_BOX_SHIPPING
    a90 = [-1] * 20
    a90[3] = rank
    return cur, a90


class _FakeKeepa:
    def __init__(self, key, timeout=None):
        self.tokens_left = 1500

    def update_status(self):
        pass

    def query(self, items, **kw):
        self.tokens_left -= len(items)
        if kw.get('product_code_is_asin'):        # Fase 2: un ASIN, un pais
            asin = items[0]
            precio, ref, fee, rank = PAIS[asin][(kw.get('domain') or 'ES').upper()]
            cur, a90 = _stats(rank, precio, True)
            return [{'asin': asin, 'title': 'T ' + asin,
                     'stats': {'current': cur, 'avg90': a90, 'buyBoxIsFBA': True,
                               'totalOfferCount': 5, 'buyBoxPrice': int(round(precio * 100))},
                     'referralFeePercentage': ref,
                     'fbaFees': ({'pickAndPackFee': int(round(fee * 100))}
                                 if fee is not None else {}),
                     'monthlySold': 40, 'images': ['x.jpg']}]
        out = []                                   # Fase 1: EAN -> ASIN + rank
        for cod in items:
            c13 = str(cod).zfill(13)
            a = ASIN.get(c13)
            if not a:
                continue
            precio, _ref, _fee, rank = PAIS[a]['ES']
            cur, a90 = _stats(rank, precio, False)
            out.append({'asin': a, 'title': 'T ' + a,
                        'stats': {'current': cur, 'avg90': a90, 'salesRankDrops30': 12},
                        'listedSince': 6000000, 'eanList': [c13], 'upcList': []})
        return out


BUZON = {'escaner/_solicitud_escaner.json': json.dumps(RECADO).encode('utf-8'),
         'escaner/mini_compras.csv': CATALOGO.encode('utf-8-sig')}
SUBIDOS = {}


class _Bucket:
    def list(self, carpeta):
        pre = carpeta.rstrip('/') + '/'
        return [{'name': n} for n in sorted({k[len(pre):].split('/')[0]
                                             for k in list(BUZON) + list(SUBIDOS)
                                             if k.startswith(pre)})]

    def download(self, ruta):
        if ruta in BUZON:
            return BUZON[ruta]
        if ruta in SUBIDOS:
            return SUBIDOS[ruta]
        raise Exception('404 ' + ruta)

    def upload(self, ruta, data, opts=None):
        SUBIDOS[ruta] = data
        return {'path': ruta}

    def remove(self, rutas):
        for r in rutas:
            BUZON.pop(r, None)
            SUBIDOS.pop(r, None)
        return []


class _Resp:
    def __init__(self, data):
        self.data = data


# 🔴 `productos` NO PUEDE VENIR VACIA. Desde el 3-sep-2026 el escaner ABORTA en
# rojo si el catalogo propio no se lee o devuelve 0 filas (era un 'AVISO' y el run
# seguia verde con todo "no propio"). Con el doble tonto devolviendo [] este banco
# ni llegaria al Excel. Dos fichas bastan, y ademas dan las dos etiquetas de la
# columna 'Origen IVA': la del Gizmo esta en el catalogo (-> 'ficha') y las otras
# cuatro del catalogo no (-> 'asumido 21%').
PRODUCTOS = [{'ean': '889698498883', 'asin': 'B0GIZMO', 'iva_pct': '0.2100',
              'stock_moloka': 3, 'stock_fba': 7},
             {'ean': '8499999999999', 'asin': None, 'iva_pct': '0.2100',
              'stock_moloka': 0, 'stock_fba': 0}]
EAN_PROPIO = '0889698498883'


class _Query:
    """Doble TONTO a proposito: acepta cualquier metodo encadenado y devuelve vacio.
    Asi un cambio en el escaner (una tabla nueva, otro filtro) no rompe este banco.
    La UNICA tabla con contenido es `productos`, porque sin ella el escaner aborta."""
    def __init__(self, tabla):
        self.tabla = tabla

    def __getattr__(self, _nombre):
        return lambda *a, **k: self

    def execute(self):
        if self.tabla == 'productos':
            return _Resp(PRODUCTOS)
        return _Resp([{'id': 1}] if self.tabla == 'escaner_resultados' else [])


class _Cliente:
    def __init__(self):
        self.storage = types.SimpleNamespace(from_=lambda _b: _Bucket())

    def table(self, nombre):
        return _Query(nombre)


sys.modules['keepa'] = types.ModuleType('keepa')
sys.modules['keepa'].Keepa = _FakeKeepa
sys.modules['supabase'] = types.ModuleType('supabase')
sys.modules['supabase'].create_client = lambda url, key: _Cliente()

os.environ['KEEPA_API_KEY'] = 'FAKE'
os.environ['SUPABASE_URL'] = 'https://doble.local'
os.environ['SUPABASE_KEY'] = 'FAKE'
for _v in ('SUPABASE_SERVICE_KEY', 'TELEGRAM_TOKEN', 'TELEGRAM_CHAT_ID', 'AUTORELANZAR_MIN'):
    os.environ.pop(_v, None)

import runpy
print('--- corriendo %s de punta a punta (con dobles, sin red) ---' % RUTA)
_ns_run = runpy.run_path(RUTA, run_name='__main__')
print('--- fin del escaner ---')
print()

CALC = {}
for _it in _ns_run['registros']:
    for _dom, _d in (_it.get('_paises_calc') or {}).items():
        CALC[(str(_it['ean']), _dom)] = _d

_xlsx = [r for r in SUBIDOS if r.endswith('.xlsx')]
eq('(C0) el escaner ha subido UN Excel', len(_xlsx), 1)
if not _xlsx:
    print('\n❌ sin Excel no hay nada que cotejar')
    sys.exit(1)

from openpyxl import load_workbook
wb = load_workbook(io.BytesIO(SUBIDOS[_xlsx[0]]), data_only=False)
ws = wb['Análisis']
CAB = [c.value for c in ws[1]]
IDX = {n: i + 1 for i, n in enumerate(CAB)}
LET = {n: ws.cell(row=1, column=i + 1).column_letter for i, n in enumerate(CAB)}
eq('(C0) la cabecera del .xlsx es COLS', CAB, COLS)

# Evaluador ARITMETICO (no eval): solo numeros y + - * / ( ). Cualquier otra cosa
# en la formula revienta aqui en vez de ejecutarse.
_OPS = {ast.Add: lambda a, b: a + b, ast.Sub: lambda a, b: a - b,
        ast.Mult: lambda a, b: a * b, ast.Div: lambda a, b: a / b}
_REF = re.compile(r'\b([A-Z]{1,3})(\d+)\b')


def _aritmetica(expr):
    def ev(n):
        if isinstance(n, ast.Expression):
            return ev(n.body)
        if isinstance(n, ast.Constant) and isinstance(n.value, (int, float)):
            return n.value
        if isinstance(n, ast.UnaryOp) and isinstance(n.op, (ast.UAdd, ast.USub)):
            return ev(n.operand) if isinstance(n.op, ast.UAdd) else -ev(n.operand)
        if isinstance(n, ast.BinOp) and type(n.op) in _OPS:
            return _OPS[type(n.op)](ev(n.left), ev(n.right))
        raise ValueError('la formula tiene algo que no es aritmetica: ' + ast.dump(n))
    return ev(ast.parse(expr, mode='eval'))


def valor(ref, pila=()):
    """Resuelve una celda; si lleva formula, la EVALUA tal cual la vera Excel."""
    if ref in pila:
        raise RuntimeError('referencia circular en ' + ref)
    v = ws[ref].value
    if isinstance(v, str) and v.startswith('='):
        return _aritmetica(_REF.sub(lambda m: '(%r)' % valor(m.group(0), pila + (ref,)), v[1:]))
    return v


def celda(nombre, fila):
    return valor('%s%d' % (LET[nombre], fila))


_cotejadas, _es, _fr, _canon = 0, 0, 0, 0
_peor_ben = _peor_mar = _peor_roi = 0.0
for _r in range(2, ws.max_row + 1):
    _ean = str(ws.cell(row=_r, column=IDX['EAN']).value or '')
    _pais = ws.cell(row=_r, column=IDX['País']).value
    _d = CALC.get((_ean, _pais))
    if not _d or _d.get('beneficio') is None:
        continue
    _cotejadas += 1
    _es += (_pais == 'ES')
    _fr += (_pais == 'FR')
    _peor_ben = max(_peor_ben, abs(celda('Beneficio (€)', _r) - _d['beneficio']))
    _peor_mar = max(_peor_mar, abs(celda('Margen', _r) - _d['margen']))
    _peor_roi = max(_peor_roi, abs(celda('ROI', _r) - _d['roi']))
    if _ean == '0889698000004' and _pais == 'FR':
        _canon = _r

eq('(C1) hay filas ES y FR que cotejar', (_cotejadas >= 6, _es >= 1, _fr >= 1), (True, True, True))
casi('(C1) 🔴 Beneficio: la CELDA y _paises_calc, al centimo (peor fila)', _peor_ben, 0.0, 0.005)
casi('(C1) 🔴 Margen: la CELDA y _paises_calc (peor fila)', _peor_mar, 0.0, 5e-5)
casi('(C1) 🔴 ROI: la CELDA y _paises_calc (peor fila)', _peor_roi, 0.0, 5e-5)

# (C2) El caso canonico frances, LEIDO DE LA CELDA del Excel generado.
eq('(C2) la fila del caso canonico esta en el Excel', _canon > 0, True)
if _canon:
    _pr = celda('Precio venta (€)', _canon)
    _fee = celda('Fee Logística (€)', _canon)
    _com_cel = celda('Com. Amazon (€)', _canon)
    _comision = _pr * 15.0 / 100
    casi('(C2) comision de la fila = 1,82 (factura)', round(_comision, 2), 1.82, 0)
    casi('(C2) fee FBA de la fila = 5,29 (factura)', round(_fee, 2), 5.29, 0)
    casi('(C2) 🔴 ISD que sale de las celdas = 0,21 (factura 404-7912092-2024339)',
         round(_com_cel - _comision, 2), 0.21, 0)
    casi('(C2) y la columna ISD s/ Fee Log. lleva el 3% de la fee',
         celda(ISD, _canon), _fee * 0.03, 1e-12)

# (C4) 🔴 SIN TARIFA FBA, LA HOJA NO ECHA LA CUENTA. `Celda 8` exige `fee is not
# None` para calcular, asi que deja el pais en 'Sin datos'; la hoja pedia menos y
# escribia igual una formula de Beneficio con la celda de fee VACIA leida como 0.
# Dos cuentas en la misma fila, por el otro lado. Aqui se comprueba que callan las dos.
_sin_fee = [_r for _r in range(2, ws.max_row + 1)
            if str(ws.cell(row=_r, column=IDX['EAN']).value or '') == '0889698000005'
            and ws.cell(row=_r, column=IDX['País']).value == 'ES']
eq('(C4) la fila sin tarifa FBA esta en el Excel', len(_sin_fee), 1)
if _sin_fee:
    _r = _sin_fee[0]
    eq('(C4) Keepa no dio fee -> la celda de Fee Logística esta vacia',
       ws.cell(row=_r, column=IDX['Fee Logística (€)']).value, None)
    eq('(C4) y _paises_calc no calculo ese pais',
       CALC[('0889698000005', 'ES')].get('beneficio'), None)
    eq("(C4) la columna Decisión dice 'Sin datos'",
       ws.cell(row=_r, column=IDX['Decisión']).value, 'Sin datos')
    eq('(C4) 🔴 y la hoja NO escribe Beneficio / ROI / Margen',
       [ws.cell(row=_r, column=IDX[c]).value for c in ('Beneficio (€)', 'ROI', 'Margen')],
       [None, None, None])
    # La otra mitad del ancla: los paises del MISMO producto que si tienen fee si
    # se calculan. Si no, este caso pasaria por estar el producto entero fuera.
    _con_fee = [_r2 for _r2 in range(2, ws.max_row + 1)
                if str(ws.cell(row=_r2, column=IDX['EAN']).value or '') == '0889698000005'
                and ws.cell(row=_r2, column=IDX['País']).value in ('IT', 'FR')]
    eq('(C4) los paises del mismo producto que SI traen fee siguen calculandose',
       [ws.cell(row=_r2, column=IDX['Beneficio (€)']).value is not None for _r2 in _con_fee],
       [True, True])

# (C3) Fuera de Francia la columna nueva es 0 (no un hueco), y la comision no cambia.
_ceros = [celda(ISD, _r) for _r in range(2, ws.max_row + 1)
          if ws.cell(row=_r, column=IDX['País']).value in ('ES', 'IT')]
eq('(C3) en ES/IT la columna del ISD es 0 en todas las filas', set(_ceros), {0})

# (C5) La columna 'Origen IVA', leida del .xlsx real. El detalle vive en
# test_escaner_catalogo_propio.py; aqui se comprueba que sigue siendo la ULTIMA y
# que distingue la ficha del 21% asumido en la MISMA hoja que se acaba de cotejar.
eq('(C5) 🔴 "Origen IVA" es la ultima columna del .xlsx', CAB[-1], NUEVA)
_origen_es = {str(ws.cell(row=_r, column=IDX['EAN']).value or ''):
              ws.cell(row=_r, column=IDX[NUEVA]).value
              for _r in range(2, ws.max_row + 1)
              if ws.cell(row=_r, column=IDX['País']).value == 'ES'}
eq('(C5) la ficha que esta en productos dice "ficha"', _origen_es.get(EAN_PROPIO), 'ficha')
eq('(C5) 🔴 y las que no estan dicen "asumido 21%"',
   sorted({v for k, v in _origen_es.items() if k != EAN_PROPIO}), ['asumido 21%'])

print()
if fallos:
    print('❌ %d FALLOS: %s' % (len(fallos), ', '.join(fallos)))
    sys.exit(1)
print('✅ TODO OK (%d filas del Excel cotejadas contra _paises_calc)' % _cotejadas)
