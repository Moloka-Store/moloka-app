# ============================================================
# PRUEBA: descargar el catalogo de TCG por SERVIDOR (sin Chrome)
# ------------------------------------------------------------
# TCG Factory es PrestaShop. Login estandar por formulario + cookie de sesion,
# y luego GET al boton "Descargar Catalogo en Excel" (modulo smcatalog).
# Mismo enfoque que BEMS: curl_cffi con impersonate (por si hay Cloudflare).
#
# Credenciales: SOLO en GitHub Secrets (TCG_USER / TCG_PASS). NUNCA en el codigo.
#
# Esto es una PRUEBA: solo descarga y verifica que el Excel es valido. No sube a
# Supabase ni escanea. Si sale OK, lo integramos al director (descarga nocturna).
# ============================================================
import os, sys, io
from curl_cffi import requests as cr
from bs4 import BeautifulSoup
import openpyxl

BASE = 'https://tcgfactory.com'
LOGIN_URLS = [f'{BASE}/es/iniciar-sesion', f'{BASE}/es/mi-cuenta']
DOWNLOAD_URL = f'{BASE}/es/module/smcatalog/downloadcatalog?format=excel'

USER = os.environ.get('TCG_USER')
PASS = os.environ.get('TCG_PASS')
if not USER or not PASS:
    print("ERROR: faltan los secrets TCG_USER / TCG_PASS.")
    sys.exit(1)

s = cr.Session(impersonate='chrome120')

# 1) Abrir la pagina de login -> cookies + token CSRF del formulario
print(">>> Abriendo pagina de login...", flush=True)
form = None
for url in LOGIN_URLS:
    try:
        r = s.get(url, timeout=60)
        print(f"   {url} -> {r.status_code} ({len(r.text)} bytes)", flush=True)
        soup = BeautifulSoup(r.text, 'html.parser')
        for f in soup.find_all('form'):
            if f.find('input', {'type': 'password'}):
                form = f; break
        if form:
            print(f"   formulario de login encontrado en {url}", flush=True)
            break
    except Exception as e:
        print(f"   AVISO {url}: {e}", flush=True)

if not form:
    print("ERROR: no encuentro el formulario de login (¿Cloudflare, captcha o URL distinta?).")
    sys.exit(1)

action = form.get('action') or LOGIN_URLS[0]
if action.startswith('/'):
    action = BASE + action
elif not action.startswith('http'):
    action = f'{BASE}/es/{action}'

# Recoger TODOS los campos del form (incluido el token CSRF oculto) y rellenar credenciales
data = {}
for inp in form.find_all('input'):
    name = inp.get('name')
    if name:
        data[name] = inp.get('value', '')
data['email'] = USER
data['password'] = PASS
data['submitLogin'] = '1'
print(f"   form action: {action} | campos: {list(data.keys())}", flush=True)

# 2) Enviar el login
print(">>> Enviando login...", flush=True)
r2 = s.post(action, data=data, timeout=60, allow_redirects=True)
print(f"   login POST -> {r2.status_code}", flush=True)

# 3) Verificar que entro (PrestaShop muestra 'Cerrar sesion' / 'logout' cuando hay sesion)
def logueado(html):
    h = html.lower()
    return ('cerrar sesión' in h) or ('cerrar sesion' in h) or ('logout' in h)

ok = logueado(r2.text)
if not ok:
    rc = s.get(f'{BASE}/es/mi-cuenta', timeout=60)
    ok = logueado(rc.text)
print(f"   ¿login OK?: {ok}", flush=True)
if not ok:
    print("ERROR: el login NO entro. Causas posibles: credenciales mal, captcha, o el")
    print("       token del formulario no es el que esperaba. Cuentame que sale arriba.")
    sys.exit(1)

# 4) Descargar el catalogo
print(">>> Descargando catalogo (boton Excel)...", flush=True)
r3 = s.get(DOWNLOAD_URL, timeout=180)
contenido = r3.content
print(f"   descarga -> {r3.status_code} | {len(contenido)} bytes | "
      f"content-type: {r3.headers.get('content-type')}", flush=True)

# 5) Verificar que es un Excel de verdad (un .xlsx empieza por 'PK', es un ZIP)
if contenido[:2] != b'PK':
    print("ERROR: lo descargado NO es un .xlsx (no empieza por 'PK').")
    print("       Primeros bytes:", contenido[:30])
    print("       Probablemente nos devolvio una pagina HTML (login caducado o error).")
    sys.exit(1)

try:
    wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    print(f"   ✅ Excel VALIDO: hoja '{wb.sheetnames[0]}', ~{ws.max_row} filas | "
          f"hojas: {wb.sheetnames}", flush=True)
except Exception as e:
    print("ERROR abriendo el Excel descargado:", e)
    sys.exit(1)

with open('catalogo_tcg.xlsx', 'wb') as fp:
    fp.write(contenido)
print(">>> PRUEBA OK: TCG se baja por servidor, sin Chrome. 🎉", flush=True)
