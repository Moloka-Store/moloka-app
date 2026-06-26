# ============================================================================
# ROBOT RANK WEB  -  Fase 1 pura para la web (NO toca el escaner de FBA)
# ----------------------------------------------------------------------------
# Objetivo: pasar los Funkos EN STOCK del catalogo de TCG por Keepa SOLO para
# sacar el RANK de Espana (Fase 1, ~1 token/producto). Sin Fase 2, sin IT/FR.
# Reutiliza VERBATIM el motor de Fase 1 del escaner (keepa_query, helpers EAN,
# registra/pasada, extraccion de rank). Salida: un Excel de ranking en Storage
# (informes/web_rank/ranking_tcg.xlsx) + la DISTRIBUCION de ranks en el log,
# para cortar el umbral viendo la curva.
#
# Entrada:  informes/web_rank/catalogo.xlsx   (el Excel de TCG, subido a Storage)
# Recado opcional: informes/web_rank/_solicitud.json -> {"rank_maximo": 150000}
# Secrets: KEEPA_API_KEY, SUPABASE_URL, SUPABASE_KEY
# ============================================================================
import os, io, json, time, datetime, sys
import keepa
from supabase import create_client
import openpyxl

sys.stdout.reconfigure(line_buffering=True)

BUCKET   = 'informes'
CARPETA  = 'web_rank'
CAT_PATH = f'{CARPETA}/catalogo.xlsx'
RECADO   = f'{CARPETA}/_solicitud.json'
OUT_PATH = f'{CARPETA}/ranking_tcg.xlsx'

api = keepa.Keepa(os.environ['KEEPA_API_KEY'])
sb  = create_client(os.environ['SUPABASE_URL'], os.environ['SUPABASE_KEY'])
print(">>> Tokens Keepa AHORA:", api.tokens_left, flush=True)

# ---- Recado (rank maximo; generoso por defecto para recolectar amplio) ----
RANK_MAXIMO = 150000
try:
    crudo = sb.storage.from_(BUCKET).download(RECADO)
    RANK_MAXIMO = int(json.loads(crudo.decode('utf-8')).get('rank_maximo') or RANK_MAXIMO)
except Exception:
    pass
print(f">>> RANK_MAXIMO (corte para marcar 'pasa'): {RANK_MAXIMO}")

# ===================== Keepa con reintentos (VERBATIM del escaner) =====================
KEEPA_MAX_INTENTOS = 4
KEEPA_ESPERAS = [5, 15, 40, 90]
def keepa_query(items, **kwargs):
    kwargs.setdefault('progress_bar', False)
    for intento in range(KEEPA_MAX_INTENTOS):
        try:
            return api.query(items, **kwargs) or []
        except Exception as ex:
            if intento < KEEPA_MAX_INTENTOS - 1:
                espera = KEEPA_ESPERAS[intento]
                print(f"  [Keepa] intento {intento+1}/{KEEPA_MAX_INTENTOS} fallo: {ex} -> reintento en {espera}s")
                time.sleep(espera)
            else:
                print(f"  [Keepa] AGOTADOS {KEEPA_MAX_INTENTOS} intentos: {ex} -> se salta")
    return None

# ===================== Helpers EAN (VERBATIM del escaner) =====================
def core_ean(e):
    e = str(e).strip().upper()
    return e[:-1] if e.endswith('C') else e
def variantes_ean(core):
    c, vs = core.strip(), set()
    if c.isdigit():
        vs.add(c); vs.add(c.lstrip('0'))
        if len(c)==12: vs.add('0'+c)
        if len(c)==13 and c.startswith('0'): vs.add(c[1:])
    return [v for v in vs if v]
def norm(code): return str(code).strip().lstrip('0')

# ===================== Filtro de entrada (VERIFICADO) + 5 categorias =====================
# Tipos que entran. Mystery Minis FUERA. Pines/camisetas/merch/Dorbz fuera (no estan).
TIPOS_OK = {
    'Merchandising-Funko POP!',
    'Merchandising-Funko Bitty POP',
    'Merchandising-Funko Pop! Keychain',
}
# GRANDES = NO enviables (10"/25cm/jumbo). Agresivo: ante la duda, fuera (logistica).
GRANDES = ['10"', '10 "', "10''", '10\u201d', '25 cm', '25cm', 'jumbo', 'super sized', 'super-sized', 'mega ']
# BASURA = no es figura suelta (packs, figura+camiseta, cajas mayoristas de 12, sorpresas).
BASURA  = ['12 unidades', '(12', 'display', '& tee', 'tee!', ' pack', 'pack ', '4pk',
           '4 pack', '3 pack', '2 pack', '3-pack', 'multipack', 'surprise', 'box set',
           'choc)', 'holiday tree', 'tree holiday', 'mystery pocket']
def es_grande(cab):
    c = cab.lower(); return any(k in c for k in GRANDES)
def es_basura(cab):
    c = cab.lower(); return any(k in c for k in BASURA)
# 5 CATEGORIAS para la web: Funko Pop! / Deluxe (incl. 6") / Diorama (Moment+Ride+Town)
# / Bitty Pop (incl. Pocket) / Llavero. Diorama manda sobre Deluxe.
def formato_de(tipo, cab):
    c = cab.lower()
    if 'Keychain' in tipo: return 'Llavero'
    if 'Bitty' in tipo or 'bitty' in c: return 'Bitty Pop'
    if 'pocket' in c: return 'Bitty Pop'
    if ' ride' in c or ' town' in c or 'moment' in c: return 'Diorama'
    if 'deluxe' in c: return 'Deluxe'
    if '6"' in c or '6 "' in c or '15 cm' in c or '15cm' in c or 'oversized' in c: return 'Deluxe'
    return 'Funko Pop!'

# ===================== Cargar catalogo TCG (Funkos enviables en stock) =====================
def cargar_tcg():
    data = sb.storage.from_(BUCKET).download(CAT_PATH)
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else '' for c in rows[0]]
    idx = {h: i for i, h in enumerate(hdr)}
    iE, iS, iP, iV = idx['EAN'], idx['Stock'], idx['Precio'], idx['PVPR']
    iTipo, iEstado, iCab = idx['Tipo Producto'], idx['Estado producto'], idx['Cabecera']
    def num(x):
        try: return float(str(x).replace(',', '.'))
        except Exception: return 0.0
    ESTADOS_OK = {'disponible', 'oferta'}   # 48h (no backorder/preorder)
    filas, vistos = [], set()
    f_tipo = f_grande = f_basura = 0
    for r in rows[1:]:
        tipo = str(r[iTipo] or '').strip()
        cab = str(r[iCab] or '')
        if tipo not in TIPOS_OK:
            f_tipo += 1; continue
        if 'Funko POP!' in tipo and es_grande(cab):
            f_grande += 1; continue
        if es_basura(cab):
            f_basura += 1; continue
        if str(r[iEstado] or '').strip().lower() not in ESTADOS_OK:
            continue
        if num(r[iS]) < 1:
            continue
        ean_in = str(r[iE] or '').strip()
        if not ean_in or ean_in in vistos:
            continue
        vistos.add(ean_in)
        core = core_ean(ean_in)
        filas.append({'ean_in': ean_in, 'core': core, 'variantes': variantes_ean(core),
                      'nombre': cab, 'pa': num(r[iP]), 'pvpr': num(r[iV]),
                      'formato': formato_de(tipo, cab)})
    print(f">>> Filtro entrada: fuera tipo {f_tipo}, fuera grande {f_grande}, fuera basura {f_basura} | ENTRAN {len(filas)}")
    return filas

# ---- Dedup contra el fisico ya en web (no re-rankear lo que ya tienes) ----
def eans_fisicos():
    s = set()
    try:
        d = sb.table('web_productos').select('ean').eq('origen', 'fabrica').execute().data or []
        s = {norm(x['ean']) for x in d if x.get('ean')}
    except Exception as e:
        print("  (aviso: no pude leer fisico para dedup:", e, ")")
    return s

# ===================== FASE 1 (VERBATIM del escaner, sin 'propio') =====================
IDX_RANK = 3
LOTE_FASE1 = 100

filas = cargar_tcg()
fisicos = eans_fisicos()
filas = [f for f in filas if norm(f['core']) not in fisicos]
print(f">>> Funkos en stock a rankear (tras dedup fisico): {len(filas)}")

candidatos, ambiguos = {}, []
var_norm = {f['ean_in']: {norm(v) for v in f['variantes']} for f in filas}
fila_por_ean = {f['ean_in']: f for f in filas}
def cod_pref(f): return f['core']
def cods_reserva(f): return [v for v in f['variantes'] if v != cod_pref(f)]
def keyrank(c): return c['r_90'] if c['r_90'] and c['r_90']>0 else 10**12

def registra(prod, pool, vistos):
    asin = prod.get('asin')
    if not asin: return
    st = prod.get('stats') or {}
    cur, a90 = st.get('current') or [], st.get('avg90') or []
    r_act = cur[IDX_RANK] if len(cur)>IDX_RANK else -1
    r_90  = a90[IDX_RANK] if len(a90)>IDX_RANK else -1
    eans = {norm(str(e)) for e in (prod.get('eanList') or [])+(prod.get('upcList') or [])}
    for ein in pool:
        if not (var_norm[ein] & eans): continue
        cand = {'ean_in':ein,'asin':asin,'r_act':r_act,'r_90':r_90,'fila':fila_por_ean[ein]}
        if ein in candidatos:
            prev = candidatos[ein]
            ambiguos.append(ein)
            if keyrank(cand)<keyrank(prev): candidatos[ein]=cand
        else: candidatos[ein]=cand
        vistos.add(ein)

def pasada(cod_por_ean, etiqueta):
    pool = list(cod_por_ean.keys())
    codigos = sorted({cod_por_ean[e] for e in pool})
    lotes = [codigos[i:i+LOTE_FASE1] for i in range(0,len(codigos),LOTE_FASE1)]
    vistos = set()
    print(f"{etiqueta}: {len(pool)} productos, {len(codigos)} codigos, {len(lotes)} lotes")
    for n,lote in enumerate(lotes,1):
        prods = keepa_query(lote, product_code_is_asin=False, domain='ES', stats=90, history=0)
        if prods is None:
            # Un lote entero falla casi siempre por UN codigo malo que envenena la
            # peticion de 100. En vez de perder los 99 buenos, lo reintentamos
            # PARTIDO (en trozos de 25, y si un trozo tambien falla, de uno en uno)
            # para salvar los buenos y aislar solo el malo.
            print(f"  lote {n}/{len(lotes)} fallo entero -> reintento PARTIDO (para no perder los buenos)")
            salvados = 0
            for j in range(0, len(lote), 25):
                trozo = lote[j:j+25]
                p2 = keepa_query(trozo, product_code_is_asin=False, domain='ES', stats=90, history=0)
                if p2 is not None:
                    for prod in p2: registra(prod, pool, vistos)
                    salvados += len(p2)
                    continue
                # el trozo de 25 tambien falla -> de uno en uno para aislar el malo
                for cod in trozo:
                    p3 = keepa_query([cod], product_code_is_asin=False, domain='ES', stats=90, history=0)
                    if p3 is None:
                        print(f"      codigo malo aislado, se salta: {cod}")
                    else:
                        for prod in p3: registra(prod, pool, vistos)
                        salvados += len(p3)
            print(f"  lote {n}/{len(lotes)} (partido) | salvados {salvados} | tokens {api.tokens_left}")
            continue
        for prod in prods: registra(prod, pool, vistos)
        print(f"  lote {n}/{len(lotes)} | tokens {api.tokens_left}")
    return vistos

if filas:
    vistos = pasada({f['ean_in']: cod_pref(f) for f in filas}, "Fase 1 (1 codigo/producto)")
    for ronda in (0, 1):
        faltan = {f['ean_in'] for f in filas} - vistos
        rint = {}
        for f in filas:
            if f['ean_in'] in faltan:
                rs = cods_reserva(f)
                if len(rs) > ronda: rint[f['ean_in']] = rs[ronda]
        if rint:
            vistos |= pasada(rint, f"Fase 1 reintento {ronda+1} (variante)")

# ===================== Salida: Excel de ranking + distribucion =====================
def mejor_rank(c):
    rs = [r for r in (c['r_act'], c['r_90']) if r and r > 0]
    return min(rs) if rs else None

filas_out = []
for ein, c in candidatos.items():
    f = c['fila']; mr = mejor_rank(c)
    pasa = (mr is not None and mr <= RANK_MAXIMO)
    filas_out.append({'ean': ein, 'nombre': f['nombre'], 'formato': f.get('formato',''), 'rank_act': c['r_act'] if c['r_act']>0 else None,
                      'rank_90': c['r_90'] if c['r_90']>0 else None, 'mejor_rank': mr,
                      'pa': f['pa'], 'pvpr': f['pvpr'], 'asin': c['asin'], 'pasa': 'SI' if pasa else 'no'})
filas_out.sort(key=lambda x: (x['mejor_rank'] is None, x['mejor_rank'] or 10**12))

# Distribucion (la curva, para cortar el umbral)
cortes = [10000, 30000, 50000, 100000, 200000, 500000]
dist = {c: 0 for c in cortes}; sin_rank = 0; total_pasa = 0
for x in filas_out:
    mr = x['mejor_rank']
    if mr is None: sin_rank += 1; continue
    for c in cortes:
        if mr <= c: dist[c] += 1
    if x['pasa'] == 'SI': total_pasa += 1
print("\n===== DISTRIBUCION DE RANK (acumulada) =====")
for c in cortes:
    print(f"  rank <= {c:>7}: {dist[c]:>5} Funkos")
print(f"  sin rank        : {sin_rank:>5}")
print(f"  >>> PASAN con RANK_MAXIMO={RANK_MAXIMO}: {total_pasa}  (de {len(filas_out)} con ASIN)")

# Excel a Storage
wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Ranking'
COLS = ['EAN','Nombre','Formato','Rank actual','Rank 90d','Mejor rank','PA (coste)','PVPR','ASIN','Pasa']
ws.append(COLS)
for x in filas_out:
    ws.append([x['ean'], x['nombre'], x.get('formato',''), x['rank_act'], x['rank_90'], x['mejor_rank'],
               x['pa'], x['pvpr'], x['asin'], x['pasa']])
buf = io.BytesIO(); wb.save(buf); buf.seek(0)
try:
    sb.storage.from_(BUCKET).upload(OUT_PATH, buf.read(),
        {'content-type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet','upsert':'true'})
    print(f"\nRanking guardado en Storage: {OUT_PATH}")
except Exception as e:
    print("ATENCION: no pude subir el Excel de ranking:", e)
print("Fin.")
