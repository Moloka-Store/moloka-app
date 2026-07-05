# ==========================================================================
# MOLOKA · Procesar el informe de resultados de Miravia
# --------------------------------------------------------------------------
# Tras subir el Excel a Miravia, Miravia da un informe ("Detalle errores...").
# El generador marcó como subidos TODOS los que metió en el Excel. Este script
# lee ese informe y DESMARCA (miravia_subido = NULL) los que fallaron o quedaron
# en borrador, para que el próximo Excel los reintente. Los que Miravia reporta
# como "ya existía" se dejan como subidos (esos SÍ están en Miravia).
#
# Lee el informe del buzón informes/miravia_resultado/ (lo deja ahí el botón
# de la app). Para GitHub Actions.
# ==========================================================================
import os, sys
import openpyxl

sys.stdout.reconfigure(line_buffering=True)

HOJA   = 'Juguetesyfigurascolecciona'
BUCKET = 'informes'
BUZON  = 'miravia_resultado'

# Si el mensaje de Miravia contiene alguno de estos, el producto YA EXISTE en
# Miravia (no es un fallo): se queda marcado como subido, NO se reintenta.
_YA_SUBIDO = ('REPETITIVE', 'YA EXISTE', 'INFORMACIÓN ES LA MISMA',
              'INFORMACION ES LA MISMA', 'MISMA DE UN PRODUCTO', 'ALREADY EXIST')


def analizar_informe(ruta):
    """Devuelve (desmarcar, ya_subidos).
       desmarcar  = [(slug, ean, nombre, motivo)]  -> fallo/borrador, reintentar (NULL)
       ya_subidos = [(slug, ean, nombre)]          -> 'ya existía', se queda subido
    Testeable sin Supabase."""
    wb = openpyxl.load_workbook(ruta, data_only=True)
    if HOJA not in wb.sheetnames:
        print(f"   (aviso: el informe no tiene la hoja {HOJA})")
        return [], []
    ws = wb[HOJA]
    desmarcar, ya = [], []
    for r in range(5, ws.max_row + 1):
        msg = ws.cell(r, 1).value          # col 1 = "Mensaje de error" (la inserta Miravia)
        if not msg:
            continue
        ean    = ws.cell(r, 35).value      # col 35 = EAN (34 en el Excel de subida + 1 por el mensaje)
        slug   = ws.cell(r, 36).value      # col 36 = SKU de vendedor = slug
        nombre = ws.cell(r, 4).value       # col 4  = nombre del producto
        M = str(msg).upper()
        if any(k in M for k in _YA_SUBIDO):
            ya.append((slug, ean, nombre))
        else:
            desmarcar.append((slug, ean, nombre, str(msg).strip()[:90]))
    return desmarcar, ya


def _bajar_ultimo_informe(sb):
    """Baja el .xlsm más reciente del buzón; devuelve (ruta_local, nombre_remoto) o (None, None)."""
    objs = sb.storage.from_(BUCKET).list(BUZON) or []
    xlsm = [o for o in objs if o['name'].lower().endswith('.xlsm')]
    if not xlsm:
        return None, None
    xlsm.sort(key=lambda o: o.get('updated_at') or o.get('created_at') or '', reverse=True)
    nombre = xlsm[0]['name']
    data = sb.storage.from_(BUCKET).download(f"{BUZON}/{nombre}")
    ruta = f"/tmp/{nombre}"
    with open(ruta, 'wb') as f:
        f.write(data)
    return ruta, nombre


def main():
    from supabase import create_client
    sb = create_client(os.environ['SUPABASE_URL'], os.environ['SUPABASE_KEY'])

    ruta, nombre = _bajar_ultimo_informe(sb)
    if not ruta:
        print(">>> No hay informe en el buzón informes/miravia_resultado/. Nada que hacer.")
        return
    print(f">>> Informe: {nombre}")

    desmarcar, ya = analizar_informe(ruta)
    print(f">>> Ya existían en Miravia (se quedan subidos): {len(ya)}")
    for s, e, n in ya:
        print(f"    · {n}  (slug={s})")
    print(f">>> Fallidos/borrador (se desmarcan para reintentar): {len(desmarcar)}")

    n_ok = 0
    for slug, ean, nombre_p, motivo in desmarcar:
        print(f"    · {nombre_p}  ->  {motivo}")
        try:
            done = False
            if slug:
                r = sb.table('web_productos').update({'miravia_subido': None}).eq('slug', slug).execute()
                done = bool(r.data)
            if not done and ean:                      # fallback por EAN (con y sin ceros a la izq.)
                e = str(ean).strip()
                r = sb.table('web_productos').update({'miravia_subido': None}).eq('ean', e).execute()
                if not r.data:
                    sb.table('web_productos').update({'miravia_subido': None}).eq('ean', e.lstrip('0')).execute()
            n_ok += 1
        except Exception as ex:
            print(f"      (aviso: no pude desmarcar {slug or ean}: {ex})")

    # Limpiar el buzón (borrado verificado, patrón de la casa)
    try:
        sb.storage.from_(BUCKET).remove([f"{BUZON}/{nombre}"])
        print(f">>> Buzón limpiado: {nombre}")
    except Exception as ex:
        print(f"   (aviso: no pude limpiar el buzón: {ex})")

    print(f"\n>>> LISTO. Desmarcados {n_ok}/{len(desmarcar)} para reintento. "
          f"{len(ya)} se quedan como subidos.")


if __name__ == "__main__":
    main()
