# robot_fotos_tcg.py - Imagenes Keepa (alta resolucion) para la linea TCG.
#
# Lee el ranking ya calculado (ranking_tcg.xlsx), coge SOLO los del corte de rank
# (los que vamos a publicar) y pide a Keepa POR ASIN su imagen 1600px + nombre.
# ~1 token por producto.
#
# REANUDABLE Y SIN UMBRAL: guarda el progreso despues de CADA tanda. Si te quedas
# sin tokens a mitad, lo sacado queda guardado; al relanzar continua por donde iba
# saltando los que ya tienen foto. Nunca se tira un token.
#
# Salida: ranking_tcg_fotos.xlsx (el corte, con Nombre Keepa / Img figura / Img caja).

import os, io, sys, re, time
import keepa
from supabase import create_client
import openpyxl

sys.stdout.reconfigure(line_buffering=True)

BUCKET    = 'informes'
CARPETA   = 'web_rank'
RANK_PATH = f'{CARPETA}/ranking_tcg.xlsx'
OUT_PATH  = f'{CARPETA}/ranking_tcg_fotos.xlsx'

# Corte fijo a 30000 a proposito (los "~580" que acordamos). Sube este numero a mano
# si algun dia quieres ampliar el corte. No se lee del recado para no acoplarlo al
# robot de rank (que usa rank_maximo=100000 para otra cosa).
CORTE = 30000
LOTE  = 50      # tanda pequena = guarda mas a menudo = mas seguro ante cortes

api = keepa.Keepa(os.environ['KEEPA_API_KEY'])
sb  = create_client(os.environ['SUPABASE_URL'], os.environ['SUPABASE_KEY'])

COLS = ['EAN','Nombre','Formato','Rank actual','Rank 90d','Mejor rank','PA (coste)','PVPR','ASIN','Pasa','Nombre Keepa','Img figura','Img caja']

# ---- Imagen Keepa en alta resolucion (verbatim de la fabrica) ----
def _nombre_el(el):
    if isinstance(el, dict):
        for k in ('l','large','hiRes','m','medium','image','name'):
            if el.get(k): return str(el[k])
    elif isinstance(el, str): return el
    return None
def _a_url(n):
    n=str(n); return n if n.startswith('http') else 'https://m.media-amazon.com/images/I/'+n
def _alta(url, px=1600):
    if not url or '/images/I/' not in url: return url
    base,_,fich=url.rpartition('/'); m=re.match(r'^([^.]+)\.',fich)
    return f"{base}/{m.group(1)}._SL{px}_.jpg" if m else url
def extraer_imagenes(prod, max_fotos=8):
    urls=[]
    for el in (prod.get('images') or []):
        n=_nombre_el(el)
        if n:
            u=_alta(_a_url(n))
            if u and u not in urls: urls.append(u)
    return urls[:max_fotos]

# ---- Keepa con reintentos ----
KEEPA_MAX_INTENTOS = 4
KEEPA_ESPERAS = [5, 15, 40, 90]
def keepa_query(items, **kwargs):
    kwargs.setdefault('progress_bar', False)
    for intento in range(KEEPA_MAX_INTENTOS):
        try:
            return api.query(items, **kwargs) or []
        except Exception as ex:
            msg = str(ex)[:60]
            if intento < KEEPA_MAX_INTENTOS - 1:
                print(f"  [Keepa] intento {intento+1}/{KEEPA_MAX_INTENTOS} fallo: {msg} -> reintento en {KEEPA_ESPERAS[intento]}s")
                time.sleep(KEEPA_ESPERAS[intento])
            else:
                print(f"  [Keepa] AGOTADOS {KEEPA_MAX_INTENTOS} intentos: {msg} -> se salta tanda")
                return None

def to_int(v):
    try: return int(v) if v not in (None,'') else None
    except Exception: return None

# ---- 1) Cargar ranking guardado ----
try:
    crudo = sb.storage.from_(BUCKET).download(RANK_PATH)
except Exception as e:
    print(f"!!! No encuentro {RANK_PATH} en Supabase: {e}")
    print("!!! Hace falta el ranking calculado antes de las fotos. Corre el Paso 1 (rank) primero.")
    sys.exit(1)

ws = openpyxl.load_workbook(io.BytesIO(crudo), read_only=True, data_only=True).active
filas = list(ws.iter_rows(values_only=True))
if not filas:
    print("!!! El ranking esta vacio."); sys.exit(1)
hdr = [str(c).strip() if c else '' for c in filas[0]]
idx = {h:i for i,h in enumerate(hdr)}
need = ['EAN','Nombre','Formato','Mejor rank','ASIN']
faltan = [c for c in need if c not in idx]
if faltan:
    print(f"!!! Al ranking le faltan columnas {faltan}. Hay: {hdr}"); sys.exit(1)

iE,iN,iF,iMR,iA = idx['EAN'],idx['Nombre'],idx['Formato'],idx['Mejor rank'],idx['ASIN']
iRA = idx.get('Rank actual'); iR90 = idx.get('Rank 90d')
iPA = idx.get('PA (coste)'); iPV = idx.get('PVPR'); iPasa = idx.get('Pasa')

# ---- 2) Filtrar al corte ----
objetivo = []
for r in filas[1:]:
    mr = to_int(r[iMR])
    if mr is None or mr > CORTE: continue
    if not r[iA]: continue
    objetivo.append(r)
print(f">>> Ranking: {len(filas)-1} filas | dentro del corte <= {CORTE} con ASIN: {len(objetivo)}")
if not objetivo:
    print("!!! Nada dentro del corte."); sys.exit(1)

# ---- 3) REANUDAR: cargar fotos ya hechas de una corrida anterior ----
por_asin = {}   # asin -> {nombre_keepa, img_figura, img_caja}
try:
    prev = sb.storage.from_(BUCKET).download(OUT_PATH)
    wsp = openpyxl.load_workbook(io.BytesIO(prev), read_only=True, data_only=True).active
    pf = list(wsp.iter_rows(values_only=True))
    ph = {str(c).strip() if c else '': i for i,c in enumerate(pf[0])}
    pA, pNk, pIf, pIc = ph.get('ASIN'), ph.get('Nombre Keepa'), ph.get('Img figura'), ph.get('Img caja')
    for r in pf[1:]:
        a = str(r[pA]).strip() if pA is not None and r[pA] else ''
        figura = r[pIf] if pIf is not None else ''
        if a and figura:    # solo cuenta como "hecho" si tiene foto de figura
            por_asin[a] = {'nombre_keepa': r[pNk] if pNk is not None else '',
                           'img_figura': figura, 'img_caja': r[pIc] if pIc is not None else ''}
    print(f">>> Reanudo: {len(por_asin)} ya tenian foto de una corrida anterior")
except Exception:
    print(">>> Sin corrida previa (empiezo de cero)")

def guarda(prod):
    a = prod.get('asin')
    if not a: return
    fotos = extraer_imagenes(prod)
    por_asin[a] = {
        'nombre_keepa': (prod.get('title') or '').strip(),
        'img_caja':   fotos[0] if fotos else '',
        'img_figura': fotos[1] if len(fotos) > 1 else (fotos[0] if fotos else ''),
    }

def escribir():
    """Reconstruye el Excel completo (los ~580) con las fotos que haya y lo sube."""
    wb = openpyxl.Workbook(); ws2 = wb.active; ws2.title = 'Ranking fotos'
    ws2.append(COLS)
    con = 0
    for r in objetivo:
        a = str(r[iA]).strip()
        k = por_asin.get(a, {})
        if k.get('img_figura'): con += 1
        ws2.append([r[iE], r[iN], r[iF],
                    r[iRA] if iRA is not None else '', r[iR90] if iR90 is not None else '', r[iMR],
                    r[iPA] if iPA is not None else '', r[iPV] if iPV is not None else '', a,
                    r[iPasa] if iPasa is not None else '',
                    k.get('nombre_keepa',''), k.get('img_figura',''), k.get('img_caja','')])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    sb.storage.from_(BUCKET).upload(OUT_PATH, buf.read(),
        {'content-type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet','upsert':'true'})
    return con

# ---- 4) Solo los que faltan, por tandas, GUARDANDO tras cada una ----
pendientes = [str(r[iA]).strip() for r in objetivo if str(r[iA]).strip() not in por_asin]
print(f">>> Pendientes de foto: {len(pendientes)} (sin umbral; gasto lo que haya y guardo cada tanda)")

if not pendientes:
    con = escribir()
    print(f">>> Ya estaban todos. {con}/{len(objetivo)} con foto. Nada que pedir.")
    sys.exit(0)

n_lotes = (len(pendientes)+LOTE-1)//LOTE
for i in range(0, len(pendientes), LOTE):
    lote = pendientes[i:i+LOTE]
    prods = keepa_query(lote, product_code_is_asin=True, domain='ES', stats=90, history=0)
    if prods is None:
        print(f"  tanda {i//LOTE+1}/{n_lotes}: sin respuesta (saldo agotado?) -> guardo lo que hay y sigo")
        escribir()
        continue
    for prod in prods:
        guarda(prod)
    con = escribir()   # guarda progreso tras CADA tanda
    print(f"  tanda {i//LOTE+1}/{n_lotes} | nuevos {len(prods)} | total con foto {con}/{len(objetivo)} | tokens {api.tokens_left}")

con = escribir()
faltan = len(objetivo) - con
print(f"\n>>> FIN. {con}/{len(objetivo)} con foto Keepa | {faltan} sin foto (fallback TCG en Paso 2)")
print(f">>> Tokens al terminar: {api.tokens_left}")
if faltan:
    print(f">>> Quedan {faltan} sin foto. Si fue por saldo, relanza cuando recargue y CONTINUA solo con esos.")
