# -*- coding: utf-8 -*-
"""Banco: si el CATALOGO PROPIO no se lee, el escaner NO se pone verde.

EL FALLO QUE CIERRA (3-sep-2026). La Celda 5 de `moloka_escaner_nube.py` cruza
el catalogo del proveedor con la tabla `productos` (el catalogo propio). Estaba
escrita asi:

    try:  ...leer productos...
    except Exception as ex:  print("AVISO sin cruce Supabase:", ex)

O sea que si la lectura fallaba -- o devolvia 0 filas, que no imprimia NADA --
el escaneo seguia con `sup = {}` y:
  - `es_propio()` decia False para las 473 fichas activas de la casa, asi que
    los propios dejaban de saltarse el filtro de rank;
  - el IVA de ES caia al 21% por defecto en TODAS las filas;
  - la columna 'En mi BD' salia vacia.
Y el run acababa VERDE, con su Excel y su fila en `escaner_resultados`. Un fallo
que se ve igual que un exito no es un fallo: es un silencio (#265, por la otra
puerta).

QUE SE PRUEBA, Y COMO:
  (A) Las funciones que deciden la etiqueta de la columna 'Origen IVA', sacadas
      del fichero real con `ast` (POR ESTRUCTURA, no con un grep) y EJECUTADAS
      con un `_sup` de mentira. El valor del IVA y su etiqueta salen de la MISMA
      funcion: si se calcularan aparte, la hoja podria decir 'ficha' en una fila
      cuyo IVA es el 21% asumido.
  (B) Que las guardas esten PUESTAS: que las dos salidas nuevas vayan por
      `abortar()`, que el `except` que se tragaba el fallo ya no este, y que la
      linea CATALOGO_PROPIO se imprima en las dos ramas.
  (C) DE PUNTA A PUNTA, tres veces, cada una en su PROCESO (el escaner es un
      script: se corre con runpy y hace sys.exit). Con `keepa` y `supabase`
      sustituidos por dobles en memoria, sin red, sin secretos y con perfil
      MIS_COMPRAS (efimero: no toca escaner_memoria):
        1. `productos` LANZA      -> exit 1, cero llamadas a Keepa, cero escrituras.
        2. `productos` da 0 filas -> exit 1, cero llamadas a Keepa, cero escrituras.
        3. normal -> exit 0, Keepa SI se llama, y en el Excel la columna
           'Origen IVA' es la ULTIMA y dice 'ficha' / 'asumido 21%' donde toca.
      El caso 3 es la otra mitad del ancla: sin el, "cero llamadas a Keepa"
      saldria verde tambien si el escaner no llamara a Keepa nunca.

LAS DOS DIRECCIONES. Cada caso se ha visto ROJO desactivando a mano lo que
protege (quitando los dos `abortar()` de la Celda 5, y quitando la columna de
COLS), comprobando antes que el cambio caia donde se pretendia.
"""
import ast
import io
import json
import os
import re
import subprocess
import sys
import tempfile
import types

RUTA = 'moloka_escaner_nube.py'

# ---------------------------------------------------------------------------
# EL CATALOGO DE MENTIRA Y LAS FICHAS PROPIAS DE MENTIRA (los usan padre e hijo)
# ---------------------------------------------------------------------------
CATALOGO = (
    "Nombre,EAN,Precio,Marca,Stock\n"
    "Funko Pop Gizmo,0889698498883,4.50,Funko,12\n"          # propio, IVA 21% de ficha
    "Chai latte polvo,8412345678905,3.10,Otra,8\n"           # propio, IVA 10% de ficha
    "Funko Pop Luke Red 5,0889698851909,5.20,Funko,4\n")     # NO esta en productos
ASIN = {'0889698498883': 'B0GIZMO', '8412345678905': 'B0CHAI', '0889698851909': 'B0LUKE5'}
# `ean` tal cual lo devolveria `productos` (sin el cero de delante, como en la casa)
PRODUCTOS = [
    {'ean': '889698498883', 'asin': 'B0GIZMO', 'iva_pct': '0.2100',
     'stock_moloka': 3, 'stock_fba': 7},
    {'ean': '8412345678905', 'asin': None, 'iva_pct': '0.1000',
     'stock_moloka': 1, 'stock_fba': 0},
    # una ficha mas que NO esta en el catalogo del proveedor: filas=3, propios=2
    {'ean': '8499999999999', 'asin': None, 'iva_pct': '0.2100',
     'stock_moloka': 0, 'stock_fba': 0},
]
EAN_PROPIO = {'0889698498883', '8412345678905'}
EAN_NO_PROPIO = {'0889698851909'}
PRECIOS = {   # asin -> pais -> (precio, ref_pct, fee FBA, rank)
    'B0GIZMO': {'ES': (16.99, 15.0, 3.51, 4200), 'IT': (17.49, 15.0, 3.62, 9100),
                'FR': (18.25, 15.0, 3.72, 6400)},
    'B0CHAI':  {'ES': (12.50, 15.0, 3.10, 22000), 'IT': (12.90, 15.0, 3.20, 41000),
                'FR': (13.10, 15.0, 3.30, 38000)},
    'B0LUKE5': {'ES': (15.50, 15.0, 3.51, 30000), 'IT': (14.90, 15.0, 3.62, 51000),
                'FR': (21.90, 15.0, 5.29, 12500)},
}
RECADO = {"proveedor": "MIS_COMPRAS", "marca": "TODAS", "modo": "todo",
          "rank_maximo": 200000, "incluir_sin_rank": False}
COL_NUEVA = 'Origen IVA'


# ===========================================================================
# EL HIJO: monta los dobles y corre el escaner de punta a punta
# ===========================================================================
def hijo(caso, destino):
    import atexit

    LLAMADAS = {'keepa': 0, 'items': 0}
    ESCRITURAS = []
    SUBIDOS = {}

    def _stats(rank, precio, con_bb):
        cur = [-1] * 20
        cur[1] = int(round(precio * 100))          # NEW
        cur[3] = rank                              # SALES RANK
        if con_bb:
            cur[18] = int(round(precio * 100))     # BUY_BOX_SHIPPING
        a90 = [-1] * 20
        a90[3] = rank
        return cur, a90

    class _FakeKeepa:
        def __init__(self, key, timeout=None):
            self.tokens_left = 1500

        def update_status(self):
            pass                                   # consultar el saldo no gasta tokens

        def query(self, items, **kw):
            LLAMADAS['keepa'] += 1
            LLAMADAS['items'] += len(items)
            self.tokens_left -= len(items)
            if kw.get('product_code_is_asin'):     # Fase 2: un ASIN, un pais
                a = items[0]
                precio, ref, fee, rank = PRECIOS[a][(kw.get('domain') or 'ES').upper()]
                cur, a90 = _stats(rank, precio, True)
                return [{'asin': a, 'title': 'T ' + a,
                         'stats': {'current': cur, 'avg90': a90, 'buyBoxIsFBA': True,
                                   'totalOfferCount': 5,
                                   'buyBoxPrice': int(round(precio * 100))},
                         'referralFeePercentage': ref,
                         'fbaFees': {'pickAndPackFee': int(round(fee * 100))},
                         'monthlySold': 40, 'images': ['x.jpg']}]
            out = []                               # Fase 1: EAN -> ASIN + rank
            for cod in items:
                c13 = str(cod).zfill(13)
                a = ASIN.get(c13)
                if not a:
                    continue
                precio, _r, _f, rank = PRECIOS[a]['ES']
                cur, a90 = _stats(rank, precio, False)
                out.append({'asin': a, 'title': 'T ' + a,
                            'stats': {'current': cur, 'avg90': a90, 'salesRankDrops30': 12},
                            'listedSince': 6000000, 'eanList': [c13], 'upcList': []})
            return out

    BUZON = {'escaner/_solicitud_escaner.json': json.dumps(RECADO).encode('utf-8'),
             'escaner/mini_compras.csv': CATALOGO.encode('utf-8-sig')}

    class _Bucket:
        def list(self, carpeta):
            pre = carpeta.rstrip('/') + '/'
            return [{'name': n} for n in sorted({k[len(pre):].split('/')[0]
                                                 for k in list(BUZON) + list(SUBIDOS)
                                                 if k.startswith(pre)})]

        def download(self, ruta):
            if ruta in BUZON:
                return BUZON[ruta]
            if ruta in SUBIDOS:
                return SUBIDOS[ruta]
            raise Exception('404 ' + ruta)

        def upload(self, ruta, data, opts=None):
            SUBIDOS[ruta] = data
            return {'path': ruta}

        def remove(self, rutas):
            for r in rutas:
                BUZON.pop(r, None)
                SUBIDOS.pop(r, None)
            return []

    class _Resp:
        def __init__(self, data):
            self.data = data

    class _Query:
        """Doble tonto a proposito: acepta cualquier metodo encadenado. Solo hace
        dos cosas de verdad: apuntar las ESCRITURAS y contestar a `productos`
        como diga el caso que se esta probando."""
        def __init__(self, tabla):
            self.tabla = tabla

        def __getattr__(self, nombre):
            if nombre in ('insert', 'upsert', 'update', 'delete'):
                ESCRITURAS.append('%s.%s' % (self.tabla, nombre))
            return lambda *a, **k: self

        def execute(self):
            if self.tabla == 'productos':
                if caso == 'falla':
                    raise Exception('doble: la lectura de productos revienta a proposito')
                if caso == 'vacio':
                    return _Resp([])
                return _Resp(PRODUCTOS)
            return _Resp([{'id': 1}] if self.tabla == 'escaner_resultados' else [])

    class _Cliente:
        def __init__(self):
            self.storage = types.SimpleNamespace(from_=lambda _b: _Bucket())

        def table(self, nombre):
            return _Query(nombre)

    sys.modules['keepa'] = types.ModuleType('keepa')
    sys.modules['keepa'].Keepa = _FakeKeepa
    sys.modules['supabase'] = types.ModuleType('supabase')
    sys.modules['supabase'].create_client = lambda url, key: _Cliente()

    os.environ['KEEPA_API_KEY'] = 'FAKE'
    os.environ['SUPABASE_URL'] = 'https://doble.local'
    os.environ['SUPABASE_KEY'] = 'FAKE'
    for v in ('SUPABASE_SERVICE_KEY', 'TELEGRAM_TOKEN', 'TELEGRAM_CHAT_ID', 'AUTORELANZAR_MIN'):
        os.environ.pop(v, None)

    # Por `atexit`: el escaner sale con sys.exit(1) al abortar, y estas cuatro
    # lineas son justo las que hay que leer en ese caso.
    @atexit.register
    def _informe():
        print('LLAMADAS_KEEPA=%d' % LLAMADAS['keepa'])
        print('ITEMS_KEEPA=%d' % LLAMADAS['items'])
        print('ESCRITURAS=%s' % ','.join(ESCRITURAS))
        xlsx = [r for r in SUBIDOS if r.endswith('.xlsx')]
        print('XLSX_SUBIDOS=%d' % len(xlsx))
        if xlsx and destino:
            io.open(destino, 'wb').write(SUBIDOS[xlsx[0]])

    import runpy
    runpy.run_path(RUTA, run_name='__main__')


if len(sys.argv) > 2 and sys.argv[1] == '--hijo':
    hijo(sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else None)
    sys.exit(0)


# ===========================================================================
# EL PADRE: los asserts
# ===========================================================================
FUENTE = io.open(RUTA, encoding='utf-8').read()
ARBOL = ast.parse(FUENTE, RUTA)
fallos = []


def eq(nombre, obtenido, esperado):
    ok = obtenido == esperado
    if not ok:
        fallos.append(nombre)
    print(('OK ' if ok else 'XX ') + nombre
          + ('' if ok else '   got=%r exp=%r' % (obtenido, esperado)))


# ---------------------------------------------------------------------------
# (A) Las funciones REALES de la etiqueta, sacadas con ast y ejecutadas
# ---------------------------------------------------------------------------
def _nodo(nombre):
    for n in ARBOL.body:
        if isinstance(n, ast.FunctionDef) and n.name == nombre:
            return n
        if isinstance(n, ast.Assign) and any(
                isinstance(t, ast.Name) and t.id == nombre for t in n.targets):
            return n
    print('XX %s ya no esta en %s (o dejo de ser de nivel superior)' % (nombre, RUTA))
    sys.exit(1)


NOMBRES = ('ORIGEN_IVA_FICHA', 'ORIGEN_IVA_ASUMIDO', 'iva_es_con_origen',
           'iva_es_de', 'origen_iva_fila')
FICHAS = {'889698498883': {'iva_pct': '0.2100'},
          '8412345678905': {'iva_pct': '0.1000'},
          '8499999999998': {'iva_pct': None},          # ficha propia SIN iva_pct
          '8499999999997': {'iva_pct': 'ni un numero'}}
_ns = {'IVA_DEFAULT_ES': 0.21, '_sup': lambda core: FICHAS.get(str(core).lstrip('0'))}
exec(compile(ast.fix_missing_locations(ast.Module(body=[_nodo(n) for n in NOMBRES],
                                                  type_ignores=[])), RUTA, 'exec'), _ns)
iva_es_con_origen = _ns['iva_es_con_origen']
iva_es_de = _ns['iva_es_de']
origen_iva_fila = _ns['origen_iva_fila']
FICHA, ASUMIDO = _ns['ORIGEN_IVA_FICHA'], _ns['ORIGEN_IVA_ASUMIDO']
print('extraidas de %s: %s' % (RUTA, ', '.join(NOMBRES)))
print()

eq('(A) la etiqueta de la ficha se llama "ficha"', FICHA, 'ficha')
eq('(A) y la del defecto lleva el numero que se asume', ASUMIDO, 'asumido 21%')
eq('(A) ficha con iva_pct -> valor de la ficha + etiqueta "ficha"',
   iva_es_con_origen('0889698498883'), (0.21, FICHA))
eq('(A) el 10% de las cuatro de alimentacion tambien sale de la ficha',
   iva_es_con_origen('8412345678905'), (0.10, FICHA))
eq('(A) 🔴 producto que NO esta en el catalogo propio -> 21% ASUMIDO',
   iva_es_con_origen('0889698851909'), (0.21, ASUMIDO))
eq('(A) ficha en el catalogo pero SIN iva_pct -> asumido (la etiqueta no miente)',
   iva_es_con_origen('8499999999998'), (0.21, ASUMIDO))
eq('(A) iva_pct que no es un numero -> asumido, y no revienta',
   iva_es_con_origen('8499999999997'), (0.21, ASUMIDO))
# iva_es_de() NO cambia de comportamiento: sigue devolviendo solo el numero.
eq('(A) iva_es_de() devuelve el MISMO numero que iva_es_con_origen()',
   [iva_es_de(c) for c in ('0889698498883', '8412345678905', '0889698851909')],
   [0.21, 0.10, 0.21])

# La etiqueta de la FILA (la columna del Excel), pais a pais.
eq('(A) fila ES de un propio -> ficha', origen_iva_fila('ES', 0.21, '0889698498883'), 'ficha')
eq('(A) fila ES de un ajeno -> asumido 21%',
   origen_iva_fila('ES', 0.21, '0889698851909'), 'asumido 21%')
# 🔴 IT y FR llevan el tipo GENERAL del pais, que no vive en la ficha. Rotularlas
# 'ficha' seria mentir: el 22% italiano no sale de `productos`.
eq('(A) 🔴 fila IT -> el tipo general italiano, no la ficha',
   origen_iva_fila('IT', 0.22, '0889698498883'), 'general IT 22%')
eq('(A) 🔴 fila FR -> el tipo general frances',
   origen_iva_fila('FR', 0.20, '0889698498883'), 'general FR 20%')
eq('(A) la etiqueta lleva el numero que la fila ha USADO de verdad',
   origen_iva_fila('IT', 0.04, '0889698498883'), 'general IT 4%')
eq('(A) fila sin IVA (el pais no dio datos) -> la fila calla',
   origen_iva_fila('ES', None, '0889698498883'), '—')


# ---------------------------------------------------------------------------
# (B) Que las guardas esten PUESTAS, no solo escritas
# ---------------------------------------------------------------------------
print()
# (B1) La columna, la ULTIMA de COLS. Por estructura (se evalua la lista), no con
#      un grep: el fichero la nombra tambien en los comentarios.
COLS = ast.literal_eval(_nodo('COLS').value)
eq('(B1) 🔴 "Origen IVA" es la ULTIMA columna de la hoja', COLS[-1], COL_NUEVA)
eq('(B1) y aparece una sola vez', COLS.count(COL_NUEVA), 1)

# (B2) Las dos salidas nuevas van por abortar(), y no valen dos `if` cualesquiera:
#      se mira que el `if` que aborta pregunte por `_cat_error` / `_rows_cat`.
_guardas = {}
for n in ast.walk(ARBOL):
    if isinstance(n, ast.If):
        _nombres = {x.id for x in ast.walk(n.test) if isinstance(x, ast.Name)}
        _aborta = any(isinstance(c, ast.Call) and isinstance(c.func, ast.Name)
                      and c.func.id == 'abortar'
                      for cuerpo in (n.body, n.orelse) for h in cuerpo
                      for c in ast.walk(h))
        if _aborta:
            for v in ('_cat_error', '_rows_cat'):
                if v in _nombres:
                    _guardas[v] = _guardas.get(v, 0) + 1
eq('(B2) 🔴 la lectura que LANZA aborta (guarda sobre _cat_error)',
   _guardas.get('_cat_error'), 1)
eq('(B2) 🔴 y las 0 filas tambien (guarda sobre _rows_cat)',
   _guardas.get('_rows_cat'), 1)
# La otra puerta: que el `except` no se vuelva a tragar el fallo con un print.
_except_prints = [h for n in ast.walk(ARBOL) if isinstance(n, ast.Try)
                  for h in n.handlers
                  if any(isinstance(c, ast.Call) and isinstance(c.func, ast.Name)
                         and c.func.id == 'print'
                         and 'AVISO sin cruce Supabase' in ast.dump(c)
                         for c in ast.walk(h))]
eq('(B2) 🔴 ya no queda el `AVISO sin cruce Supabase` que se tragaba el fallo',
   len(_except_prints), 0)
eq('(B2) el motivo del cambio sigue escrito en el fichero (la otra mitad del ancla)',
   'AVISO sin cruce Supabase' in FUENTE, True)

# (B3) La linea del log se imprime en las DOS ramas (salga o no el aborto).
_prints_cat = [n for n in ast.walk(ARBOL)
               if isinstance(n, ast.Call) and isinstance(n.func, ast.Name)
               and n.func.id == 'print' and 'CATALOGO_PROPIO:' in ast.dump(n)]
eq('(B3) 🔴 CATALOGO_PROPIO se imprime en las dos ramas', len(_prints_cat), 2)


# ---------------------------------------------------------------------------
# (C) DE PUNTA A PUNTA: el escaner real, tres veces, cada una en su proceso
# ---------------------------------------------------------------------------
print()


def correr(caso, destino=None):
    cmd = [sys.executable, '-u', os.path.abspath(__file__), '--hijo', caso]
    if destino:
        cmd.append(destino)
    p = subprocess.run(cmd, capture_output=True, text=True, encoding='utf-8',
                       errors='replace', env=dict(os.environ, PYTHONIOENCODING='utf-8'))
    out = (p.stdout or '') + (p.stderr or '')

    def cifra(clave):
        m = re.search(r'^%s=(\d+)\s*$' % clave, out, re.M)
        return int(m.group(1)) if m else -1
    esc = re.search(r'^ESCRITURAS=(.*)$', out, re.M)
    return {'codigo': p.returncode, 'salida': out,
            'keepa': cifra('LLAMADAS_KEEPA'), 'items': cifra('ITEMS_KEEPA'),
            'xlsx': cifra('XLSX_SUBIDOS'),
            'escrituras': [x for x in (esc.group(1).strip() if esc else '').split(',') if x]}


for _caso in ('falla', 'vacio'):
    _r = correr(_caso)
    eq('(C) [%s] 🔴 el run sale en ROJO (exit 1)' % _caso, _r['codigo'], 1)
    eq('(C) [%s] con la linea grepable del #265' % _caso,
       'ESCANEO_NO_EJECUTADO: catalogo propio ilegible' in _r['salida'], True)
    eq('(C) [%s] 🔴 y CERO llamadas a Keepa (ni un token)' % _caso,
       (_r['keepa'], _r['items']), (0, 0))
    eq('(C) [%s] 🔴 cero escrituras en Supabase' % _caso, _r['escrituras'], [])
    eq('(C) [%s] y ningun Excel subido' % _caso, _r['xlsx'], 0)
    eq('(C) [%s] la linea CATALOGO_PROPIO esta en el log' % _caso,
       'CATALOGO_PROPIO:' in _r['salida'], True)
    # 🔑 Y EL MOTIVO, QUE DISTINGUE LAS DOS GUARDAS. Sin esto, el caso [falla]
    #    saldria verde con la guarda de _cat_error QUITADA: una lectura que
    #    lanza deja `_rows_cat` vacia, asi que la guarda de las 0 filas lo
    #    recoge igual (medido rompiendola a mano). Son dos motivos distintos y
    #    el log tiene que decir cual de los dos ha sido.
    eq('(C) [%s] y el motivo dice cual de las dos guardas ha saltado' % _caso,
       ('la lectura de productos lanzo Exception' if _caso == 'falla'
        else 'productos devolvio 0 filas con activo=true') in _r['salida'], True)
    if _caso == 'vacio':
        eq('(C) [vacio] el log dice cuantas filas vinieron: 0',
           'CATALOGO_PROPIO: filas=0 | con_iva=0 | propios_en_catalogo=0' in _r['salida'],
           True)

# --- El caso normal: la otra mitad del ancla, y el Excel --------------------
_dest = os.path.join(tempfile.mkdtemp(), 'salida.xlsx')
_ok = correr('normal', _dest)
eq('(C) [normal] 🔴 el run sale VERDE (exit 0)', _ok['codigo'], 0)
eq('(C) [normal] 🔴 y Keepa SI se llama (si no, los ceros de arriba no probarian nada)',
   (_ok['keepa'] > 0, _ok['items'] > 0), (True, True))
eq('(C) [normal] las tres cifras del catalogo propio, al log',
   'CATALOGO_PROPIO: filas=3 | con_iva=3 | propios_en_catalogo=2' in _ok['salida'], True)
eq('(C) [normal] el escaner ha subido su Excel', _ok['xlsx'], 1)

if _ok['xlsx'] == 1 and os.path.exists(_dest):
    from openpyxl import load_workbook
    ws = load_workbook(_dest, data_only=False)['Análisis']
    CAB = [c.value for c in ws[1]]
    IDX = {n: i + 1 for i, n in enumerate(CAB)}
    eq('(C) [normal] la cabecera del .xlsx es COLS', CAB, COLS)
    eq('(C) [normal] 🔴 "Origen IVA" es la ULTIMA columna del .xlsx REAL',
       CAB[-1], COL_NUEVA)

    def col(fila, nombre):
        return ws.cell(row=fila, column=IDX[nombre]).value

    _por_pais = {}
    for _f in range(2, ws.max_row + 1):
        _por_pais.setdefault(col(_f, 'País'), {})[str(col(_f, 'EAN'))] = col(_f, COL_NUEVA)
    eq('(C) [normal] hay las 3 filas de cada pais',
       {k: len(v) for k, v in sorted(_por_pais.items())}, {'ES': 3, 'FR': 3, 'IT': 3})
    eq('(C) [normal] 🔴 los propios dicen "ficha" en su fila ES',
       {e: _por_pais['ES'].get(e) for e in sorted(EAN_PROPIO)},
       {e: 'ficha' for e in sorted(EAN_PROPIO)})
    eq('(C) [normal] 🔴 el que no esta en productos dice "asumido 21%"',
       {e: _por_pais['ES'].get(e) for e in sorted(EAN_NO_PROPIO)},
       {e: 'asumido 21%' for e in sorted(EAN_NO_PROPIO)})
    eq('(C) [normal] IT y FR llevan su tipo general en las 6 filas',
       (set(_por_pais['IT'].values()), set(_por_pais['FR'].values())),
       ({'general IT 22%'}, {'general FR 20%'}))
    # La cuenta que se verifica tambien en la pasada real: las filas 'ficha' no
    # pueden pasar del numero de fichas del catalogo propio.
    _n_ficha = sum(1 for p in _por_pais.values() for v in p.values() if v == 'ficha')
    eq('(C) [normal] 🔴 filas "ficha" <= fichas del catalogo propio',
       (_n_ficha, _n_ficha <= len(PRODUCTOS)), (len(EAN_PROPIO), True))
    # Y que el IVA del 10% se haya USADO de verdad: la hoja divide el precio por
    # (1+IVA), asi que la formula del Beneficio lo delata.
    _fila_chai = [_f for _f in range(2, ws.max_row + 1)
                  if str(col(_f, 'EAN')) == '8412345678905' and col(_f, 'País') == 'ES']
    eq('(C) [normal] la fila del 10% esta en el Excel', len(_fila_chai), 1)
    if _fila_chai:
        _ben = str(col(_fila_chai[0], 'Beneficio (€)') or '')
        eq('(C) [normal] 🔴 la formula de la fila del 10% divide por 1.1, no por 1.21',
           ('/1.1)' in _ben, '/1.21)' in _ben), (True, False))

print()
if fallos:
    print('❌ %d FALLOS: %s' % (len(fallos), ', '.join(fallos)))
    sys.exit(1)
print('✅ TODO OK')
