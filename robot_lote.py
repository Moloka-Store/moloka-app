# ============================================================================
# ROBOT LOTE  -  Fabrica de fichas para la linea "Funkos TCG bajo pedido"
# ----------------------------------------------------------------------------
# Para TCG la VERDAD del producto es TCG (lo que pides por EAN = lo que llega).
# Por eso esta linea NO usa imagen ni nombre de Keepa (que pueden venir del ASIN
# equivocado) y NO pasa por el montaje M7 de Elena (la imagen de TCG es caja+
# figura, el recorte fallaria). En su lugar:
#   - nombre + imagen salen del CATALOGO de TCG (web_rank/catalogo.xlsx, que subio
#     el Paso 1), por EAN.
#   - la descripcion la redacta Claude con el MISMO prompt de la fabrica
#     (reutilizado de robot_preparar, sin reescribirlo), a partir del dato de TCG.
#   - se escribe un BORRADOR directamente en web_productos con origen='tcg' y
#     activo=false (oculto) -> se revisa y se activa cuando esta OK.
# NO toca robot_preparar / robot_generar / motor_fotos: la fabrica "joya" de
# Elena queda 100% intacta. Solo reutiliza de robot_preparar (lectura): el
# cliente de Anthropic, el PROMPT, el MODELO, las CATEGORIAS, slugify y sb.
# Resumible: si un EAN ya esta en web_productos, lo salta.
# Secrets (ya en fabrica-lote.yml): KEEPA_API_KEY, SUPABASE_URL, SUPABASE_KEY,
#   SUPABASE_SERVICE_KEY, ANTHROPIC_API_KEY
# ============================================================================
import json, datetime, sys, io, re, math, unicodedata, requests
import openpyxl
import robot_preparar as R   # reusa (solo lectura): cliente, PROMPT_SISTEMA, MODELO, CATEGORIAS, slugify, descargar_b64, sb

sys.stdout.reconfigure(line_buffering=True)

BUCKET      = 'informes'
RECADO_LOTE = 'fabrica_lote/_solicitud_lote.json'
CAT_PATH    = 'web_rank/catalogo.xlsx'   # el catalogo que dejo el Paso 1
FOTOS_PATH  = 'web_rank/ranking_tcg_fotos.xlsx'  # fotos Keepa (alta res) del corte

# Bloque legal GPSR de Funko (verbatim de robot_generar; obligatorio en Funkos).
GPSR_WEB = ("<br><br><b>Información de seguridad del producto (GPSR)</b><br>"
            "Responsable en la UE: Funko EU BV · Zuidplein 36, 1077 XV Ámsterdam (NL) · supportEMEA@funko.com")

# Normalizacion de franquicias: unifica grafias distintas del MISMO fandom para que
# el filtro de la web salga limpio. Solo fusiona duplicados claros (y las series de
# Star Wars bajo "Star Wars"). NO fusiona submarcas tipo Deadpool/X-Men en Marvel:
# eso es decision de taxonomia, se puede ampliar cuando quieras.
FANDOM_CANON = {
    'kimetsu no yaiba': 'Demon Slayer',
    'kimetsu no yaiba (demon slayer)': 'Demon Slayer',
    'demon slayer kimetsu no yaiba': 'Demon Slayer',
    'demon slayer': 'Demon Slayer',
    'bola de dragon': 'Dragon Ball', 'bola de dragón': 'Dragon Ball',
    'dragonball': 'Dragon Ball', 'dragon ball z': 'Dragon Ball',
    'dragon ball super': 'Dragon Ball', 'dragon ball gt': 'Dragon Ball',
    'masters of the universe': 'Masters del Universo',
    'masters del universo': 'Masters del Universo',
    'nightmare before christmas': 'Pesadilla antes de Navidad',
    'pesadilla antes de navidad 30th': 'Pesadilla antes de Navidad',
    'kaiju nº8': 'Kaiju No. 8', 'kaiju nº 8': 'Kaiju No. 8',
    'kaiju no 8': 'Kaiju No. 8', 'kaiju no. 8': 'Kaiju No. 8',
    'it': 'IT', 'nlf': 'NFL',
    'arcane': 'League of Legends', 'arcane: league of legends': 'League of Legends',
    'the mandalorian': 'Star Wars', 'ahsoka': 'Star Wars',
    'star wars the acolyte': 'Star Wars', 'star wars: the acolyte': 'Star Wars',
    'the acolyte': 'Star Wars',
}
def normaliza_fandom(f):
    if not f:
        return f
    return FANDOM_CANON.get(f.strip().lower(), f.strip())

# ---------- FRENO difuso: nombre TCG vs nombre Keepa ----------
# Nunca son identicos, asi que comparamos por palabras del PERSONAJE: quitamos
# morralla y la franquicia de ambos; si comparten algun token -> mismo producto
# (foto Keepa OK). Si no comparten NADA -> freno (foto TCG, por seguridad).
_STOP = set("""funko pop vinyl figura figure de del la el los las y e and the a un una uno para con sin
coleccionable coleccionables coleccionistas coleccion idea regalo gift mercancia oficial official
merchandise juguetes toys ninos adultos kids fans tv anime animation games game movies movie video
muneco modelo model display exhibicion exposicion collectable collectible special edition exclusive
exclusivo vinilo serie bobble head bobblehead nuevo new standard std emea usa eu glow chase ride
cabeza oscilante tete figurine vol""".split())

def _sin_acentos(x):
    return ''.join(c for c in unicodedata.normalize('NFD', str(x)) if unicodedata.category(c) != 'Mn')

def _tokens_personaje(nombre, fandom=''):
    s = _sin_acentos((nombre or '').lower())
    s = re.sub(r'[^a-z0-9 ]', ' ', s)
    fset = set(_sin_acentos((fandom or '').lower()).split())
    toks = set()
    for w in s.split():
        if len(w) < 3 or w in _STOP or w in fset or w.isdigit():
            continue
        toks.add(w)
    return toks

def freno_ok(nombre_tcg, nombre_keepa, fandom):
    t1 = _tokens_personaje(nombre_tcg, fandom)
    t2 = _tokens_personaje(nombre_keepa, fandom)
    if not t1 or not t2:
        return True          # nombre demasiado corto para juzgar -> no frenamos
    return len(t1 & t2) >= 1 # comparten al menos el personaje

def cargar_fotos_keepa():
    """{ean: {nombre_keepa, img_figura, img_caja}} del ranking_tcg_fotos.xlsx. Si no
    existe, {} (todo ira por fallback TCG)."""
    data = R._bajar(BUCKET, FOTOS_PATH)
    if data is None:
        return {}
    ws = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    idx = {(str(c).strip() if c is not None else ''): i for i, c in enumerate(rows[0])}
    iE, iNk, iIf, iIc = idx.get('EAN'), idx.get('Nombre Keepa'), idx.get('Img figura'), idx.get('Img caja')
    if iE is None or iIf is None:
        return {}
    m = {}
    for r in rows[1:]:
        ean = str(r[iE] or '').strip()
        if not ean:
            continue
        m[ean] = {'nombre_keepa': str(r[iNk] or '') if iNk is not None else '',
                  'img_figura':   str(r[iIf] or '') if iIf is not None else '',
                  'img_caja':     str(r[iIc] or '') if iIc is not None else ''}
    return m

# ---------------------------------------------------------------------------
def cargar_catalogo_tcg():
    """Devuelve {ean: (cabecera, [urls_imagen], precio_tcg, estado)} del catalogo TCG.
    precio_tcg = coste actual en TCG (rebajado si esta de oferta). estado = 'Oferta',
    'Saldo', 'Disponible', etc. (lo usa la logica de ofertas)."""
    data = R.sb.storage.from_(BUCKET).download(CAT_PATH)
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else '' for c in rows[0]]
    idx = {h: i for i, h in enumerate(hdr)}
    iE   = idx['EAN']
    iCab = idx['Cabecera']
    iImg = idx.get('Imágenes', idx.get('Imagenes'))
    iPre = idx.get('Precio')
    iEst = idx.get('Estado producto')
    m = {}
    for r in rows[1:]:
        ean = str(r[iE] or '').strip()
        if not ean:
            continue
        cab = str(r[iCab] or '').strip()
        imgs = []
        if iImg is not None:
            raw = str(r[iImg] or '').strip()
            imgs = [u.strip() for u in raw.split(';') if u.strip().lower().startswith('http')]
        try:
            precio_tcg = float(r[iPre]) if (iPre is not None and r[iPre] is not None) else None
        except (TypeError, ValueError):
            precio_tcg = None
        estado = str(r[iEst] or '').strip() if iEst is not None else ''
        m[ean] = (cab, imgs, precio_tcg, estado)
    return m

# Patrones que delatan un DISPLAY/CASE de reventa (varias unidades), no un producto
# individual. NO incluye "pack" a secas para no cargarnos los Bitty Pop (que vienen
# en pack de 4 de forma legitima).
_PATRONES_PACK = ('case', '5+1', 'display', 'caja de', 'pdq', 'assortment', 'counter')

def es_pack(nombre_tcg, ean):
    n = (nombre_tcg or '').lower()
    if any(p in n for p in _PATRONES_PACK):
        return True
    if str(ean).strip().upper().endswith('C'):   # TCG sufija con 'C' los Case
        return True
    return False

# Coste normal de un Funko estandar en TCG (lo que paga Moloka). Fijo por decision de
# Fernando: 8,55 (no se toma del catalogo, que cuando hay oferta da el ya rebajado).
COSTE_NORMAL_ESTANDAR = 8.55
SUELO_OFERTA_ESTANDAR  = 11.95

def _redondea_95_arriba(x):
    """Redondea al X,95 inmediatamente >= x (precio comercial, a favor de Moloka)."""
    e = math.floor(x)
    return e + 0.95 if x <= e + 0.95 + 1e-9 else e + 1.95

def calcular_oferta(formato, es_chase, es_vaulted, es_exclusivo, precio_web, precio_tcg, estado):
    """Devuelve el precio_oferta (rojo) si procede, o None. SOLO Funko estandar (no
    rarezas), cuando TCG lo marca Oferta/Saldo. Traslada el ahorro de TCG al cliente
    con suelo 11,95 y redondeo al ,95 arriba."""
    if formato != 'Funko Pop!':
        return None
    if es_chase or es_vaulted or es_exclusivo:
        return None
    if (estado or '').strip().lower() not in ('oferta', 'saldo'):
        return None
    if precio_web is None or precio_tcg is None:
        return None
    ahorro = COSTE_NORMAL_ESTANDAR - precio_tcg
    if ahorro <= 0:
        return None                              # TCG no esta mas barato de lo normal
    bruto = precio_web - ahorro
    oferta = max(SUELO_OFERTA_ESTANDAR, _redondea_95_arriba(bruto))
    if oferta >= precio_web:
        return None                              # no hay rebaja real que mostrar
    return round(oferta, 2)

# ---------------------------------------------------------------------------
def redactar_tcg(nombre_tcg, img_url, rarezas):
    """Misma redaccion que la fabrica (MISMO PROMPT_SISTEMA), pero con datos de TCG.
    Le pasa la imagen de TCG (caja+figura) para que Claude lea el numero de la caja."""
    datos = {"titulo_origen_solo_para_identificar": nombre_tcg,
             "marca": "Funko", "tamano": "aprox. 10 cm"}
    datos.update({k: v for k, v in (rarezas or {}).items() if v})
    contenido = [{"type": "text", "text":
        "DATOS (verificados, no anadas nada que no este aqui):\n" +
        json.dumps(datos, ensure_ascii=False, indent=2) +
        "\n\nLee el numero de coleccion de la imagen de la caja (esquina sup. derecha). "
        "Si no se ve claro, no pongas numero."}]
    if img_url:
        try:
            b64 = R.descargar_b64(img_url)
            contenido.insert(0, {"type": "image",
                                 "source": {"type": "base64", "media_type": "image/jpeg", "data": b64}})
        except Exception as e:
            print(f"    (aviso: no pude bajar la imagen TCG, redacto sin ella: {e})")
    msg = R.cliente.messages.create(model=R.MODELO, max_tokens=1800,
                                    system=R.PROMPT_SISTEMA,
                                    messages=[{"role": "user", "content": contenido}])
    texto = msg.content[0].text.strip()
    if texto.startswith("```"):
        texto = texto.split("```")[1]
        if texto.startswith("json"):
            texto = texto[4:]
        texto = texto.strip()
    return json.loads(texto)

# ---------------------------------------------------------------------------
def reencuadrar_blanco(datos_jpg, margen=0.08, lienzo=1200):
    """Centra la figura sobre fondo blanco y mide su forma. Las fotos de TCG vienen
    con la figura abajo y mucho aire arriba; como el fondo es blanco, detectar el
    recuadro de lo NO-blanco es trivial y fiable. Recorta a ese recuadro, lo centra
    en un cuadrado con margen uniforme, y de paso calcula el ratio ancho/alto del
    contenido (sirve para detectar cajas apaisadas tipo Ride/Moment/pack).
    Devuelve (bytes_jpg, ratio). ratio=None si todo blanco o si falla."""
    try:
        from PIL import Image
        import numpy as np
        im = Image.open(io.BytesIO(datos_jpg)).convert('RGB')
        a = np.asarray(im)
        nf = (a < 245).any(axis=2)          # pixel "con contenido" = algun canal < 245
        ys, xs = np.where(nf)
        if len(xs) == 0:
            return datos_jpg, None          # imagen toda blanca: no toco
        x0, x1 = int(xs.min()), int(xs.max()) + 1
        y0, y1 = int(ys.min()), int(ys.max()) + 1
        fig = im.crop((x0, y0, x1, y1))
        w, h = fig.size
        ratio = (w / h) if h else None
        lado = int(max(w, h) * (1 + 2 * margen))
        cuadro = Image.new('RGB', (lado, lado), (255, 255, 255))
        cuadro.paste(fig, ((lado - w) // 2, (lado - h) // 2))
        cuadro = cuadro.resize((lienzo, lienzo), Image.LANCZOS)
        buf = io.BytesIO(); cuadro.save(buf, 'JPEG', quality=90)
        return buf.getvalue(), ratio
    except Exception as e:
        print(f"      (aviso: no pude reencuadrar, subo la original: {e})")
        return datos_jpg, None

# ---------------------------------------------------------------------------
def rehospedar_imagen(url, ean, i):
    """Descarga la imagen de TCG y la SUBE a Supabase Storage (fotos-fabrica/tcg/).
    Asi la web sirve la imagen desde Moloka y NO enlaza a tcgfactory.com (oculta el
    proveedor y no depende de ellos). Devuelve (url_publica, ratio) o (None, None)."""
    try:
        r = requests.get(url, timeout=30); r.raise_for_status()
        datos, ratio = reencuadrar_blanco(r.content)   # centra la figura y mide su forma
    except Exception as e:
        print(f"      AVISO: no pude descargar la imagen TCG ({e})")
        return None, None
    nombre = f"tcg/{ean}_{i}.jpg"
    try:
        R.sb_admin.storage.from_(R.BUCKET_FOTOS).upload(
            nombre, datos, {"content-type": "image/jpeg", "upsert": "true"})
    except Exception as e:
        print(f"      AVISO: no pude subir la imagen a Supabase ({e})")
        return None, None
    pub = R.sb_admin.storage.from_(R.BUCKET_FOTOS).get_public_url(nombre)
    if isinstance(pub, dict):
        pub = pub.get('publicUrl') or pub.get('publicURL')
    return pub, ratio

# ---------------------------------------------------------------------------
def ya_en_web(ean):
    try:
        r = R.sb.table('web_productos').select('id').eq('ean', str(ean)).limit(1).execute().data
        return bool(r)
    except Exception:
        return False

# ---------------------------------------------------------------------------
def generar_hoja_revision(tanda, revision):
    """Hoja de contactos del lote: un mosaico con cada ficha (principal + galeria +
    nombre + EAN + etiqueta). Las "raras" (figura sola o caja apaisada) salen con
    borde rojo para revisarlas de un vistazo. Cada ficha trae, plegado, el SQL listo
    para pasarla a FOTO PLANA de TCG (sin montaje) si Fernando la ve mal."""
    RARAS = {'plano-ancho', 'plano', 'montaje-sincaja'}
    def esc(s):
        return (str(s or '')).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
    cards = []
    for r in revision:
        rara  = r['fuente'] in RARAS
        borde = '#e11d48' if rara else '#e5e7eb'
        minis = ''.join(f'<img src="{esc(u)}">' for u in (r.get('planas') or [])[:4])
        planas = r.get('planas') or []
        prin = planas[0] if planas else r.get('principal')
        arr  = ", ".join("'" + esc(u) + "'" for u in planas) if planas else ("'" + esc(r.get('principal')) + "'")
        sql  = (f"update web_productos set imagen_principal='{esc(prin)}', "
                f"imagenes=array[{arr}] where ean='{esc(r['ean'])}';")
        cards.append(f"""
      <div class="card" style="border:3px solid {borde}">
        <img class="big" src="{esc(r.get('principal'))}">
        <div class="minis">{minis}</div>
        <div class="nom">{esc(r['nombre'])}</div>
        <div class="meta">{esc(r.get('fandom'))} · {esc(r['ean'])}</div>
        <span class="etq" style="background:{borde}">{esc(r['fuente'])}</span>
        <details><summary>pasar a foto plana</summary><textarea readonly onclick="this.select()">{esc(sql)}</textarea></details>
      </div>""")
    html = f"""<!doctype html><html lang="es"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Revision lote {esc(tanda)}</title>
<style>
 body{{font-family:-apple-system,BlinkMacSystemFont,sans-serif;background:#f7f7fb;margin:0;padding:20px;color:#111}}
 h1{{font-size:18px;margin:0 0 4px}}
 .ley{{color:#555;font-size:13px;margin-bottom:16px}}
 .ley b{{color:#e11d48}}
 .grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(220px,1fr));gap:16px}}
 .card{{background:#fff;border-radius:10px;padding:10px;position:relative}}
 .big{{width:100%;aspect-ratio:1;object-fit:contain;background:#fff;border-radius:6px}}
 .minis{{display:flex;gap:4px;margin-top:6px;flex-wrap:wrap}}
 .minis img{{width:38px;height:38px;object-fit:contain;border:1px solid #eee;border-radius:4px;background:#fff}}
 .nom{{font-weight:600;font-size:13px;margin-top:8px;line-height:1.25}}
 .meta{{font-size:11px;color:#888;margin-top:2px}}
 .etq{{display:inline-block;color:#fff;font-size:10px;padding:2px 6px;border-radius:4px;margin-top:6px}}
 details{{margin-top:8px}}
 summary{{font-size:11px;color:#2563eb;cursor:pointer}}
 textarea{{width:100%;height:64px;font-size:10px;margin-top:4px;border:1px solid #ddd;border-radius:4px}}
</style></head><body>
<h1>Revision lote {esc(tanda)} — {len(revision)} fichas</h1>
<div class="ley">Las de <b>borde rojo</b> son las "raras" (figura sola o caja apaisada): mira esas con lupa. Si alguna esta mal, abre "pasar a foto plana", copia el SQL y ejecutalo en Supabase.</div>
<div class="grid">{''.join(cards)}</div>
</body></html>"""
    return html.encode('utf-8')

# ---------------------------------------------------------------------------
def main():
    crudo = R._bajar(BUCKET, RECADO_LOTE)
    if crudo is None:
        print("SIN recado de lote. Nada que hacer."); return
    recado = json.loads(crudo.decode('utf-8'))
    tanda  = recado.get('tanda') or datetime.datetime.now().strftime('%Y%m%d_%H%M')
    items  = recado.get('items') or []
    print(f"Recado LOTE TCG: tanda {tanda}, {len(items)} EAN(s).")

    try:
        catalogo = cargar_catalogo_tcg()
        print(f"Catalogo TCG cargado: {len(catalogo)} EANs con nombre/imagen.")
    except Exception as e:
        print(f"ERROR: no pude cargar el catalogo TCG ({CAT_PATH}): {e}")
        return

    ok, err, saltados, sin_dato, packs = [], [], [], [], []
    revision = []   # para la hoja de contactos del lote
    for i, it in enumerate(items, 1):
        ean = str(it.get('ean') or '').strip()
        if not ean:
            continue
        if ya_en_web(ean):
            print(f"[{i}/{len(items)}] {ean} ya esta en web -> salto")
            saltados.append(ean); continue
        nombre_tcg, imgs, precio_tcg, estado_tcg = catalogo.get(ean, (None, [], None, ''))
        if not nombre_tcg or not imgs:
            print(f"[{i}/{len(items)}] {ean} sin nombre/imagen en catalogo TCG -> salto")
            sin_dato.append(ean); continue
        if es_pack(nombre_tcg, ean):
            print(f"[{i}/{len(items)}] {ean} es Case/display/pack ({nombre_tcg[:40]}) -> salto")
            packs.append(ean); continue

        print(f"\n[{i}/{len(items)}] TCG {ean} | {nombre_tcg[:55]}")
        try:
            # Re-alojar las fotos de TCG en Supabase (fondo blanco -> recorte fiable).
            # Guardamos (url, ratio_ancho_alto) de cada foto.
            imgs_web = []
            for k, u in enumerate(imgs):
                pub, ratio = rehospedar_imagen(u, ean, k)
                if pub: imgs_web.append((pub, ratio))
            if not imgs_web:
                print(f"   {ean}: no pude re-alojar ninguna imagen -> salto")
                err.append(ean); continue
            planas    = [u for u, _ in imgs_web]              # fotos TCG planas (reencuadradas)
            img_fig   = imgs_web[0][0]
            caja_url   = imgs_web[1][0] if len(imgs_web) > 1 else None
            caja_ratio = imgs_web[1][1] if len(imgs_web) > 1 else None
            # Caja APAISADA (mas ancha que alta): Ride/Moment/pack. El montaje portada
            # esta pensado para caja vertical de Funko -> con esas queda esquinado y
            # con hueco. En ese caso NO montamos: van con foto plana de TCG.
            caja_ancha = caja_ratio is not None and caja_ratio > 1.0
            # Solo hay caja "montable" si hay 2a foto Y no es apaisada (si solo hay 1
            # foto, img_caja=None -> no se monta portada, evita duplicar la figura).
            img_caja = caja_url if (caja_url and not caja_ancha) else None

            rarezas = {"es_chase": it.get('es_chase'),
                       "es_vaulted": it.get('es_vaulted'),
                       "es_exclusivo": it.get('es_exclusivo')}
            # Para leer el #numero pasamos la caja real si la hay (aunque sea apaisada).
            out = redactar_tcg(nombre_tcg, caja_url or img_fig, rarezas)
            categoria    = out.get('categoria') if out.get('categoria') in R.CATEGORIAS else None
            nombre_corto = (out.get('nombre_corto') or '').strip() or nombre_tcg
            slug         = R.slugify(out.get('slug') or nombre_corto)
            fandom_norm  = normaliza_fandom(out.get('fandom'))
            formato      = (it.get('formato') or '').strip() or None

            web_desc = (out.get('web_desc') or '').rstrip()
            if web_desc and 'GPSR' not in web_desc:
                web_desc += GPSR_WEB

            # ---- OFERTA: solo Funko estandar, cuando TCG lo marca Oferta/Saldo ----
            precio_web = it.get('precio_web')
            precio_oferta = calcular_oferta(formato, it.get('es_chase'), it.get('es_vaulted'),
                                            it.get('es_exclusivo'), precio_web, precio_tcg, estado_tcg)
            if precio_oferta is not None:
                print(f"   OFERTA TCG ({estado_tcg}, coste {precio_tcg}): {precio_web} -> {precio_oferta}")

            # ---- IMAGENES ----
            if caja_ancha:
                # Caja apaisada (Ride/Moment/pack): el montaje no le pega -> fotos planas.
                print("   caja apaisada -> fotos TCG planas (sin montaje)")
                imagen_principal = img_fig
                imagenes = planas
                fuente = 'plano-ancho'
            else:
                filaf = {'ean': ean, 'nombre_corto': nombre_corto, 'fandom': fandom_norm,
                         'formato': formato,
                         'fotos_elegidas': {'caja': img_caja, 'recorte_moloka': img_fig},
                         'con_protector': False}
                enlaces, errf = R.generar_fotos(filaf, None, None, None)
                if errf or not enlaces:
                    print(f"   sin montaje ({errf or 'sin enlaces'}) -> fotos TCG planas")
                    imagen_principal = img_fig
                    imagenes = planas
                    fuente = 'plano'
                else:
                    # CON caja: principal = portada (caja+figura). SIN caja: principal =
                    # figura sola reencuadrada (limpia), no la M7. El M7 va al final.
                    gal = [enlaces.get('portada'), enlaces.get('figura'), enlaces.get('caja'), enlaces.get('ficha')]
                    imagenes = [g for g in gal if g]
                    imagen_principal = (enlaces.get('portada') or enlaces.get('figura')
                                        or enlaces.get('ficha') or img_fig)
                    fuente = 'montaje' if enlaces.get('portada') else 'montaje-sincaja'

            fila = {
                'ean': ean, 'slug': slug,
                'origen': 'tcg', 'origen_id': ean,          # id del producto dentro de TCG = su EAN
                'seccion': 'funko',
                'titulo_seo': out.get('web_titulo'),
                'nombre': nombre_corto,
                'descripcion_html': web_desc or None,
                'licencia': 'Funko',
                'categoria': categoria, 'fandom': fandom_norm,
                'es_chase': bool(it.get('es_chase')),
                'es_vaulted': bool(it.get('es_vaulted')),
                'es_exclusivo': bool(it.get('es_exclusivo')),
                'precio': it.get('precio_web'), 'precio_web': it.get('precio_web'),
                'precio_oferta': precio_oferta,             # rojo tachando el normal (None si no hay oferta)
                'imagen_principal': imagen_principal, 'imagenes': imagenes,
                'imagenes_planas': planas,                  # fotos TCG planas (para "No vale" en la revision)
                'formato': formato,
                'activo': False,                            # BORRADOR oculto hasta aprobar
                'en_web': False, 'en_miravia': False,
                'stock': 0, 'disponibilidad': 'inmediato',  # "Disponible" (no "Bajo pedido / Encargar")
            }
            fila = {k: v for k, v in fila.items() if v is not None}
            R.sb.table('web_productos').insert(fila).execute()
            print(f"      OK [{fuente}] -> web_productos (borrador, activo=false) | {nombre_corto[:40]} "
                  f"| cat={categoria} | fmt={fila.get('formato')}")
            ok.append(ean)
            revision.append({'ean': ean, 'nombre': nombre_corto, 'fandom': fandom_norm,
                             'fuente': fuente, 'principal': imagen_principal, 'planas': planas})
        except Exception as e:
            print(f"   ERROR procesando {ean}: {e}")
            err.append(ean)

    # Borra el recado SOLO al terminar (si se corta antes, al relanzar retoma).
    try:
        R.sb.storage.from_(BUCKET).remove([RECADO_LOTE])
    except Exception:
        pass

    print(f"\n==== RESUMEN LOTE TCG {tanda} ====")
    print(f"  Creados (borrador): {len(ok)} | ya en web: {len(saltados)} "
          f"| sin dato TCG: {len(sin_dato)} | packs/Case saltados: {len(packs)} | errores: {len(err)}")
    if err:      print(f"  EAN con error: {err}")
    if sin_dato: print(f"  EAN sin nombre/imagen: {sin_dato}")
    if packs:    print(f"  EAN Case/pack excluidos: {packs}")

    # Desglose por tipo de montaje (para saber cuantas raras hay que revisar).
    if revision:
        from collections import Counter
        c = Counter(r['fuente'] for r in revision)
        print("  Montaje: " + " | ".join(f"{k}={v}" for k, v in sorted(c.items())))

    # HOJA DE CONTACTOS: mosaico del lote para revisar en borrador, antes de publicar.
    if revision:
        try:
            html = generar_hoja_revision(tanda, revision)
            path = f"revision/lote_{tanda}.html"
            R.sb_admin.storage.from_(R.BUCKET_FOTOS).upload(
                path, html, {"content-type": "text/html", "upsert": "true"})
            url = R.sb_admin.storage.from_(R.BUCKET_FOTOS).get_public_url(path)
            if isinstance(url, dict):
                url = url.get('publicUrl') or url.get('publicURL')
            print(f"\n  >>> HOJA DE REVISION: {url}")
        except Exception as e:
            print(f"  (aviso: no pude generar la hoja de revision: {e})")
    print("Fin.")

if __name__ == '__main__':
    main()
