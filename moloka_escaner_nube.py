#!/usr/bin/env python3
# ============================================================
# MOLOKA - Escaner solo-Keepa  (NUBE / GitHub Actions)
# ------------------------------------------------------------
# Generado a partir de Moloka_Escaner_soloKeepa.ipynb y adaptado para correr
# SIN Colab. FUENTE UNICA DE VERDAD: este .py se edita directo (sin notebook).
#
# QUE HACE:
#   - Lee el RECADO del buzon (informes/escaner/_solicitud_escaner.json):
#       { proveedor, marca, modo, rank_maximo, incluir_sin_rank }
#   - Carga el catalogo:
#       * TCG / DBLINE / BEMS -> el fichero crudo que la app subio al buzon.
#       * MOLOKA              -> lee el inventario propio DIRECTO de Supabase.
#   - Escanea con Keepa (Fase 1 rank + Fase 2 ES/IT/FR), calcula rentabilidad.
#   - Genera el Excel, lo SUBE a Storage (informes/resultados/) y registra el
#     escaneo en la tabla 'escaner_resultados' (la biblioteca de la app).
#   - Actualiza la memoria viva del proveedor (presentes / agotados).
#   - Limpia el buzon del escaner (VERIFICADO).
#
# Variables de entorno (GitHub Secrets): KEEPA_API_KEY, SUPABASE_URL, SUPABASE_KEY
# ============================================================

import os, sys, time, json, re
import pandas as pd
import keepa
from supabase import create_client
from datetime import datetime, timezone
from collections import Counter

# Salida SIN BUFFER: que cada print aparezca en el log de Actions al instante
# (antes los print quedaban atrapados en el buffer y el log parecia "mudo";
#  solo se veian los 'Waiting...' de la libreria keepa, que van por otro canal).
try:
    sys.stdout.reconfigure(line_buffering=True)
    sys.stderr.reconfigure(line_buffering=True)
except Exception:
    pass

# ============================================================
# CREDENCIALES (entorno, no Colab)
# ============================================================
print(">>> ARRANCANDO escaner. Creando cliente Keepa...", flush=True)
# 🔒 timeout EXPLICITO: la libreria keepa trae 10.0s por defecto (keepa_sync.py,
# __init__), insuficiente para un lote de 100 ASIN con stats=90. Ese default es
# el origen de los "Read timed out" que hacian saltar lotes enteros (run 27-jul:
# 86 de 186 productos nunca preguntados). Ajustable por entorno sin desplegar.
api = keepa.Keepa(os.environ['KEEPA_API_KEY'],
                  timeout=float(os.environ.get('KEEPA_TIMEOUT', '120')))
print(">>> Cliente Keepa creado. Conectando a Supabase...", flush=True)
sb  = create_client(os.environ['SUPABASE_URL'], os.environ['SUPABASE_KEY'])
print(">>> Supabase conectado. Consultando saldo real de tokens...", flush=True)
api.update_status()   # consulta el saldo REAL al servidor (el cliente nace con 0)
print(f">>> Tokens Keepa disponibles AHORA: {api.tokens_left}", flush=True)

# ============================================================
# LLAMADA ROBUSTA A KEEPA (reintentos + backoff)
# ------------------------------------------------------------
# Keepa a veces tiene un hipo (Read timed out, corte de red, 5xx). Antes una
# sola peticion fallida mataba TODA la corrida (caso 15-jun: timeout en Fase 2
# a los 44 min -> exit code 1, corrida entera perdida).
# Aqui reintentamos hasta KEEPA_MAX_INTENTOS veces con esperas crecientes.
# Si tras todos los intentos sigue fallando, devolvemos None y el llamador
# decide: Fase 2 SALTA ese producto; Fase 1 SALTA ese lote. NUNCA se mata la
# corrida entera por un fallo transitorio.
# progress_bar=False -> no ensucia el log de Actions con barras 0%|.
# ============================================================
KEEPA_MAX_INTENTOS = 4
KEEPA_ESPERAS = [5, 15, 40, 90]   # segundos entre intentos (backoff)

# 🔒 CONTABILIDAD DE LO QUE NO SE PUDO PREGUNTAR.
# El blindaje de arriba evita que un hipo mate la corrida, pero hasta hoy el
# lote perdido no dejaba rastro en el DATO: se etiquetaba "Keepa sin ASIN"
# (mentira: nunca se pregunto) y la memoria lo grababa como visto, asi que el
# pase diario 'nuevos' ya no volvia a mirarlo. Eso es la acumulacion.
# Regla de oro: NO SE MARCA COMO VISTO LO QUE NO SE HA PODIDO PREGUNTAR.
LOTES_PERDIDOS = []          # [{'fase','etiqueta','lote','n_codigos'}]
EANS_NO_PREGUNTADOS = set()  # ean_in cuyo lote de Fase 1 se perdio
PAISES_PERDIDOS = []         # [(asin, dominio)] de Fase 2

def keepa_query(items, **kwargs):
    """Llama a api.query con reintentos. Devuelve la lista de productos, o None
    si tras KEEPA_MAX_INTENTOS sigue fallando (el llamador lo gestiona)."""
    kwargs.setdefault('progress_bar', False)
    for intento in range(KEEPA_MAX_INTENTOS):
        try:
            return api.query(items, **kwargs) or []
        except Exception as ex:
            if intento < KEEPA_MAX_INTENTOS - 1:
                espera = KEEPA_ESPERAS[intento]
                print(f"  [Keepa] intento {intento+1}/{KEEPA_MAX_INTENTOS} fallo: {ex} "
                      f"-> reintento en {espera}s")
                time.sleep(espera)
            else:
                print(f"  [Keepa] AGOTADOS {KEEPA_MAX_INTENTOS} intentos: {ex} -> se salta")
    return None

# ============================================================
# BUZON DEL ESCANER: leer recado + descargar catalogo
# ============================================================
BUCKET = 'informes'
# Buzon y checkpoint configurables por entorno: cada director usa su PROPIA carpeta
# (escaner_tcg/, escaner_heo/, ...) para no pisarse entre ellos. Por defecto, las de
# siempre -> la app de Elena y los escaneos manuales NO cambian.
CARPETA_ESCANER = os.environ.get('CARPETA_ESCANER') or 'escaner'   # recado + catalogo del proveedor
CARPETA_CKPT    = os.environ.get('CARPETA_CKPT') or 'escaner_ckpt'  # checkpoint (carpeta aparte)
CARPETA_RESULTADOS = 'resultados'  # Excel de salida
RECADO = '_solicitud_escaner.json'

SOLICITUD = {}
catalogo_local = None
catalogo_nombre = None
N_CRUDO = None            # nº de filas del catalogo crudo (para el blindaje anti-vaciado)
UMBRAL_PARCIAL = 0.35     # si el catalogo crudo trae <35% de lo que hay en memoria -> NO marcar agotados
try:
    objs = sb.storage.from_(BUCKET).list(CARPETA_ESCANER) or []
    for o in objs:
        nombre = o.get('name')
        if not nombre or nombre.startswith('.'):
            continue
        ruta = f'{CARPETA_ESCANER}/{nombre}'
        if nombre == RECADO:
            try:
                d = sb.storage.from_(BUCKET).download(ruta)
                SOLICITUD = json.loads(d.decode('utf-8'))
            except Exception as _e:
                print('AVISO recado:', _e)
        else:
            try:
                d = sb.storage.from_(BUCKET).download(ruta)
                catalogo_nombre = nombre
                catalogo_local = f'/tmp/{nombre}'
                with open(catalogo_local, 'wb') as fp:
                    fp.write(d)
                # El boton sube los catalogos COMPRIMIDOS en gzip (para que el CSV gordo
                # de OcioStock quepa en Storage). pd.read_csv descomprime .gz solo, pero
                # pd.read_excel NO -> aqui descomprimimos cualquier gzip y dejamos el
                # fichero PLANO, asi la lectura (excel o csv) recibe siempre el original.
                with open(catalogo_local, 'rb') as _fp:
                    _magic = _fp.read(2)
                if _magic == b'\x1f\x8b':
                    import gzip as _gz, shutil as _sh
                    _plano = catalogo_local[:-3] if catalogo_local.endswith('.gz') else catalogo_local + '.plano'
                    with _gz.open(catalogo_local, 'rb') as _src, open(_plano, 'wb') as _dst:
                        _sh.copyfileobj(_src, _dst)
                    catalogo_local = _plano
                    print(">>> Catalogo descomprimido (venia en gzip).")
            except Exception as _e:
                print('AVISO catalogo:', _e)
except Exception as ex:
    print('AVISO buzon escaner:', ex)

# --- Parametros desde el recado ---
PROVEEDOR        = (SOLICITUD.get('proveedor') or '').upper()
MARCA            = SOLICITUD.get('marca') or 'Funko'
RANK_MAXIMO      = int(SOLICITUD.get('rank_maximo') or 30000)
MODO             = (SOLICITUD.get('modo') or 'nuevos').lower()
INCLUIR_SIN_RANK = bool(SOLICITUD.get('incluir_sin_rank', False))
# Motor de filtros (lo usa el DIRECTOR automatico): si el recado trae 'filtros', el
# escaner aplica reglas finas (varias marcas + idioma + estado) en vez del filtro de
# marca simple. Si NO trae 'filtros' (escaneo manual de Fernando), todo va como siempre.
FILTROS = SOLICITUD.get('filtros') or None

# --- Autorrelanzamiento (lo activa SOLO el director): si el escaneo se acerca al corte
# de GitHub (6h), guarda el progreso y se relanza solo para seguir la noche entera.
# AUTORELANZAR_MIN ausente o 0 lo DESACTIVA -> el escaneo manual NUNCA se relanza solo.
_T_INICIO = time.time()
_LIMITE_MIN = int(os.environ.get('AUTORELANZAR_MIN', '0') or '0')
_GH_PAT = os.environ.get('GH_PAT')
_GH_REPO = os.environ.get('GH_REPO', 'Moloka-Store/moloka-app')
_WF_RELANZAR = os.environ.get('AUTORELANZAR_WORKFLOW', 'director-tcg.yml')
_TIPO_RELANZAR = os.environ.get('AUTORELANZAR_TIPO', 'completo')

def _cerca_del_corte():
    return _LIMITE_MIN > 0 and (time.time() - _T_INICIO) > _LIMITE_MIN * 60

def _relanzarme():
    if not _GH_PAT:
        print("AVISO: sin GH_PAT no puedo relanzarme. Guardo y salgo; relanza a mano.")
        return False
    try:
        import requests as _rq
        url = f'https://api.github.com/repos/{_GH_REPO}/actions/workflows/{_WF_RELANZAR}/dispatches'
        r = _rq.post(url, json={'ref': 'main', 'inputs': {'tipo': _TIPO_RELANZAR}}, timeout=30,
                     headers={'Authorization': f'Bearer {_GH_PAT}',
                              'Accept': 'application/vnd.github+json',
                              'X-GitHub-Api-Version': '2022-11-28'})
        print(f">>> Autorrelanzamiento: dispatch {_WF_RELANZAR} ({_TIPO_RELANZAR}) -> {r.status_code}")
        return r.status_code in (200, 201, 204)
    except Exception as _e:
        print("AVISO: no pude relanzarme:", _e); return False

# --- Guardados de arranque: sin recado o sin catalogo no se hace nada ---
if not SOLICITUD or not PROVEEDOR:
    print('Buzon del escaner SIN recado valido: nada que escanear. Fin.')
    sys.exit(0)

PERFILES = {
    'TCG': {
        'tipo':'excel', 'sheet':'Catálogo', 'header':0,
        'col_marca':'Marca', 'col_ean':'EAN', 'col_nombre':'Cabecera',
        'col_pa':'Precio', 'col_stock':'Stock', 'col_estado':'Estado producto',
        'estados_ok':['Disponible','Oferta','Saldo'],   # PreOrder / Backorder quedan FUERA
        'precio_caja6':'caja',   # la 'C' pegada al EAN ya significa la caja 5+1 -> PA es de la CAJA
    },
    'DBLINE': {
        'tipo':'excel', 'sheet':0, 'header':2,
        'col_marca':'Publisher', 'col_ean':'EAN', 'col_nombre':'Descrizione',
        'col_pa':'Prezzo (€)', 'col_pa_promo':'Prezzo promo (€)', 'col_stock':'Disponibili',
        'col_estado':None, 'estados_ok':None,
    },
    'BEMS': {
        'tipo':'csv', 'sep':';', 'header':0,
        'col_marca':'FABRICANT', 'col_ean':'EAN', 'col_nombre':'TITRE UK',
        'col_pa':'PA', 'col_stock':'STOCK', 'col_estado':None, 'estados_ok':None,
    },
    'OSMA': {
        # Mayorista aleman de drogueria/cosmetica (primera necesidad Moloka).
        # Excel .xls, cabecera fila 1 (header=0). Precio UNITARIO con coma decimal
        # alemana (1,099 = 1,099 EUR). Stock 'verfügbar' a veces viene como '>3.000'
        # (punto de MILLAR, no decimal) -> trato especial via stock_especial.
        # Sin filtro de estado (todo lo que tenga stock>0 entra). CHASE no aplica.
        'tipo':'excel', 'sheet':0, 'header':0,
        'col_marca':'Bezeichnung', 'col_ean':'EAN 1', 'col_nombre':'Bezeichnung',
        'col_pa':'Preis_', 'col_stock':'verfügbar', 'col_estado':None, 'estados_ok':None,
        'stock_especial':'osma',        # usa _stock_osma() para parsear '>3.000'
        'col_extra_liq':'wird ausgelistet',   # 'wird ausverkauft' = se liquida (info de compra)
    },
    'BIEDRO': {
        # Mayorista aleman de drogueria (primera necesidad Moloka), misma familia que OSMA.
        # Excel .xlsx; la cabecera de datos esta en la FILA 4 (header=3): las 3 primeras
        # son el formulario de pedido (Kunden-Nr, Kundenname, Anschrift).
        # Precio UNITARIO neto con PUNTO decimal (3.40) -> _num() lo parsea directo.
        # NO trae columna de stock -> sin_columna_stock=True (se asume disponible).
        # Sin marca propia (va en el nombre) ni estado. CHASE no aplica.
        # Validado contra catalogo real: 3.316 productos con EAN+precio.
        'tipo':'excel', 'sheet':0, 'header':3,
        'col_marca':None, 'col_ean':'Stück-EAN', 'col_nombre':'Artikelbezeichnung',
        'col_pa':'Stückpreis\nnetto', 'col_stock':None, 'col_estado':None, 'estados_ok':None,
        'sin_columna_stock':True,
    },
    'OCIOSTOCK': {
        # Mayorista espanol de licencias (Funko, Banpresto, Pyramid, Cerda...).
        # Feed CSV diario con URL FIJA que se autoactualiza (el token va en GitHub
        # Secrets, NUNCA en codigo). Separador ';', campos entrecomillados, BOM
        # (utf-8-sig) -> solo afecta a la 1a columna 'id_producto', que no usamos.
        # TIENE columna de marca limpia -> se puede filtrar por marca (FUNKO, etc.).
        # Stock real en 'stock_disponible' (>0).
        # 🔒 PA = 'precio_distribuidores' (coste del distribuidor). NO usar
        # 'precio_neto'/'precio_bruto': son PVP recomendado, no el coste.
        # OJO dropshipping: el precio puede venir mas alto que el mayorista real
        # -> contrastar Funko contra BEMS/TCG antes de fiarse.
        # Validado contra feed real: 13.412 con stock+EAN+precio (3.475 Funko).
        'tipo':'csv', 'sep':';', 'header':0,
        'col_marca':'marca', 'col_ean':'ean', 'col_nombre':'nombre',
        'col_pa':'precio_distribuidores', 'col_stock':'stock_disponible',
        'col_estado':None, 'estados_ok':None,
        'col_volumen':'txt_precios_volumen',   # descuentos por volumen -> pestana "Precio por lote"
        'col_url':'product_url',   # enlace a la ficha de OcioStock (verificar volumen/precio en su web)
        'precio_caja6':'unidad',   # VERIFICADO en su feed: 11,99 €/ud, 71,94 € la caja -> NO dividir
    },
    'STOCKLIST': {
        # Mayorista nordico GENERALISTA (Toys, Games and consoles, Beauty, Movies, Pet...).
        # Stocklist Excel .xlsx, hoja 'Sheet1', cabecera fila 1 (header=0). 43.174 refs, TODAS
        # con stock real. Multimoneda (EUR/GBP/USD/DKK) -> usamos EUR como coste, con PUNTO
        # decimal (17.99) -> _num() lo parsea directo. Marca limpia en 'Brand' (contains) ->
        # filtrable: Funko, Paladone, Numskull, Nemesis Now, LEGO, Ravensburger, Nintendo...
        # Stock en 'Available' (>0). Sin columna de estado (todo lo con stock entra).
        # 🔒 PENDIENTE VERIFICAR: que 'EUR' es el COSTE de proveedor y no el PVP recomendado.
        # Validado contra catalogo real: 42.878 productos con EAN(12/13)+precio+stock.
        'tipo':'excel', 'sheet':'Sheet1', 'header':0,
        'col_marca':'Brand', 'col_ean':'CodeBars', 'col_nombre':'ItemName',
        'col_pa':'EUR', 'col_stock':'Available', 'col_estado':None, 'estados_ok':None,
    },
    # ============================================================
    # PROVEEDORES DE CLAUDE-IN-CHROME (formato VARIABLE) -> DETECCION TOLERANTE
    # ------------------------------------------------------------
    # Estos catalogos se extraen a mano y cada extraccion puede salir con columnas
    # distintas (nombres, orden, xlsx/csv). En vez de fijar nombres de columna, se
    # usa 'deteccion':'tolerante': el motor detecta solo la columna de EAN (numeros
    # de 12-13 digitos), la de precio (importe, con o sin 'EUR'), la de nombre (texto
    # largo) y, si existen, marca y stock. EAN no-12/13 -> a descartados (NO se
    # inventan EANs). Stock ausente -> se asume disponible. CHASE no aplica.
    # Anadir un proveedor nuevo de Claude-in-Chrome = una linea aqui + el desplegable.
    # ============================================================
    'DINOTOYS': {'tipo':'auto', 'deteccion':'tolerante'},   # mayorista holandes (Logic4)
    'ZENTRADA': {'tipo':'auto', 'deteccion':'tolerante'},   # marketplace mayorista (xlsx/csv)
    'MIS_COMPRAS': {'tipo':'auto', 'deteccion':'tolerante', 'efimero':True},   # compras ad-hoc: deteccion tolerante y NO toca la memoria de ningun proveedor
    'HEO': {
        # heoGATE Retailer API -> catalogo cruzado por descargar_heo.py (CSV ';', columnas
        # fijas). El director de HEO PRE-FILTRA a Funko+Ultimate Guard+ofertas y sube el
        # resultado, asi que aqui se escanea con marca=TODAS. Sin stock numerico: el estado
        # 'disponible' (availableToOrder + AVAILABLE) filtra lo servible. PA = 'precio' (coste
        # de hoy, con oferta si la hay; el escaneo diario se autocorrige si la oferta acaba).
        'tipo':'csv', 'sep':';', 'header':0,
        'col_marca':'marca', 'col_ean':'ean', 'col_nombre':'nombre',
        'col_pa':'precio', 'col_stock':None,
        'col_estado':'estado', 'estados_ok':['disponible'],
        'sin_columna_stock':True,   # HEO no da stock numerico -> estado 'disponible' ya filtra
        'precio_caja6':'caja',   # la pestana Chase_manual trae 'Precio caja' y 'Precio /6' explicitos
    },
    'MOLOKA': {'tipo':'supabase'},   # inventario propio: se lee de la tabla productos
}
if PROVEEDOR not in PERFILES:
    print(f'Proveedor desconocido: {PROVEEDOR}. Validos: {list(PERFILES)}. Fin.')
    sys.exit(0)
PERFIL = PERFILES[PROVEEDOR]
if PERFIL.get('efimero'):
    MODO = 'todo'   # las compras ad-hoc se escanean enteras (no hay memoria previa que filtre)

# ============================================================
# BEMS POR API (cliente integrado)
# ------------------------------------------------------------
# BEMS no se sube como fichero: se baja de su API. Cuando el proveedor es BEMS
# y NO hay catalogo en el buzon, pedimos a la API de BEMS los productos
# DISPONIBLES (AVAILABLE=1) de la marca elegida y los dejamos en un CSV temporal
# con EXACTAMENTE las columnas que el perfil BEMS ya espera
# (FABRICANT;EAN;TITRE UK;PA;STOCK). Asi el resto del motor (2 fases, calculo,
# Excel, memoria) NO cambia nada: BEMS pasa a comportarse como MOLOKA.
# Si SI hay fichero en el buzon (CSV manual antiguo), se respeta y NO se baja
# por API (compatibilidad hacia atras).
# Credenciales por entorno: BEMS_LOGIN, BEMS_PASSWORD, BEMS_SECRET_KEY.
# Mapeo API -> columnas del perfil BEMS:
#   FABRICANT <- NAME_MAN | EAN <- EAN | TITRE UK <- NAME_PRODUCT
#   PA        <- PRICE    | STOCK <- STOCK
# ============================================================
def descargar_catalogo_bems(marca, ruta_csv):
    """Baja de BEMS los productos DISPONIBLES de 'marca' (o de todo el catalogo si
    marca == 'TODAS') y los escribe como CSV ';' con las columnas del perfil BEMS.
    Devuelve el nº de productos escritos, o -1 si hubo error."""
    import csv as _csv
    try:
        from curl_cffi import requests as _curl
    except Exception as ex:
        print("ERROR BEMS: curl_cffi no disponible:", ex); return -1
    base = "https://www.probems.be/API"; imp = "chrome120"
    login = os.environ.get("BEMS_LOGIN"); pwd = os.environ.get("BEMS_PASSWORD"); sk = os.environ.get("BEMS_SECRET_KEY")
    if not (login and pwd and sk):
        print("ERROR BEMS: faltan credenciales BEMS_* en el entorno."); return -1
    # 1) token (24h, no cuesta tokens)
    try:
        rt = _curl.post(f"{base}/TOKEN",
                        data={"login": login, "password": pwd, "secret_key": sk},
                        headers={"Content-Type": "application/x-www-form-urlencoded"},
                        impersonate=imp, timeout=30)
    except Exception as ex:
        print("ERROR BEMS token (red):", ex); return -1
    tok = rt.json().get("access_token") if rt.status_code == 200 else None
    if not tok:
        print("ERROR BEMS token:", rt.status_code, (rt.text or "")[:150]); return -1
    H = {"accept": "application/json", "authorization": f"Bearer {tok}"}
    # 2) lista de productos DISPONIBLES de la marca (DETAILS=1 trae EAN/precio/nombre;
    #    LIMIT=0 = sin limite). Si marca == TODAS, no filtramos por fabricante.
    params = {"AVAILABLE": "1", "DETAILS": "1", "LIMIT": "0"}
    if marca and marca.strip().upper() != "TODAS":
        params["MANUFACTURER"] = marca.strip()
    try:
        r = _curl.get(f"{base}/PRODUCT-LIST-FILTER", params=params,
                      headers=H, impersonate=imp, timeout=180)
    except Exception as ex:
        print("ERROR BEMS PRODUCT-LIST-FILTER (red):", ex); return -1
    if r.status_code != 200:
        # BEMS devuelve 400 {"error":"NO RESULT"} cuando la marca no tiene productos
        # (o el filtro no casa). NO es un fallo del escaner: es "0 productos".
        # Lo tratamos como vacio limpio (return 0), no como error fatal.
        txt = (r.text or "")
        if "NO RESULT" in txt.upper():
            print(f"BEMS: '{marca}' sin resultados (NO RESULT). Se trata como 0 productos.")
            # escribir CSV solo con cabecera para que el flujo siga limpio
            with open(ruta_csv, "w", newline="", encoding="utf-8") as fp:
                _csv.writer(fp, delimiter=";").writerow(["FABRICANT", "EAN", "TITRE UK", "PA", "STOCK"])
            return 0
        print("ERROR BEMS lista:", r.status_code, txt[:150]); return -1
    try:
        prods = r.json()
    except Exception as ex:
        print("ERROR BEMS: respuesta no es JSON:", ex); return -1
    if not isinstance(prods, list):
        print("ERROR BEMS: respuesta no es una lista:", str(prods)[:150]); return -1
    # 3) escribir el CSV con el formato del perfil BEMS
    # OJO: el motor, tras leer el CSV, VUELVE a filtrar por marca sobre la columna
    # FABRICANT (contains). Como ya filtramos por marca en la propia API, ponemos en
    # FABRICANT el MISMO valor que pedimos (el id de marca) para que ese 2º filtro pase
    # trivialmente. En modo TODAS no se filtra, asi que conservamos el NAME_MAN real.
    fab_fijo = marca.strip() if (marca and marca.strip().upper() != "TODAS") else None
    n = 0
    with open(ruta_csv, "w", newline="", encoding="utf-8") as fp:
        w = _csv.writer(fp, delimiter=";", quoting=_csv.QUOTE_MINIMAL)
        w.writerow(["FABRICANT", "EAN", "TITRE UK", "PA", "STOCK"])
        for p in prods:
            ean = str(p.get("EAN") or "").strip()
            if not ean:
                continue
            w.writerow([
                fab_fijo if fab_fijo else str(p.get("NAME_MAN") or "").strip(),
                ean,
                str(p.get("NAME_PRODUCT") or "").strip(),
                str(p.get("PRICE") or "").strip(),
                str(p.get("STOCK") or "").strip(),
            ])
            n += 1
    return n

if PROVEEDOR == 'BEMS' and not catalogo_local:
    _ruta_bems = f'/tmp/BEMS_{MARCA}_{datetime.now().strftime("%Y%m%d_%H%M")}.csv'
    print(f">>> BEMS por API: bajando '{MARCA}' (solo disponible)...")
    _n_bems = descargar_catalogo_bems(MARCA, _ruta_bems)
    if _n_bems is None or _n_bems < 0:
        print("ERROR: no se pudo bajar el catalogo de BEMS por API. Fin.")
        sys.exit(1)
    if _n_bems == 0:
        print(f"BEMS: la marca '{MARCA}' no devolvio productos disponibles. Nada que escanear. Fin.")
        sys.exit(0)
    catalogo_local = _ruta_bems
    catalogo_nombre = os.path.basename(_ruta_bems)
    print(f">>> BEMS por API: {_n_bems} productos disponibles -> CSV temporal listo.")

if PROVEEDOR != 'MOLOKA' and not catalogo_local:
    print(f'Falta el catalogo de {PROVEEDOR} en el buzon. Sube el fichero y vuelve a lanzar. Fin.')
    sys.exit(0)

IVA_DEFAULT_ES, IVA_IT, IVA_FR = 0.21, 0.22, 0.20
ALMACEN, COM_DIGITALES = 0.15, 1.03
UNIDADES_CASE_TCG = 6          # CHASE en case de 6 (5+1) -> coste unitario = PA / 6. TCG y chase HEO.
LOTE_FASE1 = int(os.environ.get('LOTE_FASE1', '50'))   # 100 -> 50: si se pierde un lote, el agujero es la mitad
# activo (default) -> elige el que se parece y marca los dudosos (nada desaparece).
# off              -> vuelta atras de emergencia: elige por rank, como antes del bloque 2.
COTEJO_MODO = (os.environ.get('COTEJO_MODO') or 'activo').lower()     # activo | off
TS = datetime.now().strftime('%Y%m%d_%H%M')
ARCHIVO_SALIDA = f'/tmp/Escaneo_{PROVEEDOR}_{MARCA}_{TS}.xlsx'
print(f"{PROVEEDOR} | Marca {MARCA} | Rank max {RANK_MAXIMO} | Modo {MODO}")

# ============================================================
# Celda 0 - formula de rentabilidad (validada al centimo)
# ============================================================
def calc_rentabilidad(precio_venta, pa, ref_pct, fee_fba, iva, almacen=0.15, com_digitales=1.03):
    base       = precio_venta / (1 + iva)
    com_amazon = precio_venta * (ref_pct/100) * com_digitales
    beneficio  = base - pa - com_amazon - fee_fba - almacen
    roi        = beneficio / pa if pa else 0
    margen     = beneficio / precio_venta if precio_venta else 0
    return dict(base=base, com_amazon=com_amazon, beneficio=beneficio, roi=roi, margen=margen)

_r = calc_rentabilidad(15.99, 8.12, 15, 3.51, 0.21)
print(">>> FORMULA OK <<<" if abs(_r['beneficio']+1.04)<0.01 and abs(_r['com_amazon']-2.47)<0.01 else ">>> REVISAR FORMULA <<<")

# ============================================================
# Funciones de EAN
# ============================================================
# ============================================================
# EL SUFIJO DEL EAN NO SE TIRA, SE LEE (bloque 3). Es el dato que falta.
# OcioStock manda el MISMO EAN 3 veces: '...' (suelta), '... Chase' (suelto,
# se descarta) y '... C6' (la caja de 6, que es la que Fernando compra). Antes
# el 'C6' acababa en '6' -> "EAN forma rara" -> 66 cajas de 6 a la basura.
# ============================================================
_RE_SUFIJO = re.compile(r'^(.*?)\s+(C(\d+)|CHASE|L)\s*$', re.I)

def partir_ean(e):
    """'889698679282 C6' -> ('889698679282', 'C6', 6)
       '889698679282 Chase' -> ('889698679282', 'CHASE', None)
       '196214117020L'      -> ('196214117020', 'L', None)   (Latino, va PEGADA)
       '889698679282C'      -> ('889698679282', 'C', 6)      (convencion TCG, pegada)
       '8435507873345'      -> ('8435507873345', None, None)
    Devuelve (ean_limpio, sufijo, unidades_caja)."""
    e = str(e or '').strip().upper()
    if not e:
        return '', None, None
    m = _RE_SUFIJO.match(e)               # sufijo separado por espacio
    if m:
        base, suf = m.group(1).strip(), m.group(2)
        if suf.startswith('C') and m.group(3):
            return base, f'C{m.group(3)}', int(m.group(3))
        if suf == 'CHASE':
            return base, 'CHASE', None
        return base, 'L', None
    if e.endswith('L') and e[:-1].isdigit():      # 'L' pegada = version Latino
        return e[:-1], 'L', None
    if e.endswith('C') and e[:-1].isdigit():      # convencion TCG: la C pegada YA es la caja
        return e[:-1], 'C', 6
    return e, None, None

def clasificar(sufijo):
    """(es_caja, descartar). El chase SUELTO se sigue descartando (regla de negocio)."""
    if sufijo is None:  return False, False       # figura normal
    if sufijo == 'L':   return False, False       # version Latino: producto normal
    if sufijo == 'CHASE': return False, True      # chase suelto -> descartar (ya funcionaba)
    return True, False                            # C / C6 / C12... -> caja

# ============================================================
# GTIN-14: el proveedor manda a veces el codigo de CAJA (14 digitos) en vez del
# EAN-13 de la unidad, y a veces lo manda TRUNCADO a 13. Se reconstruye el
# EAN-13 y se deja que Keepa confirme: si no existe, se descarta como siempre.
# Coste de equivocarse: 1 token. Coste de no intentarlo: el producto no se ve.
# (_chk13 y _ean_ok los usa variantes_ean; _gtin14_ok y rescatar_gtin quedan como
#  utilidades del modulo -hoy sin llamador- que documentan la reconstruccion.)
# ============================================================
def _chk13(cuerpo12):
    d = [int(x) for x in cuerpo12][::-1]
    return str((10 - sum(v * (3 if i % 2 == 0 else 1) for i, v in enumerate(d)) % 10) % 10)

def _ean_ok(s):
    return s.isdigit() and len(s) == 13 and _chk13(s[:12]) == s[12]

def _gtin14_ok(s):
    if not (s.isdigit() and len(s) == 14):
        return False
    d = [int(x) for x in s[:13]][::-1]
    return str((10 - sum(v * (3 if i % 2 == 0 else 1) for i, v in enumerate(d)) % 10) % 10) == s[13]

def rescatar_gtin(e):
    """Devuelve el EAN-13 de la unidad, o None si no aplica. NO se usa si el
    codigo ya es un EAN-13/UPC-12 valido: solo para los que hoy se tiran."""
    e = str(e or '').strip()
    if not e.isdigit():
        return None
    if len(e) == 12 or _ean_ok(e):
        return None                              # ya vale tal cual
    if len(e) == 14 and _gtin14_ok(e):           # GTIN-14 completo
        return e[1:13] + _chk13(e[1:13])
    if len(e) == 13 and not _ean_ok(e) and e[0] in '123456789':
        return e[1:13] + _chk13(e[1:13])         # GTIN-14 truncado a 13 por el proveedor
    return None

def core_ean(e):     return partir_ean(e)[0]
def es_chase_ean(e): return clasificar(partir_ean(e)[1])[0]

# ============================================================
# GUARDARRAIL RELATIVO caja-vs-suelta (regla de Fernando: NINGUN umbral absoluto
# de precio; compra Funkos a 2,99 y llaveros a 1 €). Se compara la ficha CAJA
# contra la ficha SUELTA del MISMO EAN. Una oferta agresiva baja las DOS -> pasa
# limpia. Un fallo de parseo (÷6 mal aplicado) baja solo la caja -> salta.
# Medido sobre el feed real de OcioStock (28-jul), 105 pares caja/suelta:
#   ratio real min 0.48 | mediana 1.00 | max 1.50 ; con el ÷6 del bug: 0.079 a 0.25.
# Umbral 0.40: cero falsos positivos, caza el bug 105 de 105.
# ============================================================
UMBRAL_CAJA_VS_SUELTA = 0.40
def aviso_caja_incoherente(pa_caja_ud, pa_suelta):
    if not (pa_caja_ud and pa_suelta and pa_suelta > 0):
        return None
    ratio = pa_caja_ud / pa_suelta
    if ratio < UMBRAL_CAJA_VS_SUELTA:
        return (f'INCOHERENTE: la caja sale a {pa_caja_ud:.2f} €/ud y la unidad suelta del '
                f'MISMO EAN esta a {pa_suelta:.2f} € (ratio {ratio:.2f}). Parseo roto, no ganga.')
    return None

# ---- REGLA DE NEGOCIO DEL CHASE (todos los proveedores) --------------------
# El chase SOLO se compra en CAJA DE 6 (5+1): coste unitario = PA / 6.
# El chase SUELTO es un atraco -> se DESCARTA (no se escanea, no gasta tokens).
# OJO AL ORDEN: el sufijo del EAN manda sobre el nombre, porque en TCG la 'C'
# pegada al EAN YA significa la caja 5+1; si mirasemos el nombre primero,
# descartariamos por error los cases de TCG que se llaman "... Chase".
# De paso corta el bucle de OcioStock: alli el mismo EAN llega como figura
# normal, como caja 5+1 y como chase suelto, con precios muy distintos; al
# descartar el suelto y separar normal/caja en claves de memoria distintas,
# deja de haber 'cambio_precio' perpetuo.
_RE_CAJA6 = re.compile(r'5\s*\+\s*1', re.I)
# "chase" SOLO cuenta cuando es MARCADOR DE VARIANTE: al FINAL del nombre
# ("... Dilophosaurus Chase") o entre PARENTESIS ("... Pink Batman (CHASE) 18 cm").
# Si aparece en MEDIO es, casi siempre, un NOMBRE PROPIO: Chase es el perro de la
# Patrulla Canina, y sin esta restriccion se tiraban paraguas, gorros, mochilas y
# peluches perfectamente legitimos ("Peluche Chase Patrulla Canina Paw Patrol").
# Medido en el ensayo del 24-jul-2026 contra el feed real de OcioStock.
_RE_CHASE_NOM = re.compile(r'\bchase\b\s*$|\([^)]*\bchase\b[^)]*\)', re.I)
# "w/Chase" y "with Chase" significan CON chase: la linea INCLUYE el chase, no
# ES un chase suelto. Es la forma habitual de DBLine ("FUNKO GOLD ... w/Chase",
# a precio de unidad) y tambien de HEO ("... w/CH 9 cm Surtido (6)"). Entra
# como producto NORMAL. Medido en el ensayo del 24-jul-2026.
_RE_CON_CHASE = re.compile(r'\b(?:w/|with)\s*ch(?:ase)?\b', re.I)

def clasificar_chase(nombre, ean_in):
    """(es_case, es_caja6, descartar). El SUFIJO del EAN manda sobre el nombre."""
    _base, suf, _uds = partir_ean(ean_in)
    if suf is not None:
        es_caja, descartar = clasificar(suf)
        if descartar:   return False, False, True      # chase SUELTO -> fuera
        if es_caja:     return True, True, False        # C / C6 / C12 -> caja
        return False, False, False                      # 'L' (Latino) -> producto normal
    n = str(nombre or '')
    # 🔒 El nombre YA NO decide si es caja: eso lo dice el sufijo del EAN. "5 + 1"
    # en el nombre solo sirve para NO descartarlo como chase suelto.
    if _RE_CAJA6.search(n):     return True, True, False
    if _RE_CON_CHASE.search(n): return False, False, False
    if _RE_CHASE_NOM.search(n): return False, False, True
    return False, False, False
def variantes_ean(core):
    c, vs = str(core).strip(), set()
    if c.isdigit():
        vs.add(c); vs.add(c.lstrip('0'))
        if len(c)==12: vs.add('0'+c)
        if len(c)==13 and c.startswith('0'): vs.add(c[1:])
        # 🆕 GTIN-14: codigo de CAJA (14 digitos) o truncado a 13 por el proveedor.
        # Solo si NO es un EAN-13 valido, o si empieza por 1/2 (prefijos GS1 que no
        # son de pais). Medido sobre el catalogo real: genera variante en 5 de 1.557
        # codigos (0,32%), y son EXACTAMENTE los 5 raros. Cero falsos positivos.
        # (ningun EAN-13 legitimo empieza por 1 o 2; los buenos van por 8, 0 y 3.)
        if len(c) == 14 or (len(c) == 13 and (not _ean_ok(c) or c[0] in '12')):
            cuerpo = c[1:13]
            if len(cuerpo) == 12: vs.add(cuerpo + _chk13(cuerpo))
    return [v for v in vs if v]
def norm(code): return str(code).strip().lstrip('0')
def _num(x):
    try: return float(str(x).replace(',', '.').strip())
    except Exception: return None

def _num_eur(x):
    """Precio tipo '1,77 EUR' o '2.01 €' (ZENTRADA) -> float. Quita el sufijo de moneda."""
    s = str(x).upper().replace('EUR', '').replace('€', '').strip()
    return _num(s)

def _mejor_volumen(s):
    """Parsea los descuentos por volumen de OcioStock y devuelve (uds_minimas, precio)
    del tramo MAS BARATO. Formato: tramos separados por '|', cada tramo
    'lower:upper:precio' (upper puede ir vacio). Ej: '6:12:8.99|12::8.59' -> (12, 8.59).
    Devuelve None si no hay ningun tramo valido."""
    if not s:
        return None
    mejor = None
    for parte in str(s).split('|'):
        parte = parte.strip()
        if not parte:
            continue
        campos = parte.split(':')
        if len(campos) < 3:
            continue
        try:
            uds = int(float(campos[0]))
            precio = float(campos[-1])
        except Exception:
            continue
        if precio <= 0:
            continue
        if mejor is None or precio < mejor[1]:
            mejor = (uds, precio)
    return mejor

# Umbral anti-basura de volumen: OcioStock mete valores fijos absurdos (p.ej. 5.99 de
# volumen en un casco de 109.99). Un descuento por volumen REAL rara vez baja del 50%
# del precio suelto; por debajo lo tratamos como basura y lo ignoramos.
MIN_RATIO_LOTE = 0.5

# Salvavidas de titulo: compara el nombre del proveedor con el titulo de Amazon para
# detectar EANs mal catalogados en Amazon (p.ej. un peluche cuyo ASIN es un juego de PS4).
import unicodedata as _ud, re as _reT
_STOP_TIT = {'the','and','with','de','del','la','el','los','las','un','una','uno','con','para','por',
             'pop','figura','figure','set','pack','edition','deluxe','vinilo','vinyl',
             'peluche','plush','muneco','doll','juguete','juguetes','toy','coche','coches','car'}
def _tokens_tit(t):
    t = _ud.normalize('NFKD', str(t or '')).encode('ascii','ignore').decode()
    t = _reT.sub(r'[^a-zA-Z0-9]+',' ', t).lower()
    return {w for w in t.split() if len(w) >= 3 and w not in _STOP_TIT}
def _coincide_titulo(nombre_prov, titulo_amz):
    # '?' = sin titulo (no marcamos). SI = comparten palabra distintiva. NO = nada en comun.
    if not str(titulo_amz or '').strip(): return '?'
    a = _tokens_tit(nombre_prov); b = _tokens_tit(titulo_amz)
    if not a or not b: return '?'
    return 'SÍ' if (a & b) else '⚠ NO'

# ============================================================
# COTEJO proveedor<->Amazon (bloque 2). Elige el ASIN que SE PARECE al nombre del
# proveedor, no el que mas vende. El desempate por rank tenia 50% de acierto en el
# escaneo del 26-jul (4 de 8 con varios ASIN mal elegidos), y el sesgo es dirigido:
# mejor rank ~ mas vendido ~ precio mas alto -> siempre empuja a "mas rentable de lo
# que es". El criterio calcula que palabras distinguen A PARTIR DEL PROPIO CATALOGO
# (IDF), sin lista de stopwords que mantener.
# ============================================================
import math, unicodedata   # re y Counter ya estan importados arriba; unicodedata aqui va pelado (arriba es 'as _ud')

UMBRAL_COTEJO = 0.40        # solo se usa cuando el nombre NO tiene ninguna palabra distintiva
_DF, _NDOC = Counter(), 1
COTEJO_ACTIVO = True        # se apaga si el proveedor no trae columna de nombre real

def _tok_cot(t):
    t = unicodedata.normalize('NFKD', str(t or '')).encode('ascii', 'ignore').decode().lower()
    out = []
    for w in re.sub(r'[^a-z0-9]+', ' ', t).split():
        if len(w) < 3 and not w.isdigit():
            continue
        if len(w) > 4 and w.endswith('s'):
            w = w[:-1]                     # plural simple
        out.append(w)
    return out

def construir_idf(nombres):
    """Una palabra que sale en MUCHOS productos del catalogo (funko, pop, figura,
    disney, dragon...) no distingue nada. El propio catalogo dice cuales son:
    no hace falta mantener a mano ninguna lista de palabras a ignorar."""
    global _DF, _NDOC
    _DF, vistos = Counter(), set()
    for n in nombres:
        n = str(n or '').strip()
        if not n or n in vistos:
            continue
        vistos.add(n)
        _DF.update(set(_tok_cot(n)))
    _NDOC = max(1, len(vistos))

def _idf(w):          return math.log(_NDOC / (1 + _DF.get(w, 0)))
def _distintivo(w):   return _DF.get(w, 0) <= max(2, _NDOC * 0.01)

def cotejar(nombre_prov, titulo_amz):
    """(casa, score, motivo). casa=None -> NO SE PUEDE cotejar: el llamador debe
    comportarse como hasta hoy (desempate por rank), nunca rechazar por esto."""
    if not COTEJO_ACTIVO:
        return None, 0.0, 'n/d: proveedor sin columna de nombre'
    tp = _tok_cot(nombre_prov)
    ta = set(_tok_cot(titulo_amz))
    if not tp or not ta:
        return None, 0.0, 'n/d: sin texto que comparar'
    # 🔒 RED DE SEGURIDAD (defensa en profundidad). Con deteccion tolerante, si el
    # motor no encuentra columna de nombre usa la del EAN como nombre
    # (`det['nombre'] or det['ean']`). Un "nombre" que son solo digitos NO es un
    # nombre: sin esto, ZENTRADA/DINOTOYS rechazarian el catalogo entero.
    if not [w for w in tp if not w.isdigit()]:
        return None, 0.0, 'n/d: el nombre no tiene palabras (parece un codigo)'
    pesos = {w: _idf(w) for w in set(tp)}
    score = sum(p for w, p in pesos.items() if w in ta) / (sum(pesos.values()) or 1)
    fuertes = {w for w in pesos if _distintivo(w)}
    if fuertes:                                     # via 1 (95,5% del catalogo)
        casan = fuertes & ta
        if casan:
            return True, score, 'casa: ' + ', '.join(sorted(casan)[:3])
        return False, score, 'ningun distintivo casa (faltan: ' + ', '.join(sorted(fuertes)[:3]) + ')'
    return (score >= UMBRAL_COTEJO), score, f'cobertura {score:.0%} (nombre sin palabras distintivas)'

def elegir_candidato(nombre_prov, cands, keyrank):
    """cands: [{'asin','title','r_90',...}] -> (elegido, veredicto, detalle).
    🔒 REGLA DE FERNANDO (28-jul): NUNCA devuelve None. Si ninguno pasa el filtro
    se entrega EL MAS PARECIDO y se marca DUDOSO. Sacar un producto del informe
    porque el criterio no lo reconoce seria cambiar un error VISIBLE (un ASIN
    equivocado, que se ve y se corrige) por una AUSENCIA INVISIBLE (un producto
    que falta y del que nadie se entera). Es el mismo pecado que la fuga
    silenciosa. El criterio ORDENA y AVISA; no decide qué desaparece.
    Ademas el filtro fallara por idioma (Vengadores/Avengers) y eso no puede
    costarte dejar de ver el producto."""
    if not cands:
        return None, '—', 'sin candidatos'
    ev = [(c, cotejar(nombre_prov, c.get('title'))) for c in cands]
    if all(v[0] is None for _, v in ev):                    # no hay con que cotejar
        return min(cands, key=keyrank), 'n/d', ev[0][1][2] + ' -> elegido por rank, como siempre'
    casan = [(c, v) for c, v in ev if v[0] is True]
    if casan:
        elegido = min((c for c, _ in casan), key=keyrank)
        det = next(v[2] for c, v in casan if c is elegido)
        return elegido, 'OK', f'{len(casan)}/{len(cands)} casan; entre esos, mejor rank ({det})'
    # Ninguno pasa el filtro -> se entrega el MAS PARECIDO, nunca se descarta.
    mejor_c, mejor_v = max(ev, key=lambda cv: (cv[1][1], -keyrank(cv[0])))
    if mejor_v[1] <= 0:                                     # ni uno se parece en nada
        return (min(cands, key=keyrank), '⚠ DUDOSO',
                f'ninguno de los {len(cands)} se parece al nombre; te pongo el de mejor rank — REVISAR')
    return (mejor_c, '⚠ DUDOSO',
            f'ninguno pasa el filtro; te pongo el MAS PARECIDO ({mejor_v[1]:.0%} de {len(cands)}) — REVISAR')

# ============================================================
# GUARDARRAIL DE SISTEMA (mismo espiritu que el blindaje anti-vaciado del 35%)
# Si el cotejo rechaza a MEDIO CATALOGO, el roto es el criterio, no el catalogo.
# En ese caso NO se bloquea nada: se avisa y se elige como hasta hoy.
# ============================================================
UMBRAL_PANICO = 0.50

def cotejo_de_fiar(n_evaluados, n_rechazados):
    if n_evaluados < 20:
        return True, ''                       # muestra muy chica para juzgar
    ratio = n_rechazados / n_evaluados
    if ratio > UMBRAL_PANICO:
        return False, (f"PANICO DE COTEJO: rechaza {n_rechazados}/{n_evaluados} ({ratio:.0%}). "
                       f"El roto es el criterio, no el catalogo -> NO se bloquea nada "
                       f"esta corrida; se elige por rank como siempre.")
    return True, ''

def _es_ean_valido(s):
    """True si s es un EAN/UPC utilizable: 12 o 13 digitos."""
    s = str(s).strip()
    return s.isdigit() and len(s) in (12, 13)

def _parse_precio_libre(x):
    """Precio en cualquier formato razonable: '1,77 EUR', '2.01 €', '8,62', '11.34'."""
    s = str(x)
    tiene_moneda = ('eur' in s.lower()) or ('€' in s)
    return _num_eur(x) if tiene_moneda else _num(x)

def detectar_columnas(cat):
    """Detecta por NOMBRE (pistas multiidioma) y, si falla, por CONTENIDO las columnas
    de ean / precio / nombre / marca / stock en un catalogo de formato desconocido
    (ficheros de Claude-in-Chrome). Devuelve los nombres REALES de columna (o None).
    No inventa nada: si no encuentra EAN o precio con confianza, el llamante aborta."""
    cols = list(cat.columns)
    low  = {c: str(c).strip().lower() for c in cols}

    def por_nombre(claves, excluir=()):
        for c in cols:
            if c in excluir:
                continue
            if any(k in low[c] for k in claves):
                return c
        return None

    # ---- EAN: nombre primero, si no, columna con mas valores de 12-13 digitos ----
    ean = por_nombre(['ean', 'gtin', 'barcode', 'codigo de barras', 'código de barras', 'upc'])
    if ean is None:
        mejor, mejor_score = None, 0.0
        for c in cols:
            score = cat[c].astype(str).str.strip().apply(_es_ean_valido).mean()
            if score > mejor_score:
                mejor, mejor_score = c, score
        if mejor_score >= 0.30:          # al menos 30% parecen EAN reales
            ean = mejor

    # ---- PRECIO: nombre primero, si no, columna 'tipo importe' (con decimales/moneda) ----
    precio = por_nombre(['precio', 'price', 'preis', 'prezzo', 'pvd', 'coste', 'cost', 'tarifa', 'eur', '€'],
                        excluir=(ean,))
    if precio is None:
        mejor, mejor_score = None, 0.0
        for c in cols:
            if c == ean:
                continue
            vals = cat[c].astype(str)
            parseables   = vals.apply(lambda v: _parse_precio_libre(v) is not None).mean()
            con_decimal  = vals.apply(lambda v: (',' in v or '.' in v or 'eur' in v.lower() or '€' in v)).mean()
            score = parseables * (0.4 + 0.6 * con_decimal)   # premia decimales/moneda (evita indices N°)
            if score > mejor_score:
                mejor, mejor_score = c, score
        if mejor_score >= 0.50:
            precio = mejor

    # ---- NOMBRE: nombre primero, si no, columna de texto mas largo ----
    nombre = por_nombre(['nombre', 'name', 'descrip', 'titre', 'title', 'producto', 'product',
                         'bezeichnung', 'articolo', 'artikel', 'designation'], excluir=(ean, precio))
    if nombre is None:
        mejor, mejor_len = None, 0.0
        for c in cols:
            if c in (ean, precio):
                continue
            avg = cat[c].astype(str).str.len().mean()
            if avg > mejor_len:
                mejor, mejor_len = c, avg
        nombre = mejor

    # ---- MARCA y STOCK: opcionales (solo por nombre) ----
    marca = por_nombre(['marca', 'brand', 'licencia', 'license', 'licens', 'fabricante',
                        'manufacturer', 'publisher', 'fabricant'], excluir=(ean, precio, nombre))
    stock = por_nombre(['stock', 'disponib', 'verfüg', 'verfug', 'qty', 'cantidad', 'quantity', 'quantità'],
                       excluir=(ean, precio, nombre, marca))

    return {'ean': ean, 'precio': precio, 'nombre': nombre, 'marca': marca, 'stock': stock}

def _stock_osma(x):
    """Stock de OSMA: numeros normales (910), '>3.000' (punto de MILLAR, 112) y '0' (42).
    '>3.000' -> 3000 (mas de 3000 unidades). El punto es separador de millar, NO decimal."""
    s = str(x).strip()
    if not s: return None
    s = s.lstrip('>').strip()       # quita el '>' de '>3.000'
    s = s.replace('.', '')          # quita el punto de millar: '3.000' -> '3000'
    try: return float(s)
    except Exception: return None

# ============================================================
# Celda 4 - carga del catalogo
#   MOLOKA -> Supabase (inventario propio).  Resto -> fichero crudo del buzon.
# ============================================================
problematicos = []
chase_sueltos = []   # chase suelto descartado por la regla de negocio (va a la hoja Descartados)
filas = []
fuera_disp = 0
# Defaults para que el cotejo (mas abajo, punto comun) no reviente en la rama MOLOKA:
# la rama de ficheros los reasigna; MOLOKA los deja asi (tiene nombre real -> cotejo activo).
det = {}
_tolerante = False

if PROVEEDOR == 'MOLOKA':
    print("=== MOLOKA: leyendo inventario propio de Supabase ===")
    _rows = []; _d = 0
    while True:
        res = sb.table('productos').select('ean,nombre,marca,pvd,es_chase,asin') \
                .eq('activo', True).range(_d, _d+999).execute()
        if not res.data: break
        _rows.extend(res.data)
        if len(res.data) < 1000: break
        _d += 1000
    print(f"Inventario propio (activos): {len(_rows)} filas")
    for p in _rows:
        marca_p = str(p.get('marca','') or '')
        if MARCA and MARCA.strip().upper() != 'TODAS':
            if MARCA.lower() not in marca_p.lower():
                continue
        ean_in = str(p.get('ean') or '').strip()
        if not ean_in:
            continue
        core = core_ean(ean_in)
        if (not core.isdigit()) or len(core) not in (12, 13):
            problematicos.append({'EAN':ean_in, 'Cabecera':p.get('nombre',''),
                                  'Motivo':f'EAN forma rara (len={len(core)})'}); continue
        filas.append({'ean_in':ean_in, 'core':core, 'variantes':variantes_ean(core),
                      'nombre':p.get('nombre','') or '', 'marca':marca_p or MARCA,
                      'pa':_num(p.get('pvd')), 'es_chase':bool(p.get('es_chase'))})
    print(f"A escanear: {len(filas)} | EAN problematicos: {len(problematicos)} | "
          f"CHASE: {sum(f['es_chase'] for f in filas)}")
else:
    if PERFIL['tipo'] == 'auto':
        # ZENTRADA: el extracto puede llegar como .xlsx o como .csv. Se detecta por
        # los bytes (un .xlsx es un ZIP -> empieza por 'PK'); el resto se trata como
        # CSV. Asi no depende del nombre ni de la extension del fichero subido.
        try:
            with open(catalogo_local, 'rb') as _fp:
                _es_excel = _fp.read(2) == b'PK'
        except Exception:
            _es_excel = False
        if _es_excel:
            cat = pd.read_excel(catalogo_local, sheet_name=PERFIL.get('sheet', 0),
                                header=PERFIL.get('header', 0), dtype=str).fillna('')
        else:
            cat = pd.read_csv(catalogo_local, sep=PERFIL.get('sep', ','), dtype=str,
                              encoding='utf-8-sig', on_bad_lines='skip').fillna('')
    elif PERFIL['tipo'] == 'excel':
        cat = pd.read_excel(catalogo_local, sheet_name=PERFIL['sheet'],
                            header=PERFIL['header'], dtype=str).fillna('')
    else:
        cat = pd.read_csv(catalogo_local, sep=PERFIL.get('sep', ';'), dtype=str,
                          encoding='utf-8', on_bad_lines='skip').fillna('')
    cat.columns = [str(c).strip() for c in cat.columns]   # BEMS trae espacios en los nombres
    print(f"Catalogo crudo: {len(cat)} filas")
    N_CRUDO = len(cat)

    if PERFIL.get('deteccion') == 'tolerante':
        det = detectar_columnas(cat)
        print(f"Deteccion tolerante -> EAN={det['ean']!r} precio={det['precio']!r} "
              f"nombre={det['nombre']!r} marca={det['marca']!r} stock={det['stock']!r}")
        if not det['ean'] or not det['precio']:
            print("ERROR: no pude detectar la columna de EAN o de precio en este fichero.")
            print("Revisa el catalogo (o pasaselo a Claude para normalizarlo). Fin.")
            sys.exit(0)
        cM, cE, cN, cP, cS = (det['marca'], det['ean'], det['nombre'] or det['ean'],
                              det['precio'], det['stock'])
        _tolerante = True
    else:
        cM, cE, cN, cP, cS = (PERFIL['col_marca'], PERFIL['col_ean'], PERFIL['col_nombre'],
                              PERFIL['col_pa'], PERFIL['col_stock'])
        _tolerante = False

    # filtro 1: marca. Si el recado trae 'filtros' (director), NO filtramos aqui por
    # marca: el motor decide fila a fila mas abajo (varias marcas + idioma + estado).
    # Si no, filtro de marca simple de siempre (escaneo manual de Fernando).
    def _pasa_filtros(row):
        # Reglas del director. Para TCG: Oferta/Saldo entran SIEMPRE (cualquier marca);
        # Disponible entra solo si la marca esta en 'marcas', o en 'marcas_es' Y ademas
        # el idioma es Espanol (Magic / Yu-Gi-Oh solo en espanol explicito).
        est = str(row.get(PERFIL.get('col_estado'), '')).strip()
        if est in (FILTROS.get('incluir_estados') or []):
            return True
        marca_row = str(row.get(cM, '')).lower() if cM else ''
        for m in (FILTROS.get('marcas') or []):
            if m.lower() in marca_row:
                return True
        marcas_es = FILTROS.get('marcas_es') or []
        if marcas_es:
            col_idi = FILTROS.get('col_idioma')
            idi = str(row.get(col_idi, '')).strip().lower() if col_idi else ''
            if idi in ('español', 'espanol'):
                for m in marcas_es:
                    if m.lower() in marca_row:
                        return True
        return False

    if FILTROS:
        sel = cat.copy()
        print(f"Motor de filtros activo (director): {len(sel)} filas a evaluar")
    elif MARCA and MARCA.strip().upper() != 'TODAS' and cM is not None:
        sel = cat[cat[cM].str.contains(MARCA, case=False, na=False)].copy()
        print(f"Marca '{MARCA}': {len(sel)} filas")
    else:
        sel = cat.copy()
        print(f"Sin filtro de marca (TODAS): {len(sel)} filas")

    # filtro 2: disponibilidad (estado permitido si aplica + stock>0) + EAN valido
    for _, row in sel.iterrows():
        if PERFIL.get('estados_ok'):
            if str(row.get(PERFIL['col_estado'], '')).strip() not in PERFIL['estados_ok']:
                fuera_disp += 1; continue
        if FILTROS and not _pasa_filtros(row):
            fuera_disp += 1; continue
        if PERFIL.get('sin_columna_stock') or (_tolerante and cS is None):
            stock = 1.0     # sin columna de stock -> se asume disponible
        elif PERFIL.get('stock_especial') == 'osma':
            stock = _stock_osma(row.get(cS, ''))
        else:
            stock = _num(row.get(cS, ''))
        if stock is None or stock <= 0:
            fuera_disp += 1; continue
        ean_in = str(row[cE]).strip()
        # Regla del chase: el SUELTO se descarta (solo se compra en caja de 6).
        _es_case, _es_caja6, _descartar = clasificar_chase(row.get(cN, ''), ean_in)
        if _descartar:
            chase_sueltos.append({'EAN': ean_in, 'Cabecera': row.get(cN, ''),
                                  'Motivo': 'Chase SUELTO descartado (solo se compra en caja de 6)'})
            continue
        core = core_ean(ean_in)
        if (not core.isdigit()) or len(core) not in (12, 13):
            problematicos.append({'EAN':ean_in, 'Cabecera':row.get(cN,''),
                                  'Motivo':f'EAN forma rara (len={len(core)})'}); continue
        if _tolerante:
            pa = _parse_precio_libre(row.get(cP, ''))
        elif PERFIL.get('precio_especial') == 'eur':
            pa = _num_eur(row.get(cP, ''))
        else:
            pa = _num(row.get(cP, ''))
        if PERFIL.get('col_pa_promo'):                    # DBLine: promo si >0, si no Listino
            promo = _num(row.get(PERFIL['col_pa_promo'], ''))
            if promo and promo > 0: pa = promo
        vol = None
        if PERFIL.get('col_volumen'):
            try:
                mv = _mejor_volumen(row.get(PERFIL['col_volumen'], ''))
                # solo si es un descuento REAL: rebaja el suelto pero no es un valor basura absurdo
                if mv and pa and (pa*MIN_RATIO_LOTE) <= mv[1] < pa:
                    vol = {'uds': mv[0], 'pa': round(mv[1], 4)}
            except Exception:
                vol = None
        _cu = PERFIL.get('col_url')
        url = str(row.get(_cu,'')).strip() if _cu else ''
        filas.append({'ean_in':ean_in, 'core':core, 'variantes':variantes_ean(core),
                      'nombre':row.get(cN,''), 'marca':MARCA, 'pa':pa,
                      'es_chase':_es_case, 'es_caja6':_es_caja6,
                      'uds_caja': (partir_ean(ean_in)[2] or UNIDADES_CASE_TCG),   # C12 son 12, no 6
                      'volumen':vol, 'url':url,
                      # La ficha viaja desde la FACTURA en el catalogo de B2 (columna 'producto_id'). Los
                      # feeds de proveedor no traen esa columna -> '' -> None. Se escribe tal cual en
                      # escaner_detalle.producto_id (camino de la factura): NO se cruza EAN->ficha.
                      'producto_id': (str(row.get('producto_id','')).strip() or None)})
    print(f"Disponibles a escanear: {len(filas)} | fuera por estado/stock: {fuera_disp} | "
          f"EAN problematicos: {len(problematicos)} | CHASE: {sum(f['es_chase'] for f in filas)}")
    if chase_sueltos:
        print(f"Chase SUELTO descartado (solo se compra en caja de 6): {len(chase_sueltos)} "
              f"-> listados en la hoja 'Descartados'")

# ============================================================
# DEDUP del proveedor: UNA sola fila por clave de memoria
# ------------------------------------------------------------
# Tras la regla del chase todavia puede haber VARIAS filas con la misma clave
# (mismo EAN + misma categoria) y precios distintos. Eso ya NO es chase: es
# BASURA DE DATOS del proveedor. Caso real de OcioStock: el mismo producto,
# con el MISMO nombre, mandado 3 veces a 37,99 / 8,50 / 6,99.
# La memoria guarda un solo precio por clave, asi que sin deduplicar siempre
# habria una fila que no cuadra -> 'cambio_precio' eterno y re-escaneo.
# Nos quedamos con la MAS BARATA: es el mejor coste y, sobre todo, es
# DETERMINISTA, que es lo que corta el bucle. Las descartadas se listan.
# ============================================================
_uni, _dups = {}, []
for f in filas:
    k = (norm(f['core']), bool(f['es_chase']))
    prev = _uni.get(k)
    if prev is None:
        _uni[k] = f
        continue
    if f['pa'] is not None and (prev['pa'] is None or f['pa'] < prev['pa']):
        barato, caro = f, prev
    else:
        barato, caro = prev, f
    _uni[k] = barato
    _dups.append({'EAN': caro['ean_in'], 'Cabecera': caro.get('nombre', ''),
                  'Motivo': f"Duplicado del proveedor: me quedo con {barato['pa']} "
                            f"(esta venia a {caro['pa']})"})
if _dups:
    filas = list(_uni.values())
    print(f"Duplicados del proveedor: {len(_dups)} filas descartadas (misma clave, "
          f"varios precios) -> me quedo con la MAS BARATA. Van a la hoja 'Descartados'.")

# ============================================================
# COTEJO (bloque 2): construir el IDF y decidir si aplica. Va AQUI, en el punto
# comun tras el dedup, para que valga tanto para MOLOKA como para los proveedores
# de fichero (el sitio original ~"Disponibles a escanear" solo cubria ficheros).
# ============================================================
# 🔒 El IDF se calcula sobre el catalogo de HOY: son sus propias palabras las que
# dicen cuales distinguen y cuales no. Nada de listas de stopwords a mano.
construir_idf([f['nombre'] for f in filas])
# Con deteccion tolerante, si no hay columna de nombre el motor usa la del EAN
# (`det['nombre'] or det['ean']`). Comparar un EAN con un titulo de Amazon no
# casaria NUNCA -> el cotejo se APAGA para ese proveedor. (COTEJO_ACTIVO es el
# global del modulo de cotejo; aqui, a nivel de script, se reasigna directo.)
COTEJO_ACTIVO = not (_tolerante and not det.get('nombre'))
if not COTEJO_ACTIVO:
    print(f"COTEJO: apagado para {PROVEEDOR} (el fichero no trae columna de nombre).")
else:
    print(f"COTEJO: activo, modo '{COTEJO_MODO}' | IDF sobre {_NDOC} nombres.")

# ============================================================
# GUARDARRAIL caja-vs-suelta (ajuste bloque 3, medido 28-jul). Aqui, tras el
# dedup, ya estan todas las filas. Para cada EAN base con ficha CAJA y ficha
# SUELTA, se compara el precio POR UNIDAD de la caja (con el ÷6 que le tocaria a
# su perfil) contra el de la suelta. Solo puede saltar si un perfil aplica el ÷6
# donde NO debe: es la red de seguridad para un perfil mal declarado. MARCA, no
# borra (regla del #56: nada desaparece).
# ============================================================
aviso_caja = {}    # ean_in de la caja -> texto de aviso
_sueltas_pa = {}   # core normalizado -> PA de la ficha suelta (la mas barata si hay varias)
for f in filas:
    if not f.get('es_caja6') and f.get('pa'):
        k = norm(f['core'])
        if k not in _sueltas_pa or f['pa'] < _sueltas_pa[k]:
            _sueltas_pa[k] = f['pa']
for f in filas:
    if f.get('es_caja6') and f.get('pa'):
        _su = _sueltas_pa.get(norm(f['core']))
        if _su:
            _pa_ud = (f['pa'] / (f.get('uds_caja') or UNIDADES_CASE_TCG)
                      if PERFIL.get('precio_caja6') == 'caja' else f['pa'])
            _av = aviso_caja_incoherente(_pa_ud, _su)
            if _av:
                aviso_caja[f['ean_in']] = _av
if aviso_caja:
    print(f"GUARDARRAIL caja-vs-suelta: {len(aviso_caja)} cajas con precio/ud incoherente "
          f"vs su ficha suelta (marcadas, NO borradas).")

# ============================================================
# Celda 5 - cruce con Supabase (productos propios + stock para 'En mi BD')
# ============================================================
sup = {}
try:
    _rows = []; _d = 0
    while True:
        res = sb.table('productos').select('ean,asin,iva_pct,stock_moloka,stock_fba').eq('activo',True).range(_d, _d+999).execute()
        if not res.data: break
        _rows.extend(res.data)
        if len(res.data) < 1000: break
        _d += 1000
    for p in _rows:
        if p.get('ean'): sup[norm(p['ean'])] = p
    print(f"Supabase: {len(sup)} EANs propios")
except Exception as ex:
    print("AVISO sin cruce Supabase:", ex)

def _sup(core):
    for v in [norm(core), core, '0'+core]:
        if v in sup: return sup[v]
    return None
def iva_es_de(core):
    s = _sup(core)
    if s and s.get('iva_pct') not in (None,''):
        try: return float(s['iva_pct'])
        except Exception: pass
    return IVA_DEFAULT_ES
def es_propio(core): return _sup(core) is not None
def en_bd_txt(core):
    s = _sup(core)
    if not s: return ''
    return f"OK Alm:{s.get('stock_moloka',0)} FBA:{s.get('stock_fba',0)}"

# ============================================================
# Celda 5b - memoria viva del proveedor (nuevos / reaparicion / cambio / agotado)
# ============================================================
mem = {}   # (ean_norm, es_case) -> {'pa':, 'presente':, 'ean_db':}
try:
    _rows = []; _d = 0
    while True:
        res = (sb.table('escaner_memoria')
                 .select('ean,es_case,pa,presente')
                 .eq('proveedor', PROVEEDOR)
                 .range(_d, _d+999).execute())
        if not res.data: break
        _rows.extend(res.data)
        if len(res.data) < 1000: break
        _d += 1000
    for m in _rows:
        mem[(norm(m['ean']), bool(m['es_case']))] = {
            'pa': m.get('pa'),
            'presente': bool(m.get('presente', True)),
            'ean_db': m['ean'],
        }
    print(f"Memoria {PROVEEDOR}/{MARCA}: {len(mem)} EANs conocidos")
except Exception as ex:
    print("AVISO: no se pudo leer la memoria, se trata todo como NUEVO:", ex)

filas_hoy = list(filas)
claves_hoy = {(norm(f['core']), bool(f['es_chase'])) for f in filas_hoy}

def estado_mem(f):
    k = (norm(f['core']), bool(f['es_chase']))
    if k not in mem: return 'nuevo'
    info = mem[k]
    if not info['presente']: return 'reaparicion'
    pa_ant = info['pa']
    if f['pa'] is None or pa_ant is None: return 'sin_cambios'
    if abs(float(f['pa']) - float(pa_ant)) > 0.01: return 'cambio_precio'
    return 'sin_cambios'

for f in filas:
    f['_estado_mem'] = estado_mem(f)
cnt = Counter(f['_estado_mem'] for f in filas)
print(f"Nuevos: {cnt.get('nuevo',0)} | Reaparecidos: {cnt.get('reaparicion',0)} | "
      f"Cambio precio: {cnt.get('cambio_precio',0)} | Sin cambios: {cnt.get('sin_cambios',0)}")

ausentes = [(k, info) for k, info in mem.items() if info['presente'] and k not in claves_hoy]
print(f"Agotados/desaparecidos desde la ultima vez: {len(ausentes)}")

if MODO == 'nuevos':
    filas = [f for f in filas if f['_estado_mem'] in ('nuevo','reaparicion','cambio_precio')]
print(f"A escanear (modo '{MODO}'): {len(filas)} productos")

# ============================================================
# CHECKPOINT: id comun para la cache de rank (Fase 1) y el progreso (Fase 2).
# El id identifica ESTE escaneo (mismo catalogo + mismos parametros). Al relanzar un
# escaneo cortado reanuda; con otro catalogo empieza limpio. TODO CON RED: si la cache
# falla, se consulta Keepa normal y el escaner funciona como hoy.
# ============================================================
import hashlib
_eans_filas = sorted(str(f['ean_in']) for f in filas)
_ckpt_id = hashlib.md5(('|'.join([PROVEEDOR, str(MARCA), MODO, str(RANK_MAXIMO)] + _eans_filas)).encode()).hexdigest()[:16]

# --- Cache de rank (Fase 1): lo unico caro de la Fase 1 es consultar el rank a Keepa.
# Guardamos por lote lo minimo que usa registra() (asin, rank actual, rank 90, EANs).
# Si el escaneo se corta y se relanza, la Fase 1 se rehace LEYENDO de aqui: 0 tokens.
RANKCACHE_PATH = f'{CARPETA_CKPT}/_rankcache_{_ckpt_id}.json'
_rankcache = {}
try:
    _d = sb.storage.from_(BUCKET).download(RANKCACHE_PATH)
    _rankcache = json.loads(_d.decode('utf-8')) or {}
    if _rankcache:
        print(f">>> Cache de rank: {len(_rankcache)} lotes ya consultados (reanudo Fase 1 sin re-pagar).")
except Exception:
    _rankcache = {}

# 🔒 Subir REDUCE_VER cada vez que cambien los campos de _reduce_prod: invalida
# la cache vieja, que no traeria los campos nuevos (p.ej. 'title', anadido en el
# bloque 2 del cotejo). Sin esto, la primera corrida cotejaria contra titulos vacios.
REDUCE_VER = 'v2'
def _clave_lote(codigos, domain):
    return hashlib.md5((REDUCE_VER + '|' + domain + '|' + '|'.join(map(str, codigos))).encode()).hexdigest()[:16]

def _reduce_prod(prod):
    # asin, stats, eanList, upcList + TITLE: viene en la MISMA respuesta ya pagada
    # de Fase 1 y hasta ahora se tiraba. Sin el, registra() no puede cotejar.
    st = prod.get('stats') or {}
    return {'asin': prod.get('asin'), 'title': prod.get('title') or '',
            # salesRankDrops30 y listedSince son SEÑALES DE IDENTIDAD (#85) que ya vienen GRATIS en esta
            # misma respuesta de Fase 1 (la llamada usa stats=90) y hasta ahora se tiraban. Solo alimentan
            # sondeo_keepa; NO tocan el cotejo ni la eleccion por rank. Cache vieja sin ellas -> None (ok).
            'stats': {'current': st.get('current'), 'avg90': st.get('avg90'),
                      'salesRankDrops30': st.get('salesRankDrops30')},
            'listedSince': prod.get('listedSince'),
            'eanList': prod.get('eanList'), 'upcList': prod.get('upcList')}

def keepa_rank(codigos, domain='ES'):
    clave = _clave_lote(codigos, domain)
    if clave in _rankcache:
        return _rankcache[clave]                 # de la caja: 0 tokens
    prods = keepa_query(codigos, product_code_is_asin=False, domain=domain, stats=90, history=0)
    if prods is None:
        return None
    _rankcache[clave] = [_reduce_prod(p) for p in prods]
    return _rankcache[clave]

def _guardar_rankcache():
    try:
        sb.storage.from_(BUCKET).upload(RANKCACHE_PATH, json.dumps(_rankcache).encode(),
                                        {'upsert': 'true', 'content-type': 'application/json'})
    except Exception as _e:
        print("AVISO cache de rank (no guardada, sigo igual):", _e)

# ============================================================
# Celda 6 - FASE 1: filtro de rank (Keepa ES, 1 token)
# ============================================================
IDX_RANK, IDX_NEW, IDX_BBOX_LAND = 3, 1, 18   # 18 = BUY_BOX_SHIPPING (buy box CON envio = aterrizada, como el v1)
candidatos, ambiguos = {}, []
cands_por_ean = {}        # ein -> [todos los candidatos] (para el cotejo). candidatos guarda solo el ganador por rank.
cotejo_info = {}          # ein -> veredicto del cotejo (se rellena mas abajo; declarado aqui para que exista aunque filas este vacio)
pasan, sin_rank, no_encontrados = {}, [], []

if filas:
    def pasa_filtro(r_act, r_90):
        return any(r and r>0 and r<=RANK_MAXIMO for r in (r_act, r_90))

    var_norm = {f['ean_in']: {norm(v) for v in f['variantes']} for f in filas}
    fila_por_ean = {f['ean_in']: f for f in filas}

    def cod_pref(f): return f['core']
    def cods_reserva(f): return [v for v in f['variantes'] if v != cod_pref(f)]

    def keyrank(c): return c['r_90'] if c['r_90'] and c['r_90']>0 else 10**12
    def registra(prod, pool, vistos):
        asin = prod.get('asin')
        if not asin: return
        st = prod.get('stats') or {}
        cur, a90 = st.get('current') or [], st.get('avg90') or []
        r_act = cur[IDX_RANK] if len(cur)>IDX_RANK else -1
        r_90  = a90[IDX_RANK] if len(a90)>IDX_RANK else -1
        eans = {norm(str(e)) for e in (prod.get('eanList') or [])+(prod.get('upcList') or [])}
        for ein in pool:
            if not (var_norm[ein] & eans): continue
            f = fila_por_ean[ein]
            cand = {'ean_in':ein,'asin':asin,'r_act':r_act,'r_90':r_90,'fila':f,'propio':es_propio(f['core'])}
            cand['title'] = prod.get('title') or ''
            # señales de identidad para sondeo_keepa (#85), gratis en la respuesta de Fase 1.
            cand['rank_drops_30d'] = st.get('salesRankDrops30')
            cand['listed_since']   = prod.get('listedSince')
            cands_por_ean.setdefault(ein, []).append(cand)      # <- se guardan TODOS (para el cotejo)
            if ein in candidatos:
                prev = candidatos[ein]
                ambiguos.append({'EAN':ein,'asin_elegido':cand['asin'] if keyrank(cand)<keyrank(prev) else prev['asin']})
                if keyrank(cand)<keyrank(prev): candidatos[ein]=cand
            else: candidatos[ein]=cand
            vistos.add(ein)

    def pasada(cod_por_ean, etiqueta, lote_size=None):
        pool = list(cod_por_ean.keys())
        codigos = sorted({cod_por_ean[e] for e in pool})
        _ls = lote_size or LOTE_FASE1
        lotes = [codigos[i:i+_ls] for i in range(0,len(codigos),_ls)]
        vistos = set()
        print(f"{etiqueta}: {len(pool)} productos, {len(codigos)} codigos, {len(lotes)} lotes")
        for n,lote in enumerate(lotes,1):
            prods = keepa_rank(lote, domain='ES')
            if prods is None:
                # 🔒 Se apunta QUE EAN se ha quedado sin preguntar. No es lo mismo
                # que "Keepa no lo tiene": es que no llegamos a preguntarlo.
                _cods = set(lote)
                _perdidos = {e for e, c in cod_por_ean.items() if c in _cods}
                LOTES_PERDIDOS.append({'fase': 'F1', 'etiqueta': etiqueta,
                                       'lote': n, 'n_codigos': len(lote)})
                EANS_NO_PREGUNTADOS.update(_perdidos)
                print(f"  lote {n}/{len(lotes)} NO resuelto tras reintentos -> se salta este lote "
                      f"({len(_perdidos)} EAN quedan SIN PREGUNTAR)")
                continue
            for prod in prods: registra(prod, pool, vistos)
            if n % 5 == 0: _guardar_rankcache()
            print(f"  lote {n}/{len(lotes)} | tokens {api.tokens_left}")
            if _cerca_del_corte():
                _guardar_rankcache()
                print(">>> Cerca del corte de GitHub: guardo el rank y me relanzo para seguir.")
                _relanzarme(); sys.exit(0)
        _guardar_rankcache()
        return vistos

    vistos = pasada({f['ean_in']: cod_pref(f) for f in filas}, "Fase 1 (1 codigo/producto)")
    for ronda in (0, 1):
        faltan = {f['ean_in'] for f in filas} - vistos
        rint = {}
        for f in filas:
            if f['ean_in'] in faltan:
                rs = cods_reserva(f)
                if len(rs) > ronda: rint[f['ean_in']] = rs[ronda]
        if rint:
            vistos |= pasada(rint, f"Fase 1 reintento {ronda+1} (variante alternativa)")

    # ============================================================
    # 🔒 RONDA DE RESCATE (pedida por Fernando, 28-jul)
    # ------------------------------------------------------------
    # Un lote perdido casi siempre lo tira un hipo TRANSITORIO de Keepa. Antes
    # de darlo por no preguntado y dejarlo para manana, se reintenta AQUI, al
    # final de la Fase 1: se espera a que Keepa respire y se vuelve a preguntar
    # en lotes mas pequenos todavia. Va en Fase 1 y no al final del script a
    # proposito: lo que se rescata entra en el flujo COMPLETO (Fase 2, 3 paises,
    # Excel, memoria). Si se rescatara despues, tendriamos rank pero no informe.
    # Lo que ni asi se resuelve queda marcado y manana se reintenta solo.
    # ============================================================
    RESCATE_INTENTOS = int(os.environ.get('RESCATE_INTENTOS', '2'))
    RESCATE_ESPERA   = int(os.environ.get('RESCATE_ESPERA', '60'))   # s antes de cada rescate
    LOTE_RESCATE     = max(10, LOTE_FASE1 // 2)
    for _nr in range(1, RESCATE_INTENTOS + 1):
        _pend = EANS_NO_PREGUNTADOS - vistos
        if not _pend:
            break
        if _cerca_del_corte():
            print(f">>> RESCATE {_nr}: cerca del corte de GitHub, no lo intento (quedan {len(_pend)} EAN).")
            break
        if not api.tokens_left or api.tokens_left <= 0:
            print(f">>> RESCATE {_nr}: sin tokens Keepa, no lo intento (quedan {len(_pend)} EAN).")
            break
        print(f">>> RESCATE {_nr}/{RESCATE_INTENTOS}: {len(_pend)} EAN se quedaron sin preguntar. "
              f"Espero {RESCATE_ESPERA}s y reintento en lotes de {LOTE_RESCATE}.")
        time.sleep(RESCATE_ESPERA)
        vistos |= pasada({e: cod_pref(fila_por_ean[e]) for e in _pend},
                         f"RESCATE {_nr}", lote_size=LOTE_RESCATE)
        _rescatados = len(_pend) - len(EANS_NO_PREGUNTADOS - vistos)
        print(f">>> RESCATE {_nr}: recuperados {_rescatados} de {len(_pend)}.")

    # 🔒 Un EAN que fallo en la 1a pasada puede haberse resuelto en un reintento
    # con variante alternativa o en el rescate: solo sigue "sin preguntar" si NO
    # acabo visto. Esta resta es la que manda; LOTES_PERDIDOS es solo historial.
    EANS_NO_PREGUNTADOS.difference_update(vistos)

    # ============================================================
    # COTEJO: elegir el ASIN que SE PARECE, no el que mas vende.
    #   activo (default) -> se aplica SIEMPRE el elegido; los dudosos se marcan.
    #   off              -> como antes del bloque 2 (se elige por rank).
    # 🔒 NADA DESAPARECE. El cotejo ORDENA y AVISA; nunca saca un producto del
    # informe. Un ASIN dudoso VISIBLE (se ve y se corrige) es mejor que una
    # ausencia invisible. len(candidatos) es INVARIANTE en este bloque: solo se
    # reasigna candidatos[ein], jamas se hace pop ni se anade a no_encontrados.
    # ============================================================
    cotejo_info = {}          # ein -> {'veredicto','detalle','asin_cotejo','difiere'}
    _n_eval = _n_dud = 0
    if COTEJO_MODO != 'off' and COTEJO_ACTIVO:
        for ein, c in list(candidatos.items()):
            cands = cands_por_ean.get(ein) or [c]
            nom = fila_por_ean[ein]['nombre']
            elegido, veredicto, detalle = elegir_candidato(nom, cands, keyrank)
            if elegido is None:
                continue                                    # cands nunca vacio; defensivo
            difiere = elegido['asin'] != c['asin']
            candidatos[ein] = elegido                        # <- SIEMPRE se aplica el elegido
            cotejo_info[ein] = {'veredicto':veredicto, 'detalle':detalle,
                                'asin_cotejo':elegido['asin'], 'difiere':difiere}
            if veredicto != 'n/d':
                _n_eval += 1
                if str(veredicto).startswith('⚠'):
                    _n_dud += 1
        # Guardarrail de panico: ya no apaga nada (no hay nada destructivo). Si el
        # ratio de dudosos supera el 50%, avisa: es la senal de afinar el criterio.
        _fiar, _msg_panico = cotejo_de_fiar(_n_eval, _n_dud)
        if not _fiar:
            print("!!! " + _msg_panico)
        n_dif = sum(1 for v in cotejo_info.values() if v['difiere'])
        print(f"COTEJO ({COTEJO_MODO}): {_n_eval} evaluados | {_n_dud} dudosos | "
              f"{n_dif} donde el cotejo cambia la eleccion del rank.")

    for ein,c in candidatos.items():
        tiene = (c['r_act'] and c['r_act']>0) or (c['r_90'] and c['r_90']>0)
        if c['propio'] or pasa_filtro(c['r_act'],c['r_90']): pasan[ein]=c
        elif (not tiene) and INCLUIR_SIN_RANK: pasan[ein]=c
        elif not tiene: sin_rank.append(c)
    for f in filas:
        if f['ean_in'] not in vistos:
            # 🔒 No mentir en el motivo: "Keepa sin ASIN" solo si de verdad se
            # pregunto y no lo tenia. Si el lote se perdio, se dice.
            _motivo = ('NO PREGUNTADO (lote perdido por fallo de Keepa)'
                       if f['ean_in'] in EANS_NO_PREGUNTADOS else 'Keepa sin ASIN')
            no_encontrados.append({'EAN':f['ean_in'],'Cabecera':f['nombre'],'Motivo':_motivo})
amb_eans = {a['EAN'] for a in ambiguos}
print(f"\nCon ASIN: {len(candidatos)} | PASAN: {len(pasan)} | sin rank: {len(sin_rank)} | "
      f"no encontrados: {len(no_encontrados)} | ambiguos: {len(ambiguos)}")

# ============================================================
# Celda 7 - FASE 2: informe ES/IT/FR (3 tok/pais, buybox sin offers)
# ============================================================
def _url_imagen(prod):
    # Mismo montaje VERIFICADO que el v1 en produccion (moloka_actualizar_nube.py:1882): Keepa marco
    # imagesCSV como DEPRECATED; el campo bueno es 'images' (lista), con imagesCSV de reserva. La URL
    # es https://m.media-amazon.com/images/I/<nombre>. None si no hay imagen (hueco vacio, no icono roto).
    imgs = prod.get('images')
    if isinstance(imgs, list) and imgs:
        el = imgs[0]; nombre = None
        if isinstance(el, dict):
            for k in ('l', 'large', 'hiRes', 'm', 'medium', 'image', 'name'):
                if el.get(k): nombre = el[k]; break
        elif isinstance(el, str):
            nombre = el
        if nombre:
            nombre = str(nombre)
            return nombre if nombre.startswith('http') else ('https://m.media-amazon.com/images/I/' + nombre)
    csv = prod.get('imagesCSV')   # campo viejo, de reserva
    if csv:
        primer = str(csv).split(',')[0].strip()
        if primer:
            return 'https://m.media-amazon.com/images/I/' + primer
    return None

def datos_pais(asin, dom):
    res = keepa_query([asin], product_code_is_asin=True, domain=dom, stats=90, history=0, buybox=True)
    # 🔒 None (fallo de red tras reintentos) NO es lo mismo que [] (Keepa no
    # tiene ese ASIN en ese pais). Antes se mezclaban y el pais perdido por red
    # desaparecia en silencio del informe.
    if res is None:
        PAISES_PERDIDOS.append((asin, dom))
        return None
    if not res: return None
    p = res[0]; st = p.get('stats') or {}
    cur, a90 = st.get('current') or [], st.get('avg90') or []
    # Buy box ATERRIZADA: current[18] (BUY_BOX_SHIPPING) es el precio CON envio,
    # lo que paga el cliente. Fallback a buyBoxPrice (pelado) si el 18 no viene.
    # Mismo mecanismo que el v1 en produccion (moloka_actualizar_nube.py, 1932-1943).
    bb_land = cur[IDX_BBOX_LAND] if len(cur)>IDX_BBOX_LAND else None
    bb_pel  = st.get('buyBoxPrice')
    if bb_land and bb_land>0:
        precio = bb_land/100
        canal = 'BB-FBA' if st.get('buyBoxIsFBA') else 'BB-FBM'
    elif bb_pel and bb_pel>0:
        precio = bb_pel/100
        canal = 'BB-FBA' if st.get('buyBoxIsFBA') else 'BB-FBM'
    else:
        new = cur[IDX_NEW] if len(cur)>IDX_NEW else -1
        precio = new/100 if new and new>0 else None
        canal = 'SIN BB' if precio else 'sin precio'
    fba = p.get('fbaFees') or {}
    fee = fba.get('pickAndPackFee')
    def _pos(x): return x if (x is not None and x>=0) else None
    return {'precio':precio,'canal':canal,'ref_pct':p.get('referralFeePercentage'),
            'fee':fee/100 if fee else None,
            'rank_act':cur[IDX_RANK] if len(cur)>IDX_RANK else -1,
            'rank90':a90[IDX_RANK] if len(a90)>IDX_RANK else -1,
            'n_of':_pos(st.get('totalOfferCount')),
            'vendidos':p.get('monthlySold'),
            'imagen':_url_imagen(p),
            'titulo':(p.get('title') or '')}

lista = list(pasan.values())

# === CHECKPOINT: reanudar la Fase 2 si un escaneo grande se corto a medias ===
# Guarda el progreso cada CKPT_CADA candidatos en una carpeta APARTE (escaner_ckpt/,
# que el boton NO vacia al relanzar). Al arrancar, si hay checkpoint de ESTE mismo
# escaneo (mismos candidatos), reanuda desde donde quedo. Todo CON RED: si algo del
# checkpoint falla, el escaner sigue como siempre (empieza de cero, no se rompe).
import hashlib
CKPT_CADA = 50
CKPT_PATH = f'{CARPETA_CKPT}/_ckpt_{_ckpt_id}.json'   # _ckpt_id: el mismo de la Fase 1

infos = []
_eans_hechos = set()
try:
    _d = sb.storage.from_(BUCKET).download(CKPT_PATH)
    _prev = json.loads(_d.decode('utf-8'))
    if isinstance(_prev, list) and _prev:
        infos = _prev
        _eans_hechos = {str(it.get('ean')) for it in infos}
        print(f">>> CHECKPOINT: reanudo un escaneo a medias ({len(infos)} de {len(lista)} ya hechos).")
except Exception:
    pass   # sin checkpoint o ilegible -> empezar de cero, exactamente como hoy

def _guardar_ckpt():
    try:
        sb.storage.from_(BUCKET).upload(CKPT_PATH, json.dumps(infos).encode('utf-8'),
                                        {'upsert': 'true', 'content-type': 'application/json'})
    except Exception as _e:
        print("AVISO checkpoint (no se guardo, sigo igual):", _e)

print(f"Fase 2: {len(lista)} candidatos x 3 paises"
      + (f" | {len(infos)} ya hechos, faltan {len(lista)-len(infos)}" if infos else ""))
_nuevos = 0
for i,c in enumerate(lista,1):
    if c['ean_in'] in _eans_hechos:
        continue                       # ya escaneado en una pasada anterior -> saltar
    f = c['fila']
    item = {'nombre':f['nombre'],'ean':c['ean_in'],'asin':c['asin'],'marca':f['marca'],
            'pa':f['pa'],'core':f['core'],'es_chase':f['es_chase'],'propio':c['propio'],
            'producto_id':f.get('producto_id'),   # viaja desde la factura (B2); None en feeds de proveedor
            'volumen':f.get('volumen'),'url':f.get('url',''),'titulo_amz':'',
            'ambiguo':c['ean_in'] in amb_eans,'paises':{},
            'cotejo':(cotejo_info.get(c['ean_in']) or {}).get('veredicto','—'),   # para marcar dudosos en Telegram
            'coherencia_caja':aviso_caja.get(c['ean_in'],''),   # caja/ud incoherente vs su suelta (guardarrail)
            'case_de_6':bool(f.get('es_caja6'))}
    for dom in ('ES','IT','FR'):
        d = datos_pais(c['asin'], dom)
        if d:
            item['paises'][dom] = d
            if dom=='ES' and not item['titulo_amz']:
                item['titulo_amz'] = d.get('titulo','')
    item['coincide'] = _coincide_titulo(item['nombre'], item['titulo_amz'])
    infos.append(item)
    _nuevos += 1
    if i%50==0:
        print(f"  {i}/{len(lista)} | tokens {api.tokens_left}")
    if _nuevos % CKPT_CADA == 0:
        _guardar_ckpt()
    if _cerca_del_corte():
        _guardar_ckpt()
        print(">>> Cerca del corte de GitHub: guardo el progreso y me relanzo para seguir.")
        _relanzarme(); sys.exit(0)
print(f"Fase 2 completa: {len(infos)} productos")
# Escaneo completo: el checkpoint ya no hace falta -> borrar
try: sb.storage.from_(BUCKET).remove([CKPT_PATH, RANKCACHE_PATH])
except Exception: pass

# ============================================================
# Celda 7.5 - CHASE FUNKO de HEO: puente escaner_chase_asin (ASIN a mano)
# Los que ya tienen ASIN se ESCANEAN (valorados /6, como una caja de 6); los
# pendientes van a la pestana Chase_manual. La puente esta CERRADA (RLS): se lee
# con SERVICE_KEY, NUNCA con la anon (lleva el precio de coste).
# ============================================================
chase_pendientes = []
if PROVEEDOR == 'HEO':
    _svc = os.environ.get('SUPABASE_SERVICE_KEY')
    if not _svc:
        print("AVISO: sin SUPABASE_SERVICE_KEY -> no leo la puente de chase (ni ASIN ni pestana).")
    else:
        try:
            sb_svc = create_client(os.environ['SUPABASE_URL'], _svc)
            _rows = (sb_svc.table('escaner_chase_asin')
                     .select('producto_heo,nombre,ean_caja,precio_caja,estado,imagen,link_amazon,asin')
                     .execute().data) or []
            _con_asin = [x for x in _rows if (x.get('asin') or '').strip() and x.get('estado') == 'disponible']
            chase_pendientes = [x for x in _rows if not (x.get('asin') or '').strip()]
            print(f">>> Puente chase: {len(_rows)} total | {len(_con_asin)} con ASIN disponibles | "
                  f"{len(chase_pendientes)} pendientes de ASIN.")
            for x in _con_asin:
                _as = x['asin'].strip()
                item = {'nombre': x.get('nombre', ''), 'ean': str(x.get('ean_caja') or ''), 'asin': _as,
                        'marca': 'Funko', 'pa': _num(x.get('precio_caja')), 'core': str(x.get('ean_caja') or ''),
                        'es_chase': True, 'propio': False, 'volumen': None, 'url': x.get('link_amazon', ''),
                        'titulo_amz': '', 'ambiguo': False, 'paises': {}, 'case_de_6': True}
                for dom in ('ES', 'IT', 'FR'):
                    d = datos_pais(_as, dom)
                    if d:
                        item['paises'][dom] = d
                        if dom == 'ES' and not item['titulo_amz']:
                            item['titulo_amz'] = d.get('titulo', '')
                item['coincide'] = _coincide_titulo(item['nombre'], item['titulo_amz'])
                infos.append(item)
            if _con_asin:
                print(f">>> Chase con ASIN escaneados y anadidos a resultados: {len(_con_asin)} (valorados /6).")
        except Exception as _e:
            print("AVISO: no se pudo leer/escanear la puente de chase (sigo con el escaneo normal):", _e)

# ============================================================
# Celda 8 - calculo (decision + orden por margen ES)
# ============================================================
def decision_de(margen):
    if margen is None: return 'Sin datos'
    if margen*100 >= 10: return 'COMPRAR'
    if margen*100 >= 1:  return 'VALORAR'
    return 'NO COMPRAR'

registros = []
for item in infos:
    iva = {'ES':iva_es_de(item['core']),'IT':IVA_IT,'FR':IVA_FR}
    pa = item['pa']
    # 🔒 Solo se divide donde el proveedor da el precio de la CAJA COMPLETA.
    # OcioStock lo da POR UNIDAD (11,99 €/ud, 71,94 € la caja): dividir alli
    # convertia un Funko de 9,99 € en uno de 1,66 € y sacaba 18 COMPRAR falsos.
    # Default: NO dividir (perfil sin 'precio_caja6'). Adivinarlo es lo que rompio esto.
    if item.get('case_de_6') and pa and PERFIL.get('precio_caja6') == 'caja':
        pa = pa / (item.get('uds_caja') or UNIDADES_CASE_TCG)
    item['_pa_efectivo'] = pa
    margen_es = None; paises_out = {}
    for dom in ('ES','IT','FR'):
        d = item['paises'].get(dom)
        if d and d.get('precio') and pa and d.get('ref_pct') is not None and d.get('fee') is not None:
            r = calc_rentabilidad(d['precio'], pa, d['ref_pct'], d['fee'], iva[dom],
                                  almacen=ALMACEN, com_digitales=COM_DIGITALES)
            paises_out[dom] = {**d,'iva':iva[dom],'beneficio':r['beneficio'],
                               'roi':r['roi'],'margen':r['margen'],'decision':decision_de(r['margen'])}
            if dom == 'ES': margen_es = r['margen']
        elif d:
            paises_out[dom] = {**d,'iva':iva[dom],'beneficio':None,'roi':None,'margen':None,'decision':'Sin datos'}
    item['_paises_calc'] = paises_out
    item['_margen_es'] = margen_es
    registros.append(item)

registros.sort(key=lambda x: x['_margen_es'] if x['_margen_es'] is not None else -10**9, reverse=True)
n_mandar = sum(1 for it in registros for d in it['_paises_calc'].values() if d['decision'] == 'COMPRAR')
print(f"Productos: {len(registros)} | filas COMPRAR (algun pais): {n_mandar}")

# ============================================================
# Celda 9 - Excel final (1 fila por pais, formulas vivas, semaforo)
# ============================================================
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule, CellIsRule

COLS = ['Nombre','EAN','ASIN','Marca','PA (€)','País','Rank actual','Rank 90d','Vendidos/mes',
        'Precio venta (€)','Canal BB','Nº ofertas','% Comisión',
        'Com. Amazon (€)','Fee Logística (€)','Almacén (€)','Promo activa',
        'Beneficio (€)','ROI','Margen','Decisión','En mi BD','EAN ambiguo','Amazon (título)','Coincide',
        'Cotejo','Cotejo (detalle)','Coherencia caja','OcioStock']
L = {name:get_column_letter(i+1) for i,name in enumerate(COLS)}
DOM_AMZ = {'ES':'amazon.es','IT':'amazon.it','FR':'amazon.fr'}

wb = Workbook(); ws = wb.active; ws.title='Análisis'
ws.append(COLS)

r = 1
for item in registros:
    en_bd = en_bd_txt(item['core'])
    amb = 'AMBIGUO' if item['ambiguo'] else ''
    _cot = cotejo_info.get(item['ean']) or {}   # veredicto del cotejo para este producto ('—' si no hay)
    for dom in ('ES','IT','FR'):
        d = item['_paises_calc'].get(dom)
        if not d:
            d = {'rank_act':None,'rank90':None,'vendidos':None,'precio':None,'canal':'sin datos',
                 'n_of':None,'ref_pct':None,'fee':None,'iva':None,'decision':'Sin datos'}
        r += 1
        pct = (d['ref_pct']/100*COM_DIGITALES) if d.get('ref_pct') is not None else None
        div = (1+d['iva']) if d.get('iva') else None
        ws.append([
            item['nombre'], item['ean'], item['asin'], item['marca'], item['_pa_efectivo'], dom,
            d['rank_act'] if d['rank_act'] and d['rank_act']>0 else None,
            d['rank90'] if d['rank90'] and d['rank90']>0 else None,
            d['vendidos'], d['precio'], d['canal'], d['n_of'], pct,
            f"={L['Precio venta (€)']}{r}*{L['% Comisión']}{r}" if pct is not None else None,
            d['fee'], ALMACEN, None,
            (f"=({L['Precio venta (€)']}{r}/{div})-{L['PA (€)']}{r}-{L['Com. Amazon (€)']}{r}"
             f"-{L['Fee Logística (€)']}{r}-{L['Almacén (€)']}{r}") if (div and d['precio'] and pct is not None) else None,
            f"={L['Beneficio (€)']}{r}/{L['PA (€)']}{r}" if (div and d['precio'] and pct is not None and item['pa']) else None,
            f"={L['Beneficio (€)']}{r}/{L['Precio venta (€)']}{r}" if (div and d['precio'] and pct is not None) else None,
            d['decision'], en_bd, amb,
            item.get('titulo_amz',''), item.get('coincide','?'),
            _cot.get('veredicto','—'), _cot.get('detalle','—'),
            item.get('coherencia_caja','') or '—',
            ('Ver ficha ↗' if item.get('url') else '')])
        cell = ws.cell(row=r, column=3)
        cell.hyperlink = f"https://www.{DOM_AMZ[dom]}/dp/{item['asin']}"
        cell.font = Font(color='0563C1', underline='single')
        if item.get('url'):
            cocel = ws.cell(row=r, column=len(COLS))   # ultima columna = OcioStock
            cocel.hyperlink = item['url']
            cocel.font = Font(color='0563C1', underline='single')

last = ws.max_row
def fmt(colname, code):
    c = L[colname]
    for row in range(2,last+1):
        ws[f'{c}{row}'].number_format = code
for nm in ['PA (€)','Precio venta (€)','Com. Amazon (€)','Fee Logística (€)','Almacén (€)','Beneficio (€)']:
    fmt(nm,'0.00')
fmt('% Comisión','0.00%'); fmt('ROI','0.0%'); fmt('Margen','0.0%')

for c in range(1,len(COLS)+1):
    ws.cell(row=1,column=c).font = Font(bold=True)
anchos = {'Nombre':50,'EAN':14,'ASIN':12,'Marca':12,'En mi BD':20,'Decisión':15,
          'Amazon (título)':50,'Coincide':11,'Cotejo':16,'Cotejo (detalle)':46,
          'Coherencia caja':46,'OcioStock':13}
for nm,w in anchos.items(): ws.column_dimensions[L[nm]].width = w

ws.freeze_panes = 'A2'

def _cf_fill(hexcolor): return PatternFill(start_color=hexcolor, end_color=hexcolor, fill_type='solid')
# Tabla + semaforo SOLO si hay al menos una fila de datos. Un escaneo 'nuevos' sin
# novedades deja registros vacio -> last=1 -> rango invertido (U2:U1) que PETA openpyxl.
if last >= 2:
    tab = Table(displayName='T_Analisis', ref=f"A1:{get_column_letter(len(COLS))}{last}")
    tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=False,
                                        showColumnStripes=False, showFirstColumn=False, showLastColumn=False)
    ws.add_table(tab)
    dec = L['Decisión']; rng_dec = f'{dec}2:{dec}{last}'
    ws.conditional_formatting.add(rng_dec, FormulaRule(formula=[f'ISNUMBER(SEARCH("NO COMPRAR",{dec}2))'],
        fill=_cf_fill('FFC7CE'), font=Font(color='9C0006'), stopIfTrue=True))
    ws.conditional_formatting.add(rng_dec, FormulaRule(formula=[f'ISNUMBER(SEARCH("VALORAR",{dec}2))'],
        fill=_cf_fill('FFEB9C'), font=Font(color='9C6500'), stopIfTrue=True))
    ws.conditional_formatting.add(rng_dec, FormulaRule(formula=[f'ISNUMBER(SEARCH("COMPRAR",{dec}2))'],
        fill=_cf_fill('C6EFCE'), font=Font(color='006100'), stopIfTrue=True))
    ws.conditional_formatting.add(rng_dec, FormulaRule(formula=[f'ISNUMBER(SEARCH("Sin datos",{dec}2))'],
        fill=_cf_fill('E7E6E6'), font=Font(color='808080'), stopIfTrue=True))
    ws.conditional_formatting.add(f"{L['Margen']}2:{L['Margen']}{last}",
        CellIsRule(operator='greaterThan', formula=['0.1'], font=Font(color='006100')))
    coi = L['Coincide']; rng_coi = f'{coi}2:{coi}{last}'
    ws.conditional_formatting.add(rng_coi, FormulaRule(formula=[f'ISNUMBER(SEARCH("NO",{coi}2))'],
        fill=_cf_fill('FFC7CE'), font=Font(color='9C0006')))
    # Semaforo del COTEJO (mismo espiritu que Coincide): DUDOSO rojo (revisar el
    # ASIN), n/d gris (no habia con que cotejar), OK verde (rank y cotejo coinciden).
    cot = L['Cotejo']; rng_cot = f'{cot}2:{cot}{last}'
    ws.conditional_formatting.add(rng_cot, FormulaRule(formula=[f'ISNUMBER(SEARCH("DUDOSO",{cot}2))'],
        fill=_cf_fill('FFC7CE'), font=Font(color='9C0006'), stopIfTrue=True))
    ws.conditional_formatting.add(rng_cot, FormulaRule(formula=[f'ISNUMBER(SEARCH("n/d",{cot}2))'],
        fill=_cf_fill('E7E6E6'), font=Font(color='808080'), stopIfTrue=True))
    ws.conditional_formatting.add(rng_cot, FormulaRule(formula=[f'ISNUMBER(SEARCH("OK",{cot}2))'],
        fill=_cf_fill('C6EFCE'), font=Font(color='006100'), stopIfTrue=True))
    # Coherencia caja: rojo si el precio/ud de la caja no cuadra con el de la suelta.
    ccj = L['Coherencia caja']; rng_ccj = f'{ccj}2:{ccj}{last}'
    ws.conditional_formatting.add(rng_ccj, FormulaRule(formula=[f'ISNUMBER(SEARCH("INCOHERENTE",{ccj}2))'],
        fill=_cf_fill('FFC7CE'), font=Font(color='9C0006')))
    ws.conditional_formatting.add(f'A2:{get_column_letter(len(COLS))}{last}',
        FormulaRule(formula=['ISODD(INT((ROW()-2)/3))'], fill=_cf_fill('D9D9D9')))

def hoja(nombre, regs):
    w = wb.create_sheet(nombre)
    if regs:
        ks = list(regs[0].keys()); w.append(ks)
        for x in regs: w.append([x.get(k) for k in ks])
    else: w.append(['(vacio)'])
hoja('Descartados', problematicos + no_encontrados + chase_sueltos + _dups)
hoja('Ambiguos', ambiguos)
hoja('Sin_rank', [{'EAN':c['ean_in'],'ASIN':c['asin'],'Nombre':c['fila']['nombre'],
                   'rank_act':c['r_act'],'rank90':c['r_90']} for c in sin_rank])

# ============================================================
# Pestana "Precio por lote": escenario con descuento por VOLUMEN (OcioStock).
# La hoja Analisis se queda igual (precio unitario). Aqui recalculamos el beneficio
# con el precio del LOTE y ponemos a la derecha del todo las unidades minimas para
# lograr ese precio. Solo entran los productos cuyo lote REBAJA el precio suelto.
# ============================================================
COLS_LOTE = ['Nombre','EAN','ASIN','Marca','País','Precio venta (€)','PA suelto (€)',
             'PA lote (€)','Ahorro/ud (€)','Beneficio lote (€)','Margen lote','Decisión lote',
             'Uds. para ese precio','Coherencia']
filas_lote = []
for item in registros:
    vol = item.get('volumen')
    if not vol:
        continue
    pa_lote = vol['pa']; uds = vol['uds']
    pa_suelto = item.get('_pa_efectivo')
    # 🔒 Guardarrail RELATIVO (nunca un umbral absoluto de precio: Fernando compra
    # Funkos a 2,99 y llaveros a 1 €). Un descuento por volumen NEGATIVO es imposible
    # por definicion: el lote no puede salir MAS CARO que el suelto -> firma de un
    # parseo roto. Se MARCA, no se borra (regla del #56: nada desaparece).
    _coh = ''
    if pa_suelto and pa_lote and pa_lote > pa_suelto * 1.05:
        _coh = f'INCOHERENTE: lote ({pa_lote}) mas caro que suelto ({pa_suelto})'
    for dom in ('ES','IT','FR'):
        d = item['_paises_calc'].get(dom)
        if not d or not d.get('precio') or d.get('ref_pct') is None or d.get('fee') is None:
            continue
        rr = calc_rentabilidad(d['precio'], pa_lote, d['ref_pct'], d['fee'], d['iva'],
                               almacen=ALMACEN, com_digitales=COM_DIGITALES)
        filas_lote.append([
            item['nombre'], item['ean'], item['asin'], item['marca'], dom,
            round(d['precio'], 2),
            round(pa_suelto, 2) if pa_suelto else None,
            round(pa_lote, 2),
            round(pa_suelto - pa_lote, 2) if pa_suelto else None,
            round(rr['beneficio'], 2),
            round(rr['margen'], 4),
            decision_de(rr['margen']),
            uds,
            _coh,
        ])
filas_lote.sort(key=lambda x: (x[10] if x[10] is not None else -9), reverse=True)

wl = wb.create_sheet('Precio por lote')
wl.append(COLS_LOTE)
for fl in filas_lote:
    wl.append(fl)
for c in range(1, len(COLS_LOTE)+1):
    wl.cell(row=1, column=c).font = Font(bold=True)
if filas_lote:
    LL = {name: get_column_letter(i+1) for i, name in enumerate(COLS_LOTE)}
    lastL = wl.max_row
    for nm in ['Precio venta (€)','PA suelto (€)','PA lote (€)','Ahorro/ud (€)','Beneficio lote (€)']:
        for row in range(2, lastL+1):
            wl[f'{LL[nm]}{row}'].number_format = '0.00'
    for row in range(2, lastL+1):
        wl[f'{LL["Margen lote"]}{row}'].number_format = '0.0%'
    decL = LL['Decisión lote']; rngL = f'{decL}2:{decL}{lastL}'
    wl.conditional_formatting.add(rngL, FormulaRule(formula=[f'ISNUMBER(SEARCH("NO COMPRAR",{decL}2))'],
        fill=_cf_fill('FFC7CE'), font=Font(color='9C0006'), stopIfTrue=True))
    wl.conditional_formatting.add(rngL, FormulaRule(formula=[f'ISNUMBER(SEARCH("VALORAR",{decL}2))'],
        fill=_cf_fill('FFEB9C'), font=Font(color='9C6500'), stopIfTrue=True))
    wl.conditional_formatting.add(rngL, FormulaRule(formula=[f'ISNUMBER(SEARCH("COMPRAR",{decL}2))'],
        fill=_cf_fill('C6EFCE'), font=Font(color='006100'), stopIfTrue=True))
    wl.column_dimensions[LL['Nombre']].width = 50
    wl.column_dimensions[LL['Uds. para ese precio']].width = 18
    wl.column_dimensions[LL['Coherencia']].width = 46
    wl.freeze_panes = 'A2'
print(f"Pestana 'Precio por lote': {len(filas_lote)} filas con descuento por volumen")

# ============================================================
# Pestana "Chase_manual": Funko chase de HEO SIN ASIN todavia. Pega el ASIN en
# Supabase (tabla escaner_chase_asin) usando el enlace de busqueda; la proxima
# corrida ya lo cruza sola y desaparece de aqui.
# ============================================================
if PROVEEDOR == 'HEO':
    wc = wb.create_sheet('Chase_manual')
    COLS_CHASE = ['Nombre', 'Código HEO', 'EAN caja', 'Precio caja (€)', 'Precio /6 (€)',
                  'Estado', 'Imagen', 'Buscar en Amazon', 'ASIN (pégalo en Supabase)']
    wc.append(COLS_CHASE)
    for c in range(1, len(COLS_CHASE) + 1):
        wc.cell(row=1, column=c).font = Font(bold=True)
    rc = 1
    for x in chase_pendientes:
        rc += 1
        _pc = _num(x.get('precio_caja'))
        wc.append([x.get('nombre', ''), x.get('producto_heo', ''), str(x.get('ean_caja') or ''),
                   round(_pc, 2) if _pc else None,
                   round(_pc / UNIDADES_CASE_TCG, 2) if _pc else None,
                   x.get('estado', ''),
                   'Ver imagen ↗' if x.get('imagen') else '',
                   'Buscar ↗' if x.get('link_amazon') else '', ''])
        if x.get('imagen'):
            _ci = wc.cell(row=rc, column=7); _ci.hyperlink = x['imagen']; _ci.font = Font(color='0563C1', underline='single')
        if x.get('link_amazon'):
            _cl = wc.cell(row=rc, column=8); _cl.hyperlink = x['link_amazon']; _cl.font = Font(color='0563C1', underline='single')
    for _cw, _w in ((1, 55), (2, 16), (3, 16), (7, 14), (8, 16), (9, 26)):
        wc.column_dimensions[get_column_letter(_cw)].width = _w
    for _col in (4, 5):
        for _row in range(2, wc.max_row + 1):
            wc.cell(row=_row, column=_col).number_format = '0.00'
    wc.freeze_panes = 'A2'
    print(f"Pestana 'Chase_manual': {len(chase_pendientes)} Funko chase pendientes de ASIN.")

# NOTA: la pestana viaja dentro del Excel, y el Excel solo se guarda si hay
# algun COMPRAR (ver mas abajo). Decidido asi a proposito: los pendientes
# viven en la tabla escaner_chase_asin, que es donde se pegan los ASIN.
# Si no hay ningun COMPRAR, NO se genera ni se sube el Excel (limpieza: un escaneo
# sin chollos no aporta nada y cada Excel ocupa ~2 MB). El REGISTRO en la biblioteca
# se guarda IGUAL (n_comprar=0, fichero=NULL) para no romper la alarma de persistencia.
_sin_excel = (n_mandar == 0)
if _sin_excel:
    print(f"Sin COMPRAR en esta pasada ({n_mandar}): NO genero ni subo el Excel. "
          "El registro en la biblioteca se guarda igual (fichero vacio).")
else:
    wb.save(ARCHIVO_SALIDA)
    print("Guardado local:", ARCHIVO_SALIDA, "| filas:", last-1)

# ============================================================
# SUBIR EL EXCEL A STORAGE + REGISTRAR EN LA BIBLIOTECA (escaner_resultados)
# ============================================================
nombre_xlsx = os.path.basename(ARCHIVO_SALIDA)
ruta_storage = f'{CARPETA_RESULTADOS}/{nombre_xlsx}'
subido_ok = False
if not _sin_excel:
    try:
        with open(ARCHIVO_SALIDA, 'rb') as fp:
            sb.storage.from_(BUCKET).upload(
                ruta_storage, fp.read(),
                {'content-type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                 'upsert':'true'})
        # verificar que esta en Storage
        _res = sb.storage.from_(BUCKET).list(CARPETA_RESULTADOS) or []
        subido_ok = any(o.get('name') == nombre_xlsx for o in _res)
        print(f"Excel subido a Storage: {ruta_storage} | verificado: {subido_ok}")
    except Exception as ex:
        print("ATENCION: no se pudo subir el Excel a Storage:", ex)

# RED DE SEGURIDAD: un insert puede devolver OK sin que la fila llegue a persistir.
# No nos fiamos del insert ni del print: releemos la fila por su id (la verdad es la
# BD, no el log) y gritamos si no esta. Hoy NO se conoce ningun caso real: las filas
# que parecian faltar resultaron ser limpieza MANUAL desde la app (comprobado con
# pg_stat_user_tables: 451 insertadas / 435 borradas). Esto cubre el dia que falle
# de verdad. NO tumba la corrida: solo avisa.
_biblioteca_ok = False
_biblioteca_id = None
try:
    _fila_bib = {
        'proveedor': PROVEEDOR, 'marca': MARCA, 'modo': MODO,
        'rank_maximo': RANK_MAXIMO,
        'n_productos': len(registros), 'n_comprar': n_mandar,
        'n_nuevos': cnt.get('nuevo',0), 'n_reaparecidos': cnt.get('reaparicion',0),
        'n_cambios': cnt.get('cambio_precio',0), 'n_agotados': len(ausentes),
        'fichero': ruta_storage if subido_ok else None,
        'tokens_restantes': int(api.tokens_left),
    }
    # 🔒 EL AVISO VIVE EN EL DATO, no en el log (el log de Actions caduca).
    # Requiere la migracion: alter table escaner_resultados add column
    #   lotes_perdidos int default 0, eans_no_preguntados int default 0;
    # Si las columnas aun no existen, se reintenta SIN ellas: nunca se pierde el registro.
    _extra_bib = {'lotes_perdidos': len(LOTES_PERDIDOS),
                  'eans_no_preguntados': len(EANS_NO_PREGUNTADOS)}
    try:
        _resp_bib = sb.table('escaner_resultados').insert({**_fila_bib, **_extra_bib}).execute()
    except Exception as _e_cols:
        print(f"AVISO: no pude guardar lotes_perdidos/eans_no_preguntados ({_e_cols}). "
              f"Falta la migracion. Registro el escaneo sin esas columnas.")
        _resp_bib = sb.table('escaner_resultados').insert(_fila_bib).execute()
    _filas_bib = getattr(_resp_bib, 'data', None) or []
    _biblioteca_id = _filas_bib[0].get('id') if _filas_bib else None
    if _biblioteca_id is not None:
        # Verificacion DURA contra la BD (no el objeto del insert, no el log).
        # Va en su PROPIO try: si lo que falla es la RELECTURA, el insert ya fue bien
        # y decir "no se pudo registrar" seria mentira en el log. En ese caso avisamos
        # de que no se pudo verificar y NO damos falsa alarma.
        try:
            _chk_bib = sb.table('escaner_resultados').select('id').eq('id', _biblioteca_id).limit(1).execute()
            _biblioteca_ok = bool(getattr(_chk_bib, 'data', None))
        except Exception as _e_chk:
            print(f"AVISO: el insert SI fue bien (id={_biblioteca_id}) pero no pude releer "
                  f"la fila para verificarla: {_e_chk}")
            _biblioteca_ok = True
    if _biblioteca_ok:
        print(f"Escaneo registrado y VERIFICADO en la biblioteca (escaner_resultados), id={_biblioteca_id}.")
    else:
        print("!!! CRITICO: el insert en escaner_resultados dijo OK pero la BD NO devuelve la fila "
              f"(id={_biblioteca_id}). El escaneo NO quedo en la biblioteca: revisalo.")
except Exception as ex:
    print("!!! CRITICO: no se pudo registrar en escaner_resultados (el Excel puede estar en Storage):", ex)

# ============================================================
# La EJECUCION: el identificador de ESTA pasada (llave de escaner_detalle y de sondeo_keepa).
# ============================================================
# 🔴 CONTRATO EXPLICITO entre los dos repos (antes era un acuerdo tacito y por eso mordia):
#    - Cuando la pasada es de la APP (factura), el id lo MANDA ella en el recado (SOLICITUD['ejecucion']),
#      porque B2 ya le devolvio ?ejecucion=... a la pantalla ANTES de que el escaner corra. Si el escaner
#      se inventara otro, la pantalla de la factura saldria vacia.
#    - Si NO viene (escaneo de proveedor / manual de Fernando), el escaner se inventa uno UNICO con el
#      sello TS. JAMAS el basename a secas: la app v1 sube el catalogo con el nombre del fichero del disco
#      (index.html: nombre = file.name), asi que dos subidas de "catalogo.csv" colisionarian en el unique
#      (ejecucion, ean_norm, pais) NULLS NOT DISTINCT -> abortaria el lote de 500 y el freno se lo tragaria
#      en silencio. El TS hace unica cada pasada. Nadie depende de que dos repos coincidan por casualidad.
_EJECUCION = SOLICITUD.get('ejecucion') or (
    f'{os.path.basename(catalogo_local)}_{TS}' if catalogo_local else f'{PROVEEDOR}_{TS}')

# ============================================================
# Celda 9b - PELICULA: el DETALLE por producto y pais en escaner_detalle (informe de la factura)
# ============================================================
# 🔴 SOLO se ANADE una escritura de lo que el escaner YA calculo: NO se toca la formula, ni el
#    semaforo, ni escaner_memoria. Va en su PROPIO try y NO puede tumbar el escaneo: si la tabla no
#    esta (migracion sin aplicar), la RLS, o lo que sea, se AVISA en el log y el Excel sale IGUAL.
#    Ese fichero corre todos los dias para Elena y no puede depender de una tabla nueva.
# 🔒 Se escribe TAMBIEN en MIS_COMPRAS (perfil efimero): "no toca memoria" es escaner_memoria, NO la
#    pelicula. Sin esto el informe de la factura nace vacio, que es lo que veniamos a arreglar.
# 🔒 almacen/com_digitales son los MISMOS globales que se pasan a calc_rentabilidad (ALMACEN,
#    COM_DIGITALES): el valor que se USO, no una constante copiada al lado. Asi el guardian "no
#    cuadra" del informe (#84) no salta persiguiendo un fantasma.
try:
    # 🔒 escaner_detalle lleva pa/beneficio/margen (precio de COSTE): se escribe con la SERVICE KEY,
    #    NUNCA con la anon (`sb`, linea 50; `anon` esta REVOCADA de la tabla a proposito). Es la regla
    #    de la casa, la misma que el puente de chase (linea 1527). Sin la key: AVISO y NO se escribe
    #    (no se abre anon por un informe); el escaneo y el Excel siguen intactos.
    _svc_det = os.environ.get('SUPABASE_SERVICE_KEY')
    sb_det = create_client(os.environ['SUPABASE_URL'], _svc_det) if _svc_det else None
    _ejec_det = _EJECUCION   # id de la pasada: recado['ejecucion'] (app) o basename_TS (proveedor). Ver arriba.
    # 🔒 La fecha de la pasada sale del SELLO DE ARRANQUE (TS, linea 428), NO de now(): en un escaneo
    #    largo que cruce medianoche, now() (que se evalua AQUI, al final) fecharia la pelicula un dia
    #    por delante del Excel (que usa TS). Ademas asi la fecha viene del sello de la pasada y deja de
    #    ser una excepcion a la regla del now(). TS = 'YYYYMMDD_HHMM'.
    _fecha_det = TS[0:4] + '-' + TS[4:6] + '-' + TS[6:8]
    _filas_det = []
    for _it in registros:
        # ¿Se dividio el PA por el tamano de caja? MISMA condicion que uso el calculo; no se recalcula
        # nada, se lee de item tal cual estaba (item['pa'] es el PA crudo; _pa_efectivo, el ya dividido).
        _pa_div = bool(_it.get('case_de_6') and _it.get('pa') and PERFIL.get('precio_caja6') == 'caja')
        for _dom, _d in (_it.get('_paises_calc') or {}).items():
            _rk = _d.get('rank_act')
            _filas_det.append({
                'ejecucion': _ejec_det,
                'ean': _it.get('ean'),
                'nombre': _it.get('nombre'),
                # asin = el ELEGIDO POR RANK (candidatos[ein]): vale para ENLAZAR (Keepa) y AUDITAR, JAMAS
                # para escribir en una ficha (lo dice el comentario de la columna). imagen = foto del
                # producto. producto_id = la ficha que viajo desde la FACTURA (B2); None en feeds de proveedor.
                'asin': _it.get('asin'),
                'imagen': _d.get('imagen'),
                'producto_id': _it.get('producto_id'),
                'pais': _dom,
                'pa': _it.get('_pa_efectivo'),
                'pa_dividido': _pa_div,
                'precio_venta': _d.get('precio'),
                'canal': _d.get('canal'),
                'ref_pct': _d.get('ref_pct'),
                'fee_fba': _d.get('fee'),
                'iva': _d.get('iva'),
                'almacen': ALMACEN,
                'com_digitales': COM_DIGITALES,
                'beneficio': _d.get('beneficio'),
                'roi': _d.get('roi'),
                'margen': _d.get('margen'),
                'decision': _d.get('decision'),
                'rank': _rk if (_rk or 0) > 0 else None,   # rank_act llega -1 cuando no hay rank
                # 'vendidos' = monthlySold de Keepa. VERIFICADO en Product.java del SDK oficial (3-ago):
                # es un contador de compras (`int monthlySold = 0`), non-negative o AUSENTE (->None en
                # el cliente python), NUNCA -1. El -1 de "no disponible" es de OTROS campos (rank,
                # numberOfItems...), NO de este. Por eso va SIN guarda -1 (a diferencia del rank); None
                # es legitimo -> NULL, "no lo se". No hay negativo del que protegerse.
                'vendidos': _d.get('vendidos'),
                'n_ofertas': _d.get('n_of'),
                'fecha_ejecucion': _fecha_det,
                'fichero': _ejec_det,
            })
    if not sb_det:
        print("AVISO: sin SUPABASE_SERVICE_KEY -> NO escribo escaner_detalle (va con service key, "
              "nunca con la anon; lleva precio de coste). El escaneo y el Excel siguen intactos.")
    elif _filas_det:
        _n_det = 0
        for _j in range(0, len(_filas_det), 500):   # por lotes: un scan de proveedor son miles de filas
            sb_det.table('escaner_detalle').insert(_filas_det[_j:_j + 500]).execute()
            _n_det += len(_filas_det[_j:_j + 500])
        print(f"escaner_detalle: {_n_det} filas de detalle escritas (ejecucion={_ejec_det}).")
    else:
        print("escaner_detalle: no habia filas de detalle que escribir.")
except Exception as _ex_det:
    print("!!! CRITICO: no se pudo escribir escaner_detalle (el escaneo y el Excel SIGUEN intactos):", _ex_det)

# ============================================================
# Celda 9c - PELICULA: los CANDIDATOS del sondeo EAN->ASIN en sondeo_keepa (para el clasificador)
# ============================================================
# 🔴 El escaner YA resolvio el EAN->ASIN en Fase 1 y guardo TODOS los candidatos en cands_por_ean. Aqui
#    SOLO se guarda lo que ya se sabe: NO se escribe ASIN en ninguna ficha, NO se toca candidatos[ein]
#    (la eleccion por rank del Excel), ni la hoja Ambiguos, ni el Excel. El clasificador (clasificarSondeo,
#    en la app v2) leera esto y decidira; el rank del Excel dice quien vende mas, no que producto ES.
# 🔒 Su PROPIO try (independiente del de escaner_detalle) y la MISMA service key: sondeo_keepa tambien
#    esta cerrada a la anon. Si falla (tabla, RLS, red), AVISA y el escaneo y el Excel siguen intactos.
try:
    _svc_snd = os.environ.get('SUPABASE_SERVICE_KEY')
    sb_snd = create_client(os.environ['SUPABASE_URL'], _svc_snd) if _svc_snd else None
    _lote_snd = _EJECUCION   # MISMO id que escaner_detalle.ejecucion: ata informe y sondeo de la pasada.
    _fecha_snd = TS[0:4] + '-' + TS[4:6] + '-' + TS[6:8]   # del SELLO de arranque, no now() (regla de 9b)

    def _ean_norm_py(cod):
        # Replica de moloka_ean_norm() para DEDUP local por la clave unica de la tabla (que indexa por
        # ean_norm, columna GENERADA). Sin 're' (no esta importado): filtramos digitos a mano. NO se
        # escribe en la tabla (ean_norm la calcula Postgres); solo se usa aqui para no chocar en el lote.
        s = ''.join(ch for ch in str(cod or '') if ch.isdigit()).lstrip('0')
        return s or None

    def _km_a_fecha(km):
        # listedSince de Keepa viene en "Keepa minutes" (epoca 2011-01-01 = +21564000 min sobre Unix).
        # 0 / negativo = desconocido -> None. listado_desde es DATE, asi que devolvemos 'YYYY-MM-DD'.
        if not km or km <= 0:
            return None
        try:
            return datetime.fromtimestamp((km + 21564000) * 60, timezone.utc).date().isoformat()
        except Exception:
            return None

    _filas_snd = []
    _vistos_snd = set()   # (ean_norm, asin_candidato) — la clave unica es NULLS NOT DISTINCT
    _dups_snd = 0
    # Una fila por candidato de cada EAN (cands_por_ean guarda TODOS los que devolvio Keepa en Fase 1).
    for _ein, _cands in cands_por_ean.items():
        for _pos, _cd in enumerate(_cands):
            _asin_c = _cd.get('asin')
            _k = (_ean_norm_py(_ein), _asin_c)
            if _k in _vistos_snd:
                _dups_snd += 1; continue
            _vistos_snd.add(_k)
            _rk = _cd.get('r_act')
            _filas_snd.append({
                'lote': _lote_snd,               # la EJECUCION (nombre del catalogo), no la factura
                'ean_consultado': _ein,          # ean_norm lo calcula Postgres (columna generada)
                'dominio': 'es',                 # Fase 1 pregunta a domain='ES'
                'asin_candidato': _asin_c,
                'posicion': _pos,                # indice dentro de cands_por_ean[ein]
                'titulo': (_cd.get('title') or None),
                'rank': _rk if (_rk or 0) > 0 else None,   # r_act llega -1 cuando no hay rank
                'rank_drops_30d': _cd.get('rank_drops_30d'),
                'listado_desde': _km_a_fecha(_cd.get('listed_since')),
                'fecha_sondeo': _fecha_snd,
                'fichero': _lote_snd,
                'crudo': {'asin': _asin_c, 'title': _cd.get('title'), 'r_act': _cd.get('r_act'),
                          'r_90': _cd.get('r_90'), 'ean_in': _cd.get('ean_in'), 'propio': _cd.get('propio'),
                          'rank_drops_30d': _cd.get('rank_drops_30d'), 'listedSince': _cd.get('listed_since')},
            })
    # 🔴 CENTINELA: por cada EAN que Keepa NO devolvio (esta en no_encontrados) una fila con
    #    asin_candidato = NULL. "Se pregunto y no habia" es un dato; el silencio no. Es lo que el
    #    unique NULLS NOT DISTINCT (sobre ean_norm) permite distinguir EAN a EAN.
    for _ne in no_encontrados:
        _ein = _ne.get('EAN')
        _k = (_ean_norm_py(_ein), None)
        if _k in _vistos_snd:
            _dups_snd += 1; continue
        _vistos_snd.add(_k)
        _filas_snd.append({
            'lote': _lote_snd, 'ean_consultado': _ein, 'dominio': 'es',
            'asin_candidato': None, 'posicion': None, 'titulo': None, 'rank': None,
            'rank_drops_30d': None, 'listado_desde': None,
            'fecha_sondeo': _fecha_snd, 'fichero': _lote_snd,
            'crudo': {'centinela': True, 'motivo': _ne.get('Motivo')},
        })
    if _dups_snd:
        # GRITARLO (regla del lote, CLAUDE.md §2): fila a fila esto era invisible.
        print(f"sondeo_keepa: {_dups_snd} filas DESCARTADAS por clave repetida "
              f"(lote,ean_norm,dominio,asin) -> se queda la primera.")
    if not sb_snd:
        print("AVISO: sin SUPABASE_SERVICE_KEY -> NO escribo sondeo_keepa (va con service key, nunca con "
              "la anon). El escaneo y el Excel siguen intactos.")
    elif _filas_snd:
        _n_snd = 0
        for _j in range(0, len(_filas_snd), 500):   # por lotes, como escaner_detalle
            sb_snd.table('sondeo_keepa').insert(_filas_snd[_j:_j + 500]).execute()
            _n_snd += len(_filas_snd[_j:_j + 500])
        print(f"sondeo_keepa: {_n_snd} filas de sondeo escritas (lote={_lote_snd}).")
    else:
        print("sondeo_keepa: no habia candidatos ni centinelas que escribir.")
except Exception as _ex_snd:
    print("!!! CRITICO: no se pudo escribir sondeo_keepa (el escaneo y el Excel SIGUEN intactos):", _ex_snd)

# ============================================================
# Celda 10 - actualizar la memoria del proveedor (presentes / agotados)
# ============================================================
ahora = datetime.now(timezone.utc).isoformat()
regs = []; vistos_up = set()
if PERFIL.get('efimero'):
    print(f"Perfil EFIMERO ({PROVEEDOR}): NO se escribe en escaner_memoria; ningun proveedor real se ve afectado.")
else:
    # 🔒🔒 LA LINEA QUE CORTA LA ACUMULACION.
    # Un EAN cuyo lote se perdio NO entra en la memoria: si lo grabaramos con el
    # PA de hoy, manana saldria 'sin_cambios' y el pase diario 'nuevos' lo
    # filtraria fuera para siempre. Dejandolo fuera, manana vuelve a salir
    # 'nuevo' y se reintenta. Un hipo de red no puede condenar a un producto.
    _saltados_mem = 0
    for f in filas_hoy:
        if f['ean_in'] in EANS_NO_PREGUNTADOS:
            _saltados_mem += 1
            continue
        k = (PROVEEDOR, norm(f['core']), bool(f['es_chase']))
        if k in vistos_up: continue
        vistos_up.add(k)
        regs.append({'proveedor':PROVEEDOR, 'ean':f['core'], 'es_case':bool(f['es_chase']),
                     'marca':MARCA, 'pa': float(f['pa']) if f['pa'] is not None else None,
                     'presente':True, 'fecha':ahora})
    # Agotados SOLO si el catalogo llego COMPLETO. Blindaje anti-vaciado:
    #  - catalogo vacio (0 filas) -> no marcar (fichero equivocado / marca inexistente)
    #  - catalogo PARCIAL (crudo < UMBRAL_PARCIAL de lo que hay en memoria) -> no marcar
    #    (descarga incompleta). Una REBAJA no reduce el nº de filas crudas -> NO salta aqui.
    if _saltados_mem:
        print(f"MEMORIA: {_saltados_mem} productos NO se graban (su lote se perdio y nunca se "
              f"pregunto a Keepa). Manana volveran a salir 'nuevo' y se reintentaran.")
    if not filas_hoy:
        print("Catalogo vacio (0 productos): NO se marcan agotados (evita falso vaciado de la memoria).")
    elif EANS_NO_PREGUNTADOS:
        # 🔒 Si tras el RESCATE siguen quedando EAN sin preguntar, el escaneo es
        # PARCIAL: marcar agotados aqui seria dar por desaparecido lo que ni
        # siquiera se ha llegado a mirar. (Si el rescate los recupero todos,
        # este bloque no salta y los agotados se marcan con normalidad.)
        print(f"BLINDAJE: {len(EANS_NO_PREGUNTADOS)} EAN sin preguntar tras el rescate -> escaneo "
              f"PARCIAL: NO se marcan agotados. La memoria queda intacta en esa parte.")
    elif N_CRUDO is not None and len(mem) > 0 and N_CRUDO < UMBRAL_PARCIAL * len(mem):
        print(f"BLINDAJE: catalogo PARCIAL ({N_CRUDO} filas crudas vs {len(mem)} en memoria, "
              f"<{int(UMBRAL_PARCIAL*100)}%): NO se marcan agotados. Huele a descarga incompleta o "
              f"fichero equivocado; la memoria queda intacta.")
    else:
        for (ean_norm, es_case), info in ausentes:
            k = (PROVEEDOR, ean_norm, es_case)
            if k in vistos_up: continue
            vistos_up.add(k)
            pa_ant = info.get('pa')
            regs.append({'proveedor':PROVEEDOR, 'ean':info['ean_db'], 'es_case':es_case,
                         'marca':MARCA, 'pa': float(pa_ant) if pa_ant is not None else None,
                         'presente':False, 'fecha':ahora})
    if not regs:
        print("Memoria sin cambios.")
    else:
        n_ok = 0
        for i2 in range(0, len(regs), 500):
            lote = regs[i2:i2+500]
            try:
                sb.table('escaner_memoria').upsert(lote, on_conflict='proveedor,ean,es_case').execute()
                n_ok += len(lote)
            except Exception as ex:
                print(f"  AVISO lote memoria {i2//500+1}: {ex}")
        n_pres = sum(1 for x in regs if x['presente']); n_aus = len(regs) - n_pres
        print(f"Memoria actualizada: {n_ok}/{len(regs)} ({n_pres} presentes, {n_aus} agotados) [{PROVEEDOR}/{MARCA}]")

# ============================================================
# LIMPIAR EL BUZON DEL ESCANER (recado + catalogo) - VERIFICADO
# Solo si el Excel se subio bien (si no, se deja para reintentar).
# ============================================================
def _buzon_escaner_pendiente():
    try:
        objs = sb.storage.from_(BUCKET).list(CARPETA_ESCANER) or []
        return [o['name'] for o in objs if o.get('name') and not o['name'].startswith('.')]
    except Exception as _e:
        print('AVISO al listar el buzon del escaner:', _e)
        return None

# Se limpia el buzon si el Excel subio BIEN o si deliberadamente no se genero
# (sin COMPRAR). Lo que NO se limpia es un fallo real de subida: eso se reintenta.
if subido_ok or _sin_excel:
    pend = _buzon_escaner_pendiente()
    if pend:
        for intento in (1, 2, 3):
            try:
                sb.storage.from_(BUCKET).remove([f'{CARPETA_ESCANER}/{n}' for n in pend])
            except Exception as _e:
                print(f'AVISO remove buzon escaner (intento {intento}):', _e)
            rest = _buzon_escaner_pendiente()
            if rest is None: break
            if not rest:
                print('Buzon del escaner limpiado y VERIFICADO.')
                break
            pend = rest
        else:
            print(f'ATENCION: el buzon del escaner NO quedo limpio: {pend}. Borralos a mano.')
else:
    print('El Excel no se subio: dejo el buzon del escaner intacto para reintentar.')

print("=== ESCANER FIN ===")

# ============================================================
# AVISO TELEGRAM: chollos de la pasada (productos COMPRAR). Se dispara SOLO si el
# workflow pasa TELEGRAM_TOKEN/CHAT_ID en el paso de escaneo (o sea, en los directores
# que quieran aviso: DBLine, OcioStock). TCG no las pasa aqui -> no duplica su aviso.
# Envuelto en try/except: si Telegram falla, la corrida ya termino igual.
# ============================================================
# 🔒 El aviso de parcial tiene que ir DENTRO del Telegram, que es por donde se
# entera Fernando de verdad. Hasta ahora el veredicto de integridad corria
# DESPUES de enviarlo, asi que un escaneo al que le faltaba medio catalogo
# mandaba el 🟢 de siempre. Y peor: con 0 COMPRAR no se enviaba NADA, o sea que
# el caso mas roto era el mas silencioso.
_PARCIAL = bool(EANS_NO_PREGUNTADOS or PAISES_PERDIDOS)
_aviso_parcial = None
if _PARCIAL:
    _trozos = []
    if EANS_NO_PREGUNTADOS: _trozos.append(f"{len(EANS_NO_PREGUNTADOS)} EAN sin preguntar a Keepa")
    if PAISES_PERDIDOS:     _trozos.append(f"{len(PAISES_PERDIDOS)} pares ASIN/pais perdidos")
    _aviso_parcial = ("⚠️ <b>ESCANEO PARCIAL</b>: " + " y ".join(_trozos) +
                      ". Los COMPRAR de abajo salen de un catalogo INCOMPLETO. "
                      "Relanza cuando Keepa vaya fina.")

try:
    _tg_token = os.environ.get('TELEGRAM_TOKEN')
    _tg_chat  = os.environ.get('TELEGRAM_CHAT_ID')
    if _tg_token and _tg_chat:
        _compras = []
        for _it in registros:
            _mejor = None
            for _dom, _d in (_it.get('_paises_calc') or {}).items():
                if _d.get('decision') == 'COMPRAR' and _d.get('margen') is not None:
                    if _mejor is None or _d['margen'] > _mejor[1]:
                        _mejor = (_dom, _d['margen'], _d.get('precio'))
            if _mejor:
                _compras.append((_it, _mejor))
        # 🔒 Se envia si hay compras O si el escaneo fue parcial: un escaneo roto
        # que no encuentra nada NO puede quedarse en silencio.
        if _compras or _PARCIAL:
            _lineas = [f"{'🔴' if _PARCIAL else '🟢'} <b>Director {PROVEEDOR}</b> ({MODO}): "
                       f"{len(_compras)} para COMPRAR"]
            if _aviso_parcial:
                _lineas.append(_aviso_parcial)
            for _it, (_dom, _mg, _pv) in _compras[:20]:
                _nom = str(_it.get('nombre') or '')[:45]
                _pvs = f"{_pv:.2f}€" if _pv else "s/precio"
                _dud = ' ⚠ASIN DUDOSO' if str(_it.get('cotejo','')).startswith('⚠') else ''
                _inc = ' ⚠CAJA INCOHERENTE' if _it.get('coherencia_caja') else ''
                _lineas.append(f"• {_nom} — {_mg*100:.0f}% — {_pvs} ({_dom}) — {_it.get('marca','')}{_dud}{_inc}")
            if len(_compras) > 20:
                _lineas.append(f"…y {len(_compras)-20} más (mira el Excel de la Biblioteca).")
            import requests as _rq_tg
            _rq_tg.post(f"https://api.telegram.org/bot{_tg_token}/sendMessage",
                        data={'chat_id': _tg_chat, 'text': "\n".join(_lineas),
                              'parse_mode': 'HTML', 'disable_web_page_preview': 'true'}, timeout=20)
            print(f">>> Telegram enviado: {len(_compras)} COMPRAR"
                  f"{' + AVISO DE ESCANEO PARCIAL' if _PARCIAL else ''}.")
        else:
            print(">>> Telegram: 0 COMPRAR y escaneo COMPLETO -> no se envia aviso.")
    else:
        print(">>> Telegram: sin claves en este paso -> no se envia (normal en TCG o app).")
except Exception as _e_tg:
    print("AVISO Telegram (no se envio, la corrida ya termino igual):", _e_tg)

# ============================================================
# Celda 12 - VEREDICTO DE INTEGRIDAD (lo ultimo que corre)
# ------------------------------------------------------------
# 🔒 Un escaneo al que le falta parte del catalogo NO es un escaneo bueno con
# un aviso: es un escaneo INCOMPLETO. Hasta hoy terminaba en VERDE y nadie se
# enteraba (run 27-jul: 86 de 186 productos sin preguntar, job en success).
# Va AL FINAL a proposito: el Excel ya esta subido, la biblioteca registrada y
# la memoria actualizada. Lo unico que cambia es que el run sale ROJO.
# ============================================================
# 🔒 El veredicto mira EANS_NO_PREGUNTADOS (lo que quedo sin preguntar DESPUES
# del rescate), no LOTES_PERDIDOS: si un lote se cayo pero el rescate lo
# recupero, el escaneo esta COMPLETO y no hay motivo para salir en rojo.
if LOTES_PERDIDOS and not EANS_NO_PREGUNTADOS:
    print(f"NOTA: {len(LOTES_PERDIDOS)} lotes se cayeron durante la corrida, pero la RONDA DE "
          f"RESCATE los recupero todos. Escaneo COMPLETO.")
if EANS_NO_PREGUNTADOS or PAISES_PERDIDOS:
    print("")
    print("=" * 64)
    print("!!! ESCANEO PARCIAL - NO TE FIES DE ESTE RESULTADO COMO COMPLETO")
    if EANS_NO_PREGUNTADOS:
        print(f"  Fase 1: {len(EANS_NO_PREGUNTADOS)} EAN NUNCA preguntados a Keepa "
              f"(ni en la pasada normal ni en el rescate).")
        for _lp in LOTES_PERDIDOS:
            print(f"    - intento fallido: {_lp['etiqueta']} lote {_lp['lote']} ({_lp['n_codigos']} codigos)")
    if PAISES_PERDIDOS:
        print(f"  Fase 2: {len(PAISES_PERDIDOS)} pares ASIN/pais perdidos.")
    print("  La memoria NO los ha marcado como vistos: el proximo pase los reintenta.")
    print("  Relanza el escaneo cuando Keepa vaya fino.")
    print("=" * 64)
    sys.exit(1)
print("=== INTEGRIDAD OK: todo el catalogo se pregunto a Keepa. ===")
