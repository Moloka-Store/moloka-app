#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# moloka_escaner_pro.py  —  ESCANER PRO (lee CSV del Visualizador de Keepa, CERO tokens)
# Mismo motor que el escaner de la nube (mismos perfiles, misma formula, misma decision,
# mismo Excel de resultados). Diferencia unica: los datos de Amazon (rank, precio, comision,
# tarifa FBA) NO vienen de la API con tokens, sino de los CSV del Visualizador (1 por pais).
# Validado 22/22 decisiones contra el escaner de tokens (DBLINE/Altri/ES, 1-jul-2026).

import pandas as pd, csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule, CellIsRule

# ===== Constantes (identicas al escaner) =====
IVA = {'ES': 0.21, 'IT': 0.22, 'FR': 0.20}
ALMACEN, COM_DIGITALES = 0.15, 1.03
UNIDADES_CASE_TCG = 6
DOM_AMZ = {'ES':'amazon.es','IT':'amazon.it','FR':'amazon.fr'}

# ===== Perfiles de proveedor (copiados del escaner de la nube) =====
PERFILES = {
    'DBLINE': {'tipo':'excel','sheet':0,'header':2,'col_marca':'Publisher','col_ean':'EAN',
               'col_nombre':'Descrizione','col_pa':'Prezzo (€)','col_pa_promo':'Prezzo promo (€)',
               'col_stock':'Disponibili'},
    'TCG': {'tipo':'excel','sheet':'Catálogo','header':0,'col_marca':'Marca','col_ean':'EAN',
            'col_nombre':'Cabecera','col_pa':'Precio','col_stock':'Stock','col_estado':'Estado producto',
            'estados_ok':['Disponible','Oferta','Saldo']},
    'BEMS': {'tipo':'csv','sep':';','header':0,'col_marca':'FABRICANT','col_ean':'EAN',
             'col_nombre':'TITRE UK','col_pa':'PA','col_stock':'STOCK'},
    'OSMA': {'tipo':'excel','sheet':0,'header':0,'col_marca':'Bezeichnung','col_ean':'EAN 1',
             'col_nombre':'Bezeichnung','col_pa':'Preis_','col_stock':'verfügbar','stock_especial':'osma'},
    'BIEDRO': {'tipo':'excel','sheet':0,'header':3,'col_marca':None,'col_ean':'Stück-EAN',
               'col_nombre':'Artikelbezeichnung','col_pa':'Stückpreis\nnetto','col_stock':None,
               'sin_columna_stock':True},
    'OCIOSTOCK': {'tipo':'csv','sep':';','header':0,'col_marca':'marca','col_ean':'ean',
                  'col_nombre':'nombre','col_pa':'precio_distribuidores','col_stock':'stock_disponible',
                  'col_volumen':'txt_precios_volumen','cajas_sufijo':'C6',      # volumen + cajas (C6 = caja de 6)
                  'col_url':'product_url'},   # enlace a la ficha de OcioStock (para verificar el volumen real en su web)
    'STOCKLIST': {'tipo':'excel','sheet':'Sheet1','header':0,'col_marca':'Brand','col_ean':'CodeBars',
                  'col_nombre':'ItemName','col_pa':'EUR','col_stock':'Available'},
    # Proveedores de Claude-in-Chrome (formato variable) se anaden con deteccion tolerante.
}

# ===== Columnas del CSV del Visualizador (por NOMBRE: robusto a reordenacion) =====
CSV_COLS = {
    'ean':'Códigos de producto: EAN', 'asin':'ASIN',
    'rank':'Clasificación de Ventas: Actual', 'rank90':'Clasificación de Ventas: Promedio de 90 días',
    'nuevo':'Nuevo: Actual', 'buybox':'Caja de Compra: Actual', 'es_fba':'Caja de Compra: Es FBA',
    'fba':'Tarifa FBA Pick&Pack', 'compct':'% de comisión de referencia',
    'nof':'Recuento ofertas nuevas: Actual',
    'vendidos':'Tendencias de ventas mensuales: Ventas mensuales (Último conocido)',
    'vendidos2':'Tendencias de ventas mensuales: Comprados el mes pasado',
    'nvar':'Recuento de variaciones',
    'titulo':'Título principal',
}

# ===== Helpers (identicos al escaner) =====
def core_ean(e):
    e = str(e).strip().upper()
    return e[:-1] if e.endswith('C') else e
def es_chase_ean(e): return str(e).strip().upper().endswith('C')
def variantes_ean(core):
    c, vs = core.strip(), set()
    if c.isdigit():
        vs.add(c); vs.add(c.lstrip('0'))
        if len(c)==12: vs.add('0'+c)
        if len(c)==13 and c.startswith('0'): vs.add(c[1:])
    return [v for v in vs if v]
def norm(code): return str(code).strip().lstrip('0')
def _num(x):
    try: return float(str(x).replace(',', '.').strip())
    except Exception: return None
def _stock_osma(x):
    s=str(x).strip()
    if not s: return None
    s=s.lstrip('>').strip().replace('.','')
    try: return float(s)
    except Exception: return None
def _num_csv(x):
    s=(x or '').strip().replace('%','').strip()
    if s in ('','-','—'): return None
    try: return float(s.replace(',', '.'))
    except Exception: return None
def _mejor_volumen(s):
    if not s: return None
    mejor=None
    for parte in str(s).split('|'):
        campos=parte.strip().split(':')
        if len(campos)<3: continue
        try: uds=int(float(campos[0])); precio=float(campos[-1])
        except Exception: continue
        if precio<=0: continue
        if mejor is None or precio<mejor[1]: mejor=(uds,precio)
    return mejor

def calc_rentabilidad(precio, pa, ref_pct, fee, iva, almacen=ALMACEN, com_digitales=COM_DIGITALES):
    base=precio/(1+iva)
    com_amazon=precio*(ref_pct/100)*com_digitales
    beneficio=base-pa-com_amazon-fee-almacen
    return dict(com_amazon=com_amazon, beneficio=beneficio,
                roi=(beneficio/pa if pa else 0), margen=(beneficio/precio if precio else 0))
def decision_de(margen):
    if margen is None: return 'Sin datos'
    if margen*100>=10: return 'COMPRAR'
    if margen*100>=1:  return 'VALORAR'
    return 'NO COMPRAR'

# Umbral anti-basura para precios de volumen: OcioStock mete valores fijos absurdos
# (p.ej. 5.99 de volumen en un casco de 109.99). Un descuento por volumen REAL rara vez
# baja del 50% del precio suelto; por debajo lo tratamos como dato basura y lo ignoramos.
MIN_RATIO_LOTE = 0.5

# ===== Lectura del catalogo del proveedor (misma logica que el escaner) =====
def leer_proveedor(prov, marca, ruta):
    P=PERFILES[prov]
    if P['tipo']=='excel':
        cat=pd.read_excel(ruta, sheet_name=P['sheet'], header=P['header'], dtype=str).fillna('')
    else:
        cat=pd.read_csv(ruta, sep=P.get('sep',';'), dtype=str, on_bad_lines='skip').fillna('')
    cM=P.get('col_marca')
    if marca and marca.strip().upper()!='TODAS' and cM:
        cat=cat[cat[cM].str.contains(marca, case=False, na=False)]
    # ---- CAJAS (OcioStock mete la caja como "EAN base + C6"): coste el mas barato ----
    import re as _re2
    cajas={}
    _suf=P.get('cajas_sufijo')
    if _suf:
        _patc=_re2.compile(r'^(\d{12,13})\s+'+_re2.escape(_suf)+r'$')
        for _,row in cat.iterrows():
            e=str(row.get(P['col_ean'],'')).strip()
            mm=_patc.match(e)
            if not mm: continue
            base=mm.group(1)
            pd_=_num(row.get(P['col_pa'],''))
            mv=_mejor_volumen(row.get(P['col_volumen'],'')) if P.get('col_volumen') else None
            cand=[]
            if pd_ and pd_>0: cand.append((6, pd_))          # comprar 1 caja (6 ud)
            if mv: cand.append((mv[0], mv[1]))               # tramo de volumen (uds, precio)
            if not cand: continue
            barato=min(cand, key=lambda x:x[1])              # el mas barato
            if base not in cajas or barato[1]<cajas[base]['pa']:
                cajas[base]={'pa':round(barato[1],4),'uds':barato[0],'nombre':row.get(P['col_nombre'],''),
                             'url':str(row.get(P['col_url'],'')).strip() if P.get('col_url') else ''}

    filas=[]; problematicos=[]; vistos=set()
    for _,row in cat.iterrows():
        if P.get('estados_ok') and str(row.get(P['col_estado'],'')).strip() not in P['estados_ok']:
            continue
        if P.get('sin_columna_stock') or not P.get('col_stock'):
            stock=1.0
        elif P.get('stock_especial')=='osma':
            stock=_stock_osma(row.get(P['col_stock'],''))
        else:
            stock=_num(row.get(P['col_stock'],''))
        if stock is None or stock<=0: continue
        ean_in=str(row[P['col_ean']]).strip()
        core=core_ean(ean_in)
        if (not core.isdigit()) or len(core) not in (12,13):
            problematicos.append({'EAN':ean_in,'Cabecera':row.get(P['col_nombre'],''),
                                  'Motivo':f'EAN forma rara (len={len(core)})'}); continue
        pa=_num(row.get(P['col_pa'],''))
        if P.get('col_pa_promo'):
            promo=_num(row.get(P['col_pa_promo'],''))
            if promo and promo>0: pa=promo
        vol=None
        if P.get('col_volumen'):
            mv=_mejor_volumen(row.get(P['col_volumen'],''))
            # solo si es un descuento REAL (por debajo del suelto pero no un valor basura absurdo)
            if mv and pa and (pa*MIN_RATIO_LOTE) <= mv[1] < pa:
                vol={'uds':mv[0],'pa':round(mv[1],4)}
        es_caja=False; uds_caja=None
        if core in cajas and (pa is None or cajas[core]['pa'] < pa):
            pa=cajas[core]['pa']; es_caja=True; uds_caja=cajas[core]['uds']; vol=None
        vistos.add(core)
        _cu=P.get('col_url')
        url=str(row.get(_cu,'')).strip() if _cu else ''
        filas.append({'ean_in':ean_in,'core':core,'nombre':row.get(P['col_nombre'],''),
                      'marca':marca,'pa':pa,'es_chase':es_chase_ean(ean_in),
                      'variantes':variantes_ean(core),'volumen':vol,
                      'es_caja':es_caja,'uds_caja':uds_caja,'url':url})
    # cajas cuyo individual NO esta (OcioStock solo vende la caja): usar el EAN base igual
    for base,info in cajas.items():
        if base in vistos: continue
        filas.append({'ean_in':base,'core':base,'nombre':info['nombre'],'marca':marca,
                      'pa':info['pa'],'es_chase':False,'variantes':variantes_ean(base),
                      'volumen':None,'es_caja':True,'uds_caja':info['uds'],'url':info.get('url','')})
    return filas, problematicos

# ===== Lectura del CSV del Visualizador (indexado por CADA EAN; celdas multi-EAN) =====
def leer_csv_visualizador(rutas):
    # Acepta UNA ruta (str) o VARIAS (list) -> las funde en un solo diccionario.
    # Asi un catalogo grande exportado de Keepa de 5.000 en 5.000 (6-7 CSV por pais)
    # se junta en un unico dataset y sale UN solo Excel.
    if isinstance(rutas, str): rutas=[rutas]
    data={}
    for ruta in rutas:
      with open(ruta, encoding='utf-8-sig', newline='') as f:
        rr=csv.reader(f); H=next(rr); rows=list(rr)
      idx={c:i for i,c in enumerate(H)}
      # Columna del titulo de Amazon (robusto al nombre exacto del export de Keepa)
      _tit=None
      for _cand in ('Título principal','Titulo principal','Título','Titulo','Title'):
        if _cand in idx: _tit=_cand; break
      if _tit is None:
        for _c in H:
          _cl=str(_c).lower()
          if 'tulo' in _cl or 'title' in _cl: _tit=_c; break
      def col(row,key):
        c=CSV_COLS.get(key); return row[idx[c]] if (c and c in idx) else ''
      for row in rows:
        es_fba=col(row,'es_fba').strip().lower()
        rec=dict(asin=(col(row,'asin').strip() or None),
                 rank=_num_csv(col(row,'rank')), rank90=_num_csv(col(row,'rank90')),
                 nuevo=_num_csv(col(row,'nuevo')), buybox=_num_csv(col(row,'buybox')),
                 es_fba=(es_fba in ('true','verdadero','sí','si','1')),
                 fba=_num_csv(col(row,'fba')), compct=_num_csv(col(row,'compct')),
                 nof=_num_csv(col(row,'nof')),
                 vendidos=(_num_csv(col(row,'vendidos')) or _num_csv(col(row,'vendidos2'))),
                 nvar=_num_csv(col(row,'nvar')),
                 titulo=((row[idx[_tit]].strip() if (_tit and idx[_tit]<len(row)) else '')))
        for e in col(row,'ean').split(','):
            e=norm(e)
            if e: data[e]=rec
    return data

def buscar(fila, csvdata):
    for v in fila['variantes']:
        vn=norm(v)
        if vn in csvdata: return csvdata[vn]
    return None

import unicodedata, re as _re
_STOP_TIT={'the','and','with','de','del','la','el','los','las','un','una','uno','con','para','por',
           'pop','figura','figure','set','pack','edition','deluxe','vinilo','vinyl',
           'peluche','plush','muneco','doll','juguete','juguetes','toy','coche','coches','car'}
def _tokens_tit(t):
    t=unicodedata.normalize('NFKD',str(t or '')).encode('ascii','ignore').decode()
    t=_re.sub(r'[^a-zA-Z0-9]+',' ',t).lower()
    return {w for w in t.split() if len(w)>=3 and w not in _STOP_TIT}
def _coincide_titulo(nombre_prov, titulo_amz):
    # '?' = sin titulo de Amazon (no marcamos). SI = comparten palabra distintiva. NO = nada en comun.
    if not str(titulo_amz or '').strip(): return '?'
    a=_tokens_tit(nombre_prov); b=_tokens_tit(titulo_amz)
    if not a or not b: return '?'
    return 'SÍ' if (a & b) else '⚠ NO'

# ===== Motor Pro =====
def escanear_pro(prov, marca, ruta_excel, paises, rank_maximo=30000, sup=None):
    filas, problematicos = leer_proveedor(prov, marca, ruta_excel)
    csvs={p:leer_csv_visualizador(ruta) for p,ruta in paises.items()}
    dom_base='ES' if 'ES' in csvs else list(csvs)[0]   # el rank filtra por el pais base
    doms=[d for d in ('ES','IT','FR') if d in csvs]

    registros=[]; no_encontrados=[]; sin_rank=[]; ambiguos=[]
    for f in filas:
        base=buscar(f, csvs[dom_base])
        if not base or not base['asin']:
            no_encontrados.append({'EAN':f['ean_in'],'Cabecera':f['nombre'],'Motivo':'Keepa sin datos/ASIN'})
            continue
        r_act, r_90 = base['rank'], base['rank90']
        tiene=(r_act and r_act>0) or (r_90 and r_90>0)
        pasa=any(r and r>0 and r<=rank_maximo for r in (r_act,r_90))
        if not pasa:
            if not tiene:
                sin_rank.append({'EAN':f['ean_in'],'ASIN':base['asin'],'Nombre':f['nombre'],
                                 'rank_act':r_act or -1,'rank90':r_90 or -1})
            continue   # rank>maximo -> se descarta en silencio (igual que el escaner)

        pa=f['pa']
        if prov=='TCG' and f['es_chase'] and pa: pa=pa/UNIDADES_CASE_TCG
        nvar=base.get('nvar') or 0
        if nvar and nvar>0:
            ambiguos.append({'EAN':f['ean_in'],'asin_elegido':base['asin']})

        paises_calc={}
        for dom in doms:
            c = base if dom==dom_base else buscar(f, csvs[dom])
            if not c or not c.get('asin'):
                paises_calc[dom]={'precio':None,'canal':'sin datos','ref_pct':None,'fee':None,
                                  'iva':IVA[dom],'rank_act':None,'rank90':None,'vendidos':None,
                                  'n_of':None,'decision':'Sin datos'}
                continue
            if c['buybox'] and c['buybox']>0:
                precio=c['buybox']; canal='BB-FBA' if c['es_fba'] else 'BB-FBM'
            elif c['nuevo'] and c['nuevo']>0:
                precio=c['nuevo']; canal='SIN BB'
            else:
                precio=None; canal='sin precio'
            reg={'precio':precio,'canal':canal,'ref_pct':c['compct'],'fee':c['fba'],'iva':IVA[dom],
                 'rank_act':c['rank'],'rank90':c['rank90'],'vendidos':c['vendidos'],'n_of':c['nof']}
            if precio and pa and c['compct'] is not None and c['fba'] is not None:
                rr=calc_rentabilidad(precio, pa, c['compct'], c['fba'], IVA[dom])
                reg.update({'beneficio':rr['beneficio'],'roi':rr['roi'],'margen':rr['margen'],
                            'decision':decision_de(rr['margen'])})
            else:
                reg.update({'beneficio':None,'roi':None,'margen':None,'decision':'Sin datos'})
            paises_calc[dom]=reg

        enbd=''
        if sup:
            s=sup.get(norm(f['core']))
            if s: enbd=f"OK Alm:{s.get('stock_moloka',0)} FBA:{s.get('stock_fba',0)}"
        registros.append({'nombre':f['nombre'],'ean':f['ean_in'],'asin':base['asin'],'marca':marca,
                          'pa':pa,'volumen':f['volumen'],'ambiguo':(nvar and nvar>0),'nvar':nvar,
                          'en_bd':enbd,'paises':paises_calc,
                          'titulo_amz':base.get('titulo',''),
                          'coincide':_coincide_titulo(f['nombre'], base.get('titulo','')),
                          'es_caja':f.get('es_caja',False),'uds_caja':f.get('uds_caja'),
                          'url':f.get('url','')})
    registros.sort(key=lambda x:(x['paises'].get(dom_base,{}).get('margen')
                                 if x['paises'].get(dom_base,{}).get('margen') is not None else -9e9), reverse=True)
    return dict(registros=registros, doms=doms, dom_base=dom_base,
                descartados=problematicos+no_encontrados, ambiguos=ambiguos, sin_rank=sin_rank,
                filas=filas, n_filas=len(filas))

# ===== Excel de resultados (CLAVADO al del escaner de tokens) =====
COLS=['Nombre','EAN','ASIN','Marca','PA (€)','País','Rank actual','Rank 90d','Vendidos/mes',
      'Precio venta (€)','Canal BB','Nº ofertas','% Comisión',
      'Com. Amazon (€)','Fee Logística (€)','Almacén (€)','Promo activa',
      'Beneficio (€)','ROI','Margen','Decisión','En mi BD','EAN ambiguo','Amazon (título)','Coincide','Compra','OcioStock']
COLS_LOTE=['Nombre','EAN','ASIN','Marca','País','Precio venta (€)','PA suelto (€)','PA lote (€)',
           'Ahorro/ud (€)','Beneficio lote (€)','Margen lote','Decisión lote','Uds. para ese precio']
def _cf_fill(h): return PatternFill(start_color=h, end_color=h, fill_type='solid')

def escribir_excel(res, ruta_salida):
    L={n:get_column_letter(i+1) for i,n in enumerate(COLS)}
    doms=res['doms']; nd=max(len(doms),1)
    wb=Workbook(); ws=wb.active; ws.title='Análisis'; ws.append(COLS)
    r=1
    for item in res['registros']:
        amb=f"VARIAC. ({item['nvar']})" if item['ambiguo'] else ''
        for dom in doms:
            d=item['paises'].get(dom) or {}
            r+=1
            pct=(d['ref_pct']/100*COM_DIGITALES) if d.get('ref_pct') is not None else None
            div=(1+d['iva']) if d.get('iva') else None
            ws.append([
                item['nombre'], item['ean'], item['asin'], item['marca'], item['pa'], dom,
                d.get('rank_act') if d.get('rank_act') and d['rank_act']>0 else None,
                d.get('rank90') if d.get('rank90') and d['rank90']>0 else None,
                d.get('vendidos'), d.get('precio'), d.get('canal'), d.get('n_of'), pct,
                f"={L['Precio venta (€)']}{r}*{L['% Comisión']}{r}" if pct is not None else None,
                d.get('fee'), ALMACEN, None,
                (f"=({L['Precio venta (€)']}{r}/{div})-{L['PA (€)']}{r}-{L['Com. Amazon (€)']}{r}"
                 f"-{L['Fee Logística (€)']}{r}-{L['Almacén (€)']}{r}") if (div and d.get('precio') and pct is not None) else None,
                f"={L['Beneficio (€)']}{r}/{L['PA (€)']}{r}" if (div and d.get('precio') and pct is not None and item['pa']) else None,
                f"={L['Beneficio (€)']}{r}/{L['Precio venta (€)']}{r}" if (div and d.get('precio') and pct is not None) else None,
                d.get('decision'), item.get('en_bd',''), amb,
                item.get('titulo_amz',''), item.get('coincide','?'),
                (f"CAJA x{item['uds_caja']}" if item.get('es_caja') else 'individual'),
                ('Ver ficha ↗' if item.get('url') else '')])
            cell=ws.cell(row=r, column=3)
            cell.hyperlink=f"https://www.{DOM_AMZ[dom]}/dp/{item['asin']}"
            cell.font=Font(color='0563C1', underline='single')
            if item.get('url'):
                cocel=ws.cell(row=r, column=len(COLS))   # ultima columna = OcioStock
                cocel.hyperlink=item['url']
                cocel.font=Font(color='0563C1', underline='single')
    last=ws.max_row
    def fmt(nm,code):
        for row in range(2,last+1): ws[f'{L[nm]}{row}'].number_format=code
    for nm in ['PA (€)','Precio venta (€)','Com. Amazon (€)','Fee Logística (€)','Almacén (€)','Beneficio (€)']:
        fmt(nm,'0.00')
    fmt('% Comisión','0.00%'); fmt('ROI','0.0%'); fmt('Margen','0.0%')
    for c in range(1,len(COLS)+1): ws.cell(row=1,column=c).font=Font(bold=True)
    for nm,w in {'Nombre':50,'EAN':14,'ASIN':12,'Marca':12,'En mi BD':20,'Decisión':15,'Amazon (título)':50,'Coincide':11,'Compra':14,'OcioStock':13}.items():
        ws.column_dimensions[L[nm]].width=w
    ws.freeze_panes='A2'
    if last>=2:
        tab=Table(displayName='T_Analisis', ref=f"A1:{get_column_letter(len(COLS))}{last}")
        tab.tableStyleInfo=TableStyleInfo(name='TableStyleMedium2', showRowStripes=False,
                                          showColumnStripes=False, showFirstColumn=False, showLastColumn=False)
        ws.add_table(tab)
        coi=L['Coincide']; rng_coi=f'{coi}2:{coi}{last}'
        ws.conditional_formatting.add(rng_coi, FormulaRule(formula=[f'ISNUMBER(SEARCH("NO",{coi}2))'],
                                      fill=_cf_fill('FFC7CE'), font=Font(color='9C0006')))
        dec=L['Decisión']; rng=f'{dec}2:{dec}{last}'
        for txt,fill,fnt in [('NO COMPRAR','FFC7CE','9C0006'),('VALORAR','FFEB9C','9C6500'),
                             ('COMPRAR','C6EFCE','006100'),('Sin datos','E7E6E6','808080')]:
            ws.conditional_formatting.add(rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("{txt}",{dec}2))'],
                fill=_cf_fill(fill), font=Font(color=fnt), stopIfTrue=True))
        ws.conditional_formatting.add(f"{L['Margen']}2:{L['Margen']}{last}",
            CellIsRule(operator='greaterThan', formula=['0.1'], font=Font(color='006100')))
        ws.conditional_formatting.add(f'A2:{get_column_letter(len(COLS))}{last}',
            FormulaRule(formula=[f'ISODD(INT((ROW()-2)/{nd}))'], fill=_cf_fill('D9D9D9')))

    def hoja(nombre, regs):
        w=wb.create_sheet(nombre)
        if regs:
            ks=list(regs[0].keys()); w.append(ks)
            for x in regs: w.append([x.get(k) for k in ks])
        else: w.append(['(vacio)'])
    hoja('Descartados', res['descartados'])
    hoja('Ambiguos', res['ambiguos'])
    hoja('Sin_rank', res['sin_rank'])

    # ----- Pestana "Precio por lote" (descuentos por volumen, p.ej. OcioStock) -----
    filas_lote=[]
    for item in res['registros']:
        vol=item.get('volumen')
        if not vol: continue
        pa_lote=vol['pa']; uds=vol['uds']; pa_suelto=item.get('pa')
        for dom in doms:
            d=item['paises'].get(dom)
            if not d or not d.get('precio') or d.get('ref_pct') is None or d.get('fee') is None: continue
            rr=calc_rentabilidad(d['precio'], pa_lote, d['ref_pct'], d['fee'], d['iva'])
            filas_lote.append([item['nombre'],item['ean'],item['asin'],item['marca'],dom,
                               round(d['precio'],2), round(pa_suelto,2) if pa_suelto else None,
                               round(pa_lote,2), round(pa_suelto-pa_lote,2) if pa_suelto else None,
                               round(rr['beneficio'],2), round(rr['margen'],4),
                               decision_de(rr['margen']), uds])
    filas_lote.sort(key=lambda x:(x[10] if x[10] is not None else -9), reverse=True)
    wl=wb.create_sheet('Precio por lote'); wl.append(COLS_LOTE)
    for fl in filas_lote: wl.append(fl)
    for c in range(1,len(COLS_LOTE)+1): wl.cell(row=1,column=c).font=Font(bold=True)
    if filas_lote:
        LL={n:get_column_letter(i+1) for i,n in enumerate(COLS_LOTE)}; lastL=wl.max_row
        for nm in ['Precio venta (€)','PA suelto (€)','PA lote (€)','Ahorro/ud (€)','Beneficio lote (€)']:
            for row in range(2,lastL+1): wl[f'{LL[nm]}{row}'].number_format='0.00'
        for row in range(2,lastL+1): wl[f'{LL["Margen lote"]}{row}'].number_format='0.0%'
        decL=LL['Decisión lote']; rngL=f'{decL}2:{decL}{lastL}'
        for txt,fill,fnt in [('NO COMPRAR','FFC7CE','9C0006'),('VALORAR','FFEB9C','9C6500'),('COMPRAR','C6EFCE','006100')]:
            wl.conditional_formatting.add(rngL, FormulaRule(formula=[f'ISNUMBER(SEARCH("{txt}",{decL}2))'],
                fill=_cf_fill(fill), font=Font(color=fnt), stopIfTrue=True))
        wl.column_dimensions[LL['Nombre']].width=50
        wl.column_dimensions[LL['Uds. para ese precio']].width=18
        wl.freeze_panes='A2'
    wb.save(ruta_salida)
    return dict(analisis_filas=last-1, lote_filas=len(filas_lote))


if __name__ == '__main__':
    import sys, json
    prov, marca, excel = sys.argv[1], sys.argv[2], sys.argv[3]
    paises = json.loads(sys.argv[4])
    salida = sys.argv[5] if len(sys.argv)>5 else 'Escaneo_PRO.xlsx'
    res = escanear_pro(prov, marca, excel, paises)
    info = escribir_excel(res, salida)
    print(f"Filas proveedor (stock>0): {res['n_filas']}")
    print(f"Análisis: {len(res['registros'])} productos ({info['analisis_filas']} filas) | "
          f"Descartados: {len(res['descartados'])} | Ambiguos(variac.): {len(res['ambiguos'])} | "
          f"Sin_rank: {len(res['sin_rank'])} | Precio por lote: {info['lote_filas']}")
    print(f"Excel -> {salida}")
