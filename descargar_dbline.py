#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# descargar_dbline.py  —  Descarga el catalogo de DBLine POR SERVIDOR (sin Chrome).
# DBLine (shop.dbline.it) es web a medida con buzones AJAX. Cazado con DevTools:
#   LOGIN:    POST /include/login_ajax.php           action=ESEGUI_LOGIN + codclifor + password(claro) + checkricorda + otp
#   DESCARGA: POST /include/Servizi/listini_ajax.php action=DOWNLOAD_CATALOGO_GENERALE + formato=xlsx
# Credenciales SOLO en Secrets (DBLINE_USER = codclifor, DBLINE_PASS). NUNCA en el codigo.
#
# Ejecutable suelto = PRUEBA (baja y verifica). El director importa descargar_catalogo_dbline().
import os, sys, io, re
from curl_cffi import requests as cr
import openpyxl

BASE = 'https://shop.dbline.it'
LOGIN_URL = f'{BASE}/include/login_ajax.php'
DOWNLOAD_URL = f'{BASE}/include/Servizi/listini_ajax.php'


# Mapa cabeceras INGLES -> ITALIANO. DBLine exporta por servidor en ingles y por navegador
# en italiano; el perfil del escaner usa los nombres italianos. Traducimos para que el
# escaner lea igual venga como venga (sin tocar el escaner).
_MAP_ITA = {
    'Genre': 'Genere', 'Price List ID': 'ID Listino', 'Image Link': 'Link immagine',
    'Code/Link': 'Codice/Link', 'Description': 'Descrizione', 'Notes': 'Note',
    'Release date': 'Data uscita', 'Available': 'Disponibili', 'List Price (\u20ac)': 'Listino (\u20ac)',
    'Discount 1 (%)': 'Sconto 1 (%)', 'Discount 2 (%)': 'Sconto 2 (%)', 'Price (\u20ac)': 'Prezzo (\u20ac)',
    'VAT (%)': 'Iva (%)', 'Promo Expiration': 'Scadenza promo', 'Promo Price (\u20ac)': 'Prezzo promo (\u20ac)',
    'Weight (gr)': 'Peso (gr)',
}


def _normalizar_cabeceras(cont):
    import openpyxl, io as _io
    wb = openpyxl.load_workbook(_io.BytesIO(cont))
    ws = wb[wb.sheetnames[0]]
    hdr = None
    for r in range(1, 7):
        vals = [str(ws.cell(row=r, column=c).value or '').strip() for c in range(1, ws.max_column + 1)]
        if 'Publisher' in vals and 'EAN' in vals:
            hdr = r; break
    if hdr is None:
        print('   AVISO: no encuentro la fila de cabecera; dejo el Excel tal cual.', flush=True)
        return cont
    cambiadas = 0
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=hdr, column=c)
        v = str(cell.value or '').strip()
        if v in _MAP_ITA:
            cell.value = _MAP_ITA[v]; cambiadas += 1
    if cambiadas:
        print(f'   Cabeceras traducidas ingles->italiano: {cambiadas} columnas (fila {hdr}).', flush=True)
        out = _io.BytesIO(); wb.save(out); return out.getvalue()
    print('   Cabeceras ya en italiano; no toco nada.', flush=True)
    return cont


def descargar_catalogo_dbline():
    USER, PASS = os.environ.get('DBLINE_USER'), os.environ.get('DBLINE_PASS')
    if not USER or not PASS:
        raise RuntimeError('Faltan los secrets DBLINE_USER / DBLINE_PASS.')
    # verify=False: DBLine tiene la cadena SSL mal montada (le falta el intermedio) y el
    # runner no puede validar su certificado (curl error 60). La conexion sigue cifrada por
    # HTTPS; solo nos saltamos la validacion de la cadena, que es lo que su servidor rompe.
    s = cr.Session(impersonate='chrome120', verify=False)
    ajax = {'X-Requested-With': 'XMLHttpRequest', 'Origin': BASE, 'Referer': BASE + '/'}

    # 0) Home -> cookies de sesion iniciales
    print('>>> Abriendo la home de DBLine (cookies)...', flush=True)
    r0 = s.get(BASE + '/', timeout=60)
    print(f'   home -> {r0.status_code} ({len(r0.text)} bytes)', flush=True)

    # 1) Login (clave en claro, POST AJAX)
    print('>>> Enviando login...', flush=True)
    r1 = s.post(LOGIN_URL, data={'action': 'ESEGUI_LOGIN', 'codclifor': USER,
                                 'password': PASS, 'checkricorda': 'N', 'otp': ''},
                headers=ajax, timeout=60)
    print(f'   login -> {r1.status_code} | respuesta: {r1.text[:300]}', flush=True)
    low = r1.text.lower()
    if ('errata' in low) or ('non valido' in low) or ('errore' in low and 'ok' not in low):
        print('   AVISO: el login parece NO haber entrado (revisa la respuesta de arriba).', flush=True)

    # 2) Descargar catalogo (POST AJAX). A veces el AJAX responde con la URL del fichero.
    print('>>> Descargando catalogo (DOWNLOAD_CATALOGO_GENERALE)...', flush=True)
    r2 = s.post(DOWNLOAD_URL, data={'action': 'DOWNLOAD_CATALOGO_GENERALE', 'formato': 'xlsx'},
                headers=ajax, timeout=300)
    cont = r2.content
    print(f'   descarga -> {r2.status_code} | {len(cont)} bytes | content-type: {r2.headers.get("content-type","")}', flush=True)

    if cont[:2] != b'PK':
        txt = cont[:600].decode('utf-8', 'replace')
        print('   No es un .xlsx directo. Respuesta (primeros 600):', txt, flush=True)
        m = re.search(r'(https?://[^\s"\'<>]+\.xlsx[^\s"\'<>]*)', txt) or re.search(r'([\w./\-]+\.xlsx)', txt)
        if m:
            url = m.group(1)
            if url.startswith('/'): url = BASE + url
            elif not url.startswith('http'): url = BASE + '/' + url
            print('   Intento bajar el fichero enlazado:', url, flush=True)
            r3 = s.get(url, headers={'Referer': BASE + '/'}, timeout=300)
            cont = r3.content
            print(f'   fichero -> {r3.status_code} | {len(cont)} bytes', flush=True)

    if cont[:2] != b'PK':
        raise RuntimeError('La descarga NO es un .xlsx (login caducado, cabecera distinta o el action devuelve otra cosa). Mira el log de arriba.')

    wb = openpyxl.load_workbook(io.BytesIO(cont), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    print(f'   Excel VALIDO: hoja "{wb.sheetnames[0]}", ~{ws.max_row} filas | hojas: {wb.sheetnames}', flush=True)
    cont = _normalizar_cabeceras(cont)   # ingles -> italiano para que el escaner lo lea
    return cont


if __name__ == '__main__':
    cont = descargar_catalogo_dbline()
    with open('dblinecatalog.xlsx', 'wb') as f:
        f.write(cont)
    print('>>> PRUEBA OK: DBLine se baja por servidor, sin Chrome. 🎉', flush=True)
